# -*- coding: utf-8 -*-
"""
B-3 옵션 A 실행 — 행안부 단어 3건 DELETE + 영향 용어 108건 DELETE + RAMP 단어 3건 N 등록.

작업:
  1. 백업 xlsx 생성 (Phase1_행안부_변경_관리_2026-05-23.xlsx)
  2. SQL 파일 생성 (검토용)
  3. 사용자 검토 후 별도 단계에서 실행
"""
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from pathlib import Path
import subprocess, csv, io, sys

BASE = Path(r"C:\Users\장재영\Desktop\dataQ\CAMS표준화")
RAMP_DICT = BASE / "01_원본자료" / "RMS4-DE14-04.데이터표준화_단어사전-V2.0.xlsx"
OUT_XLSX = BASE / "04_RAMP분석_2026-05-21" / "Phase1_행안부_변경_관리_2026-05-23.xlsx"
OUT_SQL = BASE / "04_RAMP분석_2026-05-21" / "Phase1_B3_apply.sql"

TARGETS = ("SEQ", "ELCT", "SQL")

# ============ RAMP 단어 메타 추출 ============
ramp_meta = {}  # abrv -> {nm, eng, desc, frmt, dmn}
wb = load_workbook(RAMP_DICT, read_only=True, data_only=True)
ws = wb["단어사전"]
for r in ws.iter_rows(min_row=2, values_only=True):
    if r and r[2] and r[3]:
        nm = str(r[2]).strip(); abrv = str(r[3]).strip()
        if nm in ("순번","전자","질의"):
            ramp_meta[abrv] = {
                "nm": nm, "abrv": abrv,
                "eng": str(r[4] or "").strip().replace("_x000D_",""),
                "desc": str(r[5] or "").strip().replace("_x000D_",""),
                "frmt": str(r[6] or "").strip(),
                "dmn": str(r[7] or "").strip(),
            }
wb.close()
print("RAMP 단어 메타:")
for k,v in ramp_meta.items():
    print(f"  {k}: {v}")

# ============ DB에서 삭제 대상 추출 ============
def psql(sql):
    r = subprocess.run(["docker","exec","-i","dataq-db","psql","-U","admin","-d","postgres",
                        "-c", f"COPY ({sql}) TO STDOUT WITH (FORMAT csv, FORCE_QUOTE *)"],
                       capture_output=True, encoding="utf-8")
    if r.returncode != 0: print("ERR:",r.stderr); sys.exit(1)
    return list(csv.reader(io.StringIO(r.stdout)))

# 행안부 단어 3건
del_words = psql(f"""
SELECT word_id, word_nm, word_eng_abrv_nm, word_eng_nm, word_desc,
       coalesce(word_clsf_yn,''), coalesce(domain_clsf_nm,''),
       coalesce(comm_stnd_yn,''), coalesce(aprv_yn,''), coalesce(cret_dt,'')
FROM quality.tb_word
WHERE comm_stnd_yn='Y' AND word_eng_abrv_nm IN ('SEQ','ELCT','SQL')
ORDER BY word_eng_abrv_nm
""")
print(f"\n행안부 단어 삭제 대상: {len(del_words)}")

# 행안부 용어 — 매핑 기반 추출 (단어 ID 사용분)
del_terms = psql("""
SELECT DISTINCT t.terms_id, t.terms_nm, t.terms_eng_abrv_nm, t.terms_desc,
       coalesce(t.domain_nm,''), coalesce(t.code_grp,''), coalesce(t.chrg_org,''),
       w.word_eng_abrv_nm AS via_word
FROM quality.tb_terms t
JOIN quality.tb_terms_words tw ON t.terms_id = tw.terms_id
JOIN quality.tb_word w ON (tw.word_id, tw.word_nm) = (w.word_id, w.word_nm)
WHERE t.comm_stnd_yn='Y' AND w.comm_stnd_yn='Y' AND w.word_eng_abrv_nm IN ('SEQ','ELCT','SQL')
ORDER BY w.word_eng_abrv_nm, t.terms_nm
""")
print(f"행안부 용어 삭제 대상: {len(del_terms)}")

# 매핑 건수
map_cnt = psql("""
SELECT count(*)::text FROM quality.tb_terms_words tw
JOIN quality.tb_word w ON (tw.word_id, tw.word_nm) = (w.word_id, w.word_nm)
WHERE w.comm_stnd_yn='Y' AND w.word_eng_abrv_nm IN ('SEQ','ELCT','SQL')
""")
n_map = int(map_cnt[0][0]) if map_cnt else 0
print(f"매핑 (CASCADE 자동 삭제): {n_map}")

# ============ xlsx 백업 ============
print("\n=== xlsx 백업 생성 ===")
wb = Workbook()
THIN = Side(border_style="thin", color="808080")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
HDR_FILL = PatternFill("solid", fgColor="C00000")  # 빨간 — 위험 작업 강조
HDR_FONT = Font(name="맑은 고딕", size=10, color="FFFFFF", bold=True)
CELL_FONT = Font(name="맑은 고딕", size=10)
SECTION_FILL = PatternFill("solid", fgColor="FFE6E6")

# 시트 1: 표지
ws = wb.active; ws.title = "표지"
ws.column_dimensions["A"].width = 24; ws.column_dimensions["B"].width = 100
ws.merge_cells("A1:B1")
t = ws.cell(row=1, column=1, value="Phase 1 — 행안부 표준 변경 관리 (B-3 예외 처리)")
t.font = Font(name="맑은 고딕", size=18, bold=True, color="C00000")
t.alignment = Alignment(horizontal="center", vertical="center")
ws.row_dimensions[1].height = 36

def cr(r, label, value):
    ws.cell(row=r, column=1, value=label).fill = SECTION_FILL
    ws.cell(row=r, column=1).font = Font(name="맑은 고딕", size=10, bold=True)
    ws.cell(row=r, column=1).border = BORDER
    c = ws.cell(row=r, column=2, value=value)
    c.font = Font(name="맑은 고딕", size=11); c.border = BORDER
    c.alignment = Alignment(vertical="center", wrap_text=True)

cr(3, "작성일", "2026-05-23")
cr(4, "작성자", "장재영")
cr(5, "작업 분류", "Phase 1 B-3 예외 처리 — 행안부 단어 의미와 RAMP 의미 완전 충돌 3건")
cr(6, "결정", "옵션 A 채택 — 행안부 단어/용어 DELETE, RAMP 단어를 같은 영문약어로 N 등록")
cr(7, "사유", "RAMP에서 SEQ/ELCT/SQL을 다른 의미로 사용 중. KEY 컬럼 변경 위험 회피 위해 단어 사전 수준에서 처리. R1 일부 훼손 감수")
cr(8, "정책 위반", "R1 (행안부 표준사전 불변) — 본 케이스에 한해 예외")
ws.row_dimensions[9].height = 8
cr(10, "영향 1 — tb_word DELETE", f"{len(del_words)}건 (순서/SEQ, 전자계약/ELCT, SQL/SQL)")
cr(11, "영향 2 — tb_terms DELETE", f"{len(del_terms)}건 (SEQ 104 + ELCT 2 + SQL 2)")
cr(12, "영향 3 — tb_terms_words 자동 CASCADE", f"{n_map}건 (용어 DELETE 시 자동)")
cr(13, "영향 4 — RAMP 신규 등록", "tb_word 3건 N 등록 (순번/SEQ, 전자/ELCT, 질의/SQL)")
ws.row_dimensions[14].height = 8
cr(15, "RAMP 컬럼 영향", "변경 0 — KEY 안전 회피 위해 컬럼명/한글 유지")
cr(16, "향후 행안부 갱신 시", "행안부 9차 배포본 등에서 SEQ/ELCT/SQL 단어가 다시 들어오면 본 결정 재검토 필요")
cr(17, "RAMP 의미 사용 인지", "dataQ DB 내에서 SEQ='순번', ELCT='전자', SQL='질의' 로 해석. 행안부 의미(순서/전자계약/SQL)는 본 시스템에서 사용 안 함")

# 시트 2: 삭제된 행안부 단어
ws2 = wb.create_sheet("행안부단어_삭제")
H = ["No","단어명","영문약어","영문명","설명","형식단어여부","도메인분류","comm_stnd_yn","aprv_yn","cret_dt","RAMP 대체 단어"]
W = [4,14,10,22,55,8,12,8,8,16,18]
for i, w in enumerate(W, 1): ws2.column_dimensions[get_column_letter(i)].width = w
for i, h in enumerate(H, 1):
    c = ws2.cell(row=1, column=i, value=h); c.fill = HDR_FILL; c.font = HDR_FONT
    c.alignment = Alignment(horizontal="center"); c.border = BORDER
ws2.freeze_panes = "A2"
for i, r in enumerate(del_words, 1):
    abrv = r[2]
    ramp_alt = ramp_meta.get(abrv,{}).get("nm","")
    values = [i, r[1], r[2], r[3], r[4], r[5], r[6], r[7], r[8], r[9], ramp_alt]
    for j, v in enumerate(values, 1):
        c = ws2.cell(row=i+1, column=j, value=v); c.font = CELL_FONT; c.border = BORDER
        c.alignment = Alignment(vertical="top", wrap_text=True)

# 시트 3: 삭제된 행안부 용어
ws3 = wb.create_sheet("행안부용어_삭제")
H = ["No","용어명","영문약어","설명","도메인","코드그룹","소관기관","구성 단어 (via)"]
W = [4,22,28,55,14,14,14,12]
for i, w in enumerate(W, 1): ws3.column_dimensions[get_column_letter(i)].width = w
for i, h in enumerate(H, 1):
    c = ws3.cell(row=1, column=i, value=h); c.fill = HDR_FILL; c.font = HDR_FONT
    c.alignment = Alignment(horizontal="center"); c.border = BORDER
ws3.freeze_panes = "A2"
for i, r in enumerate(del_terms, 1):
    values = [i, r[1], r[2], r[3], r[4], r[5], r[6], r[7]]
    for j, v in enumerate(values, 1):
        c = ws3.cell(row=i+1, column=j, value=v); c.font = CELL_FONT; c.border = BORDER
        c.alignment = Alignment(vertical="top", wrap_text=True)

# 시트 4: 신규 등록 RAMP 단어
ws4 = wb.create_sheet("RAMP신규등록")
H = ["No","RAMP 한글","영문약어","영문명","설명","형식단어여부","도메인분류","comm_stnd_yn","비고"]
W = [4,14,10,22,55,8,12,8,30]
for i, w in enumerate(W, 1): ws4.column_dimensions[get_column_letter(i)].width = w
for i, h in enumerate(H, 1):
    c = ws4.cell(row=1, column=i, value=h); c.fill = HDR_FILL; c.font = HDR_FONT
    c.alignment = Alignment(horizontal="center"); c.border = BORDER
ws4.freeze_panes = "A2"
for i, abrv in enumerate(TARGETS, 1):
    m = ramp_meta.get(abrv, {})
    values = [i, m.get("nm",""), abrv, m.get("eng",""), m.get("desc",""),
              m.get("frmt",""), m.get("dmn",""), "N",
              "행안부 삭제분과 같은 영문약어. RAMP 의미 우선 적용"]
    for j, v in enumerate(values, 1):
        c = ws4.cell(row=i+1, column=j, value=v); c.font = CELL_FONT; c.border = BORDER
        c.alignment = Alignment(vertical="top", wrap_text=True)

wb.save(OUT_XLSX)
print(f"\n→ 백업 xlsx: {OUT_XLSX}")

# ============ SQL 파일 ============
sql_lines = [
    "-- Phase 1 B-3 옵션 A 실행 SQL",
    "-- 작성: 2026-05-23",
    "-- 1) 행안부 용어 108건 DELETE (CASCADE로 tb_terms_words 자동 삭제)",
    "-- 2) 행안부 단어 3건 DELETE",
    "-- 3) RAMP 신규 단어 3건 INSERT (comm_stnd_yn='N')",
    "",
    "BEGIN;",
    "",
    "-- 1) 행안부 용어 삭제 (SEQ/ELCT/SQL 단어 사용분)",
    "DELETE FROM quality.tb_terms",
    "WHERE comm_stnd_yn='Y'",
    "  AND terms_id IN (",
    "    SELECT DISTINCT tw.terms_id",
    "    FROM quality.tb_terms_words tw",
    "    JOIN quality.tb_word w ON (tw.word_id, tw.word_nm) = (w.word_id, w.word_nm)",
    "    WHERE w.comm_stnd_yn='Y' AND w.word_eng_abrv_nm IN ('SEQ','ELCT','SQL')",
    "  );",
    "",
    "-- 2) 행안부 단어 삭제",
    "DELETE FROM quality.tb_word",
    "WHERE comm_stnd_yn='Y' AND word_eng_abrv_nm IN ('SEQ','ELCT','SQL');",
    "",
    "-- 3) RAMP 신규 단어 등록",
]
import os, base64
for abrv in TARGETS:
    m = ramp_meta[abrv]
    nm = m["nm"].replace("'","''")
    eng = m["eng"].replace("'","''")
    desc = (m["desc"] or m["nm"]).replace("'","''")
    frmt = m["frmt"] or "N"
    dmn = m["dmn"].replace("'","''")
    word_id = base64.urlsafe_b64encode(os.urandom(17)).rstrip(b'=').decode()[:22]
    sql_lines.append(f"""INSERT INTO quality.tb_word (
  word_id, word_nm, word_eng_abrv_nm, word_eng_nm, word_desc,
  word_clsf_yn, domain_clsf_nm, comm_stnd_yn, aprv_yn,
  cret_dt, cret_user_id, use_yn
) VALUES (
  '{word_id}', '{nm}', '{abrv}', '{eng}', '{desc}',
  '{frmt}', '{dmn}', 'N', 'Y',
  to_char(now(),'YYYYMMDDHH24MISS'), 'admin', 'Y'
);""")

sql_lines.append("")
sql_lines.append("-- 검증")
sql_lines.append("SELECT word_nm, word_eng_abrv_nm, comm_stnd_yn FROM quality.tb_word WHERE word_eng_abrv_nm IN ('SEQ','ELCT','SQL');")
sql_lines.append("SELECT count(*) FROM quality.tb_terms WHERE comm_stnd_yn='Y';")
sql_lines.append("SELECT count(*) FROM quality.tb_word WHERE comm_stnd_yn='Y';")
sql_lines.append("")
sql_lines.append("-- 문제 없으면 COMMIT, 아니면 ROLLBACK")
sql_lines.append("-- COMMIT;")
sql_lines.append("-- ROLLBACK;")

OUT_SQL.write_text("\n".join(sql_lines), encoding="utf-8")
print(f"→ SQL: {OUT_SQL}")

print("\n=== 요약 ===")
print(f"행안부 단어 DELETE: {len(del_words)}")
print(f"행안부 용어 DELETE: {len(del_terms)} (매핑 {n_map}건 CASCADE)")
print(f"RAMP 단어 INSERT:   {len(TARGETS)}")
