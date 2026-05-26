# -*- coding: utf-8 -*-
"""
보정본 BEFORE → 깨끗한 AFTER (컬럼 양식 v3) + 매핑정의서 + 용어사전
처음부터 모든 룰 일괄 적용.

룰:
  1. 보정본 cols 로드 (5,774)
  2. Phase 2-1 단어변환 (영문/한글)
  3. D3 시간 룰 (한글)
  4. D5 R8 보충 (한글)
  5. 형식단어 545종 영문 정합 + D5 fallback (한글 미종결시)
  6. CHAR→VARCHAR + UPPER
  7. ord 1~N 부여 (테이블별)
"""
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from pathlib import Path
from collections import defaultdict, Counter
import subprocess, csv, io

BASE = Path(r"C:\Users\장재영\Desktop\dataQ\CAMS표준화")
ROOT = BASE / "04_RAMP분석_2026-05-21" / "00_핵심산출"
SRC = BASE / "CAMS_RAMP_통합" / "ramp기관스키마정보_보정.xlsx"
COL_OUT = ROOT / "RAMP_업로드_컬럼_2026-05-26_v3.xlsx"
TERMS_OUT = ROOT / "RAMP_업로드_용어_2026-05-26.xlsx"
MAP_OUT = ROOT / "RAMP_컬럼_매핑정의서_2026-05-26.xlsx"
MOIS_SEED = BASE.parent / "q-center" / "src" / "main" / "resources" / "seed" / "행안부_공통표준"
MOIS_WORDS = MOIS_SEED / "행정안전부_공공데이터 공통표준단어.xlsx"
MOIS_TERMS_UPLOAD = [MOIS_SEED / "용어사전_일괄등록_1.xlsx", MOIS_SEED / "용어사전_일괄등록_2.xlsx"]

EXTRA_FW = {"구분":"SE", "값":"VL", "기한":None, "빈도":None, "주기":None,
            "횟수":"NMTM", "등급":"GRD", "단계":None, "순번":"SEQ",
            "총수":"TCNT"}

def s(v): return "" if v is None else str(v).strip()

# ============ 0. 행안부 형식단어 + 단어 사전 ============
print("=== 0. 행안부 단어 로드 ===")
wb = load_workbook(MOIS_WORDS, read_only=True, data_only=True)
ws = wb.active
all_words = {}; fw_y = {}
for r in ws.iter_rows(min_row=2, values_only=True):
    if r and r[0] and r[1]:
        kr, en = s(r[0]), s(r[1])
        all_words[kr] = en
        if r[4] and s(r[4]) == "Y":
            fw_y[kr] = en
wb.close()
for kr, en in EXTRA_FW.items():
    if kr not in fw_y:
        fw_y[kr] = en if en else all_words.get(kr, "")
fw_sorted = sorted([k for k in fw_y if fw_y[k]], key=len, reverse=True)
print(f"  형식단어 (545 + 추가): {len(fw_sorted)}")

def match_fw(kr):
    for f in fw_sorted:
        if kr.endswith(f):
            return (f, fw_y[f])
    return (None, None)

# ============ 1. 보정본 BEFORE 로드 ============
print("\n=== 1. 보정본 BEFORE 로드 ===")
wb = load_workbook(SRC, read_only=True, data_only=True)
ws_t = wb["테이블 목록"]
tables = {}
for r in ws_t.iter_rows(min_row=2, values_only=True):
    if r and r[0]:
        tables[s(r[0])] = s(r[1]) if len(r) > 1 and r[1] else ""
ws_c = wb["컬럼"]
cols = []
seen = set()
for r in ws_c.iter_rows(min_row=2, values_only=True):
    if not r or not r[0] or not r[1]: continue
    key = (s(r[0]), s(r[1]))
    if key in seen: continue
    seen.add(key)
    cols.append({
        "tbl": key[0], "bef_en": key[1], "bef_kr": s(r[2]),
        "bef_dtype": s(r[5]).upper() if r[5] else "",
        "bef_dlen": s(r[6]),
        "bef_null": s(r[4]),
        "bef_pk": s(r[7]),
        "bef_ord": s(r[10]),
        "desc": s(r[3]),
    })
wb.close()
print(f"  보정본 cols: {len(cols)}  (테이블 {len(tables)})")

# 컬럼시트엔 있는데 테이블목록 없는 케이스 보강
for c in cols:
    if c["tbl"] not in tables:
        tables[c["tbl"]] = ""

# ============ 2. Phase 2-1 / D3 / D5 룰 로드 ============
print("\n=== 2. 변환 룰 로드 ===")
phase21 = {}
wb = load_workbook(ROOT/"RAMP_컬럼_BEFORE_AFTER_2026-05-23.xlsx", read_only=True, data_only=True)
for r in wb["컬럼BEFORE_AFTER"].iter_rows(min_row=2, values_only=True):
    if r and r[1] and r[2]:
        phase21[(s(r[1]), s(r[2]))] = (s(r[3] or r[2]), s(r[5] or r[4] or ""))
wb.close()
time_rule = {}
wb = load_workbook(ROOT/"Phase2_시간컬럼_BEFORE_AFTER_2026-05-23.xlsx", read_only=True, data_only=True)
for r in wb["적용_BEFORE_AFTER"].iter_rows(min_row=2, values_only=True):
    if r and r[1] and r[2]:
        time_rule[(s(r[1]), s(r[2]))] = s(r[4] or r[3] or "")
wb.close()
d5_kr = {}
wb = load_workbook(ROOT/"Phase2_D5_R8형식단어_2026-05-23.xlsx", read_only=True, data_only=True)
for r in wb["새용어_BEFORE_AFTER"].iter_rows(min_row=2, values_only=True):
    if r and r[1] and r[3]: d5_kr[s(r[1])] = s(r[3])
wb.close()
print(f"  Phase2-1: {len(phase21)}, D3: {len(time_rule)}, D5: {len(d5_kr)}")

# ============ 3. AFTER 적용 ============
print("\n=== 3. AFTER 적용 ===")
char_cnt = 0
en_patched = 0
fb_cnt = 0
over30 = []
for c in cols:
    key = (c["tbl"], c["bef_en"])
    # 단어변환
    aft_en, aft_kr = phase21.get(key, (c["bef_en"], c["bef_kr"]))
    if not aft_kr: aft_kr = c["bef_kr"]
    # D3 시간
    if key in time_rule:
        tk = time_rule[key]
        if tk: aft_kr = tk
    # D5 한글+영문 동시 보충 (형식단어 6종: CN/SE/NM/VL/YN/CD)
    D5_FW_EN = {"내용":"CN", "구분":"SE", "명":"NM", "값":"VL", "여부":"YN", "코드":"CD"}
    if aft_kr in d5_kr:
        new_kr = d5_kr[aft_kr]
        diff = new_kr[len(aft_kr):] if new_kr.startswith(aft_kr) else ""
        if diff in D5_FW_EN:
            sfx = "_" + D5_FW_EN[diff]
            if not aft_en.upper().endswith(sfx):
                aft_en = aft_en + sfx
        aft_kr = new_kr
    aft_kr = aft_kr.replace(" ", "")
    aft_en_up = aft_en.upper()

    # 형식단어 매칭: 한글이 형식단어로 끝나면 영문 그대로 (추가 X)
    # R8 미종결만 D5 fallback (한글+영문 동시 보충)
    fw_kr, fw_en = match_fw(aft_kr)
    if fw_kr:
        pass  # 이미 형식단어로 끝남 → 그대로
    else:
        # R8 미종결: D5 fallback (한글+영문 동시 보충)
        dt_u = c["bef_dtype"].upper()
        if dt_u in ("NUMERIC","INTEGER","DECIMAL"):
            fb_k, fb_e = "값", "VL"
        elif dt_u == "DATETIME":
            fb_k, fb_e = "일시", "TS"
        elif dt_u == "DATE":
            fb_k, fb_e = "일자", "DT"
        else:
            fb_k, fb_e = "내용", "CN"
        aft_kr = aft_kr + fb_k
        sfx = "_" + fb_e
        if not aft_en_up.endswith(sfx):
            aft_en_up = aft_en_up + sfx
        fb_cnt += 1

    if len(aft_en_up) > 30:
        over30.append((c["tbl"], aft_en_up, len(aft_en_up), aft_kr))

    # CHAR→VARCHAR + Cubrid→Oracle
    dt_u = c["bef_dtype"].upper()
    if dt_u == "CHAR":
        aft_dtype = "VARCHAR"; char_cnt += 1
    elif dt_u in ("STRING","VARCHAR2"):
        aft_dtype = "VARCHAR" if (c["bef_dlen"].isdigit() and int(c["bef_dlen"]) <= 4000) else "CLOB"
    elif dt_u in ("NUMERIC","INTEGER","DECIMAL"):
        aft_dtype = "NUMBER"
    elif dt_u == "DATETIME":
        aft_dtype = "TIMESTAMP"
    else:
        aft_dtype = dt_u
    aft_dlen = c["bef_dlen"] if (c["bef_dlen"].isdigit() and aft_dtype != "CLOB") else "0"

    c["aft_en"] = aft_en_up
    c["aft_kr"] = aft_kr
    c["aft_dtype"] = aft_dtype
    c["aft_dlen"] = aft_dlen
    c["null"] = "N" if c["bef_null"].upper() in ("N","NO","NOT NULL") else "Y"
    c["pk"] = "Y" if c["bef_pk"].upper() in ("Y","PK","TRUE","1") else "N"

# ord 1~N 부여
ord_by = defaultdict(int)
for c in cols:
    ord_by[c["tbl"]] += 1
    c["ord"] = ord_by[c["tbl"]]

print(f"  CHAR→VARCHAR: {char_cnt}, 영문 정합 patched: {en_patched}, D5 fallback: {fb_cnt}")
print(f"  30자 초과: {len(over30)}")

# ============ 4. AFTER 컬럼 양식 v3 ============
print(f"\n=== 4. AFTER 컬럼 양식 v3 ===")
wb = Workbook(); ws = wb.active; ws.title = "컬럼"
H = ["소유자","테이블명(영문)","테이블명(한글)","컬럼명(영문)","컬럼명(한글)",
     "데이터타입","길이","소수점자리","컬럼 순서","NULL여부","PK여부","FK여부",
     "디폴트값","참조 테이블(한글)","참조 컬럼(한글)","삭제 규칙"]
W = [12, 28, 30, 24, 28, 12, 8, 8, 8, 8, 6, 6, 12, 24, 22, 10]
THIN = Side(border_style="thin", color="808080")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
HDR_FILL = PatternFill("solid", fgColor="1F4E78")
HDR_FONT = Font(name="맑은 고딕", size=10, color="FFFFFF", bold=True)
CELL_FONT = Font(name="맑은 고딕", size=10)
for i, w in enumerate(W, 1): ws.column_dimensions[get_column_letter(i)].width = w
for i, h in enumerate(H, 1):
    cell = ws.cell(row=1, column=i, value=h)
    cell.fill = HDR_FILL; cell.font = HDR_FONT
    cell.alignment = Alignment(horizontal="center"); cell.border = BORDER
ws.freeze_panes = "C2"
for i, c in enumerate(cols, 2):
    vals = ["RAMP", c["tbl"], tables.get(c["tbl"], ""),
            c["aft_en"], c["aft_kr"], c["aft_dtype"], c["aft_dlen"], "0",
            c["ord"], c["null"], c["pk"], "N", "", "", "", ""]
    for j, v in enumerate(vals, 1):
        cell = ws.cell(row=i, column=j, value=v)
        cell.font = CELL_FONT; cell.border = BORDER
        cell.alignment = Alignment(vertical="top", wrap_text=True)
wb.save(COL_OUT)
print(f"  → {COL_OUT}  ({len(cols)}행)")

# ============ 4.5. DB tb_word 로드 (영문약어→한글 매핑) ============
print(f"\n=== 4.5 tb_word 영문→한글 매핑 로드 ===")
cmd = ["docker","exec","-i","dataq-db","psql","-U","admin","-d","postgres",
       "-c","COPY (SELECT word_eng_abrv_nm, word_nm FROM quality.tb_word) TO STDOUT WITH (FORMAT csv)"]
r = subprocess.run(cmd, capture_output=True, encoding="utf-8")
en2kr = {}   # 영문약어 UPPER → 한글
for row in csv.reader(io.StringIO(r.stdout)):
    if row and row[0] and row[1]:
        en2kr.setdefault(row[0].upper(), row[1])
print(f"  tb_word 매핑: {len(en2kr)}")

def en_to_kr(en):
    """AFTER 영문 → '_' split → 토큰별 한글 조합"""
    parts = en.split("_")
    out = []
    unknown = []
    for p in parts:
        if p in en2kr:
            out.append(en2kr[p])
        else:
            out.append(f"?{p}?")
            unknown.append(p)
    return "".join(out), unknown

def map_domain(kr, dt, dl):
    dt_u = dt.upper()
    if dt_u == "VARCHAR":
        if kr.endswith("여부"): return f"여부V{dl}"
        if kr.endswith("코드"): return f"코드V{dl}"
        if kr.endswith("구분"): return f"구분V{dl}"
        if kr.endswith("명"):   return f"명V{dl}"
        if kr.endswith("값"):   return f"값V{dl}"
        if kr.endswith("내용"): return f"내용V{dl}"
        if kr.endswith("일자") and dl == "8":  return "일자V8"
        if kr.endswith("일시") and dl == "14": return "일시V14"
        if kr.endswith("연도"): return "연도V4"
        return f"내용V{dl}" if dl else "내용V"
    if dt_u == "NUMBER":
        if kr.endswith("값"): return f"값N{dl}"
        if kr.endswith("수"): return f"수N{dl}"
        if kr.endswith("순번"): return f"순번N{dl}"
        return f"값N{dl}"
    if dt_u == "DATE":     return "일자DT"
    if dt_u == "TIMESTAMP": return "일시TS"
    if dt_u == "CLOB":     return "내용L"
    return f"{dt_u}{dl}"

# ============ 5. 매핑정의서 ============
print(f"\n=== 5. 매핑정의서 ===")
wb = Workbook()
ws = wb.active; ws.title = "표지"
ws.column_dimensions["A"].width = 24; ws.column_dimensions["B"].width = 100
ws.merge_cells("A1:B1")
t = ws.cell(row=1, column=1, value="RAMP 컬럼 BEFORE/AFTER 매핑정의서 (보정본 기준)")
t.font = Font(name="맑은 고딕", size=18, bold=True, color="1F4E78")
t.alignment = Alignment(horizontal="center", vertical="center")
ws.row_dimensions[1].height = 36
SECTION_FILL = PatternFill("solid", fgColor="D9E1F2")
def cr(r, label, value):
    c1 = ws.cell(row=r, column=1, value=label); c1.fill = SECTION_FILL
    c1.font = Font(name="맑은 고딕", size=10, bold=True); c1.border = BORDER
    c2 = ws.cell(row=r, column=2, value=value); c2.font = Font(name="맑은 고딕", size=11)
    c2.border = BORDER; c2.alignment = Alignment(vertical="center", wrap_text=True)
cr(3,"작성일","2026-05-26")
cr(4,"대상",f"RAMP 컬럼 {len(cols)}건 (보정본 기준)")
cr(5,"BEFORE","ramp기관스키마정보_보정.xlsx (영문 2건 보정 + 중복 3건 정리)")
cr(6,"AFTER",f"RAMP_업로드_컬럼_2026-05-26_v3.xlsx — 깨끗한 5,774 컬럼")
cr(7,"적용 룰","Phase2-1 단어 / D3 시간 / D5 한글 보충 / 형식단어 545 정합 / CHAR→VARCHAR / 대문자")
cr(8,"매칭 키","(테이블, ord) — 보정본 ord = AFTER ord")
cr(9,"30자 초과",f"{len(over30)}건")

# 변경 유형 + 검증 카운터
type_cnt = Counter()
kr_match_cnt = 0; kr_mismatch_cnt = 0
fw_end_ok_cnt = 0; fw_end_ng_cnt = 0
len_shrink_cnt = 0
unknown_tokens = Counter()
len_shrink_list = []
ws2 = wb.create_sheet("매핑정의서")
H2 = ["No","테이블","ord","BEFORE 영문","AFTER 영문","BEFORE 한글","AFTER 한글",
      "BEFORE 타입","BEFORE 길이","AFTER 타입","AFTER 길이","PK","변경유형",
      "도메인","영문→한글 조합","KR 검증","영문 형식단어 종결","길이 축소"]
W2 = [5,26,5,24,24,28,28,12,8,12,8,4,16, 16, 32, 10, 14, 10]
for i, w in enumerate(W2, 1): ws2.column_dimensions[get_column_letter(i)].width = w
for i, h in enumerate(H2, 1):
    cc = ws2.cell(row=1, column=i, value=h); cc.fill = HDR_FILL; cc.font = HDR_FONT
    cc.alignment = Alignment(horizontal="center", wrap_text=True); cc.border = BORDER
ws2.row_dimensions[1].height = 30
ws2.freeze_panes = "D2"
EN_FILL = PatternFill("solid", fgColor="FFE699")
KR_FILL = PatternFill("solid", fgColor="C6E0B4")
TYPE_FILL = PatternFill("solid", fgColor="F8CBAD")
KEEP_FILL = PatternFill("solid", fgColor="F2F2F2")
NG_FILL = PatternFill("solid", fgColor="FFC7CE")
OK_FILL = PatternFill("solid", fgColor="D9F2D9")
fw_abrvs = set(fw_y.values())
for i, c in enumerate(cols, 1):
    rn = i + 1
    types = []
    if c["bef_en"].upper() != c["aft_en"]: types.append("EN")
    if c["bef_kr"] != c["aft_kr"]: types.append("KR")
    if c["bef_dtype"] != c["aft_dtype"]: types.append("TYPE")
    if str(c["bef_dlen"]) != str(c["aft_dlen"]): types.append("LEN")
    chg = "+".join(types) if types else "KEEP"
    type_cnt[chg] += 1

    dom = map_domain(c["aft_kr"], c["aft_dtype"], c["aft_dlen"])
    kr_combined, unk = en_to_kr(c["aft_en"])
    for u in unk: unknown_tokens[u] += 1
    kr_ok = "OK" if kr_combined == c["aft_kr"] else "NG"
    if kr_ok == "OK": kr_match_cnt += 1
    else: kr_mismatch_cnt += 1

    last_tok = c["aft_en"].rsplit("_",1)[-1]
    fw_end = "OK" if last_tok in fw_abrvs else "NG"
    if fw_end == "OK": fw_end_ok_cnt += 1
    else: fw_end_ng_cnt += 1

    # 길이 축소 검증
    bef_len = int(c["bef_dlen"]) if c["bef_dlen"].isdigit() else 0
    aft_len = int(c["aft_dlen"]) if c["aft_dlen"].isdigit() else 0
    if c["aft_dtype"] == "CLOB":
        len_shrink = "-"   # CLOB 는 길이 무한대 의미상 축소 아님
    elif bef_len > 0 and aft_len > 0 and aft_len < bef_len:
        len_shrink = f"축소 ({bef_len}→{aft_len})"
        len_shrink_cnt += 1
        len_shrink_list.append((c["tbl"], c["aft_en"], c["aft_kr"], bef_len, aft_len))
    else:
        len_shrink = "OK"

    vals = [i, c["tbl"], c["ord"], c["bef_en"], c["aft_en"], c["bef_kr"], c["aft_kr"],
            c["bef_dtype"], c["bef_dlen"], c["aft_dtype"], c["aft_dlen"], c["pk"], chg,
            dom, kr_combined, kr_ok, fw_end, len_shrink]
    for j, v in enumerate(vals, 1):
        cc = ws2.cell(row=rn, column=j, value=v); cc.font = CELL_FONT; cc.border = BORDER
        cc.alignment = Alignment(vertical="top", wrap_text=True)
    if chg == "KEEP":
        ws2.cell(row=rn, column=13).fill = KEEP_FILL
    if "EN" in chg:
        ws2.cell(row=rn, column=4).fill = EN_FILL; ws2.cell(row=rn, column=5).fill = EN_FILL
    if "KR" in chg:
        ws2.cell(row=rn, column=6).fill = KR_FILL; ws2.cell(row=rn, column=7).fill = KR_FILL
    if "TYPE" in chg or "LEN" in chg:
        for j in [8,9,10,11]: ws2.cell(row=rn, column=j).fill = TYPE_FILL
    # 검증 컬럼 색
    ws2.cell(row=rn, column=16).fill = OK_FILL if kr_ok == "OK" else NG_FILL
    ws2.cell(row=rn, column=17).fill = OK_FILL if fw_end == "OK" else NG_FILL
    if len_shrink.startswith("축소"):
        ws2.cell(row=rn, column=18).fill = NG_FILL
    elif len_shrink == "OK":
        ws2.cell(row=rn, column=18).fill = OK_FILL

# 요약
ws3 = wb.create_sheet("요약")
ws3.column_dimensions["A"].width = 28; ws3.column_dimensions["B"].width = 12
ws3.cell(row=1, column=1, value="변경유형").fill = HDR_FILL
ws3.cell(row=1, column=1).font = HDR_FONT
ws3.cell(row=1, column=2, value="건수").fill = HDR_FILL
ws3.cell(row=1, column=2).font = HDR_FONT
rr = 2
for k, v in type_cnt.most_common():
    ws3.cell(row=rr, column=1, value=k)
    ws3.cell(row=rr, column=2, value=v)
    rr += 1
rr += 1
ws3.cell(row=rr, column=1, value="=== 검증 결과 ===").font = Font(bold=True); rr += 1
ws3.cell(row=rr, column=1, value="KR 검증 OK"); ws3.cell(row=rr, column=2, value=kr_match_cnt); rr += 1
ws3.cell(row=rr, column=1, value="KR 검증 NG"); ws3.cell(row=rr, column=2, value=kr_mismatch_cnt); rr += 1
ws3.cell(row=rr, column=1, value="영문 형식단어 종결 OK"); ws3.cell(row=rr, column=2, value=fw_end_ok_cnt); rr += 1
ws3.cell(row=rr, column=1, value="영문 형식단어 종결 NG"); ws3.cell(row=rr, column=2, value=fw_end_ng_cnt); rr += 1
ws3.cell(row=rr, column=1, value="길이 축소 발생"); ws3.cell(row=rr, column=2, value=len_shrink_cnt); rr += 1
rr += 1
ws3.cell(row=rr, column=1, value="=== 미등록 영문 토큰 (Top 30) ===").font = Font(bold=True); rr += 1
for tok, n in unknown_tokens.most_common(30):
    ws3.cell(row=rr, column=1, value=tok)
    ws3.cell(row=rr, column=2, value=n)
    rr += 1

# 길이 축소 시트
if len_shrink_list:
    ws4 = wb.create_sheet("길이축소")
    H4 = ["테이블","AFTER 영문","AFTER 한글","BEFORE 길이","AFTER 길이"]
    W4 = [30, 26, 28, 12, 12]
    for i, w in enumerate(W4, 1): ws4.column_dimensions[get_column_letter(i)].width = w
    for i, h in enumerate(H4, 1):
        cc = ws4.cell(row=1, column=i, value=h); cc.fill = HDR_FILL; cc.font = HDR_FONT
        cc.alignment = Alignment(horizontal="center"); cc.border = BORDER
    for i, (tbl, en, kr, bl, al) in enumerate(len_shrink_list, 2):
        for j, v in enumerate([tbl, en, kr, bl, al], 1):
            cc = ws4.cell(row=i, column=j, value=v); cc.font = CELL_FONT; cc.border = BORDER

wb.save(MAP_OUT)
print(f"  → {MAP_OUT}  ({len(cols)}행, 유형={dict(type_cnt)})")

# ============ 6. 용어사전 RAMP N ============
print(f"\n=== 6. 용어사전 RAMP N ===")
mois = {}
for p in MOIS_TERMS_UPLOAD:
    wb = load_workbook(p, read_only=True, data_only=True)
    ws = wb.active
    for r in ws.iter_rows(min_row=2, values_only=True):
        if r and r[2] and r[4]:
            mois[(s(r[2]), s(r[4]))] = s(r[5])
    wb.close()
print(f"  행안부 용어: {len(mois)}")

def map_domain(kr, dt, dl):
    dt_u = dt.upper()
    if dt_u == "VARCHAR":
        if kr.endswith("여부"): return f"여부V{dl}"
        if kr.endswith("코드"): return f"코드V{dl}"
        if kr.endswith("구분"): return f"구분V{dl}"
        if kr.endswith("명"):   return f"명V{dl}"
        if kr.endswith("값"):   return f"값V{dl}"
        if kr.endswith("내용"): return f"내용V{dl}"
        if kr.endswith("일자") and dl == "8":  return "일자V8"
        if kr.endswith("일시") and dl == "14": return "일시V14"
        if kr.endswith("연도"): return "연도V4"
        return f"내용V{dl}" if dl else "내용V"
    if dt_u == "NUMBER":
        if kr.endswith("값"): return f"값N{dl}"
        if kr.endswith("수"): return f"수N{dl}"
        if kr.endswith("순번"): return f"순번N{dl}"
        return f"값N{dl}"
    if dt_u == "DATE":     return "일자DT"
    if dt_u == "TIMESTAMP": return "일시TS"
    if dt_u == "CLOB":     return "내용L"
    return f"{dt_u}{dl}"

terms_cand = {}
for c in cols:
    key = (c["aft_en"], c["aft_kr"])
    if key not in terms_cand:
        terms_cand[key] = (c["aft_dtype"], c["aft_dlen"])

ramp_n = [(en, kr, dt, dl) for (en, kr), (dt, dl) in sorted(terms_cand.items(), key=lambda x: (x[0][1], x[0][0]))
          if (kr, en) not in mois]
print(f"  컬럼 distinct: {len(terms_cand)}, RAMP N: {len(ramp_n)}")

wb = Workbook(); ws = wb.active; ws.title = "Sheet1"
H3 = ["No","제정차수","용어명","용어설명","용어영문약어명","도메인명","허용값","저장형식","표현형식","코드그룹명","소관기관명","이음동의어목록","요청시스템","표준여부"]
W3 = [5, 8, 30, 40, 22, 16, 10, 18, 12, 12, 14, 20, 12, 8]
for i, w in enumerate(W3, 1): ws.column_dimensions[get_column_letter(i)].width = w
for i, h in enumerate(H3, 1):
    cc = ws.cell(row=1, column=i, value=h); cc.fill = HDR_FILL; cc.font = HDR_FONT
    cc.alignment = Alignment(horizontal="center"); cc.border = BORDER
ws.freeze_panes = "A2"
for i, (en, kr, dt, dl) in enumerate(ramp_n, 2):
    dom = map_domain(kr, dt, dl)
    vals = [i-1, "RAMP", kr, "", en, dom, "", "", "", "", "RAMP", "", "RAMP", "N"]
    for j, v in enumerate(vals, 1):
        cc = ws.cell(row=i, column=j, value=v); cc.font = CELL_FONT; cc.border = BORDER
        cc.alignment = Alignment(vertical="top", wrap_text=True)
wb.save(TERMS_OUT)
print(f"  → {TERMS_OUT}  ({len(ramp_n)}건)")

print("\n=== 30자 초과 ===")
for tbl, en, L, kr in over30[:20]:
    print(f"  {L}자  {tbl}.{en}  ({kr})")
