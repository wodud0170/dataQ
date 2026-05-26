# -*- coding: utf-8 -*-
"""
v2: CSV 안전 추출 + 제어문자/공백 노이즈 별도 점검 + normalize 후 정확 비교.

분류:
  [DEAD_IN_DB]      행안부 폐기인데 DB에 있음
  [MISSING]         살아있어야 하는데 DB에 누락
  [UNKNOWN_IN_DB]   행안부에 없는 단어가 DB에 있음
  [NOISE_TAB]       값에 TAB(\\t) 포함
  [NOISE_LF]        값에 LF(\\n) 포함
  [NOISE_CR]        값에 CR(\\r) 포함
  [NOISE_NBSP]      값에 NBSP(U+00A0) 포함
  [NOISE_WS_EDGE]   값 앞/뒤 공백
  [VAL_DIFF]        normalize 후에도 실제 값이 다름 (진짜 불일치)
"""
from openpyxl import load_workbook
from pathlib import Path
from collections import defaultdict, Counter
import subprocess, csv, sys, io, re

ORIG = Path(r"C:\Users\장재영\Desktop\dataQ\q-center\src\main\resources\seed\행안부_공통표준\행정안전부_공공데이터 공통표준단어.xlsx")
OUT_DIR = Path(r"C:\Users\장재영\Desktop\dataQ\CAMS표준화\04_RAMP분석_2026-05-21\_분석산출물_tsv")
OUT_DIR.mkdir(parents=True, exist_ok=True)

# ============ normalize ============
def has_tab(s):  return "\t" in s
def has_lf(s):   return "\n" in s
def has_cr(s):   return "\r" in s
def has_nbsp(s): return " " in s
def has_ws_edge(s): return s != s.strip() and s.strip() != ""

def normalize(s):
    """제어문자·NBSP 제거 + 양끝 공백 제거 + 내부 다중공백 정리. 콤마 뒤 공백 정규화."""
    if s is None: return ""
    s = s.replace("\t", "").replace("\r", "").replace("\n", "")
    s = s.replace(" ", " ")
    s = s.strip()
    # 콤마 뒤 공백 통일 (이음동의어/금칙어용)
    s = re.sub(r",\s*", ",", s)
    return s

# ============ 1) 원본 xlsx ============
print("=== 1. 원본 xlsx 파싱 ===")
wb = load_workbook(ORIG, read_only=True, data_only=True)
ws = wb["Sheet"]
rows = []
for r in ws.iter_rows(min_row=2, values_only=True):
    if r and r[0]:
        차수 = (r[8] or "").strip()
        구분 = (r[9] or "").strip()
        is_dead = ("폐기" in 차수 and "폐기후제정" not in 차수) or 구분 == "폐기"
        rows.append({
            "nm": (r[0] or "").strip(),
            "abrv": (r[1] or "").strip(),
            "eng": (r[2] or "").strip(),
            "desc": (r[3] or "").strip(),
            "frmt": (r[4] or "").strip(),
            "dmn": (r[5] or "").strip(),
            "syn": (r[6] or "").strip(),
            "fbd": (r[7] or "").strip(),
            "차수": 차수, "구분": 구분,
            "dead": is_dead,
        })
wb.close()
print(f"  총 행: {len(rows)}")

# 그룹화 — 살아있는 대표 행
groups = defaultdict(list)
for r in rows:
    groups[r["nm"]].append(r)
expected = {}
dead_only = []
for nm, lst in groups.items():
    alive = [r for r in lst if not r["dead"]]
    if alive:
        expected[nm] = alive[0]
    else:
        dead_only.append(nm)
print(f"  DB에 있어야 할: {len(expected)}")
print(f"  완전 폐기:      {len(dead_only)} → {dead_only}")

# ============ 2) DB CSV 안전 추출 ============
print("\n=== 2. DB CSV 추출 ===")
sql = """COPY (
  SELECT
    word_nm,
    coalesce(word_eng_abrv_nm,''),
    coalesce(word_eng_nm,''),
    coalesce(word_desc,''),
    coalesce(word_clsf_yn,''),
    coalesce(domain_clsf_nm,''),
    coalesce(array_to_string(alloph_synm_lst, ','), ''),
    coalesce(array_to_string(forbdn_word_lst, ','), '')
  FROM quality.tb_word
  WHERE comm_stnd_yn='Y'
  ORDER BY word_nm
) TO STDOUT WITH (FORMAT csv, FORCE_QUOTE *)
"""
res = subprocess.run(
    ["docker", "exec", "-i", "dataq-db", "psql", "-U", "admin", "-d", "postgres", "-c", sql],
    capture_output=True, encoding="utf-8"
)
if res.returncode != 0:
    print("ERR:", res.stderr); sys.exit(1)

db_raw = {}  # 원본 그대로 (노이즈 포함)
reader = csv.reader(io.StringIO(res.stdout))
for row in reader:
    if len(row) >= 8 and row[0]:
        db_raw[row[0]] = {
            "nm": row[0], "abrv": row[1], "eng": row[2], "desc": row[3],
            "frmt": row[4], "dmn": row[5], "syn": row[6], "fbd": row[7],
        }
print(f"  DB 단어 수: {len(db_raw)}")

# ============ 3) 노이즈 점검 (DB 측 전수) ============
print("\n=== 3. DB 측 제어문자/공백 노이즈 점검 ===")
FIELDS = [("abrv","영문약어"),("eng","영문명"),("desc","설명"),("frmt","형식단어여부"),
          ("dmn","도메인분류명"),("syn","이음동의어"),("fbd","금칙어")]
noise = defaultdict(list)  # (단어, 컬럼, 노이즈종류, raw값)
for nm, d in db_raw.items():
    for k, label in FIELDS:
        v = d[k] or ""
        if has_tab(v):     noise[("TAB", label)].append((nm, repr(v)[:60]))
        if has_lf(v):      noise[("LF", label)].append((nm, repr(v)[:60]))
        if has_cr(v):      noise[("CR", label)].append((nm, repr(v)[:60]))
        if has_nbsp(v):    noise[("NBSP", label)].append((nm, repr(v)[:60]))
        if has_ws_edge(v): noise[("WS_EDGE", label)].append((nm, repr(v)[:60]))

print(f"  노이즈 종류별 건수 (총 {sum(len(v) for v in noise.values())}건):")
for (kind, label), items in sorted(noise.items(), key=lambda x: -len(x[1])):
    print(f"    [{kind:8s}] {label:10s}: {len(items)}건  예: {items[0]}")

# ============ 4) 폐기/누락/미지 분류 ============
print("\n=== 4. 적재 정합성 분류 ===")
exp_set = set(expected.keys())
db_set = set(db_raw.keys())
xlsx_all = set(groups.keys())
dead_set = set(dead_only)

E_DEAD = sorted(dead_set & db_set)
E_MISSING = sorted(exp_set - db_set)
E_UNKNOWN = sorted(db_set - xlsx_all)

print(f"  [DEAD_IN_DB]    폐기 단어 DB 적재: {len(E_DEAD)} → {E_DEAD}")
print(f"  [MISSING]       살아야할 단어 누락: {len(E_MISSING)} → {E_MISSING[:10]}")
print(f"  [UNKNOWN_IN_DB] 행안부에 없음:      {len(E_UNKNOWN)} → {E_UNKNOWN[:10]}")

# ============ 5) normalize 후 진짜 값 차이 ============
print("\n=== 5. normalize 후 실제 값 차이 ===")
common = exp_set & db_set
val_diff = []
for nm in common:
    e = expected[nm]; d = db_raw[nm]
    for k, label in FIELDS:
        en = normalize(e[k])
        dn = normalize(d[k])
        if en != dn:
            val_diff.append((nm, label, e[k], d[k], en, dn))

print(f"  진짜 값 다른 케이스: {len(val_diff)}건")
변경_set = set(r["nm"] for r in rows if r["구분"] == "변경")
mm_names = set(x[0] for x in val_diff)
print(f"  └ 변경 53건과 겹침: {len(mm_names & 변경_set)}")
if val_diff[:5]:
    print("  샘플 5건:")
    for nm, lbl, ev, dv, en, dn in val_diff[:5]:
        print(f"    {nm} / {lbl}")
        print(f"      원본 norm: {en[:80]}")
        print(f"      DB norm:   {dn[:80]}")

# ============ 6) 산출 ============
print("\n=== 6. 산출 ===")
OUT = OUT_DIR / "_G_mois_db_verify_v2.tsv"
with open(OUT, "w", encoding="utf-8", newline="") as f:
    w = csv.writer(f, delimiter="\t")
    w.writerow(["분류","단어명","컬럼","원본값(raw)","DB값(raw)","원본norm","DB norm","비고"])
    for nm in E_DEAD:
        r = next(x for x in rows if x["nm"]==nm and x["dead"])
        w.writerow(["DEAD_IN_DB", nm, "", "", "(DB에 있음)", "", "", f"차수={r['차수']}"])
    for nm in E_MISSING:
        e = expected[nm]
        w.writerow(["MISSING", nm, "", e["abrv"], "(없음)", "", "", ""])
    for nm in E_UNKNOWN:
        w.writerow(["UNKNOWN_IN_DB", nm, "", "(없음)", db_raw[nm]["abrv"], "", "", ""])
    for (kind, label), items in sorted(noise.items()):
        for nm, raw in items:
            w.writerow([f"NOISE_{kind}", nm, label, "", raw, "", "", ""])
    for nm, lbl, ev, dv, en, dn in val_diff:
        w.writerow(["VAL_DIFF", nm, lbl, ev[:300], dv[:300], en[:300], dn[:300], ""])

total = len(E_DEAD)+len(E_MISSING)+len(E_UNKNOWN)+sum(len(v) for v in noise.values())+len(val_diff)
print(f"  → {OUT}")
print(f"  총 행: {total} (헤더 제외)")

# ============ 요약 ============
print("\n" + "="*60)
print("=== 최종 요약 ===")
print(f"DB 단어 수: {len(db_raw)}  /  살아있어야 할 수: {len(expected)}")
print(f"")
print(f"🔴 즉시 조치 필요:")
print(f"  - DEAD_IN_DB (폐기 적재):  {len(E_DEAD)}건 → DELETE")
print(f"  - MISSING (살아야할 누락): {len(E_MISSING)}건 → INSERT")
print(f"  - UNKNOWN_IN_DB:          {len(E_UNKNOWN)}건")
print(f"  - VAL_DIFF (진짜 값 차이): {len(val_diff)}건 → UPDATE")
print(f"")
print(f"🟠 데이터 품질 노이즈 (값은 맞지만 제어문자 섞임):")
total_noise = sum(len(v) for v in noise.values())
print(f"  - 총 {total_noise}건 (자세히는 산출 TSV 참조)")
