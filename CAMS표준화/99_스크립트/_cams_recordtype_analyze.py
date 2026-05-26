"""
CAMS 기록물 유형 분석 (1단계: 식별·카운트).

수행:
  1) 전체 테이블·컬럼 중에서 RG_* 테이블 식별
  2) 컬럼 코멘트에서 유형 키워드 (유형, TYPE, KIND, 분류 등) 발견되는 컬럼 추출
  3) RG_DOCUMENT / RG_DETAIL 의 유형 분기 컬럼 식별
  4) 유형별 전용 테이블 후보 (RG_GOVTINFO, RG_ADMIN_ADDINFO, RG_VIDEOTAPE 등) 식별
  5) "철·건" (BSID/DSID) 키 패턴 보유 테이블 목록화

출력: 05_CAMS기록물유형_2026-05-21/_01_식별결과.xlsx
"""
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from pathlib import Path
from collections import defaultdict, Counter
import re

BASE = Path(r"C:\Users\장재영\Desktop\dataQ\CAMS표준화")
SCHEMA = BASE / "01_원본자료" / "CAMS_SCHEMA_원본.xlsx"
OUT = BASE / "05_CAMS기록물유형_2026-05-21" / "_01_식별결과.xlsx"

# === Load ===
wb = load_workbook(SCHEMA, read_only=True, data_only=True)

# 테이블 목록
tables = []
for r in wb["테이블"].iter_rows(min_row=2, values_only=True):
    if not r or not r[1]: continue
    tables.append({"owner": (r[0] or "").strip(),
                   "tbl": (r[1] or "").strip(),
                   "cmt": (r[2] or "").strip()})

# 컬럼 정의
cols = []
for r in wb["컬럼정의"].iter_rows(min_row=2, values_only=True):
    if not r or not r[0]: continue
    cols.append({
        "tbl": (r[0] or "").strip(),
        "tbl_cmt": (r[1] or "").strip(),
        "col": (r[2] or "").strip(),
        "col_cmt": (r[3] or "").strip(),
        "type": (r[4] or "").strip(),
        "len": r[5],
        "notnull": (r[6] or "").strip(),
        "idx": (r[7] or "").strip(),
        "pk": (r[8] or "").strip(),
        "fk": (r[9] or "").strip(),
    })
wb.close()

print(f"전체 테이블: {len(tables)}")
print(f"전체 컬럼: {len(cols)}")

# === RG_* 테이블 ===
rg_tables = [t for t in tables if t["tbl"].upper().startswith("RG_")]
print(f"\nRG_* 테이블: {len(rg_tables)}")

# 컬럼 데이터를 테이블별로 그룹
cols_by_tbl = defaultdict(list)
for c in cols:
    cols_by_tbl[c["tbl"]].append(c)

# === 유형 분기 컬럼 식별 ===
# 컬럼 코멘트나 컬럼명에 유형/TYPE/종류/분류 등이 들어가면서 RG_* 또는 메인 테이블에 있는 것
type_keywords = ["유형", "종류", "TYPE", "KIND", "분류", "RECORD_TYPE", "DOC_TYPE", "RG_TYPE"]
type_cols = []
for c in cols:
    cm = c["col_cmt"]
    cn = c["col"].upper()
    if any(k.upper() in cn for k in ["TYPE", "KIND"]) or any(k in cm for k in ["유형", "종류"]):
        if "유형" in cm or "TYPE" in cn or "종류" in cm:
            type_cols.append(c)

print(f"유형/TYPE 키워드 컬럼: {len(type_cols)}")

# RG_DOCUMENT, RG_DETAIL 의 모든 컬럼
target_main = ["RG_DOCUMENT", "RG_DETAIL"]
main_cols = {t: cols_by_tbl.get(t, []) for t in target_main}
for t, cs in main_cols.items():
    print(f"\n{t} 컬럼수: {len(cs)}")
    # 유형 관련
    for c in cs:
        if "유형" in c["col_cmt"] or "TYPE" in c["col"].upper() or "종류" in c["col_cmt"]:
            print(f"  ★ {c['col']}: {c['col_cmt']} ({c['type']}{c['len']})")

# === 유형별 전용 테이블 후보 ===
# 키워드 매칭으로 후보 추출
type_kw = {
    "정부간행물": ["GOVT", "GOVTINFO", "PUBLICATION"],
    "행정박물": ["ADMIN_ADD", "ADMINISTRATIVE", "ADMINOBJ"],
    "총독부": ["KOREAN_GOV", "JAPANESE_GOV", "JEONGRYUNG"],
    "해외기록": ["FOREIGN", "OVERSEAS"],
    "구술": ["ORAL"],
    "시청각": ["AV_", "VIDEO", "AUDIO", "VIDEOTAPE", "PHOTOGRAPH"],
    "사진": ["PHOTO", "PICTURE", "IMAGE"],
    "비디오": ["VIDEO", "TAPE"],
    "오디오": ["AUDIO", "SOUND"],
    "도서": ["BOOK"],
    "지도": ["MAP"],
    "도면": ["DRAWING", "BLUEPRINT"],
    "박물": ["OBJECT", "ARTIFACT"],
    "회의록": ["MEETING"],
}
candidates_by_type = defaultdict(list)
for t in tables:
    tn = t["tbl"].upper()
    cm = t["cmt"]
    for type_nm, kws in type_kw.items():
        if type_nm in cm:
            candidates_by_type[type_nm].append(t)
        else:
            for kw in kws:
                if kw in tn:
                    candidates_by_type[type_nm].append(t)
                    break

print(f"\n=== 유형별 테이블 후보 (이름·코멘트 매칭) ===")
for tp, lst in candidates_by_type.items():
    print(f"\n[{tp}] {len(lst)} 후보")
    for t in lst[:10]:
        print(f"  {t['tbl']} — {t['cmt']}")
    if len(lst) > 10:
        print(f"  ... 외 {len(lst)-10}")

# === BSID/DSID 키 보유 테이블 ===
bsid_tables = set()
dsid_tables = set()
mbsid_tables = set()
for c in cols:
    cu = c["col"].upper()
    if cu == "BSID":
        bsid_tables.add(c["tbl"])
    elif cu == "DSID":
        dsid_tables.add(c["tbl"])
    elif cu == "MBSID":
        mbsid_tables.add(c["tbl"])

print(f"\nBSID 보유 테이블: {len(bsid_tables)}")
print(f"DSID 보유 테이블: {len(dsid_tables)}")
print(f"MBSID 보유 테이블 (이관 영역): {len(mbsid_tables)}")

# === xlsx 작성 ===
wb = Workbook()
hdr_font = Font(bold=True, color="FFFFFF")
hdr_fill = PatternFill("solid", fgColor="305496")
hdr_align = Alignment(horizontal="center", vertical="center")

def style_header(ws, ncol):
    for c in range(1, ncol+1):
        cell = ws.cell(row=1, column=c)
        cell.font = hdr_font; cell.fill = hdr_fill; cell.alignment = hdr_align

def auto_width(ws, widths):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

# 시트 1: 요약
ws = wb.active
ws.title = "요약"
ws.append(["항목", "값", "비고"])
ws.append(["전체 CAMS 테이블", len(tables), ""])
ws.append(["전체 CAMS 컬럼", len(cols), ""])
ws.append(["RG_* 테이블", len(rg_tables), "기록물 영역"])
ws.append(["BSID 보유 테이블", len(bsid_tables), "철 키"])
ws.append(["DSID 보유 테이블", len(dsid_tables), "건 키"])
ws.append(["MBSID 보유 테이블", len(mbsid_tables), "이관 영역"])
ws.append([])
ws.append(["[유형 분기 컬럼]", "", ""])
for c in cols:
    if c["tbl"] in target_main and ("유형" in c["col_cmt"] or "TYPE" in c["col"].upper() or "종류" in c["col_cmt"]):
        ws.append([f"  {c['tbl']}.{c['col']}", f"{c['type']}({c['len']})", c["col_cmt"]])
ws.append([])
ws.append(["[유형별 테이블 후보 수]", "", ""])
for tp, lst in sorted(candidates_by_type.items(), key=lambda x: -len(x[1])):
    ws.append([f"  {tp}", len(lst), "; ".join(t["tbl"] for t in lst[:5])])
style_header(ws, 3)
auto_width(ws, [40, 12, 70])

# 시트 2: RG_* 테이블 전수
ws = wb.create_sheet("RG_테이블_전수")
ws.append(["테이블명", "코멘트", "컬럼수", "BSID 보유", "DSID 보유", "MBSID 보유"])
for t in sorted(rg_tables, key=lambda x: x["tbl"]):
    tn = t["tbl"]
    ws.append([tn, t["cmt"], len(cols_by_tbl.get(tn, [])),
               "Y" if tn in bsid_tables else "",
               "Y" if tn in dsid_tables else "",
               "Y" if tn in mbsid_tables else ""])
style_header(ws, 6)
auto_width(ws, [32, 45, 10, 10, 10, 12])
ws.freeze_panes = "A2"

# 시트 3: 유형 분기 컬럼 전수
ws = wb.create_sheet("유형분기컬럼_전수")
ws.append(["테이블", "테이블 코멘트", "컬럼명", "컬럼 코멘트", "타입", "길이", "PK"])
type_cols_sorted = sorted(type_cols, key=lambda c: (c["tbl"], c["col"]))
for c in type_cols_sorted:
    ws.append([c["tbl"], c["tbl_cmt"], c["col"], c["col_cmt"], c["type"], c["len"], c["pk"]])
style_header(ws, 7)
auto_width(ws, [28, 40, 22, 30, 12, 8, 6])
ws.freeze_panes = "A2"

# 시트 4: 유형별 후보 테이블
ws = wb.create_sheet("유형별_후보테이블")
ws.append(["유형 키워드", "후보 테이블명", "코멘트", "컬럼수"])
for tp, lst in candidates_by_type.items():
    for t in lst:
        ws.append([tp, t["tbl"], t["cmt"], len(cols_by_tbl.get(t["tbl"], []))])
style_header(ws, 4)
auto_width(ws, [15, 32, 50, 10])
ws.freeze_panes = "A2"

# 시트 5: RG_DOCUMENT 컬럼 전수
ws = wb.create_sheet("RG_DOCUMENT_컬럼")
ws.append(["컬럼명", "컬럼 코멘트", "타입", "길이", "NOT NULL", "PK", "FK"])
for c in cols_by_tbl.get("RG_DOCUMENT", []):
    ws.append([c["col"], c["col_cmt"], c["type"], c["len"], c["notnull"], c["pk"], c["fk"]])
style_header(ws, 7)
auto_width(ws, [25, 35, 12, 8, 10, 6, 6])
ws.freeze_panes = "A2"

# 시트 6: RG_DETAIL 컬럼 전수
ws = wb.create_sheet("RG_DETAIL_컬럼")
ws.append(["컬럼명", "컬럼 코멘트", "타입", "길이", "NOT NULL", "PK", "FK"])
for c in cols_by_tbl.get("RG_DETAIL", []):
    ws.append([c["col"], c["col_cmt"], c["type"], c["len"], c["notnull"], c["pk"], c["fk"]])
style_header(ws, 7)
auto_width(ws, [25, 35, 12, 8, 10, 6, 6])
ws.freeze_panes = "A2"

wb.save(OUT)
print(f"\nSAVED: {OUT}")
print(f"size: {OUT.stat().st_size:,} bytes")
