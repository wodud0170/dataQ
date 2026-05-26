# -*- coding: utf-8 -*-
# 02_BEFORE_AFTER_컬럼.xlsx 생성 — 행안부 표준단어(word_nm) + 단어_등록안.tsv 기준
# 이음동의어 미사용: 모든 변형어는 단어_등록안 A규칙으로 행안부 표준어에 매핑
import openpyxl, re, csv
from openpyxl import Workbook

# 1) 행안부 표준단어(word_nm) — _w.tsv 에서 적재 (psql 추출본)
res = {}   # surface -> canonical word_nm
abrv = {}  # canonical -> 영문약어
with open('_w.tsv', encoding='utf-8') as f:
    for ln in f:
        p = ln.rstrip('\n').split('\t')
        if len(p) == 2 and p[0]:
            res.setdefault(p[0], p[0]); abrv[p[0]] = p[1]
assert len(res) > 3000, '_w.tsv 적재 실패 (행안부 표준단어 %d개)' % len(res)

# 2) 코멘트정정 규칙 — 변형어를 행안부 표준어로 치환 (이음동의어 미사용)
with open('코멘트정정_규칙.tsv', encoding='utf-8') as f:
    for row in csv.DictReader(f, delimiter='\t'):
        res[row['변형어']] = row['표준어(행안부)']
# 3) 기관표준 신규 단어 등록안
with open('단어_등록안.tsv', encoding='utf-8') as f:
    for row in csv.DictReader(f, delimiter='\t'):
        han = row['한글']
        res.setdefault(han, han)
        abrv[han] = row['영문약어']
maxlen = max(len(w) for w in res)

def seg(ch):
    i = 0; toks = []; runs = []; buf = ''
    while i < len(ch):
        hit = None
        for L in range(min(maxlen, len(ch) - i), 0, -1):
            if ch[i:i+L] in res:
                hit = ch[i:i+L]; break
        if hit:
            if buf: runs.append(buf); buf = ''
            toks.append(res[hit]); i += len(hit)
        else:
            buf += ch[i]; i += 1
    if buf: runs.append(buf)
    return toks, runs

# 3) 컬럼명 표준규칙 — 같은 영문컬럼명은 코멘트 무시하고 표준한글명으로 통일
colmap = {}
with open('컬럼명_표준규칙.tsv', encoding='utf-8') as f:
    for r in csv.DictReader(f, delimiter='\t'):
        colmap[r['영문컬럼명'].strip()] = r['표준한글명'].strip()

# 4) 원본 적재 + AFTER 생성
wb = openpyxl.load_workbook('CAMS_SCHEMA_원본.xlsx', read_only=True, data_only=True)
rows = list(wb['컬럼정의'].iter_rows(values_only=True))[1:]
o = Workbook(); ws = o.active; ws.title = 'BEFORE_AFTER'
ws.append(['테이블명', '테이블코멘트', 'BEFORE_컬럼명', 'BEFORE_컬럼코멘트', '데이터타입',
           '데이터길이', 'AFTER_컬럼명', 'AFTER_한글명(표준)', '적용표준단어', '변경여부', '비고'])
stat = {'동일': 0, '변경': 0, '미처리': 0, '보류': 0}
for r in rows:
    tbl, tcmt, col, ccmt, dt, dl = r[0], r[1], r[2], r[3], r[4], r[5]
    ovr = colmap.get((col or '').strip())
    c = ovr if ovr else (ccmt or '').strip()
    ac = an = aw = ''; chg = ''; note = '컬럼명 표준규칙 적용' if ovr else ''
    if not c:
        chg = '보류'; note = '코멘트 없음'
    else:
        toks = []; runs = []
        for ch in [x for x in re.split(r'[\s()\[\]{}/,~\-_·:;.]+', c) if x]:
            t, rn = seg(ch); toks += t; runs += rn
        if runs:
            chg = '미처리'; note = '미존재단어: ' + ','.join(sorted(set(runs)))[:120]
        elif any(not abrv.get(t) for t in toks):
            chg = '미처리'; note = '약어없는 표준단어: ' + ','.join(t for t in toks if not abrv.get(t))
        else:
            ac = '_'.join(abrv[t] for t in toks)
            an = ''.join(toks)
            aw = ' + '.join(toks)
            chg = '동일' if ac == (col or '') else '변경'
    stat[chg] += 1
    ws.append([tbl, tcmt, col, ccmt, dt, dl, ac, an, aw, chg, note])
for i, w in enumerate([26, 34, 24, 26, 12, 8, 26, 24, 30, 8, 40], 1):
    ws.column_dimensions[chr(64 + i)].width = w
ws.freeze_panes = 'A2'
o.save('02_BEFORE_AFTER_컬럼.xlsx')
tot = sum(stat.values())
print('02_BEFORE_AFTER_컬럼.xlsx 재생성 — 총 %d행' % tot)
for k in ['동일', '변경', '미처리', '보류']:
    print('  %s: %d (%.1f%%)' % (k, stat[k], stat[k] * 100 / tot))
print('  표준화 완료(동일+변경): %d (%.1f%%)' % (stat['동일'] + stat['변경'],
      (stat['동일'] + stat['변경']) * 100 / tot))
