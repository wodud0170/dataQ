# -*- coding: utf-8 -*-
"""
RAMP 용어사전 생성 (Phase 2-1/2-2 결정 반영, AFTER 기준).

흐름:
  1. RAMP 컬럼 로드 + Phase 2-1 단어 변환 적용 (영문약어·한글)
  2. Phase 2-2 도메인 매핑 (시간 D3 / 다중도메인 D1 / R8 D5)
  3. AFTER 한글명 그룹화
  4. tb_terms 신규 INSERT (comm_stnd_yn='N')
"""
from openpyxl import load_workbook
from pathlib import Path
from collections import defaultdict, Counter
import subprocess, csv, io, sys, os, base64, re

BASE = Path(r"C:\Users\장재영\Desktop\dataQ\CAMS표준화")
COL_BA = BASE / "04_RAMP분석_2026-05-21" / "00_핵심산출" / "RAMP_컬럼_BEFORE_AFTER_2026-05-23.xlsx"
TIME_BA = BASE / "04_RAMP분석_2026-05-21" / "00_핵심산출" / "Phase2_시간컬럼_BEFORE_AFTER_2026-05-23.xlsx"
D1_XLSX = BASE / "04_RAMP분석_2026-05-21" / "00_핵심산출" / "Phase2_D1_도메인통일_2026-05-23.xlsx"
D5_XLSX = BASE / "04_RAMP분석_2026-05-21" / "00_핵심산출" / "Phase2_D5_R8형식단어_2026-05-23.xlsx"
RAMP_SCHEMA = BASE / "CAMS_RAMP_통합" / "ramp기관스키마정보.xlsx"
OUT_SQL = BASE / "04_RAMP분석_2026-05-21" / "99_실행SQL_기록" / "Phase2_terms_insert.sql"

# ============ 1. RAMP 컬럼 로드 + 단어 변환 적용 ============
print("=== 1. RAMP 컬럼 + Phase 2-1 단어 적용 ===")

# Phase 2-1 컬럼별 변환 정보 로드
col_map = {}  # (tbl, col_en BEFORE) → (col_en AFTER, col_kr AFTER)
wb = load_workbook(COL_BA, read_only=True, data_only=True)
ws = wb["컬럼BEFORE_AFTER"]
for r in ws.iter_rows(min_row=2, values_only=True):
    if r and r[1] and r[2]:
        col_map[(str(r[1]), str(r[2]))] = (str(r[3] or r[2]), str(r[5] or r[4] or ""))
wb.close()
print(f"  Phase 2-1 변환된 컬럼: {len(col_map)}")

# D5 결과 (BEFORE → AFTER 한글) 로드
D5_XLSX = BASE / "04_RAMP분석_2026-05-21" / "00_핵심산출" / "Phase2_D5_R8형식단어_2026-05-23.xlsx"
d5_after = {}; d5_dom = {}
wb = load_workbook(D5_XLSX, read_only=True, data_only=True)
for r in wb["새용어_BEFORE_AFTER"].iter_rows(min_row=2, values_only=True):
    if r and r[1] and r[3]:
        d5_after[str(r[1])] = str(r[3])
        d5_dom[str(r[3])] = str(r[6] or "")
wb.close()
print(f"  D5 매핑 (R8 보충): {len(d5_after)}")

# RAMP 스키마 전체 컬럼
wb = load_workbook(RAMP_SCHEMA, read_only=True, data_only=True)
ws = wb["컬럼"]
all_cols = []
seen = set()
for r in ws.iter_rows(min_row=2, values_only=True):
    if r and r[0] and r[1]:
        key = (str(r[0]).strip(), str(r[1]).strip())
        if key in seen: continue
        seen.add(key)
        tbl, col_en, col_kr = key[0], key[1], str(r[2] or "").strip()
        if key in col_map:
            after_en, after_kr = col_map[key]
            col_en = after_en or col_en
            col_kr = after_kr or col_kr
        # D5 보충 한글 적용
        if col_kr in d5_after:
            col_kr = d5_after[col_kr]
        # 한글 공백 제거 (정규화)
        col_kr = col_kr.replace(" ", "")
        # 영문약어 대문자 (Oracle 정책)
        col_en = col_en.upper()
        all_cols.append({
            "tbl": tbl, "col_en": col_en, "col_kr": col_kr,
            "dtype": str(r[5] or "").strip().upper() if len(r) > 5 else "",
            "dlen": str(r[6] or "").strip() if len(r) > 6 else "",
            "desc": str(r[3] or "").strip() if len(r) > 3 else "",
        })
wb.close()
print(f"  RAMP 컬럼 (dedupe): {len(all_cols)}")

# ============ 2. 시간 컬럼 도메인 매핑 ============
print("\n=== 2. 시간 컬럼 도메인 매핑 ===")
time_dom_map = {}  # (tbl, col_en UPPER) → domain_nm
wb = load_workbook(TIME_BA, read_only=True, data_only=True)
ws = wb["적용_BEFORE_AFTER"]
for r in ws.iter_rows(min_row=2, values_only=True):
    if r and r[1]:
        tbl = str(r[1]); col = str(r[2]).upper(); dom = str(r[7] or "")  # 대문자 통일
        if dom: time_dom_map[(tbl, col)] = dom
wb.close()
print(f"  시간 도메인 매핑: {len(time_dom_map)}")

# ============ 3. 도메인 사전 로드 ============
print("\n=== 3. 도메인 사전 로드 ===")
r = subprocess.run(["docker","exec","-i","dataq-db","psql","-U","admin","-d","postgres",
    "-c", """COPY (SELECT domain_nm, domain_clsf_nm, data_type, coalesce(data_len::text,'0')
             FROM quality.tb_domain) TO STDOUT WITH (FORMAT csv, FORCE_QUOTE *)"""],
    capture_output=True, encoding="utf-8")
all_doms = []
for row in csv.reader(io.StringIO(r.stdout)):
    if row: all_doms.append({"nm": row[0], "clsf": row[1], "type": row[2], "len": row[3]})
print(f"  도메인 전체: {len(all_doms)}")

# 매칭 함수: (분류, 타입, 길이) → 도메인명
def find_domain(clsf, dt, dl):
    if not clsf: return None
    L = int(dl) if str(dl).isdigit() else 0
    dt_u = dt.upper()
    # CLOB 매핑
    if dt_u in ("VARCHAR","VARCHAR2","STRING") and L > 4000:
        return next((d["nm"] for d in all_doms if d["clsf"]==clsf and d["type"].upper()=="CLOB"), None)
    # 일반 매칭 (분류 + 타입 + 길이 정확 일치)
    target_dt = "VARCHAR" if dt_u in ("STRING","VARCHAR2") else dt_u
    for d in all_doms:
        if d["clsf"] == clsf and d["type"].upper() == target_dt and (int(d["len"]) if d["len"].isdigit() else 0) == L:
            return d["nm"]
    return None

# ============ 4. 한글명 그룹화 + 용어 산출 ============
print("\n=== 4. 한글명 그룹화 ===")
grp = defaultdict(list)
for c in all_cols:
    if c["col_kr"]:
        grp[c["col_kr"]].append(c)
print(f"  unique 용어 (AFTER 한글): {len(grp)}")

# 형식단어 추출 — 한글 끝의 형식단어
r = subprocess.run(["docker","exec","-i","dataq-db","psql","-U","admin","-d","postgres",
    "-c", """COPY (SELECT word_nm, coalesce(domain_clsf_nm,'') FROM quality.tb_word
             WHERE word_clsf_yn='Y') TO STDOUT WITH (FORMAT csv, FORCE_QUOTE *)"""],
    capture_output=True, encoding="utf-8")
fw_map = {}
for row in csv.reader(io.StringIO(r.stdout)):
    if row and row[0]: fw_map[row[0]] = row[1]
EXTRA_CLSF = {"구분":"구분","값":"값","경로":"경로","상태":"상태","메시지":"메시지","식별자":"식별자",
              "파일":"파일","순번":"순번","번호":"번호","자":"자","이름":"이름","내역":"내역",
              "유형":"유형","범위":"범위","위치":"위치","방법":"방법","수단":"수단","목록":"목록",
              "사항":"사항","항목":"항목","기간":"기간","결과":"결과","대상":"대상","주체":"주체",
              "주소":"주소","권한":"권한","역할":"역할","사유":"사유","율":"율","률":"률","비율":"비율"}
fw_map.update(EXTRA_CLSF)

def extract_clsf(kr):
    cand = [(fw, fw_map[fw]) for fw in fw_map if kr.endswith(fw)]
    valid = [c for c in cand if c[1]]
    if valid: return max(valid, key=lambda x: len(x[0]))[1]
    # 분류 추출 실패 시 "값" 부여 (D8 결정)
    return "값"

# 용어 정보 산출
def parse_int(s):
    try: return int(s)
    except: return 0

terms = []
no_domain = 0
for kr, lst in grp.items():
    # 영문약어 (최빈)
    en_cnt = Counter(c["col_en"] for c in lst)
    eng_abrv = en_cnt.most_common(1)[0][0]
    # 설명 (가장 긴 desc)
    desc = max((c["desc"] for c in lst), key=len, default="") or kr
    # 도메인: 시간 매핑 우선 → 최장 도메인 매칭
    time_doms = set()
    for c in lst:
        d = time_dom_map.get((c["tbl"], c["col_en"]))
        if d: time_doms.add(d)
    if len(time_doms) == 1:
        dom = time_doms.pop()
    else:
        # D5 도메인 우선
        dom = d5_dom.get(kr) if kr in d5_dom else None
        if not dom:
            longest = max(lst, key=lambda c: parse_int(c["dlen"]))
            clsf = extract_clsf(kr)
            dom = find_domain(clsf, longest["dtype"], longest["dlen"]) if clsf else None
        # 시간 suffix fallback (D3 매핑 누락 보완)
        if not dom and (kr.endswith("일자") or kr.endswith("일시")):
            longest = max(lst, key=lambda c: parse_int(c["dlen"]))
            dt = longest["dtype"].upper(); L = parse_int(longest["dlen"])
            if dt == "DATETIME": dom = "일시TS"
            elif dt == "DATE":   dom = "일자DT" if kr.endswith("일자") else "일자DT"
            elif kr.endswith("일자") and L <= 8: dom = "일자V8"
            elif kr.endswith("일시") or L >= 9:  dom = "일시V14"
    if not dom: no_domain += 1
    terms.append({
        "nm": kr,
        "abrv": eng_abrv,
        "desc": desc[:1000],
        "dom": dom or "",
        "col_cnt": len(lst),
    })

terms.sort(key=lambda x: -x["col_cnt"])
print(f"  용어: {len(terms)}  / 도메인 매칭 실패: {no_domain}")

# ============ 5. 기존 DB 용어 점검 + dedup ============
print("\n=== 5. 기존 DB 용어 점검 ===")
r = subprocess.run(["docker","exec","-i","dataq-db","psql","-U","admin","-d","postgres",
    "-c", "COPY (SELECT terms_nm, terms_eng_abrv_nm FROM quality.tb_terms) TO STDOUT WITH (FORMAT csv, FORCE_QUOTE *)"],
    capture_output=True, encoding="utf-8")
db_terms = set()  # 한글명 set
db_pair = set()
db_abrv = set()   # 영문약어 set (UNIQUE)
for row in csv.reader(io.StringIO(r.stdout)):
    if row:
        db_terms.add(row[0])
        db_pair.add((row[0], row[1]))
        db_abrv.add(row[1])
print(f"  DB 기존 용어: {len(db_terms)} / 영문약어 unique: {len(db_abrv)}")

to_insert = []
collision = []
no_dom_skip = []
abrv_collision = []
inserted_abrv = set()  # 이번 batch 안에서도 중복 방지
for t in terms:
    if (t["nm"], t["abrv"]) in db_pair:
        continue
    if t["nm"] in db_terms:
        db_abrv_for_kr = next((a for (n,a) in db_pair if n==t["nm"]), None)
        if db_abrv_for_kr and db_abrv_for_kr.upper() == t["abrv"].upper():
            continue
        collision.append(t); continue
    if not t["dom"]:
        no_dom_skip.append(t); continue
    if t["abrv"] in db_abrv or t["abrv"] in inserted_abrv:
        abrv_collision.append(t); continue
    to_insert.append(t)
    inserted_abrv.add(t["abrv"])
print(f"  INSERT 대상: {len(to_insert)} / 충돌(한글): {len(collision)} / 도메인NULL: {len(no_dom_skip)} / 영문약어충돌: {len(abrv_collision)}")

# ============ 6. SQL 생성 ============
print(f"\n=== 6. SQL 생성 ===")
def esc(s):
    if s is None: return ""
    return str(s).replace("'", "''")

sql = ["-- RAMP 용어사전 INSERT (Phase 2 종합)", "BEGIN;", ""]
for t in to_insert:
    tid = base64.urlsafe_b64encode(os.urandom(17)).rstrip(b'=').decode()[:22]
    dom_str = f"'{esc(t['dom'])}'" if t["dom"] else "NULL"
    sql.append(
        f"INSERT INTO quality.tb_terms (terms_id, terms_nm, terms_eng_abrv_nm, terms_desc, domain_nm, comm_stnd_yn, aprv_yn, cret_dt, cret_user_id, use_yn) "
        f"VALUES ('{tid}','{esc(t['nm'])}','{esc(t['abrv'])}','{esc(t['desc'])}',{dom_str},'N','Y',to_char(now(),'YYYYMMDDHH24MISS'),'admin','Y');"
    )

sql += ["",
        "SELECT 'tb_terms Y' as t, count(*) FROM quality.tb_terms WHERE comm_stnd_yn='Y';",
        "SELECT 'tb_terms N' as t, count(*) FROM quality.tb_terms WHERE comm_stnd_yn='N';",
        "",
        "COMMIT;"]
OUT_SQL.write_text("\n".join(sql), encoding="utf-8")
print(f"  → {OUT_SQL}")
print(f"  INSERT 대상: {len(to_insert)}")
print(f"  충돌 (한글 동일, 약어 다름): {len(collision)}")
print(f"  도메인 매칭 실패: {no_domain}")
