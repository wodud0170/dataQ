# -*- coding: utf-8 -*-
"""
행안부 원본 xlsx vs dataQ DB tb_word(comm_stnd_yn='Y') 정합성 검증.

규칙:
  - 단어명 그룹화. 한 단어명에 여러 행이 있을 수 있음 (폐기+제정).
  - 살아있는 행 = 차수에 '폐기' 없고 개정구분명 != '폐기'
  - DB에는 "그 단어명에 살아있는 행이 1개라도 있으면" 그 행이 들어가야 함.
  - 단어명의 모든 행이 폐기이면 DB에 그 단어명은 없어야 함.

비교 항목 (살아있는 행 기준):
  단어명 / 영문약어 / 영문명 / 설명 / 형식단어여부 / 도메인분류명 / 이음동의어 / 금칙어
"""
from openpyxl import load_workbook
from pathlib import Path
import subprocess, csv, sys, io, json

ORIG = Path(r"C:\Users\장재영\Desktop\dataQ\q-center\src\main\resources\seed\행안부_공통표준\행정안전부_공공데이터 공통표준단어.xlsx")
OUT_DIR = Path(r"C:\Users\장재영\Desktop\dataQ\CAMS표준화\04_RAMP분석_2026-05-21\_분석산출물_tsv")
OUT_DIR.mkdir(parents=True, exist_ok=True)

# ============ 1) 원본 xlsx 파싱 ============
print("=== 1. 원본 xlsx 파싱 ===")
wb = load_workbook(ORIG, read_only=True, data_only=True)
ws = wb["Sheet"]
rows = []
for r in ws.iter_rows(min_row=2, values_only=True):
    if r and r[0]:
        차수 = (r[8] or "").strip()
        구분 = (r[9] or "").strip()
        is_dead = ("폐기" in 차수 and "폐기후제정" not in 차수) or 구분 == "폐기"
        rows.append({
            "nm": (r[0] or "").strip(),
            "abrv": (r[1] or "").strip(),
            "eng": (r[2] or "").strip(),
            "desc": (r[3] or "").strip(),
            "frmt": (r[4] or "").strip(),
            "dmn": (r[5] or "").strip(),
            "syn": (r[6] or "").strip(),
            "fbd": (r[7] or "").strip(),
            "차수": 차수,
            "구분": 구분,
            "항목": (r[10] or "").strip(),
            "사유": (r[11] or "").strip(),
            "dead": is_dead,
        })
wb.close()
print(f"  총 행 수: {len(rows)}")
print(f"  폐기 행:  {sum(1 for r in rows if r['dead'])}")

# ============ 2) 단어명 그룹 — 살아있는 단어 셋 산출 ============
from collections import defaultdict
groups = defaultdict(list)
for r in rows:
    groups[r["nm"]].append(r)

expected = {}   # 단어명 -> 살아있는 대표 행 (없으면 제외)
dead_only = []  # 모든 행이 폐기인 단어명
dup_alive = []  # 살아있는 행이 2+ 인 단어명 (드물어야 함)
for nm, lst in groups.items():
    alive = [r for r in lst if not r["dead"]]
    if not alive:
        dead_only.append(nm)
    else:
        expected[nm] = alive[0]
        if len(alive) > 1:
            dup_alive.append((nm, len(alive)))

print(f"\n  단어명 종류: {len(groups)}")
print(f"  DB에 있어야 할 단어 (살아있음): {len(expected)}")
print(f"  완전 폐기 단어 (DB에 없어야 함): {len(dead_only)} — {dead_only}")
print(f"  살아있는 행 2개 이상 단어: {len(dup_alive)} — {dup_alive[:5]}")

# ============ 3) DB tb_word(comm_stnd_yn='Y') 로드 ============
print("\n=== 3. DB tb_word(comm_stnd_yn='Y') 로드 ===")
sql = """\\COPY (SELECT
    coalesce(word_nm,''),
    coalesce(word_eng_abrv_nm,''),
    coalesce(word_eng_nm,''),
    coalesce(word_desc,''),
    coalesce(word_clsf_yn,''),
    coalesce(domain_clsf_nm,''),
    coalesce(alloph_synm_lst,''),
    coalesce(forbdn_word_lst,''),
    coalesce(comm_stnd_yn,''),
    coalesce(aprv_yn,''),
    coalesce(use_yn,'')
FROM quality.tb_word WHERE comm_stnd_yn='Y'
ORDER BY word_nm) TO STDOUT WITH (FORMAT csv, DELIMITER E'\\t');
"""
res = subprocess.run(
    ["docker", "exec", "-i", "dataq-db", "psql", "-U", "admin", "-d", "postgres",
     "-At", "-F", "\t", "-c",
     "SELECT coalesce(word_nm,'')||E'\\t'||coalesce(word_eng_abrv_nm,'')||E'\\t'||"
     "coalesce(word_eng_nm,'')||E'\\t'||coalesce(word_desc,'')||E'\\t'||"
     "coalesce(word_clsf_yn,'')||E'\\t'||coalesce(domain_clsf_nm,'')||E'\\t'||"
     "coalesce(array_to_string(alloph_synm_lst, ','), '')||E'\\t'||"
     "coalesce(array_to_string(forbdn_word_lst, ','), '') "
     "FROM quality.tb_word WHERE comm_stnd_yn='Y' ORDER BY word_nm"],
    capture_output=True, text=True, encoding="utf-8"
)
if res.returncode != 0:
    print("ERR:", res.stderr); sys.exit(1)

db_rows = {}
for ln in res.stdout.splitlines():
    p = ln.split("\t")
    if len(p) >= 8 and p[0]:
        db_rows[p[0]] = {
            "nm": p[0], "abrv": p[1], "eng": p[2], "desc": p[3],
            "frmt": p[4], "dmn": p[5], "syn": p[6], "fbd": p[7],
        }
print(f"  DB 단어 수: {len(db_rows)}")

# ============ 4) 비교 ============
print("\n=== 4. 비교 ===")
exp_set = set(expected.keys())
db_set = set(db_rows.keys())

E1_dead_in_db = sorted(set(dead_only) & db_set)         # 🔴 폐기인데 DB에 있음
E2_missing = sorted(exp_set - db_set)                    # 🟡 있어야 하는데 DB에 없음
E3_unknown_in_db = sorted(db_set - set(groups.keys()))   # 🟠 행안부에 아예 없는 단어가 DB에 있음

print(f"  🔴 폐기 단어인데 DB에 적재됨: {len(E1_dead_in_db)} — {E1_dead_in_db}")
print(f"  🟡 살아있어야 하는데 DB에 누락: {len(E2_missing)} — {E2_missing[:10]}")
print(f"  🟠 행안부 사전에 없는 단어가 DB에 있음: {len(E3_unknown_in_db)} — {E3_unknown_in_db[:10]}")

# ============ 5) 데이터 불일치 (단어명은 같은데 컬럼값 다른 경우) ============
print("\n=== 5. 컬럼값 불일치 (변경 53건 영향 확인) ===")
common = exp_set & db_set
mismatches = []
fields = [("abrv","영문약어"),("eng","영문명"),("desc","설명"),("frmt","형식단어여부"),
          ("dmn","도메인분류명"),("syn","이음동의어"),("fbd","금칙어")]
for nm in common:
    e = expected[nm]; d = db_rows[nm]
    diffs = []
    for k, label in fields:
        ev = (e[k] or "").strip()
        dv = (d[k] or "").strip()
        if ev != dv:
            diffs.append((label, ev, dv))
    if diffs:
        mismatches.append((nm, diffs))

print(f"  컬럼값 불일치 단어수: {len(mismatches)}")

# 변경 표기 단어가 그 안에 있는지
변경_nm = set(r["nm"] for r in rows if r["구분"] == "변경")
mm_set = set(m[0] for m in mismatches)
print(f"  └ 그 중 행안부에서 '변경' 표기된 단어: {len(mm_set & 변경_nm)} / 53")

# ============ 6) 산출 ============
print("\n=== 6. 산출 ===")
OUT = OUT_DIR / "_G_mois_db_verify.tsv"
with open(OUT, "w", encoding="utf-8", newline="") as f:
    w = csv.writer(f, delimiter="\t")
    w.writerow(["분류","단어명","항목","원본값","DB값","비고"])
    for nm in E1_dead_in_db:
        r = next(x for x in rows if x["nm"]==nm and x["dead"])
        w.writerow(["🔴폐기인데DB있음", nm, "", "", "(DB에 적재됨)", f"차수={r['차수']} 사유={r['사유']}"])
    for nm in E2_missing:
        e = expected[nm]
        w.writerow(["🟡누락", nm, "", e["abrv"], "(없음)", f"차수={e['차수']}"])
    for nm in E3_unknown_in_db:
        w.writerow(["🟠행안부에없음", nm, "", "(없음)", db_rows[nm]["abrv"], ""])
    for nm, diffs in sorted(mismatches):
        is_change = nm in 변경_nm
        for label, ev, dv in diffs:
            w.writerow(["🟠컬럼불일치"+("(변경표기)" if is_change else ""), nm, label, ev[:200], dv[:200], ""])
print(f"  → {OUT}")
print(f"  총 행: {1 + len(E1_dead_in_db) + len(E2_missing) + len(E3_unknown_in_db) + sum(len(d) for _,d in mismatches)}")
