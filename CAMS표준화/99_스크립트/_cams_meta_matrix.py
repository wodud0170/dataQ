"""
유형별 메타 컬럼 비교 매트릭스.

확정된 유형 메인테이블 + 부속 메타 테이블의 컬럼을 비교해서:
  - 공통 컬럼 (모든 유형이 가짐 — RG_DOCUMENT/DETAIL 차원)
  - 유형 특화 컬럼 (특정 유형에만 있음)
  - 어디에 어떤 메타가 들어있는지 매트릭스화

출력: 05_CAMS기록물유형_2026-05-21/_03_유형별_메타매트릭스.xlsx
"""
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from pathlib import Path
from collections import defaultdict

BASE = Path(r"C:\Users\장재영\Desktop\dataQ\CAMS표준화")
SCHEMA = BASE / "01_원본자료" / "CAMS_SCHEMA_원본.xlsx"
OUT = BASE / "05_CAMS기록물유형_2026-05-21" / "_03_유형별_메타매트릭스.xlsx"

# Load
wb = load_workbook(SCHEMA, read_only=True, data_only=True)
cols = []
for r in wb["컬럼정의"].iter_rows(min_row=2, values_only=True):
    if r and r[0]:
        cols.append({
            "tbl": (r[0] or "").strip(),
            "tbl_cmt": (r[1] or "").strip(),
            "col": (r[2] or "").strip().upper(),
            "col_cmt": (r[3] or "").strip(),
            "type": (r[4] or "").strip(),
            "len": r[5],
            "notnull": (r[6] or "").strip(),
            "pk": (r[8] or "").strip(),
            "fk": (r[9] or "").strip(),
        })
wb.close()

cols_by_tbl = defaultdict(list)
for c in cols:
    cols_by_tbl[c["tbl"]].append(c)

# === 확정 유형별 등록 메인 + 부속 메타 ===
# 사용자 가설 + 검증 결과 반영. W3 생산현황 영역은 별개로 제외.
TYPE_TABLES = {
    "01_일반문서_메인": {
        "main": ["RG_DOCUMENT", "RG_DETAIL"],
        "어태치": ["RG_APPENDFILE"],
        "특수파일": ["RG_DOCTUK", "RG_DETTUK"],
    },
    "02_시청각_부속": {
        "공통메타_BSID": ["SV_ARCHIVE_ADDITION_ITEM"],
        "건단위_언어": ["SV_ADDITION_LANGUAGE"],
    },
    "04_총독부_부속": {
        "메타_BSID": ["RG_OLD_GOVTINFO"],
        "목차_BSID": ["RG_OLD_GOVTINFO_CR"],
    },
    "05_정부간행물_부속": {
        "메타_BSID": ["RG_GOVTINFO"],
        "목차_BSID": ["RG_GOVTINFO_CR"],
        "복본": ["RG_GOVTMASTER"],
    },
    "06_해외기록물_별도": {
        "매체": ["RG_POSS_MEDIAINFO"],
    },
    "08_행정박물_부속": {
        "메타_BSID": ["RG_ADMIN_ADDINFO"],
        "임시": ["RG_TADMIN_ADDINFO"],
    },
    "구술기록_별도시스템": {
        "메인_자체키": ["RG_ORAL_ARCHIVE"],
        "건": ["RG_ORAL_DOCUMENT"],
        "키워드": ["RG_ORAL_KEYWORD"],
        "구성자료": ["RG_ORAL_DOCUMENT_ITEM"],
        "관련자": ["RG_ORAL_CONCERNED"],
    },
}

# === 전 컬럼명 수집 (BSID/DSID 제외, 메타만) ===
EXCLUDE_KEYS = {"BSID", "DSID", "MBSID", "TBSID"}
type_cols_map = {}   # 유형 → {컬럼명: (tbl, col_cmt, type+len, role)}
for grp, sub in TYPE_TABLES.items():
    merged = {}
    for role, tbls in sub.items():
        for tn in tbls:
            for c in cols_by_tbl.get(tn, []):
                if c["col"] in EXCLUDE_KEYS:
                    continue
                if c["col"] not in merged:  # 첫번째 등장만 기록 (충돌은 거의 없음)
                    merged[c["col"]] = {
                        "tbl": tn,
                        "cmt": c["col_cmt"],
                        "type": f"{c['type']}({c['len']})" if c['len'] else c["type"],
                        "role": role,
                        "pk": c["pk"],
                    }
    type_cols_map[grp] = merged

# === 공통 / 특화 컬럼 식별 ===
all_cols = set()
for m in type_cols_map.values():
    all_cols |= set(m.keys())

# 각 컬럼이 어느 유형에 등장하는지
col_presence = {c: [] for c in all_cols}
for grp, m in type_cols_map.items():
    for c in m.keys():
        col_presence[c].append(grp)

common_cols = sorted([c for c, g in col_presence.items() if len(g) >= len(TYPE_TABLES)-1])  # 거의 모든 유형
type_only = defaultdict(list)   # 유형 → 그 유형에만 있는 컬럼
for c, g in col_presence.items():
    if len(g) == 1:
        type_only[g[0]].append(c)

print(f"총 unique 컬럼: {len(all_cols)}")
print(f"공통 컬럼 (대부분 유형): {len(common_cols)}")
print(f"유형 특화 컬럼:")
for g, cs in type_only.items():
    print(f"  {g}: {len(cs)}")

# === xlsx ===
wb = Workbook()
hdr_font = Font(bold=True, color="FFFFFF")
hdr_fill = PatternFill("solid", fgColor="305496")
hdr_align = Alignment(horizontal="center", vertical="center")
yes_fill = PatternFill("solid", fgColor="C6E0B4")
no_fill  = PatternFill("solid", fgColor="F4B084")

def style_header(ws, ncol):
    for c in range(1, ncol+1):
        cell = ws.cell(row=1, column=c)
        cell.font = hdr_font; cell.fill = hdr_fill; cell.alignment = hdr_align

def auto_width(ws, widths):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

# 시트 1: 유형별 메인 구성 요약
ws = wb.active
ws.title = "유형별_메인구성"
ws.append(["유형", "역할", "테이블", "컬럼 수", "메타 컬럼 수 (키 제외)", "코멘트"])
for grp, sub in TYPE_TABLES.items():
    for role, tbls in sub.items():
        for tn in tbls:
            cs = cols_by_tbl.get(tn, [])
            meta_cnt = sum(1 for c in cs if c["col"] not in EXCLUDE_KEYS)
            tbl_cmt = cs[0]["tbl_cmt"] if cs else ""
            ws.append([grp, role, tn, len(cs), meta_cnt, tbl_cmt])
style_header(ws, 6)
auto_width(ws, [22, 20, 28, 10, 16, 40])
ws.freeze_panes = "A2"

# 시트 2: 유형 × 컬럼 매트릭스
grps = list(TYPE_TABLES.keys())
ws = wb.create_sheet("유형x컬럼_매트릭스")
header_row = ["컬럼명", "코멘트(대표)", "타입(대표)"] + grps
ws.append(header_row)
# 컬럼명 alphabetical
for col_nm in sorted(all_cols):
    presence = col_presence[col_nm]
    # 대표 cmt/type 가져오기 (첫번째 등장 유형)
    rep = None
    for g in grps:
        if col_nm in type_cols_map[g]:
            rep = type_cols_map[g][col_nm]
            break
    row = [col_nm, rep["cmt"] if rep else "", rep["type"] if rep else ""]
    for g in grps:
        if col_nm in type_cols_map[g]:
            info = type_cols_map[g][col_nm]
            mark = "●" + (f"({info['tbl']})" if len(presence) <= 3 else "")
            row.append(mark)
        else:
            row.append("")
    ws.append(row)
    # 색상
    last = ws.max_row
    for i, g in enumerate(grps):
        cell = ws.cell(row=last, column=4+i)
        if cell.value:
            cell.fill = yes_fill
style_header(ws, len(header_row))
widths = [22, 35, 16] + [16]*len(grps)
auto_width(ws, widths)
ws.freeze_panes = "D2"

# 시트 3: 유형 특화 컬럼만
ws = wb.create_sheet("유형별_특화컬럼")
ws.append(["유형", "컬럼명", "테이블", "코멘트", "타입"])
for g, cs in type_only.items():
    for col_nm in sorted(cs):
        info = type_cols_map[g][col_nm]
        ws.append([g, col_nm, info["tbl"], info["cmt"], info["type"]])
style_header(ws, 5)
auto_width(ws, [22, 25, 28, 40, 15])
ws.freeze_panes = "A2"

# 시트 4: 공통 컬럼 (대부분 유형 보유)
ws = wb.create_sheet("공통컬럼_대부분유형")
ws.append(["컬럼명", "등장 유형 수", "코멘트"])
for col_nm in sorted(common_cols, key=lambda x: -len(col_presence[x])):
    rep = None
    for g in grps:
        if col_nm in type_cols_map[g]:
            rep = type_cols_map[g][col_nm]
            break
    ws.append([col_nm, len(col_presence[col_nm]), rep["cmt"] if rep else ""])
style_header(ws, 3)
auto_width(ws, [25, 14, 40])

# 시트 5: 통계 요약
ws = wb.create_sheet("통계")
ws.append(["항목", "값", "비고"])
ws.append(["전체 unique 컬럼 (메타만, 키 제외)", len(all_cols), ""])
ws.append(["공통 컬럼 (거의 모든 유형)", len(common_cols), "통합 시 RG_DOCUMENT/DETAIL에 흡수"])
total_specialized = sum(len(v) for v in type_only.values())
ws.append(["유형 특화 컬럼 총합", total_specialized, ""])
ws.append([])
ws.append(["[유형별 특화 컬럼 수]", "", ""])
for g, cs in sorted(type_only.items(), key=lambda x: -len(x[1])):
    ws.append([f"  {g}", len(cs), ""])
style_header(ws, 3)
auto_width(ws, [40, 12, 40])

wb.save(OUT)
print(f"\nSAVED: {OUT}")
