# -*- coding: utf-8 -*-
"""
RAMP 컬럼 BEFORE/AFTER 통합 매핑정의서 (메인 산출).
DB 의 모든 변환 결과 (단어 + 도메인 + R8 + 시간 + 대문자 + Oracle 타입) 반영.
"""
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from pathlib import Path
from collections import defaultdict, Counter
import subprocess, csv, io

BASE = Path(r"C:\Users\장재영\Desktop\dataQ\CAMS표준화")
RAMP_SCHEMA = BASE / "CAMS_RAMP_통합" / "ramp기관스키마정보.xlsx"
OUT = BASE / "04_RAMP분석_2026-05-21" / "00_핵심산출" / "RAMP_컬럼_매핑정의서_2026-05-26.xlsx"

# ============ 1. BEFORE (원본 RAMP) 로드 ============
print("=== BEFORE (원본 RAMP) ===")
wb = load_workbook(RAMP_SCHEMA, read_only=True, data_only=True)
ws = wb["컬럼"]
before = {}  # (tbl, col_en_lower) → {kr, dtype, dlen, pk, ...}
seen = set()
for r in ws.iter_rows(min_row=2, values_only=True):
    if r and r[0] and r[1]:
        key = (str(r[0]).strip(), str(r[1]).strip().lower())
        if key in seen: continue
        seen.add(key)
        before[key] = {
            "col_en": str(r[1]).strip(),
            "col_kr": str(r[2] or "").strip(),
            "dtype":  str(r[5] or "").strip(),
            "dlen":   str(r[6] or "").strip(),
            "pk":     str(r[7] or "").strip(),
        }
wb.close()
print(f"  BEFORE 컬럼: {len(before)}")

# ============ 2. AFTER (DB 통합 적재 결과) 로드 ============
print("\n=== AFTER (DB 통합) ===")
r = subprocess.run(["docker","exec","-i","dataq-db","psql","-U","admin","-d","postgres",
    "-c", """COPY (SELECT a.obj_nm, a.attr_nm, coalesce(a.attr_nm_kr,''),
             a.data_type, coalesce(a.data_len::text,''), coalesce(a.data_decimal_len::text,'0'),
             coalesce(a.nullable_yn,'Y'), coalesce(a.pk_yn,'N'), coalesce(a.fk_yn,'N'),
             coalesce(t.domain_nm,'')
             FROM quality.tb_data_model_attr a
             LEFT JOIN quality.tb_terms t ON t.terms_nm=a.attr_nm_kr AND t.terms_eng_abrv_nm=a.attr_nm
             WHERE a.dm_id=(SELECT dm_id FROM quality.tb_data_model WHERE dm_nm='RAMP' LIMIT 1)
             ORDER BY a.obj_nm, a.attr_ord) TO STDOUT WITH (FORMAT csv, FORCE_QUOTE *)"""],
    capture_output=True, encoding="utf-8")
after_rows = []
for row in csv.reader(io.StringIO(r.stdout)):
    if row:
        after_rows.append({
            "tbl": row[0], "col_en": row[1], "col_kr": row[2],
            "dtype": row[3], "dlen": row[4], "ddec": row[5],
            "null": row[6], "pk": row[7], "fk": row[8],
            "dom": row[9],
        })
print(f"  AFTER 컬럼: {len(after_rows)}")

# ============ 3. 매핑 결합 + 변경 유형 산출 ============
print("\n=== 매핑 결합 ===")
rows = []
no_match = 0
for a in after_rows:
    key = (a["tbl"], a["col_en"].lower())
    b = before.get(key)
    if not b:
        no_match += 1
        b = {"col_en": "", "col_kr": "", "dtype": "", "dlen": "", "pk": ""}
    types = []
    if b["col_en"] != a["col_en"]: types.append("EN")
    if b["col_kr"] != a["col_kr"]: types.append("KR")
    if b["dtype"].upper() != a["dtype"].upper(): types.append("TYPE")
    if str(b["dlen"]) != str(a["dlen"]): types.append("LEN")
    if not types: types = ["KEEP"]
    rows.append({**a, "bef_en": b["col_en"], "bef_kr": b["col_kr"],
                 "bef_dtype": b["dtype"], "bef_dlen": b["dlen"], "bef_pk": b["pk"],
                 "chg_type": "+".join(types)})
print(f"  매칭 안 됨 (BEFORE 없음): {no_match}")

# 통계
type_cnt = Counter(r["chg_type"] for r in rows)
print(f"  변경 유형 분포: {dict(type_cnt)}")
# 대문자 검증
upper_check = sum(1 for r in rows if r["col_en"] == r["col_en"].upper())
print(f"  AFTER 영문 대문자: {upper_check}/{len(rows)}")

# ============ 4. xlsx 생성 ============
print("\n=== xlsx 생성 ===")
wb = Workbook()
THIN = Side(border_style="thin", color="808080")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
HDR_FILL = PatternFill("solid", fgColor="1F4E78")
HDR_FONT = Font(name="맑은 고딕", size=10, color="FFFFFF", bold=True)
CELL_FONT = Font(name="맑은 고딕", size=10)
SECTION_FILL = PatternFill("solid", fgColor="D9E1F2")
EN_FILL = PatternFill("solid", fgColor="FFE699")
KR_FILL = PatternFill("solid", fgColor="C6E0B4")
TYPE_FILL = PatternFill("solid", fgColor="F8CBAD")
KEEP_FILL = PatternFill("solid", fgColor="F2F2F2")

# 표지
ws = wb.active; ws.title = "표지"
ws.column_dimensions["A"].width = 24; ws.column_dimensions["B"].width = 100
ws.merge_cells("A1:B1")
t = ws.cell(row=1, column=1, value="RAMP 컬럼 BEFORE/AFTER 매핑정의서 (통합)")
t.font = Font(name="맑은 고딕", size=18, bold=True, color="1F4E78")
t.alignment = Alignment(horizontal="center", vertical="center")
ws.row_dimensions[1].height = 36

def cr(r, label, value):
    ws.cell(row=r, column=1, value=label).fill = SECTION_FILL
    ws.cell(row=r, column=1).font = Font(name="맑은 고딕", size=10, bold=True)
    ws.cell(row=r, column=1).border = BORDER
    c = ws.cell(row=r, column=2, value=value); c.font = Font(name="맑은 고딕", size=11)
    c.border = BORDER; c.alignment = Alignment(vertical="center", wrap_text=True)

cr(3, "작성일", "2026-05-26")
cr(4, "대상", f"RAMP 컬럼 {len(rows)}건 (DB tb_data_model_attr 통합 상태)")
cr(5, "BEFORE", "RAMP 원본 (Cubrid, ramp기관스키마정보.xlsx)")
cr(6, "AFTER", "Oracle 적재 기준 — 단어+도메인+R8+시간+대문자+Oracle타입 모두 반영")
cr(7, "변경 유형", "EN(영문) / KR(한글) / TYPE(타입) / LEN(길이) / KEEP(변경 없음)")
cr(8, "정책", "행안부 우선 / 최장 도메인 / R8 형식단어 보충 / Cubrid→Oracle 타입(STRING→VARCHAR, NUMERIC/INTEGER→NUMBER, DATETIME→TIMESTAMP) / 모든 영문 대문자")

# 요약
ws2 = wb.create_sheet("요약")
ws2.column_dimensions["A"].width = 30
ws2.column_dimensions["B"].width = 12
ws2.column_dimensions["C"].width = 60
t = ws2.cell(row=1, column=1, value="변경 유형별 통계")
t.font = Font(name="맑은 고딕", size=14, bold=True, color="1F4E78")
ws2.merge_cells("A1:C1")
rr = 3
for col, h in enumerate(["변경 유형","건수","설명"], 1):
    cell = ws2.cell(row=rr, column=col, value=h); cell.fill = HDR_FILL; cell.font = HDR_FONT
    cell.alignment = Alignment(horizontal="center"); cell.border = BORDER
rr += 1
explains = {
    "KEEP": "변경 없음",
    "EN": "영문만 변경 (단어 또는 대문자화)",
    "KR": "한글만 변경 (D5 R8 보충 또는 D3 suffix)",
    "TYPE": "데이터타입 변경 (Cubrid→Oracle)",
    "LEN": "길이 변경 (D1 최장)",
    "EN+KR": "영문+한글",
    "EN+TYPE": "영문+타입",
    "EN+LEN": "영문+길이",
    "KR+TYPE": "한글+타입",
    "TYPE+LEN": "타입+길이",
    "EN+KR+TYPE": "영문+한글+타입",
    "EN+KR+LEN": "영문+한글+길이",
    "EN+TYPE+LEN": "영문+타입+길이",
    "KR+TYPE+LEN": "한글+타입+길이",
    "EN+KR+TYPE+LEN": "모두 변경",
}
for ct, n in type_cnt.most_common():
    ws2.cell(row=rr, column=1, value=ct).border = BORDER
    ws2.cell(row=rr, column=2, value=n).border = BORDER
    ws2.cell(row=rr, column=2).alignment = Alignment(horizontal="right")
    ws2.cell(row=rr, column=3, value=explains.get(ct, "")).border = BORDER
    rr += 1
rr += 1
ws2.cell(row=rr, column=1, value="AFTER 영문 대문자 일치").border = BORDER
ws2.cell(row=rr, column=2, value=f"{upper_check}/{len(rows)}").border = BORDER

# 매핑정의서 메인
ws3 = wb.create_sheet("매핑정의서")
H = ["No","테이블","BEFORE 영문","AFTER 영문","BEFORE 한글","AFTER 한글",
     "BEFORE 타입","BEFORE 길이","AFTER 타입","AFTER 길이","AFTER 소수점",
     "도메인(AFTER)","PK","FK","NULL","변경유형"]
W = [5,26,24,24,28,28,12,8,12,8,8,18,4,4,5,16]
for i, w in enumerate(W, 1): ws3.column_dimensions[get_column_letter(i)].width = w
for i, h in enumerate(H, 1):
    c = ws3.cell(row=1, column=i, value=h); c.fill = HDR_FILL; c.font = HDR_FONT
    c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True); c.border = BORDER
ws3.row_dimensions[1].height = 30
ws3.freeze_panes = "C2"

for i, r in enumerate(rows, 1):
    rn = i + 1
    vals = [i, r["tbl"], r["bef_en"], r["col_en"], r["bef_kr"], r["col_kr"],
            r["bef_dtype"], r["bef_dlen"], r["dtype"], r["dlen"], r["ddec"],
            r["dom"], r["pk"], r["fk"], r["null"], r["chg_type"]]
    for j, v in enumerate(vals, 1):
        c = ws3.cell(row=rn, column=j, value=v); c.font = CELL_FONT; c.border = BORDER
        c.alignment = Alignment(vertical="top", wrap_text=True)
    # 색 표시
    if r["chg_type"] == "KEEP":
        ws3.cell(row=rn, column=16).fill = KEEP_FILL
    if "EN" in r["chg_type"]:
        ws3.cell(row=rn, column=3).fill = EN_FILL
        ws3.cell(row=rn, column=4).fill = EN_FILL
    if "KR" in r["chg_type"]:
        ws3.cell(row=rn, column=5).fill = KR_FILL
        ws3.cell(row=rn, column=6).fill = KR_FILL
    if "TYPE" in r["chg_type"] or "LEN" in r["chg_type"]:
        ws3.cell(row=rn, column=7).fill = TYPE_FILL
        ws3.cell(row=rn, column=8).fill = TYPE_FILL
        ws3.cell(row=rn, column=9).fill = TYPE_FILL
        ws3.cell(row=rn, column=10).fill = TYPE_FILL

wb.save(OUT)
print(f"\n→ {OUT}")
print(f"  시트: 표지 / 요약 / 매핑정의서 ({len(rows)}행)")
