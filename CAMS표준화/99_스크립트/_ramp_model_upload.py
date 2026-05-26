# -*- coding: utf-8 -*-
"""
RAMP 스키마를 dataQ 데이터모델로 업로드.
  - tb_data_model: 모델 1건 (RAMP)
  - tb_data_model_obj: 테이블 365건
  - tb_data_model_attr: 컬럼 5,740건
"""
from openpyxl import load_workbook
from pathlib import Path
from collections import defaultdict
import subprocess, os, base64, sys

BASE = Path(r"C:\Users\장재영\Desktop\dataQ\CAMS표준화")
RAMP_SCHEMA = BASE / "CAMS_RAMP_통합" / "ramp기관스키마정보.xlsx"
OUT_SQL = BASE / "04_RAMP분석_2026-05-21" / "99_실행SQL_기록" / "Phase2_RAMP_model_upload.sql"

# ============ 1. RAMP 데이터 로드 ============
print("=== 1. RAMP 스키마 로드 ===")
wb = load_workbook(RAMP_SCHEMA, read_only=True, data_only=True)

# 테이블 목록
ws_t = wb["테이블 목록"] if "테이블 목록" in wb.sheetnames else wb["테이블목록"]
tables = []
for r in ws_t.iter_rows(min_row=2, values_only=True):
    if r and r[0]:
        tables.append({
            "tbl": str(r[0]).strip(),
            "cmt": str(r[1] or "").strip() if len(r) > 1 else "",
        })
print(f"  테이블: {len(tables)}")

# 컬럼
ws_c = wb["컬럼"]
cols = []
seen = set()
for r in ws_c.iter_rows(min_row=2, values_only=True):
    if r and r[0] and r[1]:
        key = (str(r[0]).strip(), str(r[1]).strip())
        if key in seen: continue
        seen.add(key)
        cols.append({
            "tbl": str(r[0]).strip(),
            "col_en": str(r[1]).strip(),
            "col_kr": str(r[2] or "").strip(),
            "desc": str(r[3] or "").strip() if len(r) > 3 else "",
            "null": str(r[4] or "").strip().upper() if len(r) > 4 else "",
            "dtype": str(r[5] or "").strip().upper() if len(r) > 5 else "",
            "dlen": str(r[6] or "").strip() if len(r) > 6 else "",
            "pk": str(r[7] or "").strip() if len(r) > 7 else "",
        })
wb.close()
print(f"  컬럼: {len(cols)}")

# ============ AFTER 적용 (Phase 2-1 + D3 + D5) ============
print("\n=== AFTER 한글/영문 적용 ===")
ROOT = BASE / "04_RAMP분석_2026-05-21" / "00_핵심산출"
col_after = {}  # (tbl, col_en) → (en_AFTER, kr_AFTER)
wb = load_workbook(ROOT/"RAMP_컬럼_BEFORE_AFTER_2026-05-23.xlsx", read_only=True, data_only=True)
for r in wb["컬럼BEFORE_AFTER"].iter_rows(min_row=2, values_only=True):
    if r and r[1] and r[2]:
        col_after[(str(r[1]), str(r[2]))] = (str(r[3] or r[2]), str(r[5] or r[4] or ""))
wb.close()
time_after = {}
wb = load_workbook(ROOT/"Phase2_시간컬럼_BEFORE_AFTER_2026-05-23.xlsx", read_only=True, data_only=True)
for r in wb["적용_BEFORE_AFTER"].iter_rows(min_row=2, values_only=True):
    if r and r[1] and r[2]:
        time_after[(str(r[1]), str(r[2]))] = str(r[4] or r[3] or "")
wb.close()
d5_kr = {}
wb = load_workbook(ROOT/"Phase2_D5_R8형식단어_2026-05-23.xlsx", read_only=True, data_only=True)
for r in wb["새용어_BEFORE_AFTER"].iter_rows(min_row=2, values_only=True):
    if r and r[1] and r[3]: d5_kr[str(r[1])] = str(r[3])
wb.close()
print(f"  Phase2-1 변환: {len(col_after)}  D3 시간: {len(time_after)}  D5 R8: {len(d5_kr)}")

for c in cols:
    key = (c["tbl"], c["col_en"])
    bef_kr = c["col_kr"]
    # Phase 2-1
    aft_en, aft_kr = col_after.get(key, (c["col_en"], bef_kr))
    if not aft_kr: aft_kr = bef_kr
    # D5 보충
    if aft_kr in d5_kr: aft_kr = d5_kr[aft_kr]
    # D3 시간
    if key in time_after:
        t_kr = time_after[key]
        if t_kr: aft_kr = t_kr
    # 정규화
    aft_kr = aft_kr.replace(" ", "")
    aft_en = aft_en.upper()
    c["col_en"] = aft_en
    c["col_kr"] = aft_kr

# 컬럼 순번 (테이블별)
attr_ord = defaultdict(int)

# ============ 2. SQL 생성 ============
print("\n=== 2. SQL 생성 ===")
def gid(): return base64.urlsafe_b64encode(os.urandom(17)).rstrip(b'=').decode()[:22]

DM_ID = gid()
DM_NM = "RAMP"
OWNER = "RAMP"

def esc(s):
    if s is None: return ""
    return str(s).replace("'", "''")

sql = [
    "-- RAMP 데이터모델 업로드",
    "BEGIN;",
    "",
    "-- 1) tb_data_model",
    f"INSERT INTO quality.tb_data_model (dm_id, dm_nm, ver, model_type, cret_dt, cret_user_id, use_yn) "
    f"VALUES ('{DM_ID}','{DM_NM}','1.0','RDB',to_char(now(),'YYYYMMDDHH24MISS'),'admin','Y');",
    "",
    f"-- 2) tb_data_model_obj ({len(tables)}건)",
]
for t in tables:
    sql.append(
        f"INSERT INTO quality.tb_data_model_obj (dm_id, obj_nm, obj_nm_kr, obj_owner, obj_comment, use_yn) "
        f"VALUES ('{DM_ID}','{esc(t['tbl'])}','{esc(t['cmt'])}','{OWNER}','{esc(t['cmt'])}','Y');"
    )

sql.append("")
sql.append(f"-- 3) tb_data_model_attr ({len(cols)}건)")
for c in cols:
    attr_ord[c["tbl"]] += 1
    ord_n = attr_ord[c["tbl"]]
    nullable = "N" if c["null"] in ("N","NO","NOT NULL") else "Y"
    pk = "Y" if c["pk"] and c["pk"].upper() in ("Y","PK","TRUE","1") else "N"
    dlen = c["dlen"] if c["dlen"].isdigit() else "0"
    sql.append(
        f"INSERT INTO quality.tb_data_model_attr (dm_id, obj_nm, attr_nm, attr_nm_kr, data_type, data_len, "
        f"nullable_yn, pk_yn, fk_yn, attr_ord, obj_owner, attr_comment, use_yn) "
        f"VALUES ('{DM_ID}','{esc(c['tbl'])}','{esc(c['col_en'])}','{esc(c['col_kr'])}','{esc(c['dtype'])}',{dlen},"
        f"'{nullable}','{pk}','N',{ord_n},'{OWNER}','{esc(c['desc'])}','Y');"
    )

sql += ["",
        f"SELECT count(*) AS model FROM quality.tb_data_model WHERE dm_id='{DM_ID}';",
        f"SELECT count(*) AS objs FROM quality.tb_data_model_obj WHERE dm_id='{DM_ID}';",
        f"SELECT count(*) AS attrs FROM quality.tb_data_model_attr WHERE dm_id='{DM_ID}';",
        "",
        "COMMIT;"]

OUT_SQL.write_text("\n".join(sql), encoding="utf-8")
print(f"  → {OUT_SQL}")
print(f"  dm_id: {DM_ID}")
print(f"  INSERT: 1 model + {len(tables)} obj + {len(cols)} attr")
