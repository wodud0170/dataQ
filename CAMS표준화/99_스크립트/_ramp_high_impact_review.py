# -*- coding: utf-8 -*-
"""
영향 큰 단어 변환 검토 xlsx 산출 — 회의용.
Phase 2-1 결과의 영문 변경 단어 108종 전수 + 메타 정보.

시트:
  1. 표지
  2. 요약 (영향 구간별)
  3. 단어별 검토 (108행, 영향 컬럼수 내림차순)
  4. 사용 컬럼 샘플 (Top 영향 단어들의 컬럼 일부)
"""
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from pathlib import Path
from collections import Counter, defaultdict
import re

BASE = Path(r"C:\Users\장재영\Desktop\dataQ\CAMS표준화")
BA = BASE / "04_RAMP분석_2026-05-21" / "RAMP_컬럼_BEFORE_AFTER_2026-05-23.xlsx"
RAMP_DICT = BASE / "01_원본자료" / "RMS4-DE14-04.데이터표준화_단어사전-V2.0.xlsx"
MOIS_DICT = BASE.parent / "q-center" / "src" / "main" / "resources" / "seed" / "행안부_공통표준" / "행정안전부_공공데이터 공통표준단어.xlsx"
OUT = BASE / "04_RAMP분석_2026-05-21" / "RAMP_단어변환_영향검토_2026-05-23.xlsx"

# ============ 변환 정보 + 영향 컬럼 ============
print("=== 로드 ===")
wb = load_workbook(BA, read_only=True, data_only=True)
ws = wb["컬럼BEFORE_AFTER"]
# 변환별 영향 컬럼 모음
word_cols = defaultdict(list)  # before_abrv → [(tbl, col_en, col_kr, type, len)]
for r in ws.iter_rows(min_row=2, values_only=True):
    if r and r[0] is not None:
        sig = r[7] or ""  # 영문 변경 단어
        if sig:
            # 단일 또는 복합 (예: "A→B; C→D")
            for s in sig.split("; "):
                if "→" in s:
                    bef = s.split("→")[0].upper()
                    word_cols[bef].append((r[1], r[2], r[4], r[9], r[10]))  # tbl, en, kr, dtype, dlen
wb.close()
print(f"  영향 단어: {len(word_cols)}")

# 변환 매핑 (영문 변경 단어 시트)
wb = load_workbook(BA, read_only=True, data_only=True)
ws = wb["변경단어매핑_영문"]
word_change = []  # [(sig, n, reason)]
for r in ws.iter_rows(min_row=2, values_only=True):
    if r and r[0] is not None:
        word_change.append((r[1], r[2], r[3] or ""))
wb.close()

# RAMP 사전
ramp_meta = {}  # abrv → {nm, eng, desc}
wb = load_workbook(RAMP_DICT, read_only=True, data_only=True)
for r in wb["단어사전"].iter_rows(min_row=2, values_only=True):
    if r and r[2] and r[3]:
        abrv = str(r[3]).strip().upper()
        ramp_meta[abrv] = {
            "nm": str(r[2]).strip(),
            "eng": str(r[4] or "").strip().replace("_x000D_",""),
            "desc": str(r[5] or "").strip().replace("_x000D_",""),
        }
wb.close()

# MOIS 사전
mois_meta = {}
wb = load_workbook(MOIS_DICT, read_only=True, data_only=True)
for r in wb["Sheet"].iter_rows(min_row=2, values_only=True):
    if r and r[0] and r[1]:
        abrv = str(r[1]).strip().upper()
        mois_meta[abrv] = {
            "nm": str(r[0]).strip(),
            "eng": str(r[2] or "").strip(),
            "desc": str(r[3] or "").strip(),
        }
wb.close()

# ============ xlsx 생성 ============
print("\n=== xlsx 생성 ===")
wb = Workbook()
THIN = Side(border_style="thin", color="808080")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
HDR_FILL = PatternFill("solid", fgColor="1F4E78")
HDR_FONT = Font(name="맑은 고딕", size=10, color="FFFFFF", bold=True)
CELL_FONT = Font(name="맑은 고딕", size=10)
SECTION_FILL = PatternFill("solid", fgColor="D9E1F2")
DECISION_FILL = PatternFill("solid", fgColor="FFF2CC")
HIGH_FILL = PatternFill("solid", fgColor="F8CBAD")  # 100+
MID_FILL = PatternFill("solid", fgColor="FFE699")   # 50~99
LOW_FILL = PatternFill("solid", fgColor="C6E0B4")   # 30~49

def impact_level(n):
    if n >= 100: return ("높음", HIGH_FILL)
    if n >= 50:  return ("중간", MID_FILL)
    if n >= 30:  return ("낮음", LOW_FILL)
    return ("미미", None)

# 시트 1: 표지
ws = wb.active; ws.title = "표지"
ws.column_dimensions["A"].width = 22; ws.column_dimensions["B"].width = 100
ws.merge_cells("A1:B1")
t = ws.cell(row=1, column=1, value="RAMP 단어 변환 — 영향 컬럼 검토 (회의용)")
t.font = Font(name="맑은 고딕", size=18, bold=True, color="1F4E78")
t.alignment = Alignment(horizontal="center", vertical="center")
ws.row_dimensions[1].height = 36

def cr(r, label, value):
    ws.cell(row=r, column=1, value=label).fill = SECTION_FILL
    ws.cell(row=r, column=1).font = Font(name="맑은 고딕", size=10, bold=True)
    ws.cell(row=r, column=1).border = BORDER
    c = ws.cell(row=r, column=2, value=value); c.font = Font(name="맑은 고딕", size=11)
    c.border = BORDER; c.alignment = Alignment(vertical="center", wrap_text=True)

cr(3, "작성일", "2026-05-23")
cr(4, "작성자", "장재영")
cr(5, "대상", f"Phase 2-1 영문 변경 단어 {len(word_change)}종")
cr(6, "기본 방향", "행안부 우선 (고객 요구사항·팀 방향)")
cr(7, "검토 목적", "영향 컬럼이 큰 단어는 변경 비용 대비 행안부 흡수 이득을 평가. RAMP 우선(보류) 결정 시 B-3 옵션 A 패턴 적용")
cr(8, "결정 라인", "[흡수] 행안부 약어로 변경 (기본) / [보류] RAMP 약어 유지 (FLS·기록물철 패턴)")
ws.row_dimensions[9].height = 8
cr(10, "영향 — 높음 (100+)", "변경 비용 큼. RAMP 우선 검토 필요 (5종 예상)")
cr(11, "영향 — 중간 (50~99)", "기본 흡수, 단 운영 영향 확인 (6종 내외)")
cr(12, "영향 — 낮음 (30~49)", "기본 흡수 (8종 내외)")
cr(13, "영향 — 미미 (<30)", "자동 흡수 (89종 내외)")
ws.row_dimensions[14].height = 8
cr(15, "이미 보류 안건", "기록물철 (FLS→RCDSF, 254컬럼) — Phase1_단어결정_결과.md 참조")
cr(16, "B-3 처리 (확정)", "순번/SEQ, 전자/ELCT, 질의/SQL — 사용자 결정")

# 시트 2: 요약
ws2 = wb.create_sheet("요약")
ws2.column_dimensions["A"].width = 22; ws2.column_dimensions["B"].width = 10
ws2.column_dimensions["C"].width = 12; ws2.column_dimensions["D"].width = 55
t = ws2.cell(row=1, column=1, value="영향 구간별 분포"); t.font = Font(name="맑은 고딕", size=14, bold=True, color="1F4E78")
ws2.merge_cells("A1:D1")
r = 3
for c, h in enumerate(["구간","단어 종수","컬럼 합계","비고"], 1):
    cell = ws2.cell(row=r, column=c, value=h); cell.fill = HDR_FILL; cell.font = HDR_FONT
    cell.alignment = Alignment(horizontal="center"); cell.border = BORDER
r += 1
buckets = {"100+ (높음)": 0, "50~99 (중간)": 0, "30~49 (낮음)": 0, "<30 (미미)": 0}
col_buckets = {k:0 for k in buckets}
for sig, n, _ in word_change:
    if n >= 100: k = "100+ (높음)"
    elif n >= 50: k = "50~99 (중간)"
    elif n >= 30: k = "30~49 (낮음)"
    else: k = "<30 (미미)"
    buckets[k] += 1; col_buckets[k] += n
for k, v in buckets.items():
    ws2.cell(row=r, column=1, value=k).border = BORDER
    ws2.cell(row=r, column=2, value=v).border = BORDER
    ws2.cell(row=r, column=2).alignment = Alignment(horizontal="right")
    ws2.cell(row=r, column=3, value=col_buckets[k]).border = BORDER
    ws2.cell(row=r, column=3).alignment = Alignment(horizontal="right")
    ws2.cell(row=r, column=4, value="" if "미미" in k else ("검토 권장" if "높음" in k else "")).border = BORDER
    r += 1

# 시트 3: 단어별 검토
ws3 = wb.create_sheet("단어별검토")
HEADERS = ["순위","영향","변환","BEFORE 약어","RAMP 한글","RAMP 영문명","RAMP 설명",
           "AFTER 약어","MOIS 한글","MOIS 영문명","MOIS 설명",
           "영향 컬럼수","영향 테이블수","사유","결정","결정사유"]
WIDTHS = [5,8,18,12,16,22,46,12,16,22,46,10,10,28,12,28]
for i, w in enumerate(WIDTHS, 1):
    ws3.column_dimensions[get_column_letter(i)].width = w
for i, h in enumerate(HEADERS, 1):
    c = ws3.cell(row=1, column=i, value=h); c.fill = HDR_FILL; c.font = HDR_FONT
    c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True); c.border = BORDER
ws3.row_dimensions[1].height = 32
ws3.freeze_panes = "D2"

for i, (sig, n, reason) in enumerate(word_change, 1):
    row = i + 1
    bef_aft = sig.split("→") if "→" in sig else [sig, ""]
    bef = bef_aft[0].upper(); aft = bef_aft[1].upper() if len(bef_aft)>1 else ""
    rmeta = ramp_meta.get(bef, {})
    mmeta = mois_meta.get(aft, {})
    impact, fill = impact_level(n)
    tbl_cnt = len(set(c[0] for c in word_cols.get(bef, [])))
    values = [i, impact, sig, bef, rmeta.get("nm",""), rmeta.get("eng",""), rmeta.get("desc",""),
              aft, mmeta.get("nm",""), mmeta.get("eng",""), mmeta.get("desc",""),
              n, tbl_cnt, reason, "", ""]
    for j, v in enumerate(values, 1):
        c = ws3.cell(row=row, column=j, value=v); c.font = CELL_FONT; c.border = BORDER
        c.alignment = Alignment(vertical="top", wrap_text=True)
    if fill:
        ws3.cell(row=row, column=2).fill = fill
    ws3.cell(row=row, column=15).fill = DECISION_FILL
    ws3.cell(row=row, column=16).fill = DECISION_FILL

# 결정 드롭다운
dv = DataValidation(type="list", formula1='"흡수,보류,Skip"', allow_blank=True)
dv.add(f"O2:O{len(word_change)+1}")
ws3.add_data_validation(dv)

# 시트 4: 사용 컬럼 상세 (Top 영향 단어들의 컬럼)
ws4 = wb.create_sheet("사용컬럼상세")
H4 = ["BEFORE 약어","RAMP 한글","영향","변환","테이블","BEFORE 영문","BEFORE 한글","타입","길이"]
W4 = [12,14,8,18,24,28,28,10,8]
for i, w in enumerate(W4, 1): ws4.column_dimensions[get_column_letter(i)].width = w
for i, h in enumerate(H4, 1):
    c = ws4.cell(row=1, column=i, value=h); c.fill = HDR_FILL; c.font = HDR_FONT
    c.alignment = Alignment(horizontal="center"); c.border = BORDER
ws4.freeze_panes = "A2"
r = 2
# 영향 30+ 단어들만 상세 (4 시트 너무 커지지 않게)
TOP_THRESHOLD = 30
top_words = [(sig.split("→")[0].upper(), sig, n) for sig, n, _ in word_change if n >= TOP_THRESHOLD]
print(f"  Top 상세 단어 (30+): {len(top_words)}")
for bef, sig, n in top_words:
    rmeta = ramp_meta.get(bef, {})
    impact, _ = impact_level(n)
    for col in word_cols.get(bef, []):
        vals = [bef, rmeta.get("nm",""), impact, sig, col[0], col[1], col[2], col[3], col[4]]
        for j, v in enumerate(vals, 1):
            c = ws4.cell(row=r, column=j, value=v); c.font = CELL_FONT; c.border = BORDER
            c.alignment = Alignment(vertical="top", wrap_text=True)
        r += 1

wb.save(OUT)
print(f"\n→ {OUT}")
print(f"  시트: 표지 / 요약 / 단어별검토 ({len(word_change)}) / 사용컬럼상세 ({r-2})")
print(f"\n=== 영향 구간 분포 ===")
for k, v in buckets.items():
    print(f"  {k:15s}: 단어 {v:>4}종 / 컬럼 {col_buckets[k]:>5}건")
