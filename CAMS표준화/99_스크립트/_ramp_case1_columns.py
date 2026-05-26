"""
Case 1 (한글동일/영문약어다름) 170건 — RAMP 스키마에서 변경 영향받는 테이블·컬럼 전수 추출.
RAMP 영문약어 X 토큰을 가진 컬럼 → MOIS 영문약어 Y 로 치환한 신규 컬럼명까지 산출.

출력:
  CAMS표준화/RAMP_Case1_컬럼변경_2026-05-21.xlsx
    시트1: 요약 (단어별 영향 컬럼 수)
    시트2: 변경 컬럼 전수
    시트3: 변경 테이블 (DISTINCT)
"""
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from pathlib import Path
from collections import defaultdict, Counter
import re

BASE = Path(r"C:\Users\장재영\Desktop\dataQ\CAMS표준화")
DICT = BASE / "01_원본자료" / "RMS4-DE14-04.데이터표준화_단어사전-V2.0.xlsx"
SCHEMA = BASE / "CAMS_RAMP_통합" / "ramp기관스키마정보.xlsx"
MOIS_WORD = BASE.parent / "q-center" / "src" / "main" / "resources" / "seed" / "행안부_공통표준" / "행정안전부_공공데이터 공통표준단어.xlsx"
OUT = BASE / "04_RAMP분석_2026-05-21" / "RAMP_Case1_컬럼변경_2026-05-21.xlsx"

# === RAMP 사전 ===
wb = load_workbook(DICT, read_only=True, data_only=True)
ws = wb["단어사전"]
ramp = []
for r in ws.iter_rows(min_row=2, values_only=True):
    if not r or not r[2]: continue
    ramp.append({"단어명": (r[2] or "").strip(),
                 "영문약어": (r[3] or "").strip().upper(),
                 "영문명": (r[4] or "").strip()})
wb.close()

# === MOIS 사전 ===
wb = load_workbook(MOIS_WORD, read_only=True, data_only=True)
ws = wb[wb.sheetnames[0]]
head = list(next(ws.iter_rows(min_row=1, max_row=1, values_only=True)))
def col_idx(patterns):
    for i, h in enumerate(head):
        for p in patterns:
            if p in str(h or ""): return i
    return -1
i_nm = col_idx(["공통표준단어명", "단어명"])
i_ab = col_idx(["영문약어"])
i_en = col_idx(["영문명"])
mois = []
for r in ws.iter_rows(min_row=2, values_only=True):
    if not r or not r[i_nm]: continue
    mois.append({"단어명": str(r[i_nm]).strip(),
                 "영문약어": str(r[i_ab] or "").strip().upper(),
                 "영문명": str(r[i_en] or "").strip()})
wb.close()

mois_by_abbr = {m["영문약어"]: m for m in mois if m["영문약어"]}
mois_by_name = defaultdict(list)  # 한글명 → MOIS 후보들
for m in mois:
    mois_by_name[m["단어명"]].append(m)

mois_abbrs = set(mois_by_abbr.keys())
mois_names = set(mois_by_name.keys())

# === Case 1 식별 — 한글이 MOIS에 있고 RAMP 영문약어가 MOIS에 없음 ===
case1 = []
for d in ramp:
    a = d["영문약어"]; n = d["단어명"]
    if not a: continue
    if a in mois_abbrs:
        continue  # Case 2 또는 완전일치
    if n in mois_names:
        # RAMP 영문약어가 MOIS에 없고, 한글이 MOIS에 있음 = Case 1
        # MOIS 후보 중 첫번째 영문약어 채택 (대부분 1개일 것)
        mm = mois_by_name[n][0]
        case1.append({
            "ramp_단어명": n,
            "ramp_영문약어": a,
            "ramp_영문명": d["영문명"],
            "mois_영문약어": mm["영문약어"],
            "mois_영문명": mm["영문명"],
        })

print(f"Case 1 단어: {len(case1)}")

# === RAMP 스키마 로드 ===
wb = load_workbook(SCHEMA, read_only=True, data_only=True)
ws = wb["컬럼"]
schema = []
for r in ws.iter_rows(min_row=2, values_only=True):
    if not r or not r[1]: continue
    schema.append({
        "tbl": (r[0] or "").strip(),
        "en": (r[1] or "").strip().lower(),
        "kr": (r[2] or "").strip(),
        "desc": (r[3] or "").strip() if len(r) > 3 else "",
        "null": r[4] or "",
        "type": r[5] or "",
        "len": r[6] or "",
        "pk": r[7] or "",
        "fk": r[8] or "",
        "order": r[10] or "",
    })
wb.close()

# 테이블 코멘트 시트 (선택)
tbl_comment = {}
wb = load_workbook(SCHEMA, read_only=True, data_only=True)
if "테이블 목록" in wb.sheetnames:
    ws = wb["테이블 목록"]
    for r in ws.iter_rows(min_row=2, values_only=True):
        if r and r[0]:
            tbl_comment[(r[0] or "").strip()] = (r[1] or "").strip() if len(r) > 1 else ""
wb.close()

# === 변경 영향 컬럼 추출 ===
# 각 컬럼의 영문명을 _ 로 토큰 분해 → 토큰 중 case1.ramp_영문약어 가 있으면 영향
case1_abbrs = {c["ramp_영문약어"]: c for c in case1}

affected = []  # (단어, before_col, after_col, ...)
for c in schema:
    tokens = [t for t in re.split(r"_+", c["en"]) if t]
    tokens_up = [t.upper() for t in tokens]
    # 영향받는 토큰 위치
    changes = []
    new_tokens = list(tokens)
    for i, tk in enumerate(tokens_up):
        if tk in case1_abbrs:
            changes.append((i, tk, case1_abbrs[tk]["mois_영문약어"]))
            new_tokens[i] = case1_abbrs[tk]["mois_영문약어"].lower()
    if not changes:
        continue
    new_en = "_".join(new_tokens)
    # 변경 단어 요약
    chg_summary = "; ".join(f"{old}→{new}" for _, old, new in changes)
    affected.append({
        "tbl": c["tbl"],
        "tbl_cmt": tbl_comment.get(c["tbl"], ""),
        "before_en": c["en"],
        "after_en": new_en,
        "kr": c["kr"],
        "변경단어": chg_summary,
        "변경단어수": len(changes),
        "type": c["type"],
        "len": c["len"],
        "null": c["null"],
        "pk": c["pk"],
        "fk": c["fk"],
    })

print(f"영향 컬럼: {len(affected)}")

# 단어별 영향 컬럼수
per_word = Counter()
for r in affected:
    # 변경단어 요약에서 OLD→NEW 패턴 다중 가능. 단어별 카운트
    for piece in r["변경단어"].split("; "):
        old = piece.split("→")[0]
        per_word[old] += 1

# 변경 테이블 (DISTINCT)
tbl_changes = defaultdict(lambda: {"cols": 0, "words": set()})
for r in affected:
    tbl_changes[r["tbl"]]["cols"] += 1
    for piece in r["변경단어"].split("; "):
        tbl_changes[r["tbl"]]["words"].add(piece.split("→")[0])

print(f"영향 테이블 (unique): {len(tbl_changes)}")

# === xlsx 작성 ===
wb = Workbook()
hdr_font = Font(bold=True, color="FFFFFF")
hdr_fill = PatternFill("solid", fgColor="305496")
hdr_align = Alignment(horizontal="center", vertical="center")

def style_header(ws, ncol):
    for c in range(1, ncol+1):
        cell = ws.cell(row=1, column=c)
        cell.font = hdr_font; cell.fill = hdr_fill; cell.alignment = hdr_align

def auto_width(ws, widths):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

# 시트 1: 요약 (단어별)
ws = wb.active
ws.title = "단어별_요약"
ws.append(["RAMP 영문약어", "MOIS 영문약어 (변경 후)", "한글 단어명", "영향 컬럼 수", "영향 테이블 수"])
# 단어별 영향 테이블 수도 계산
word_tbls = defaultdict(set)
for r in affected:
    for piece in r["변경단어"].split("; "):
        old = piece.split("→")[0]
        word_tbls[old].add(r["tbl"])

case1_sorted = sorted(case1, key=lambda c: -per_word.get(c["ramp_영문약어"], 0))
for c in case1_sorted:
    a = c["ramp_영문약어"]
    ws.append([a, c["mois_영문약어"], c["ramp_단어명"],
               per_word.get(a, 0), len(word_tbls.get(a, set()))])
style_header(ws, 5)
auto_width(ws, [18, 22, 18, 14, 14])

# 시트 2: 변경 컬럼 전수
ws = wb.create_sheet("변경컬럼_전수")
ws.append(["테이블명", "테이블 설명", "변경전 컬럼명", "변경후 컬럼명", "컬럼 한글명",
           "변경 단어 (OLD→NEW)", "데이터타입", "길이", "NULL", "PK", "FK"])
affected_sorted = sorted(affected, key=lambda r: (r["tbl"], r["before_en"]))
for r in affected_sorted:
    ws.append([r["tbl"], r["tbl_cmt"], r["before_en"], r["after_en"], r["kr"],
               r["변경단어"], r["type"], r["len"], r["null"], r["pk"], r["fk"]])
style_header(ws, 11)
auto_width(ws, [28, 30, 38, 38, 25, 30, 12, 8, 8, 6, 6])
ws.freeze_panes = "A2"

# 시트 3: 변경 테이블 DISTINCT
ws = wb.create_sheet("변경테이블_DISTINCT")
ws.append(["테이블명", "테이블 설명", "영향 컬럼 수", "변경 단어 (unique)"])
tbl_sorted = sorted(tbl_changes.items(), key=lambda x: -x[1]["cols"])
for tn, info in tbl_sorted:
    ws.append([tn, tbl_comment.get(tn, ""), info["cols"], ", ".join(sorted(info["words"]))])
style_header(ws, 4)
auto_width(ws, [30, 40, 14, 50])
ws.freeze_panes = "A2"

wb.save(OUT)
print(f"\nSAVED: {OUT}")
print(f"size: {OUT.stat().st_size:,} bytes")
print(f"\n=== 요약 ===")
print(f"  변경 대상 단어: {len(case1)}")
print(f"  영향 컬럼: {len(affected)}")
print(f"  영향 테이블: {len(tbl_changes)}")
