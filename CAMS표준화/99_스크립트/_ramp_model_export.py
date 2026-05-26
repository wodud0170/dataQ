# -*- coding: utf-8 -*-
"""RAMP 모델 DB → 테이블/컬럼 업로드 양식 xlsx"""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from pathlib import Path
import subprocess, csv, io

OUT_TBL = Path(r"C:\Users\장재영\Desktop\dataQ\CAMS표준화\04_RAMP분석_2026-05-21\00_핵심산출\RAMP_업로드_테이블_2026-05-26.xlsx")
OUT_COL = Path(r"C:\Users\장재영\Desktop\dataQ\CAMS표준화\04_RAMP분석_2026-05-21\00_핵심산출\RAMP_업로드_컬럼_2026-05-26.xlsx")

def psql(sql):
    r = subprocess.run(["docker","exec","-i","dataq-db","psql","-U","admin","-d","postgres",
                        "-c", f"COPY ({sql}) TO STDOUT WITH (FORMAT csv, FORCE_QUOTE *)"],
                       capture_output=True, encoding="utf-8")
    if r.returncode != 0: print("ERR:", r.stderr); raise Exception(r.stderr)
    return list(csv.reader(io.StringIO(r.stdout)))

THIN = Side(border_style="thin", color="808080")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
HDR_FILL = PatternFill("solid", fgColor="1F4E78")
HDR_FONT = Font(name="맑은 고딕", size=10, color="FFFFFF", bold=True)
CELL_FONT = Font(name="맑은 고딕", size=10)

# ============ 1. 테이블 ============
# 양식 헤더: 소유자 / 업무영역 / 테이블명(한글) / 테이블명(영문) / 테이블스페이스 / 설명
print("=== 테이블 xlsx ===")
tbl_rows = psql("""
SELECT obj_owner, '', coalesce(obj_nm_kr,''), obj_nm, '', coalesce(obj_comment,'')
FROM quality.tb_data_model_obj
WHERE dm_id=(SELECT dm_id FROM quality.tb_data_model WHERE dm_nm='RAMP' LIMIT 1)
ORDER BY obj_nm
""")

wb = Workbook()
ws = wb.active; ws.title = "테이블"
H = ["소유자","업무영역","테이블명(한글)","테이블명(영문)","테이블스페이스","설명"]
W = [12, 14, 36, 32, 14, 60]
for i, w in enumerate(W, 1): ws.column_dimensions[get_column_letter(i)].width = w
for i, h in enumerate(H, 1):
    c = ws.cell(row=1, column=i, value=h); c.fill = HDR_FILL; c.font = HDR_FONT
    c.alignment = Alignment(horizontal="center"); c.border = BORDER
ws.freeze_panes = "A2"
for i, row in enumerate(tbl_rows, 2):
    for j, v in enumerate(row, 1):
        c = ws.cell(row=i, column=j, value=v); c.font = CELL_FONT; c.border = BORDER
        c.alignment = Alignment(vertical="top", wrap_text=True)
wb.save(OUT_TBL)
print(f"  → {OUT_TBL}  ({len(tbl_rows)}건)")

# ============ 2. 컬럼 ============
# 양식: 소유자 / 테이블명(영문) / 테이블명(한글) / 컬럼명(영문) / 컬럼명(한글) /
#       데이터타입 / 길이 / 소수점자리 / 컬럼 순서 / NULL여부 / PK여부 / FK여부 /
#       디폴트값 / 참조 테이블(한글) / 참조 컬럼(한글) / 삭제 규칙
print("\n=== 컬럼 xlsx ===")
col_rows = psql("""
SELECT a.obj_owner, a.obj_nm, coalesce(o.obj_nm_kr,''),
       a.attr_nm, coalesce(a.attr_nm_kr,''),
       a.data_type, coalesce(a.data_len::text,''), coalesce(a.data_decimal_len::text,'0'),
       coalesce(a.attr_ord::text,''), coalesce(a.nullable_yn,'Y'), coalesce(a.pk_yn,'N'), coalesce(a.fk_yn,'N'),
       coalesce(a.default_val,''),
       coalesce((SELECT o2.obj_nm_kr FROM quality.tb_data_model_obj o2 WHERE o2.dm_id=a.dm_id AND o2.obj_nm=a.fk_parent_obj_nm),''),
       coalesce(a.fk_parent_attr_nm,''),
       ''
FROM quality.tb_data_model_attr a
LEFT JOIN quality.tb_data_model_obj o ON a.dm_id=o.dm_id AND a.obj_nm=o.obj_nm
WHERE a.dm_id=(SELECT dm_id FROM quality.tb_data_model WHERE dm_nm='RAMP' LIMIT 1)
ORDER BY a.obj_nm, a.attr_ord
""")

wb = Workbook()
ws = wb.active; ws.title = "컬럼"
H2 = ["소유자","테이블명(영문)","테이블명(한글)","컬럼명(영문)","컬럼명(한글)",
      "데이터타입","길이","소수점자리","컬럼 순서","NULL여부","PK여부","FK여부",
      "디폴트값","참조 테이블(한글)","참조 컬럼(한글)","삭제 규칙"]
W2 = [12, 28, 30, 24, 28, 12, 8, 8, 8, 8, 6, 6, 12, 24, 22, 10]
for i, w in enumerate(W2, 1): ws.column_dimensions[get_column_letter(i)].width = w
for i, h in enumerate(H2, 1):
    c = ws.cell(row=1, column=i, value=h); c.fill = HDR_FILL; c.font = HDR_FONT
    c.alignment = Alignment(horizontal="center"); c.border = BORDER
ws.freeze_panes = "C2"
for i, row in enumerate(col_rows, 2):
    for j, v in enumerate(row, 1):
        c = ws.cell(row=i, column=j, value=v); c.font = CELL_FONT; c.border = BORDER
        c.alignment = Alignment(vertical="top", wrap_text=True)
wb.save(OUT_COL)
print(f"  → {OUT_COL}  ({len(col_rows)}건)")
