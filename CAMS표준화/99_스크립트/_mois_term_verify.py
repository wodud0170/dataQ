# -*- coding: utf-8 -*-
"""
행안부 용어 원본 xlsx vs dataQ DB tb_terms(comm_stnd_yn='Y') 검증.
단어 검증 v2와 동일 패턴 — 폐기 처리·노이즈·정합성·normalize 후 값 차이.

key: (용어명, 영문약어) — 용어명만으로는 중복 가능성.
"""
from openpyxl import load_workbook
from pathlib import Path
from collections import defaultdict, Counter
import subprocess, csv, sys, io, re

ORIG = Path(r"C:\Users\장재영\Desktop\dataQ\q-center\src\main\resources\seed\행안부_공통표준\행정안전부_공공데이터 공통표준용어.xlsx")
OUT_DIR = Path(r"C:\Users\장재영\Desktop\dataQ\CAMS표준화\04_RAMP분석_2026-05-21\_분석산출물_tsv")

def normalize(s):
    if s is None: return ""
    s = s.replace("\t","").replace("\r","").replace("\n","").replace(" "," ").strip()
    return re.sub(r",\s*", ",", s)

# ============ 1) 원본 파싱 ============
print("=== 1. 원본 용어 xlsx 파싱 ===")
wb = load_workbook(ORIG, read_only=True, data_only=True)
ws = wb["Sheet"]
rows = []
for r in ws.iter_rows(min_row=2, values_only=True):
    if r and r[0]:
        차수 = (r[10] or "").strip()
        구분 = (r[11] or "").strip()
        is_dead = ("폐기" in 차수 and "폐기후제정" not in 차수) or 구분 == "폐기"
        rows.append({
            "nm": (r[0] or "").strip(),
            "desc": (r[1] or "").strip(),
            "abrv": (r[2] or "").strip(),
            "dmn": (r[3] or "").strip(),
            "code_grp": (r[7] or "").strip(),
            "chrg": (r[8] or "").strip(),
            "syn": (r[9] or "").strip(),
            "차수": 차수, "구분": 구분,
            "dead": is_dead,
        })
wb.close()
print(f"  원본 행: {len(rows)}")
print(f"  폐기 행: {sum(1 for r in rows if r['dead'])}")
print(f"  변경 행: {sum(1 for r in rows if r['구분']=='변경')}")

# 중복 키 점검
cnt_nm = Counter(r["nm"] for r in rows)
dup_nm = [(nm,c) for nm,c in cnt_nm.items() if c>1]
cnt_nm_abrv = Counter((r["nm"], r["abrv"]) for r in rows)
dup_nm_abrv = [(k,c) for k,c in cnt_nm_abrv.items() if c>1]
print(f"  용어명 중복: {len(dup_nm)} 예: {dup_nm[:3]}")
print(f"  (용어명,약어) 중복: {len(dup_nm_abrv)} 예: {dup_nm_abrv[:3]}")

# 그룹화 (용어명, 영문약어) key
groups = defaultdict(list)
for r in rows:
    groups[(r["nm"], r["abrv"])].append(r)
expected = {}
dead_only = []
for k, lst in groups.items():
    alive = [r for r in lst if not r["dead"]]
    if alive:
        expected[k] = alive[0]
    else:
        dead_only.append(k)
print(f"  살아있어야 할 (용어명,약어): {len(expected)}")
print(f"  완전 폐기 (용어명,약어):     {len(dead_only)} 예: {dead_only[:3]}")

# 폐기인 용어명들
dead_nms = sorted(set(k[0] for k in dead_only))
print(f"  완전 폐기 용어명: {len(dead_nms)} 예: {dead_nms[:10]}")

# ============ 2) DB CSV 추출 ============
print("\n=== 2. DB tb_terms 추출 ===")
sql = """COPY (
  SELECT
    terms_nm,
    coalesce(terms_eng_abrv_nm,''),
    coalesce(terms_desc,''),
    coalesce(domain_nm,''),
    coalesce(code_grp,''),
    coalesce(chrg_org,''),
    coalesce(array_to_string(alloph_synm_lst, ','), '')
  FROM quality.tb_terms
  WHERE comm_stnd_yn='Y'
  ORDER BY terms_nm, terms_eng_abrv_nm
) TO STDOUT WITH (FORMAT csv, FORCE_QUOTE *)
"""
res = subprocess.run(
    ["docker", "exec", "-i", "dataq-db", "psql", "-U", "admin", "-d", "postgres", "-c", sql],
    capture_output=True, encoding="utf-8"
)
if res.returncode != 0:
    print("ERR:", res.stderr); sys.exit(1)

db_raw = {}  # (용어명, 약어) -> dict
db_dup = []
reader = csv.reader(io.StringIO(res.stdout))
for row in reader:
    if len(row) >= 7 and row[0]:
        k = (row[0], row[1])
        if k in db_raw:
            db_dup.append(k)
        db_raw[k] = {
            "nm": row[0], "abrv": row[1], "desc": row[2], "dmn": row[3],
            "code_grp": row[4], "chrg": row[5], "syn": row[6],
        }
print(f"  DB 건수: {len(db_raw)} (중복 키 {len(db_dup)})")

# ============ 3) 노이즈 점검 ============
print("\n=== 3. DB 측 노이즈 점검 ===")
FIELDS = [("nm","용어명"),("abrv","영문약어"),("desc","설명"),("dmn","도메인명"),
          ("code_grp","코드그룹"),("chrg","소관기관"),("syn","이음동의어")]
def has_tab(s): return "\t" in s
def has_lf(s):  return "\n" in s
def has_cr(s):  return "\r" in s
def has_nbsp(s): return " " in s
def has_ws_edge(s): return s.strip()!="" and s!=s.strip()

noise = defaultdict(int)
noise_samples = defaultdict(list)
for k, d in db_raw.items():
    for fk, label in FIELDS:
        v = d[fk] or ""
        for kind, fn in [("TAB",has_tab),("LF",has_lf),("CR",has_cr),("NBSP",has_nbsp),("WS_EDGE",has_ws_edge)]:
            if fn(v):
                noise[(kind, label)] += 1
                if len(noise_samples[(kind, label)]) < 3:
                    noise_samples[(kind, label)].append((k, repr(v)[:55]))

print(f"  노이즈 종류별 (총 {sum(noise.values())}건):")
for (kind, label), n in sorted(noise.items(), key=lambda x: -x[1]):
    print(f"    [{kind:8s}] {label:10s}: {n:5d}건  예: {noise_samples[(kind,label)][0]}")

# ============ 4) 적재 정합성 ============
print("\n=== 4. 적재 정합성 ===")
exp_set = set(expected.keys())
db_set = set(db_raw.keys())
xlsx_all = set(groups.keys())
dead_set = set(dead_only)

E_DEAD = sorted(dead_set & db_set)
E_MISSING = sorted(exp_set - db_set)
E_UNKNOWN = sorted(db_set - xlsx_all)

print(f"  [DEAD_IN_DB]    폐기인데 DB에 있음: {len(E_DEAD)}")
for k in E_DEAD[:10]: print(f"    - {k}")
if len(E_DEAD) > 10: print(f"    ... +{len(E_DEAD)-10}")

print(f"  [MISSING]       살아야할 누락: {len(E_MISSING)}")
for k in E_MISSING[:10]: print(f"    - {k}")
if len(E_MISSING) > 10: print(f"    ... +{len(E_MISSING)-10}")

print(f"  [UNKNOWN_IN_DB] 행안부에 없음:    {len(E_UNKNOWN)}")
for k in E_UNKNOWN[:10]: print(f"    - {k}")
if len(E_UNKNOWN) > 10: print(f"    ... +{len(E_UNKNOWN)-10}")

# ============ 5) normalize 후 값 차이 (용어명·영문약어 매핑 검증) ============
print("\n=== 5. normalize 후 값 차이 ===")
common = exp_set & db_set
val_diff = defaultdict(int)
val_diff_samples = defaultdict(list)
for k in common:
    e = expected[k]; d = db_raw[k]
    for fk, label in FIELDS:
        en = normalize(e[fk]); dn = normalize(d[fk])
        if en != dn:
            val_diff[label] += 1
            if len(val_diff_samples[label]) < 3:
                val_diff_samples[label].append((k, en[:50], dn[:50]))

print(f"  실제 값 차이 (컬럼별):")
for label, cnt in sorted(val_diff.items(), key=lambda x: -x[1]):
    print(f"    {label:10s}: {cnt:5d}건  예: {val_diff_samples[label][0]}")

# 용어명·영문약어 컬럼은 검증의 key 라 항상 같음 — 다른 컬럼 차이가 중요
# ============ 6) 산출 ============
print("\n=== 6. 산출 ===")
OUT = OUT_DIR / "_H_mois_term_verify.tsv"
with open(OUT, "w", encoding="utf-8", newline="") as f:
    w = csv.writer(f, delimiter="\t")
    w.writerow(["분류","용어명","영문약어","컬럼","원본","DB","비고"])
    for k in E_DEAD:
        nm, abrv = k
        r = next(x for x in rows if (x["nm"],x["abrv"])==k and x["dead"])
        w.writerow(["DEAD_IN_DB", nm, abrv, "", "", "(DB에있음)", f"차수={r['차수']}"])
    for k in E_MISSING:
        nm, abrv = k
        w.writerow(["MISSING", nm, abrv, "", expected[k]["dmn"], "(없음)", ""])
    for k in E_UNKNOWN:
        nm, abrv = k
        w.writerow(["UNKNOWN_IN_DB", nm, abrv, "", "(없음)", db_raw[k]["dmn"], ""])
print(f"  → {OUT}  (DEAD/MISSING/UNKNOWN만)")

print("\n" + "="*60)
print("=== 최종 요약 ===")
print(f"DB 용어 수: {len(db_raw)}  /  살아야할 수: {len(expected)}  /  폐기: {len(dead_only)}")
print(f"")
print(f"🔴 적재 사고:")
print(f"  - DEAD_IN_DB:    {len(E_DEAD)}건")
print(f"  - MISSING:       {len(E_MISSING)}건")
print(f"  - UNKNOWN_IN_DB: {len(E_UNKNOWN)}건")
print(f"")
print(f"🟡 정합성:")
print(f"  - 용어명·영문약어 매핑 차이: 0 (key 자체) — 다른 컬럼 차이는 {sum(val_diff.values())}건")
print(f"  - 노이즈 총 {sum(noise.values())}건")
