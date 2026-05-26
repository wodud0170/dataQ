# -*- coding: utf-8 -*-
"""
사용자가 채운 결정 영문약어 검증.

검증 항목:
  1. R5 형식: 대문자+숫자, 3~6자
  2. MOIS(행안부) 영문약어와 중복 금지
  3. RAMP 기존 영문약어와 중복 금지 (자기 자신 제외)
  4. 같은 시트 내 결정칸끼리 중복 금지
  5. 빈칸 (미결정) 카운트
"""
from openpyxl import load_workbook
from pathlib import Path
import re

BASE = Path(r"C:\Users\장재영\Desktop\dataQ\CAMS표준화")
DEC = BASE / "04_RAMP분석_2026-05-21" / "RAMP_단어결정_2026-05-23.xlsx"
RAMP_DICT = BASE / "01_원본자료" / "RMS4-DE14-04.데이터표준화_단어사전-V2.0.xlsx"
MOIS_DICT = BASE.parent / "q-center" / "src" / "main" / "resources" / "seed" / "행안부_공통표준" / "행정안전부_공공데이터 공통표준단어.xlsx"

# MOIS 약어 set
mois_abrv = set()
mois_meta = {}  # abrv → (단어명)
wb = load_workbook(MOIS_DICT, read_only=True, data_only=True)
for r in wb["Sheet"].iter_rows(min_row=2, values_only=True):
    if r and r[1]:
        abrv = str(r[1]).strip()
        mois_abrv.add(abrv)
        mois_meta[abrv] = str(r[0] or "").strip()
wb.close()

# RAMP 약어 set (기존)
ramp_abrv = set()
ramp_meta = {}
wb = load_workbook(RAMP_DICT, read_only=True, data_only=True)
for r in wb["단어사전"].iter_rows(min_row=2, values_only=True):
    if r and r[2] and r[3]:
        abrv = str(r[3]).strip()
        ramp_abrv.add(abrv)
        ramp_meta[abrv] = str(r[2]).strip()
wb.close()

# 결정 xlsx 읽기 — Case2_신규약어 시트
wb = load_workbook(DEC, read_only=True, data_only=True)
ws = wb["Case2_신규약어"]
header = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
print(f"=== 헤더 ===")
for i, h in enumerate(header):
    print(f"  col{i+1}: {h!r}")

# 컬럼 인덱스 찾기
def col_idx(name):
    for i, h in enumerate(header):
        if h and name in str(h): return i
    return -1

ic_no = col_idx("No")
ic_nm = col_idx("RAMP 한글")
ic_old = col_idx("RAMP 기존약어")
ic_suggest = col_idx("신규 약어 후보")
ic_decide = col_idx("확정 영문약어")
ic_note = col_idx("비고")
print(f"\n  no={ic_no} 한글={ic_nm} 기존약어={ic_old} 후보={ic_suggest} 확정={ic_decide} 비고={ic_note}")

decisions = []
for r in ws.iter_rows(min_row=2, values_only=True):
    if r and r[ic_no]:
        decisions.append({
            "no": r[ic_no],
            "nm": str(r[ic_nm] or "").strip(),
            "old": str(r[ic_old] or "").strip(),
            "suggest": str(r[ic_suggest] or "").strip(),
            "decide": str(r[ic_decide] or "").strip(),
            "note": str(r[ic_note] or "").strip() if ic_note >= 0 else "",
        })
wb.close()
print(f"\n=== Case2 결정 입력 ({len(decisions)}건) ===")

# ============ 검증 ============
print("\n=== 검증 ===")

# 1. 빈칸 (미결정)
empty = [d for d in decisions if not d["decide"]]
print(f"\n[1] 미결정 (빈칸): {len(empty)}건")
for d in empty[:5]:
    print(f"  - {d['no']}. {d['nm']:18s} (기존 {d['old']}, 후보 {d['suggest']})")
if len(empty) > 5: print(f"    ... 외 {len(empty)-5}")

filled = [d for d in decisions if d["decide"]]
print(f"\n[2] 결정 완료: {len(filled)}건")

# 2. R5 형식 위반
R5 = re.compile(r"^[A-Z0-9]+$")
r5_fail = []
for d in filled:
    v = d["decide"]
    if not R5.match(v) or not (3 <= len(v) <= 6):
        r5_fail.append((d, f"형식 위반 (길이 {len(v)}, 패턴 {v!r})"))
print(f"\n[3] R5 형식 위반 (대문자+숫자, 3~6자): {len(r5_fail)}건")
for d, reason in r5_fail:
    print(f"  🔴 {d['no']}. {d['nm']:18s} → {d['decide']:8s} : {reason}")

# 3. MOIS 충돌
mois_conflict = []
for d in filled:
    if d["decide"] in mois_abrv:
        mois_conflict.append((d, mois_meta[d["decide"]]))
print(f"\n[4] 행안부(MOIS) 약어와 중복: {len(mois_conflict)}건")
for d, mois_nm in mois_conflict:
    print(f"  🔴 {d['no']}. {d['nm']:18s} → {d['decide']:8s} : 행안부 '{mois_nm}' 와 중복")

# 4. RAMP 기존 약어 충돌 (자기 자신 제외 — 사용자가 같은 약어 유지하는 경우는 의미 없으므로 검출)
ramp_conflict = []
for d in filled:
    v = d["decide"]
    if v in ramp_abrv:
        owner = ramp_meta[v]
        if owner != d["nm"]:  # 본인 자기 약어가 아니면 충돌
            ramp_conflict.append((d, owner))
print(f"\n[5] RAMP 기존 약어와 중복: {len(ramp_conflict)}건")
for d, ramp_nm in ramp_conflict:
    print(f"  🔴 {d['no']}. {d['nm']:18s} → {d['decide']:8s} : RAMP '{ramp_nm}' 와 중복")

# 5. 시트 내 결정칸 중복
dec_seen = {}
intra_dup = []
for d in filled:
    v = d["decide"]
    if v in dec_seen:
        intra_dup.append((d, dec_seen[v]))
    else:
        dec_seen[v] = d
print(f"\n[6] 결정칸 내부 중복: {len(intra_dup)}건")
for d, prev in intra_dup:
    print(f"  🔴 {d['no']}. {d['nm']:18s} → {d['decide']:8s} : {prev['no']}. {prev['nm']} 와 중복")

# ============ 종합 ============
print("\n" + "="*60)
print("=== 종합 ===")
total_err = len(r5_fail) + len(mois_conflict) + len(ramp_conflict) + len(intra_dup)
print(f"결정 완료: {len(filled)} / {len(decisions)}")
print(f"미결정:    {len(empty)}")
print(f"오류:      {total_err}")
print(f"  - R5 위반:        {len(r5_fail)}")
print(f"  - MOIS 충돌:      {len(mois_conflict)}")
print(f"  - RAMP 기존 충돌: {len(ramp_conflict)}")
print(f"  - 시트내 중복:    {len(intra_dup)}")

if total_err == 0 and len(empty) == 0:
    print("\n✅ 모두 통과 — 다음 단계 진행 가능")
else:
    print(f"\n🔴 위 항목 보정 필요. 결정 시트의 [확정 영문약어] 칸 재검토.")
