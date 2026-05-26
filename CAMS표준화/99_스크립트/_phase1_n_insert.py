# -*- coding: utf-8 -*-
"""
Phase 1 — A 30건 + RAMP only 529건 = 559건 tb_word 신규 등록 (comm_stnd_yn='N').

순서:
  1. 결정 xlsx (Case2_신규약어) → A 30건 추출 (R5 통과 행)
  2. RAMP 사전 → RAMP only 529건 추출 (한글·약어 둘 다 행안부와 다름)
  3. DB 기존 한글 단어 충돌 점검
  4. INSERT SQL 생성 (BEGIN/INSERT/검증/COMMIT)
  5. 실행
  6. 검증
"""
from openpyxl import load_workbook
from pathlib import Path
from collections import defaultdict
import subprocess, csv, io, sys, re, os, base64

BASE = Path(r"C:\Users\장재영\Desktop\dataQ\CAMS표준화")
DEC = BASE / "04_RAMP분석_2026-05-21" / "RAMP_단어결정_2026-05-23.xlsx"
RAMP_DICT = BASE / "01_원본자료" / "RMS4-DE14-04.데이터표준화_단어사전-V2.0.xlsx"
MOIS_DICT = BASE.parent / "q-center" / "src" / "main" / "resources" / "seed" / "행안부_공통표준" / "행정안전부_공공데이터 공통표준단어.xlsx"
OUT_SQL = BASE / "04_RAMP분석_2026-05-21" / "Phase1_N_insert.sql"

R5 = re.compile(r"^[A-Z0-9]+$")

# ============ 1) A 30건 추출 ============
print("=== 1. A 30건 추출 ===")
wb = load_workbook(DEC, read_only=True, data_only=True)
ws = wb["Case2_신규약어"]
A_rows = []
for r in ws.iter_rows(min_row=2, values_only=True):
    if not r or not r[0]: continue
    nm = (str(r[1] or "").strip())
    eng = str(r[3] or "").strip()
    desc = str(r[4] or "").strip()
    decide = str(r[12] or "").strip()
    # R5 통과 + 길이 3~6 = A 분류
    if decide and R5.match(decide) and 3 <= len(decide) <= 6:
        A_rows.append({"nm": nm, "abrv": decide, "eng": eng, "desc": desc, "frmt":"", "dmn":""})
wb.close()
print(f"  A: {len(A_rows)}건")

# ============ 2) RAMP only 529건 추출 ============
print("\n=== 2. RAMP only 추출 ===")
# MOIS abrv + nm sets
mois_abrv = set(); mois_nm = set()
wb = load_workbook(MOIS_DICT, read_only=True, data_only=True)
for r in wb["Sheet"].iter_rows(min_row=2, values_only=True):
    if r and r[0] and r[1]:
        mois_abrv.add(str(r[1]).strip())
        mois_nm.add(str(r[0]).strip())
wb.close()

ramp_only = []
wb = load_workbook(RAMP_DICT, read_only=True, data_only=True)
for r in wb["단어사전"].iter_rows(min_row=2, values_only=True):
    if r and r[2] and r[3]:
        nm = str(r[2]).strip(); abrv = str(r[3]).strip()
        if nm not in mois_nm and abrv not in mois_abrv:
            ramp_only.append({
                "nm": nm, "abrv": abrv,
                "eng": str(r[4] or "").strip().replace("_x000D_",""),
                "desc": str(r[5] or "").strip().replace("_x000D_",""),
                "frmt": str(r[6] or "").strip(),
                "dmn": str(r[7] or "").strip(),
            })
wb.close()
print(f"  RAMP only: {len(ramp_only)}건")

# A와 RAMP only 결합 (한글 중복 없는지 확인 위해)
all_insert = A_rows + ramp_only

# 영문약어 충돌 5건 처리 (사용자 결정 2026-05-23)
ABRV_OVERRIDE = {"디렉토리": "DIR", "분": "MINUTE"}
SKIP_NMS = {"CLASS", "스캔", "컷"}  # 다음 단계에서 한글 통일
all_insert = [x for x in all_insert if x["nm"] not in SKIP_NMS]
for x in all_insert:
    if x["nm"] in ABRV_OVERRIDE:
        x["abrv"] = ABRV_OVERRIDE[x["nm"]]
print(f"  합계: {len(all_insert)}건")

# A 30 vs RAMP only 한글 중복 점검
a_nms = set(x["nm"] for x in A_rows)
ro_nms = set(x["nm"] for x in ramp_only)
intersect = a_nms & ro_nms
print(f"  A∩RAMPonly 한글 중복: {len(intersect)} → {list(intersect)[:5]}")

# ============ 3) DB 기존 한글 충돌 점검 ============
print("\n=== 3. DB 기존 한글 충돌 점검 ===")
def psql_csv(sql):
    r = subprocess.run(["docker","exec","-i","dataq-db","psql","-U","admin","-d","postgres",
                        "-c", f"COPY ({sql}) TO STDOUT WITH (FORMAT csv, FORCE_QUOTE *)"],
                       capture_output=True, encoding="utf-8")
    if r.returncode != 0: print("ERR:",r.stderr); sys.exit(1)
    return list(csv.reader(io.StringIO(r.stdout)))

db_nms = set()
db_meta = {}
for row in psql_csv("SELECT word_nm, comm_stnd_yn, word_eng_abrv_nm FROM quality.tb_word"):
    if row and row[0]:
        db_nms.add(row[0])
        db_meta[row[0]] = (row[1], row[2])
print(f"  DB 한글 단어 총: {len(db_nms)}")

insert_nms = set(x["nm"] for x in all_insert)
already = sorted(insert_nms & db_nms)
print(f"  DB와 한글 중복: {len(already)}")
for nm in already[:10]:
    yn, abrv = db_meta[nm]
    print(f"  - '{nm}' (DB: {yn} / {abrv})")

# 중복 단어는 INSERT 제외 (또는 사용자 결정)
to_insert = [x for x in all_insert if x["nm"] not in already]
print(f"\n  → INSERT 대상: {len(to_insert)} (중복 {len(already)} 제외)")

# ============ 4) INSERT SQL 생성 ============
print("\n=== 4. SQL 생성 ===")
sql_lines = [
    "-- Phase 1 — A + RAMP only 신규 단어 N 적재",
    f"-- 대상: {len(to_insert)}건 (중복 {len(already)}건 제외)",
    "BEGIN;",
    "",
]
for w in to_insert:
    nm = w["nm"].replace("'","''")
    abrv = w["abrv"].replace("'","''")
    eng = w["eng"].replace("'","''")
    desc = (w["desc"] or w["nm"]).replace("'","''")
    frmt = w["frmt"] if w["frmt"] in ("Y","N") else "N"
    dmn = w["dmn"].replace("'","''")
    wid = base64.urlsafe_b64encode(os.urandom(17)).rstrip(b'=').decode()[:22]
    sql_lines.append(f"INSERT INTO quality.tb_word (word_id,word_nm,word_eng_abrv_nm,word_eng_nm,word_desc,word_clsf_yn,domain_clsf_nm,comm_stnd_yn,aprv_yn,cret_dt,cret_user_id,use_yn) VALUES "
                    f"('{wid}','{nm}','{abrv}','{eng}','{desc}','{frmt}','{dmn}','N','Y',to_char(now(),'YYYYMMDDHH24MISS'),'admin','Y');")

sql_lines += [
    "",
    "-- 검증",
    "SELECT 'tb_word Y', count(*) FROM quality.tb_word WHERE comm_stnd_yn='Y';",
    "SELECT 'tb_word N', count(*) FROM quality.tb_word WHERE comm_stnd_yn='N';",
    f"SELECT 'inserted (N, admin, today)', count(*) FROM quality.tb_word WHERE comm_stnd_yn='N' AND cret_user_id='admin' AND cret_dt LIKE '20260523%';",
    "",
    "COMMIT;",
]

OUT_SQL.write_text("\n".join(sql_lines), encoding="utf-8")
print(f"  → {OUT_SQL}  ({len(to_insert)}건 INSERT)")
