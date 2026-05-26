"""CAMS 스키마에서 RG_* 테이블 및 유형 분기 컬럼 식별 (1단계 탐색)."""
from openpyxl import load_workbook
from pathlib import Path
from collections import defaultdict, Counter
import re

BASE = Path(r"C:\Users\장재영\Desktop\dataQ\CAMS표준화")
SCHEMA = BASE / "01_원본자료" / "CAMS_SCHEMA_원본.xlsx"

wb = load_workbook(SCHEMA, read_only=True, data_only=True)
print(f"=== CAMS 스키마 시트 목록 ===")
for sn in wb.sheetnames:
    ws = wb[sn]
    rows = list(ws.iter_rows(min_row=1, max_row=3, values_only=True))
    print(f"\n  '{sn}'  (max_row: {ws.max_row}, max_col: {ws.max_column})")
    for i, r in enumerate(rows, 1):
        cells = [str(c)[:30] if c is not None else "" for c in (r or [])[:12]]
        print(f"    r{i}: {cells}")
wb.close()
