# -*- coding: utf-8 -*-
"""
원본 엑셀 보정 (영문 비어있는 37행 채우기).
출력: ramp기관스키마정보_보정.xlsx (원본은 그대로)
"""
from openpyxl import load_workbook
from pathlib import Path
import shutil

BASE = Path(r"C:\Users\장재영\Desktop\dataQ\CAMS표준화")
SRC = BASE / "CAMS_RAMP_통합" / "ramp기관스키마정보.xlsx"
DST = BASE / "CAMS_RAMP_통합" / "ramp기관스키마정보_보정.xlsx"

# 한글 → 영문 보정 룰 (사람 결정)
KR_TO_EN = {
    "검수내용": "INSP_CN",
    "인수요청자명": "ACPTN_RQSTR_NM",
}

print(f"=== 원본 복사 → {DST.name} ===")
shutil.copyfile(SRC, DST)

print("=== 영문 비어있는 행 보정 ===")
wb = load_workbook(DST)
ws = wb["컬럼"]
fixed = []
no_rule = []
for row in ws.iter_rows(min_row=2):
    a, b, c = row[0].value, row[1].value, row[2].value
    if not a or b: continue   # 영문 있음 또는 테이블명 없음 → 스킵
    if not c: continue
    tbl = str(a).strip()
    kr = str(c).strip()
    if kr in KR_TO_EN:
        new_en = KR_TO_EN[kr]
        # 같은 테이블 내 영문 충돌 검사
        existing = []
        for rr in ws.iter_rows(min_row=2):
            if rr[0].value and str(rr[0].value).strip() == tbl and rr[1].value:
                existing.append(str(rr[1].value).strip().upper())
        if new_en in existing:
            # 충돌 시 _2 suffix (드물게)
            i = 2
            while f"{new_en}_{i}" in existing:
                i += 1
            new_en = f"{new_en}_{i}"
        row[1].value = new_en
        fixed.append((tbl, kr, new_en, row[10].value))
    else:
        no_rule.append((tbl, kr))

print(f"  보정: {len(fixed)}건")
for tbl, kr, en, ord_n in fixed:
    print(f"    {tbl}  kr={kr}  → en={en}  (ord_col={ord_n})")

# 중복 3건 (같은 테이블+영문 2번 나옴) - PK 있는 행 유지, 다른 행 삭제
DUP_DELETE_KEYS = {
    # (tbl, en, ord_col)
    ("tb_mgfolder", "prdctn_sys_fls_id", "65"),  # PK 없음, 행1529
    ("tb_stqna", "pstg_seq", "3"),               # PK 없음, 행4245
    ("tb_streporthist", "seq", "2"),             # PK 없음, 행4344 (4347은 PK=Y 살림)
}
print("\n=== 중복 행 삭제 (PK 없는 쪽) ===")
to_delete = []
for ridx, row in enumerate(ws.iter_rows(min_row=2), 2):
    a = str(row[0].value or "").strip()
    b = str(row[1].value or "").strip()
    k = str(row[10].value or "").strip() if row[10].value is not None else ""
    if (a, b, k) in DUP_DELETE_KEYS:
        to_delete.append((ridx, a, b, k))

# 아래에서 위로 삭제 (행 인덱스 변동 방지)
for ridx, a, b, k in sorted(to_delete, reverse=True):
    print(f"  삭제: 행{ridx} {a}/{b} ord={k}")
    ws.delete_rows(ridx, 1)

wb.save(DST)
print(f"\n→ {DST}")
