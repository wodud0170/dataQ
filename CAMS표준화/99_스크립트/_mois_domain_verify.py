# -*- coding: utf-8 -*-
"""행안부 도메인/도메인분류/도메인그룹 vs DB 검증. 단어/용어와 동일 패턴."""
from openpyxl import load_workbook
from pathlib import Path
from collections import defaultdict, Counter
import subprocess, csv, sys, io, re

BASE = Path(r"C:\Users\장재영\Desktop\dataQ\q-center\src\main\resources\seed\행안부_공통표준")

def has_tab(s): return "\t" in s
def has_nbsp(s): return "\xa0" in s
def has_ws_edge(s): return s.strip()!="" and s!=s.strip()

def normalize(s):
    if s is None: return ""
    s = s.replace("\t","").replace("\r","").replace("\n","").replace("\xa0"," ").strip()
    return re.sub(r",\s*", ",", s)

def fetch_db(table, cols, where="comm_stnd_yn='Y'", key_idx=0):
    def conv(c):
        if c in ("expr_fmt_lst","allow_val_lst"):
            return f"coalesce(array_to_string({c}, ','),'')"
        if c in ("data_len","data_decimal_len"):
            return f"coalesce({c}::text,'')"
        return f"coalesce({c},'')"
    select = ", ".join(conv(c) for c in cols)
    sql = f"COPY (SELECT {select} FROM quality.{table} WHERE {where}) TO STDOUT WITH (FORMAT csv, FORCE_QUOTE *)"
    r = subprocess.run(["docker","exec","-i","dataq-db","psql","-U","admin","-d","postgres","-c",sql],
                       capture_output=True, encoding="utf-8")
    if r.returncode != 0: print("ERR:", r.stderr); sys.exit(1)
    out = {}
    for row in csv.reader(io.StringIO(r.stdout)):
        if row and row[key_idx]:
            out[row[key_idx]] = row
    return out

# ============ A. 도메인그룹 ============
print("="*60); print("A. 도메인그룹"); print("="*60)
wb = load_workbook(BASE/"도메인그룹_일괄등록.xlsx", read_only=True, data_only=True)
ws = wb["Sheet1"]
xlsx_grp = set()
for r in ws.iter_rows(min_row=2, values_only=True):
    if r and r[1]: xlsx_grp.add(r[1].strip())
wb.close()
db_grp = fetch_db("tb_domain_grp", ["domain_grp_nm"])
db_grp_set = set(db_grp.keys())
print(f"  xlsx: {len(xlsx_grp)}  /  DB(Y): {len(db_grp_set)}")
print(f"  DB only: {db_grp_set - xlsx_grp}")
print(f"  xlsx only: {xlsx_grp - db_grp_set}")

# ============ B. 도메인분류 ============
print("\n"+"="*60); print("B. 도메인분류"); print("="*60)
wb = load_workbook(BASE/"도메인분류_일괄등록.xlsx", read_only=True, data_only=True)
ws = wb["Sheet1"]
xlsx_clsf = {}  # (grp, clsf_nm) -> row
for r in ws.iter_rows(min_row=2, values_only=True):
    if r and r[2]:
        k = ((r[1] or "").strip(), (r[2] or "").strip())
        xlsx_clsf[k] = r
wb.close()
db_clsf_rows = fetch_db("tb_domain_clsf", ["domain_clsf_nm","domain_grp_nm"])
db_clsf = {(row[1].strip(), row[0].strip()) for row in db_clsf_rows.values()}
xlsx_clsf_set = set(xlsx_clsf.keys())
print(f"  xlsx: {len(xlsx_clsf_set)}  /  DB(Y): {len(db_clsf)}")
print(f"  DB only: {sorted(db_clsf - xlsx_clsf_set)[:10]}")
print(f"  xlsx only: {sorted(xlsx_clsf_set - db_clsf)[:10]}")

# ============ C. 도메인 ============
print("\n"+"="*60); print("C. 도메인사전"); print("="*60)
ORIG = BASE / "행정안전부_공공데이터 공통표준도메인.xlsx"
wb = load_workbook(ORIG, read_only=True, data_only=True)
ws = wb["Sheet"]
rows = []
for r in ws.iter_rows(min_row=2, values_only=True):
    if r and r[2]:
        차수 = (r[11] or "").strip()
        구분 = (r[12] or "").strip()
        is_dead = ("폐기" in 차수 and "폐기후제정" not in 차수) or 구분 == "폐기"
        rows.append({
            "grp": (r[0] or "").strip(),
            "clsf": (r[1] or "").strip(),
            "nm": (r[2] or "").strip(),
            "desc": (r[3] or "").strip(),
            "dtype": (r[4] or "").strip(),
            "dlen": (r[5] or ""),
            "ddec": (r[6] or ""),
            "stor": (r[7] or "").strip(),
            "expr": (r[8] or "").strip(),
            "unit": (r[9] or "").strip(),
            "allow": (r[10] or "").strip(),
            "차수": 차수, "구분": 구분,
            "dead": is_dead,
        })
wb.close()
print(f"  원본 행: {len(rows)}  /  폐기: {sum(1 for r in rows if r['dead'])}  /  변경: {sum(1 for r in rows if r['구분']=='변경')}")

# 키: 도메인명 (unique 추정)
cnt_nm = Counter(r["nm"] for r in rows)
dup = [(nm,c) for nm,c in cnt_nm.items() if c>1]
print(f"  도메인명 중복: {len(dup)} 예: {dup[:3]}")

groups = defaultdict(list)
for r in rows: groups[r["nm"]].append(r)
expected = {}; dead_only = []
for nm, lst in groups.items():
    alive = [r for r in lst if not r["dead"]]
    if alive: expected[nm] = alive[0]
    else: dead_only.append(nm)
print(f"  살아야할: {len(expected)}  /  완전폐기: {len(dead_only)} → {dead_only}")

db_dom = fetch_db("tb_domain", ["domain_nm","domain_grp_nm","domain_clsf_nm","domain_desc","data_type","data_len","data_decimal_len","data_unit","stor_fmt","expr_fmt_lst","allow_val_lst"])
print(f"  DB(Y): {len(db_dom)}")

exp_set = set(expected.keys()); db_set = set(db_dom.keys())
xlsx_all = set(groups.keys())
E_DEAD = sorted(set(dead_only) & db_set)
E_MISSING = sorted(exp_set - db_set)
E_UNKNOWN = sorted(db_set - xlsx_all)
print(f"  🔴 DEAD_IN_DB:    {len(E_DEAD)} → {E_DEAD}")
print(f"  🔴 MISSING:       {len(E_MISSING)} → {E_MISSING[:10]}")
print(f"  🔴 UNKNOWN_IN_DB: {len(E_UNKNOWN)} → {E_UNKNOWN[:10]}")

# normalize 후 비교
common = exp_set & db_set
val_diff = defaultdict(int)
val_diff_sample = defaultdict(list)
field_map = [("grp","도메인그룹",1),("clsf","도메인분류",2),("desc","설명",3),
             ("dtype","데이터타입",4),("dlen","데이터길이",5),("ddec","소수점",6),
             ("unit","단위",7),("stor","저장형식",8),("expr","표현형식",9),
             ("allow","허용값",10)]
for nm in common:
    e = expected[nm]; d = db_dom[nm]
    for ek, label, di in field_map:
        ev = normalize(str(e[ek] or ""))
        dv = normalize(d[di])
        if ev != dv:
            val_diff[label] += 1
            if len(val_diff_sample[label]) < 3:
                val_diff_sample[label].append((nm, ev[:50], dv[:50]))

print(f"\n  normalize 후 값 차이:")
for label, cnt in sorted(val_diff.items(), key=lambda x: -x[1]):
    print(f"    {label:10s}: {cnt:4d}건  예: {val_diff_sample[label][0]}")

# 노이즈
print(f"\n  DB 노이즈:")
NFIELDS = [(1,"그룹"),(2,"분류"),(3,"설명"),(4,"타입"),(7,"단위"),(8,"저장형식"),(9,"표현형식"),(10,"허용값")]
nz = defaultdict(int)
for d in db_dom.values():
    for i, label in NFIELDS:
        v = d[i] or ""
        if has_tab(v): nz[("TAB", label)] += 1
        if has_nbsp(v): nz[("NBSP", label)] += 1
        if has_ws_edge(v): nz[("WS_EDGE", label)] += 1
for (k, l), n in sorted(nz.items(), key=lambda x: -x[1]):
    print(f"    [{k:8s}] {l:10s}: {n}건")
if not nz: print("    (없음)")

print("\n" + "="*60)
print("=== 최종 요약 ===")
print(f"도메인그룹: xlsx {len(xlsx_grp)} / DB(Y) {len(db_grp_set)} / 차이 {(db_grp_set ^ xlsx_grp)}")
print(f"도메인분류: xlsx {len(xlsx_clsf_set)} / DB(Y) {len(db_clsf)}")
print(f"도메인사전: 원본 {len(groups)} (살아야할 {len(expected)}) / DB(Y) {len(db_dom)}")
print(f"  - DEAD_IN_DB: {len(E_DEAD)}")
print(f"  - MISSING:    {len(E_MISSING)}")
print(f"  - UNKNOWN:    {len(E_UNKNOWN)}")
print(f"  - 값차이 총: {sum(val_diff.values())}건 (컬럼별)")
print(f"  - 노이즈 총: {sum(nz.values())}건")
