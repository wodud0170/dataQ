# -*- coding: utf-8 -*-
"""
tb_terms WHERE comm_stnd_yn='N' 전체 교체:
  1. 현 N 1,786건 → 백업 xlsx
  2. DELETE (CASCADE 로 tb_terms_words 자동 정리)
  3. RAMP_업로드_용어_2026-05-26.xlsx → INSERT
  4. 검증
"""
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from pathlib import Path
import subprocess, csv, io, os, base64

BASE = Path(r"C:\Users\장재영\Desktop\dataQ\CAMS표준화")
ROOT = BASE / "04_RAMP분석_2026-05-21" / "00_핵심산출"
TERMS_XLSX = ROOT / "RAMP_업로드_용어_2026-05-26.xlsx"
BACKUP_XLSX = ROOT / "tb_terms_N_백업_2026-05-26.xlsx"
SQL_OUT = BASE / "04_RAMP분석_2026-05-21" / "99_실행SQL_기록" / "Phase3_terms_replace.sql"

def psql(sql, copy=False):
    if copy:
        cmd = ["docker","exec","-i","dataq-db","psql","-U","admin","-d","postgres",
               "-c", f"COPY ({sql}) TO STDOUT WITH (FORMAT csv, FORCE_QUOTE *, HEADER true)"]
    else:
        cmd = ["docker","exec","-i","dataq-db","psql","-U","admin","-d","postgres","-c", sql]
    r = subprocess.run(cmd, capture_output=True, encoding="utf-8")
    if r.returncode != 0:
        print("ERR:", r.stderr); raise Exception(r.stderr)
    return r.stdout

def esc(s):
    if s is None: return ""
    return str(s).replace("'", "''")

def gid():
    return base64.urlsafe_b64encode(os.urandom(17)).rstrip(b"=").decode()[:22]

# ============ 1. 현 N 백업 ============
print("=== 1. 현 N 백업 ===")
out = psql("""SELECT terms_id, terms_nm, terms_eng_abrv_nm, terms_desc, domain_nm,
              code_grp, chrg_org, comm_stnd_yn, aprv_yn, cret_dt, cret_user_id
              FROM quality.tb_terms WHERE comm_stnd_yn='N'
              ORDER BY terms_nm""", copy=True)
rows = list(csv.reader(io.StringIO(out)))
print(f"  백업 대상: {len(rows)-1}건")

wb = Workbook()
ws = wb.active; ws.title = "tb_terms_N_백업"
H = rows[0]
for i, h in enumerate(H, 1):
    c = ws.cell(row=1, column=i, value=h)
    c.font = Font(name="맑은 고딕", size=10, bold=True, color="FFFFFF")
    c.fill = PatternFill("solid", fgColor="1F4E78")
    c.alignment = Alignment(horizontal="center")
W = [22, 30, 20, 40, 16, 12, 14, 6, 6, 16, 12]
for i, w in enumerate(W, 1): ws.column_dimensions[get_column_letter(i)].width = w
for r_idx, row in enumerate(rows[1:], 2):
    for c_idx, v in enumerate(row, 1):
        ws.cell(row=r_idx, column=c_idx, value=v)
ws.freeze_panes = "A2"
wb.save(BACKUP_XLSX)
print(f"  → {BACKUP_XLSX}")

# ============ 2. 새 용어 xlsx 로드 ============
print("\n=== 2. 새 용어 xlsx 로드 ===")
wb = load_workbook(TERMS_XLSX, read_only=True, data_only=True)
ws = wb.active
new_terms = []
for r in ws.iter_rows(min_row=2, values_only=True):
    if not r or not r[2] or not r[4]: continue
    new_terms.append({
        "kr": str(r[2]).strip(), "en": str(r[4]).strip(),
        "desc": str(r[3] or "").strip(), "dom": str(r[5] or "").strip(),
    })
wb.close()
print(f"  새 용어: {len(new_terms)}")

# ============ 3. 도메인 존재 여부 검증 ============
print("\n=== 3. 도메인 존재 여부 검증 ===")
out = psql("SELECT domain_nm, domain_clsf_nm FROM quality.tb_domain", copy=True)
existing_dom_map = {}
for r in csv.reader(io.StringIO(out)):
    if r and r[0] != "domain_nm":
        existing_dom_map[r[0]] = r[1]
existing_doms = set(existing_dom_map.keys())
print(f"  DB 도메인 수: {len(existing_doms)}")

# 도메인 분류 기존
out = psql("SELECT domain_clsf_nm, domain_grp_nm FROM quality.tb_domain_clsf", copy=True)
clsf_grp_by_nm = {}
for r in csv.reader(io.StringIO(out)):
    if r and r[0] != "domain_clsf_nm":
        clsf_grp_by_nm[r[0]] = r[1]
print(f"  DB 도메인분류 수: {len(clsf_grp_by_nm)}")

# 도메인 그룹 기존
out = psql("SELECT domain_grp_nm FROM quality.tb_domain_grp", copy=True)
existing_grps = set(r[0] for r in csv.reader(io.StringIO(out)) if r and r[0] != "domain_grp_nm")
print(f"  DB 도메인그룹 수: {len(existing_grps)}")

# 새 도메인 분류명 추출 (도메인명에서 분류 prefix)
def parse_clsf(dom):
    """'내용V100' → '내용', '값N5' → '값', '일자V8' → '일자', '일자DT' → '일자'"""
    import re
    m = re.match(r'^([가-힣]+)', dom)
    return m.group(1) if m else ""

missing_doms = {}  # dom → clsf
for t in new_terms:
    if t["dom"] and t["dom"] not in existing_doms:
        missing_doms[t["dom"]] = parse_clsf(t["dom"])
print(f"  누락 도메인: {len(missing_doms)}")
if missing_doms:
    print(f"  샘플: {sorted(missing_doms.keys())[:10]}")

# 누락 도메인분류
missing_clsf = set(missing_doms.values()) - set(clsf_grp_by_nm.keys())
print(f"  누락 도메인분류: {len(missing_clsf)}")
if missing_clsf:
    print(f"  분류 샘플: {sorted(missing_clsf)[:10]}")

# ============ 4. UNIQUE 충돌 검증 ============
print("\n=== 4. UNIQUE 충돌 검증 ===")
# 행안부 Y 와 충돌
out = psql("SELECT terms_nm FROM quality.tb_terms WHERE comm_stnd_yn='Y'", copy=True)
y_nm = set(r[0] for r in csv.reader(io.StringIO(out)) if r and r[0] != "terms_nm")
out = psql("SELECT terms_eng_abrv_nm FROM quality.tb_terms WHERE comm_stnd_yn='Y'", copy=True)
y_en = set(r[0] for r in csv.reader(io.StringIO(out)) if r and r[0] != "terms_eng_abrv_nm")

conflict_nm = [t for t in new_terms if t["kr"] in y_nm]
conflict_en = [t for t in new_terms if t["en"] in y_en]
print(f"  한글 충돌 (행안부 Y와 동일): {len(conflict_nm)}")
print(f"  영문 충돌 (행안부 Y와 동일): {len(conflict_en)}")
if conflict_nm:
    print("  샘플 한글 충돌:")
    for t in conflict_nm[:5]: print(f"    {t['kr']} ({t['en']})")
if conflict_en:
    print("  샘플 영문 충돌:")
    for t in conflict_en[:5]: print(f"    {t['en']} ({t['kr']})")

# 내부 중복 (new_terms 내)
from collections import Counter
nm_cnt = Counter(t["kr"] for t in new_terms)
en_cnt = Counter(t["en"] for t in new_terms)
dup_nm = [k for k,v in nm_cnt.items() if v>1]
dup_en = [k for k,v in en_cnt.items() if v>1]
print(f"  내부 한글 중복: {len(dup_nm)}")
print(f"  내부 영문 중복: {len(dup_en)}")
if dup_en[:5]:
    print(f"    예: {dup_en[:5]}")

# ============ 5. SQL 생성 ============
print("\n=== 5. SQL 생성 ===")
lines = [
    "-- Phase 3: tb_terms N 전체 교체",
    "-- 1) 백업 완료, 2) DELETE, 3) 누락 도메인분류/도메인 등록, 4) INSERT 새 용어",
    "BEGIN;",
    "",
    "-- DELETE 현 N (CASCADE 로 tb_terms_words 자동 정리)",
    "DELETE FROM quality.tb_terms WHERE comm_stnd_yn='N';",
    "",
]

# 누락 도메인분류 INSERT (기본 그룹 '기타')
if missing_clsf:
    lines.append(f"-- 누락 도메인분류 ({len(missing_clsf)}건)")
    for clsf in sorted(missing_clsf):
        cid = gid()
        grp = "기타" if "기타" in existing_grps else next(iter(existing_grps))
        clsf_grp_by_nm[clsf] = grp
        lines.append(
            f"INSERT INTO quality.tb_domain_clsf "
            f"(domain_clsf_id, domain_clsf_nm, domain_grp_nm, comm_stnd_yn, cret_dt, cret_user_id) "
            f"VALUES ('{cid}','{esc(clsf)}','{esc(grp)}','N',to_char(now(),'YYYYMMDDHH24MISS'),'admin');"
        )
    lines.append("")

# 누락 도메인 INSERT
import re
def parse_dom(dom):
    """ '내용V500' → (분류=내용, 타입=VARCHAR, 길이=500) """
    m = re.match(r'^([가-힣]+)([A-Z]+)(\d*)$', dom)
    if not m: return ("", "", "")
    clsf, type_tok, ln = m.group(1), m.group(2), m.group(3) or ""
    type_map = {"V":"VARCHAR","N":"NUMBER","DT":"DATE","TS":"TIMESTAMP","L":"CLOB"}
    return (clsf, type_map.get(type_tok, type_tok), ln)

if missing_doms:
    lines.append(f"-- 누락 도메인 ({len(missing_doms)}건)")
    for dom in sorted(missing_doms.keys()):
        clsf = missing_doms[dom]
        grp = clsf_grp_by_nm.get(clsf, "기타")
        _, dtype, dlen = parse_dom(dom)
        dlen_v = dlen if dlen.isdigit() else "0"
        did = gid()
        lines.append(
            f"INSERT INTO quality.tb_domain "
            f"(domain_id, domain_nm, domain_grp_nm, domain_clsf_nm, domain_desc, data_type, data_len, "
            f"comm_stnd_yn, aprv_yn, cret_dt, cret_user_id, use_yn) "
            f"VALUES ('{did}','{esc(dom)}','{esc(grp)}','{esc(clsf)}','','{esc(dtype)}',{dlen_v},"
            f"'N','Y',to_char(now(),'YYYYMMDDHH24MISS'),'admin','Y');"
        )
    lines.append("")

lines.append(f"-- INSERT 새 용어 ({len(new_terms)}건)")
seen_nm = set(y_nm)
seen_en = set(y_en)
skipped = 0
for t in new_terms:
    if t["kr"] in seen_nm or t["en"] in seen_en:
        skipped += 1
        lines.append(f"-- SKIP (충돌): {t['kr']} / {t['en']}")
        continue
    seen_nm.add(t["kr"]); seen_en.add(t["en"])
    tid = gid()
    lines.append(
        f"INSERT INTO quality.tb_terms "
        f"(terms_id, terms_nm, terms_eng_abrv_nm, terms_desc, domain_nm, "
        f"comm_stnd_yn, aprv_yn, cret_dt, cret_user_id, use_yn) "
        f"VALUES ('{tid}','{esc(t['kr'])}','{esc(t['en'])}','{esc(t['desc'])}','{esc(t['dom'])}',"
        f"'N','Y',to_char(now(),'YYYYMMDDHH24MISS'),'admin','Y');"
    )
print(f"  SKIP (충돌): {skipped}")
print(f"  실제 INSERT: {len(new_terms) - skipped}")

lines += [
    "",
    "-- 검증",
    "SELECT 'tb_terms Y' as t, count(*) FROM quality.tb_terms WHERE comm_stnd_yn='Y';",
    "SELECT 'tb_terms N' as t, count(*) FROM quality.tb_terms WHERE comm_stnd_yn='N';",
    "",
    "COMMIT;",
]
SQL_OUT.write_text("\n".join(lines), encoding="utf-8")
print(f"  → {SQL_OUT}")
