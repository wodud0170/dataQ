"""
RAMP 단어사전 ↔ RAMP 스키마 컬럼 커버율 정밀 분석.

레벨 3단계:
  1) 토큰 레벨 : unique 토큰 555개 중 사전 등재 비율
  2) 컬럼 레벨 : 컬럼의 모든 토큰이 사전 등재 = 표준 / 일부라도 미등재 = 비표준
  3) 테이블 레벨 : 테이블의 비표준 컬럼 비율로 비표준 테이블 식별

출력: CAMS표준화/04_RAMP분석_2026-05-21/RAMP_커버율_2026-05-21.xlsx
  시트 1. 요약
  시트 2. 미등재 토큰 (스키마 only)
  시트 3. 비표준 컬럼 전수
  시트 4. 비표준 테이블 (커버율 낮은 순)
"""
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from pathlib import Path
from collections import defaultdict, Counter
import re

BASE = Path(r"C:\Users\장재영\Desktop\dataQ\CAMS표준화")
DICT = BASE / "01_원본자료" / "RMS4-DE14-04.데이터표준화_단어사전-V2.0.xlsx"
SCHEMA = BASE / "CAMS_RAMP_통합" / "ramp기관스키마정보.xlsx"
OUT = BASE / "04_RAMP분석_2026-05-21" / "RAMP_커버율_2026-05-21.xlsx"

# === 사전 ===
wb = load_workbook(DICT, read_only=True, data_only=True)
ws = wb["단어사전"]
dict_abbrs = set()
for r in ws.iter_rows(min_row=2, values_only=True):
    if not r or not r[3]: continue
    a = (r[3] or "").strip().upper()
    if a: dict_abbrs.add(a)
wb.close()
print(f"사전 영문약어 unique: {len(dict_abbrs)}")

# === 스키마 ===
wb = load_workbook(SCHEMA, read_only=True, data_only=True)
ws = wb["컬럼"]
schema = []
for r in ws.iter_rows(min_row=2, values_only=True):
    if not r or not r[1]: continue
    schema.append({
        "tbl": (r[0] or "").strip(),
        "en": (r[1] or "").strip().lower(),
        "kr": (r[2] or "").strip(),
        "type": r[5] or "",
        "len": r[6] or "",
        "pk": r[7] or "",
    })
tbl_comment = {}
if "테이블 목록" in wb.sheetnames:
    ws = wb["테이블 목록"]
    for r in ws.iter_rows(min_row=2, values_only=True):
        if r and r[0]:
            tbl_comment[(r[0] or "").strip()] = (r[1] or "").strip() if len(r) > 1 else ""
wb.close()

# === 토큰 분해 + 컬럼별 분석 ===
col_data = []   # 컬럼별 토큰/매칭 결과
token_all = Counter()
token_missing = Counter()
missing_token_cols = defaultdict(list)   # 토큰 → 등장 컬럼

for c in schema:
    raw_tokens = [t for t in re.split(r"[_]+", c["en"]) if t]
    tokens_up = [t.upper() for t in raw_tokens]
    for tk in tokens_up:
        token_all[tk] += 1
    if not tokens_up:
        continue
    missing = [tk for tk in tokens_up if tk not in dict_abbrs]
    cov_rate = (len(tokens_up) - len(missing)) / len(tokens_up)
    if missing:
        for m in missing:
            token_missing[m] += 1
            missing_token_cols[m].append(f"{c['tbl']}.{c['en']}({c['kr']})")
    col_data.append({
        "tbl": c["tbl"], "en": c["en"], "kr": c["kr"],
        "type": c["type"], "len": c["len"], "pk": c["pk"],
        "tokens_n": len(tokens_up),
        "missing_n": len(missing),
        "missing_tokens": ", ".join(missing),
        "cov_rate": cov_rate,
    })

n_total = len(col_data)
n_full = sum(1 for r in col_data if r["missing_n"] == 0)
n_partial = sum(1 for r in col_data if 0 < r["missing_n"] < r["tokens_n"])
n_none = sum(1 for r in col_data if r["missing_n"] == r["tokens_n"])

print(f"\n=== 컬럼 레벨 ===")
print(f"전체 컬럼: {n_total}")
print(f"완전 매칭 (표준): {n_full} ({n_full/n_total*100:.2f}%)")
print(f"일부 매칭 (부분 표준): {n_partial} ({n_partial/n_total*100:.2f}%)")
print(f"모두 미매칭 (완전 비표준): {n_none} ({n_none/n_total*100:.2f}%)")

# === 테이블 레벨 ===
tbl_stats = defaultdict(lambda: {"total": 0, "full": 0, "partial": 0, "none": 0, "missing_tokens": Counter()})
for r in col_data:
    s = tbl_stats[r["tbl"]]
    s["total"] += 1
    if r["missing_n"] == 0: s["full"] += 1
    elif r["missing_n"] == r["tokens_n"]: s["none"] += 1
    else: s["partial"] += 1
    for tk in r["missing_tokens"].split(", "):
        if tk: s["missing_tokens"][tk] += 1

# 비표준 테이블 = (partial + none) / total 비율로 정렬
tbl_list = []
for tn, s in tbl_stats.items():
    nonstd = s["partial"] + s["none"]
    rate = nonstd / s["total"] if s["total"] > 0 else 0
    tbl_list.append({
        "tbl": tn,
        "cmt": tbl_comment.get(tn, ""),
        "total": s["total"],
        "full": s["full"],
        "partial": s["partial"],
        "none": s["none"],
        "nonstd_rate": rate,
        "missing_tokens_top": ", ".join(f"{tk}({cnt})" for tk, cnt in s["missing_tokens"].most_common(5)),
    })

n_tbl = len(tbl_list)
n_tbl_clean = sum(1 for t in tbl_list if t["nonstd_rate"] == 0)
n_tbl_partial = sum(1 for t in tbl_list if 0 < t["nonstd_rate"] < 0.5)
n_tbl_dirty = sum(1 for t in tbl_list if t["nonstd_rate"] >= 0.5)

print(f"\n=== 테이블 레벨 ===")
print(f"전체 테이블: {n_tbl}")
print(f"완전 표준 (비표준 0%): {n_tbl_clean}")
print(f"부분 비표준 (0~50%): {n_tbl_partial}")
print(f"비표준 우세 (50%+): {n_tbl_dirty}")

# === xlsx ===
wb = Workbook()
hdr_font = Font(bold=True, color="FFFFFF")
hdr_fill = PatternFill("solid", fgColor="305496")
hdr_align = Alignment(horizontal="center", vertical="center")
warn_fill = PatternFill("solid", fgColor="FFD966")
err_fill = PatternFill("solid", fgColor="F4B084")

def style_header(ws, ncol):
    for c in range(1, ncol+1):
        cell = ws.cell(row=1, column=c)
        cell.font = hdr_font; cell.fill = hdr_fill; cell.alignment = hdr_align

def auto_width(ws, widths):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

# 시트 1: 요약
ws = wb.active
ws.title = "요약"
ws.append(["레벨", "구분", "건수", "비율", "비고"])
# 토큰
ws.append(["토큰", "스키마 unique 토큰", len(token_all), "100%", ""])
ws.append(["토큰", "사전 등재 토큰", len(token_all) - len(token_missing), f"{(len(token_all)-len(token_missing))/len(token_all)*100:.2f}%", "사전 커버"])
ws.append(["토큰", "사전 미등재 토큰", len(token_missing), f"{len(token_missing)/len(token_all)*100:.2f}%", "↓ 시트2 미등재 토큰"])
ws.append([])
# 컬럼
ws.append(["컬럼", "전체 컬럼", n_total, "100%", ""])
ws.append(["컬럼", "완전 표준 (모든 토큰 등재)", n_full, f"{n_full/n_total*100:.2f}%", "표준 준수"])
ws.append(["컬럼", "부분 비표준 (일부 토큰 미등재)", n_partial, f"{n_partial/n_total*100:.2f}%", "↓ 시트3"])
ws.append(["컬럼", "완전 비표준 (모든 토큰 미등재)", n_none, f"{n_none/n_total*100:.2f}%", "↓ 시트3, 매우 적음 예상"])
ws.append([])
# 테이블
ws.append(["테이블", "전체 테이블", n_tbl, "100%", ""])
ws.append(["테이블", "완전 표준 (비표준 컬럼 0%)", n_tbl_clean, f"{n_tbl_clean/n_tbl*100:.2f}%", ""])
ws.append(["테이블", "부분 비표준 (0~50%)", n_tbl_partial, f"{n_tbl_partial/n_tbl*100:.2f}%", ""])
ws.append(["테이블", "비표준 우세 (50%+)", n_tbl_dirty, f"{n_tbl_dirty/n_tbl*100:.2f}%", "↓ 시트4 상단"])
style_header(ws, 5)
auto_width(ws, [10, 35, 12, 12, 35])
for row in ws.iter_rows(min_row=2, max_col=5):
    for cell in row:
        cell.alignment = Alignment(vertical="center", wrap_text=True)

# 시트 2: 미등재 토큰
ws = wb.create_sheet("미등재토큰_스키마only")
ws.append(["미등재 토큰", "등장 컬럼 수", "샘플 컬럼 (최대 5)"])
for tk, cnt in token_missing.most_common():
    sample = "; ".join(missing_token_cols[tk][:5])
    ws.append([tk, cnt, sample])
style_header(ws, 3)
auto_width(ws, [20, 14, 100])
ws.freeze_panes = "A2"

# 시트 3: 비표준 컬럼 전수
ws = wb.create_sheet("비표준컬럼_전수")
ws.append(["테이블명", "컬럼 영문", "컬럼 한글", "타입", "길이", "PK",
           "전체 토큰", "미등재 토큰 수", "미등재 토큰", "커버율"])
nonstd_cols = [r for r in col_data if r["missing_n"] > 0]
nonstd_cols.sort(key=lambda r: (r["cov_rate"], r["tbl"], r["en"]))
for r in nonstd_cols:
    ws.append([r["tbl"], r["en"], r["kr"], r["type"], r["len"], r["pk"],
               r["tokens_n"], r["missing_n"], r["missing_tokens"],
               f"{r['cov_rate']*100:.0f}%"])
    last = ws.max_row
    if r["missing_n"] == r["tokens_n"]:
        for col in range(1, 11):
            ws.cell(row=last, column=col).fill = err_fill
    else:
        for col in range(1, 11):
            ws.cell(row=last, column=col).fill = warn_fill
style_header(ws, 10)
auto_width(ws, [25, 32, 22, 12, 8, 6, 10, 12, 25, 10])
ws.freeze_panes = "A2"

# 시트 4: 비표준 테이블 (비표준율 내림차순)
ws = wb.create_sheet("비표준테이블_랭킹")
ws.append(["테이블명", "테이블 설명", "총 컬럼", "표준 컬럼", "부분 비표준", "완전 비표준",
           "비표준율 (%)", "미등재 토큰 Top5 (빈도)"])
tbl_sorted = sorted(tbl_list, key=lambda t: (-t["nonstd_rate"], -t["total"]))
for t in tbl_sorted:
    ws.append([t["tbl"], t["cmt"], t["total"], t["full"], t["partial"], t["none"],
               round(t["nonstd_rate"]*100, 1), t["missing_tokens_top"]])
    last = ws.max_row
    if t["nonstd_rate"] >= 0.5:
        for col in range(1, 9):
            ws.cell(row=last, column=col).fill = err_fill
    elif t["nonstd_rate"] > 0:
        for col in range(1, 9):
            ws.cell(row=last, column=col).fill = warn_fill
style_header(ws, 8)
auto_width(ws, [28, 35, 10, 10, 12, 12, 12, 50])
ws.freeze_panes = "A2"

wb.save(OUT)
print(f"\nSAVED: {OUT}")
print(f"size: {OUT.stat().st_size:,} bytes")
