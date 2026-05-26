"""
행안부 표준 우선 적용 시 RAMP 스키마 BEFORE-AFTER (단어 차원).

원칙:
  - Case 1 (한글동일/영문약어다름): RAMP 컬럼 영문약어 → MOIS 영문약어 (영문 변경)
  - Case 2 (영문동일/한글다름): RAMP 컬럼 한글 코멘트 → MOIS 한글명 (한글 변경)
  - Case 1+2 동시 (한 컬럼에 양쪽 단어): 둘 다 적용
  - 데이터 타입·길이는 변경 없음 (용어/도메인 차원이라 보류)

출력: CAMS표준화/04_RAMP분석_2026-05-21/RAMP_BEFORE_AFTER_2026-05-21.xlsx
  시트1 요약: 변경 통계
  시트2 BEFORE_AFTER 전수: 컬럼별 변경 내역 + 변경 종류 (영문/한글/양쪽)
  시트3 ALTER 가이드: 테이블별 변경 컬럼 묶음 (SQL 작성 베이스)
"""
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from pathlib import Path
from collections import defaultdict, Counter
import re

BASE = Path(r"C:\Users\장재영\Desktop\dataQ\CAMS표준화")
DICT = BASE / "01_원본자료" / "RMS4-DE14-04.데이터표준화_단어사전-V2.0.xlsx"
SCHEMA = BASE / "CAMS_RAMP_통합" / "ramp기관스키마정보.xlsx"
MOIS_WORD = BASE.parent / "q-center" / "src" / "main" / "resources" / "seed" / "행안부_공통표준" / "행정안전부_공공데이터 공통표준단어.xlsx"
OUT = BASE / "04_RAMP분석_2026-05-21" / "RAMP_BEFORE_AFTER_2026-05-21.xlsx"

# === RAMP 사전 ===
wb = load_workbook(DICT, read_only=True, data_only=True)
ws = wb["단어사전"]
ramp_words = []
for r in ws.iter_rows(min_row=2, values_only=True):
    if not r or not r[2]: continue
    ramp_words.append({"단어명": (r[2] or "").strip(),
                       "영문약어": (r[3] or "").strip().upper()})
wb.close()
ramp_by_abbr = {w["영문약어"]: w for w in ramp_words if w["영문약어"]}

# === MOIS 사전 ===
wb = load_workbook(MOIS_WORD, read_only=True, data_only=True)
ws = wb[wb.sheetnames[0]]
head = list(next(ws.iter_rows(min_row=1, max_row=1, values_only=True)))
def col_idx(patterns):
    for i, h in enumerate(head):
        for p in patterns:
            if p in str(h or ""): return i
    return -1
i_nm = col_idx(["공통표준단어명", "단어명"])
i_ab = col_idx(["영문약어"])
mois = []
for r in ws.iter_rows(min_row=2, values_only=True):
    if not r or not r[i_nm]: continue
    mois.append({"단어명": str(r[i_nm]).strip(),
                 "영문약어": str(r[i_ab] or "").strip().upper()})
wb.close()

mois_by_abbr = {m["영문약어"]: m for m in mois if m["영문약어"]}
mois_by_name = defaultdict(list)
for m in mois:
    mois_by_name[m["단어명"]].append(m)
mois_abbrs = set(mois_by_abbr.keys())
mois_names = set(mois_by_name.keys())

# === Case 1 / Case 2 단어 분류 ===
# Case 1: RAMP 영문약어가 MOIS에 없고 RAMP 한글이 MOIS에 있음 → 영문 변경
# Case 2: RAMP 영문약어가 MOIS에 있고 한글이 다름 → 한글 변경
case1_map = {}   # RAMP_abbr → MOIS_abbr (영문 변경 매핑)
case2_map = {}   # 영문약어 → (RAMP_한글, MOIS_한글)  (한글 변경 매핑)

for w in ramp_words:
    a = w["영문약어"]; n = w["단어명"]
    if not a: continue
    if a in mois_abbrs:
        mm = mois_by_abbr[a]
        if mm["단어명"] != n:
            # Case 2 — 영문 같고 한글 다름
            case2_map[a] = (n, mm["단어명"])
    else:
        if n in mois_names:
            # Case 1
            mm = mois_by_name[n][0]
            case1_map[a] = mm["영문약어"]

print(f"Case 1 단어: {len(case1_map)}")
print(f"Case 2 단어: {len(case2_map)}")

# === RAMP 스키마 로드 ===
wb = load_workbook(SCHEMA, read_only=True, data_only=True)
ws = wb["컬럼"]
schema = []
for r in ws.iter_rows(min_row=2, values_only=True):
    if not r or not r[1]: continue
    schema.append({
        "tbl": (r[0] or "").strip(),
        "en": (r[1] or "").strip().lower(),
        "kr": (r[2] or "").strip(),
        "desc": (r[3] or "").strip() if len(r) > 3 else "",
        "null": r[4] or "",
        "type": r[5] or "",
        "len": r[6] or "",
        "pk": r[7] or "",
        "fk": r[8] or "",
        "order": r[10] or "",
    })
wb.close()

# 테이블 코멘트
tbl_comment = {}
wb = load_workbook(SCHEMA, read_only=True, data_only=True)
if "테이블 목록" in wb.sheetnames:
    ws = wb["테이블 목록"]
    for r in ws.iter_rows(min_row=2, values_only=True):
        if r and r[0]:
            tbl_comment[(r[0] or "").strip()] = (r[1] or "").strip() if len(r) > 1 else ""
wb.close()

# === BEFORE-AFTER 변환 ===
rows_out = []   # 영향받은 컬럼들
for c in schema:
    tokens = [t for t in re.split(r"_+", c["en"]) if t]
    tokens_up = [t.upper() for t in tokens]

    # 영문 변경 (Case 1)
    en_changes = []   # (token_idx, old_abbr, new_abbr)
    new_tokens = list(tokens)
    for i, tk in enumerate(tokens_up):
        if tk in case1_map:
            new_abbr = case1_map[tk]
            new_tokens[i] = new_abbr.lower()
            en_changes.append((i, tk, new_abbr))

    # 한글 변경 (Case 2) — 컬럼 영문에 case2_map 영문약어 토큰이 있으면 한글 코멘트가 RAMP 의미일 가능성
    # 단, 한글 코멘트 자체가 RAMP 단어인지 확인 어려우니, 컬럼 한글 코멘트에 RAMP 한글이 포함돼 있으면 치환
    kr_changes = []
    new_kr = c["kr"]
    for tk in tokens_up:
        if tk in case2_map:
            ramp_kr, mois_kr = case2_map[tk]
            if ramp_kr and ramp_kr in new_kr:
                new_kr = new_kr.replace(ramp_kr, mois_kr)
                kr_changes.append((tk, ramp_kr, mois_kr))

    if not en_changes and not kr_changes:
        continue

    new_en = "_".join(new_tokens)
    chg_type = ("EN" if en_changes else "") + ("+" if en_changes and kr_changes else "") + ("KR" if kr_changes else "")
    en_reason = "; ".join(f"{old}→{new}({ramp_by_abbr.get(old,{}).get('단어명','')})"
                          for _, old, new in en_changes)
    kr_reason = "; ".join(f"{tk}: {old}→{new}" for tk, old, new in kr_changes)

    rows_out.append({
        "tbl": c["tbl"],
        "tbl_cmt": tbl_comment.get(c["tbl"], ""),
        "type": c["type"], "len": c["len"], "null": c["null"], "pk": c["pk"], "fk": c["fk"],
        "before_en": c["en"], "after_en": new_en,
        "before_kr": c["kr"], "after_kr": new_kr,
        "chg_type": chg_type,
        "en_reason": en_reason, "kr_reason": kr_reason,
    })

# 통계
n_total = len(rows_out)
n_en_only = sum(1 for r in rows_out if r["chg_type"] == "EN")
n_kr_only = sum(1 for r in rows_out if r["chg_type"] == "KR")
n_both = sum(1 for r in rows_out if r["chg_type"] == "EN+KR")
tbl_affected = set(r["tbl"] for r in rows_out)
print(f"영향 컬럼: {n_total} (영문만 {n_en_only}, 한글만 {n_kr_only}, 양쪽 {n_both})")
print(f"영향 테이블: {len(tbl_affected)}")

# === xlsx 작성 ===
wb = Workbook()
hdr_font = Font(bold=True, color="FFFFFF")
hdr_fill = PatternFill("solid", fgColor="305496")
both_fill = PatternFill("solid", fgColor="FFE699")   # EN+KR 강조
en_fill   = PatternFill("solid", fgColor="DDEBF7")
kr_fill   = PatternFill("solid", fgColor="E2EFDA")
hdr_align = Alignment(horizontal="center", vertical="center")
thin = Side(border_style="thin", color="CCCCCC")
border = Border(left=thin, right=thin, top=thin, bottom=thin)

def style_header(ws, ncol):
    for c in range(1, ncol+1):
        cell = ws.cell(row=1, column=c)
        cell.font = hdr_font; cell.fill = hdr_fill; cell.alignment = hdr_align; cell.border = border

def auto_width(ws, widths):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

# 시트 1: 요약
ws = wb.active
ws.title = "요약"
ws.append(["항목", "값", "비고"])
ws.append(["원칙", "행안부 표준 우선 (단어 차원)", "타입/길이는 보류 — 용어/도메인 차원에서 별도"])
ws.append([])
ws.append(["RAMP 사전 단어", len(ramp_words), ""])
ws.append(["Case 1 단어 (한글동일/영문약어다름)", len(case1_map), "→ RAMP 컬럼 영문약어 변경"])
ws.append(["Case 2 단어 (영문동일/한글다름)", len(case2_map), "→ RAMP 컬럼 한글 코멘트 변경 (RAMP 의미가 MOIS 의미로 덮임 → 위험 가능)"])
ws.append([])
ws.append(["RAMP 스키마 컬럼", len(schema), ""])
ws.append(["변경 영향 컬럼", n_total, f"{n_total/len(schema)*100:.1f}%"])
ws.append(["  └ 영문약어만 변경", n_en_only, "Case 1 단독"])
ws.append(["  └ 한글만 변경", n_kr_only, "Case 2 단독"])
ws.append(["  └ 영문+한글 동시 변경", n_both, "Case 1·2 단어가 한 컬럼에 함께"])
ws.append(["변경 영향 테이블", len(tbl_affected), f"{len(tbl_affected)}/365 ({len(tbl_affected)/365*100:.1f}%)"])
style_header(ws, 3)
auto_width(ws, [35, 12, 60])
for row in ws.iter_rows(min_row=2, max_col=3):
    for cell in row:
        cell.alignment = Alignment(vertical="center", wrap_text=True)
ws.row_dimensions[1].height = 22

# 시트 2: BEFORE-AFTER 전수
ws = wb.create_sheet("BEFORE_AFTER_전수")
ws.append([
    "테이블명", "테이블 설명",
    "BEFORE 컬럼 영문", "AFTER 컬럼 영문",
    "BEFORE 컬럼 한글", "AFTER 컬럼 한글",
    "변경유형", "영문 변경 사유 (단어)", "한글 변경 사유 (단어)",
    "데이터타입", "길이", "NULL", "PK", "FK"
])
rows_sorted = sorted(rows_out, key=lambda r: (r["tbl"], r["before_en"]))
for r in rows_sorted:
    ws.append([
        r["tbl"], r["tbl_cmt"],
        r["before_en"], r["after_en"],
        r["before_kr"], r["after_kr"],
        r["chg_type"], r["en_reason"], r["kr_reason"],
        r["type"], r["len"], r["null"], r["pk"], r["fk"]
    ])
    # 색상 표시 (마지막 행)
    last = ws.max_row
    if r["chg_type"] == "EN+KR":
        for col in (3, 4, 5, 6, 7):
            ws.cell(row=last, column=col).fill = both_fill
    elif r["chg_type"] == "EN":
        for col in (3, 4, 7):
            ws.cell(row=last, column=col).fill = en_fill
    elif r["chg_type"] == "KR":
        for col in (5, 6, 7):
            ws.cell(row=last, column=col).fill = kr_fill
style_header(ws, 14)
auto_width(ws, [25, 28, 32, 32, 22, 22, 10, 40, 30, 12, 8, 8, 6, 6])
ws.freeze_panes = "C2"

# 시트 3: 테이블별 ALTER 가이드
ws = wb.create_sheet("테이블별_ALTER가이드")
ws.append(["테이블명", "테이블 설명", "변경 컬럼 수", "영문만", "한글만", "양쪽",
           "변경 컬럼 목록 (BEFORE → AFTER 요약)"])
tbl_grp = defaultdict(list)
for r in rows_out:
    tbl_grp[r["tbl"]].append(r)
for tn, items in sorted(tbl_grp.items(), key=lambda x: -len(x[1])):
    en_c = sum(1 for r in items if r["chg_type"] == "EN")
    kr_c = sum(1 for r in items if r["chg_type"] == "KR")
    both_c = sum(1 for r in items if r["chg_type"] == "EN+KR")
    summary_lines = []
    for r in items[:30]:  # 최대 30개까지 인라인 표시
        line = f"{r['before_en']}"
        if r['before_en'] != r['after_en']:
            line += f"→{r['after_en']}"
        if r['before_kr'] != r['after_kr']:
            line += f" [한글: {r['before_kr']}→{r['after_kr']}]"
        summary_lines.append(line)
    if len(items) > 30:
        summary_lines.append(f"... 외 {len(items)-30}건")
    ws.append([tn, tbl_comment.get(tn, ""), len(items), en_c, kr_c, both_c, "\n".join(summary_lines)])
style_header(ws, 7)
auto_width(ws, [25, 30, 14, 10, 10, 10, 80])
ws.freeze_panes = "A2"

wb.save(OUT)
print(f"\nSAVED: {OUT}")
print(f"size: {OUT.stat().st_size:,} bytes")
