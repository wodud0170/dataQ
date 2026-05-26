# -*- coding: utf-8 -*-
"""
Phase 2-2 시간 도메인 적용:
  1. 시간 도메인 5개 INSERT (tb_domain)
  2. RAMP 컬럼 정리 BEFORE/AFTER xlsx
     - 길이 ≤8 + suffix(일자/일시) → 일자V8, suffix '일자' 통일
     - 길이 9~14 + suffix(일자/일시) → 일시V14, suffix '일시' 통일
     - 그 외 (V17, V50, DATE, DATETIME 등) → 보류
"""
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from pathlib import Path
import subprocess, os, base64, csv, io

BASE = Path(r"C:\Users\장재영\Desktop\dataQ\CAMS표준화")
RAMP_SCHEMA = BASE / "CAMS_RAMP_통합" / "ramp기관스키마정보.xlsx"
APPLY_SQL = BASE / "04_RAMP분석_2026-05-21" / "99_실행SQL_기록" / "Phase2_time_domain_insert.sql"
OUT = BASE / "04_RAMP분석_2026-05-21" / "00_핵심산출" / "Phase2_시간컬럼_BEFORE_AFTER_2026-05-23.xlsx"

# ============ 1. 도메인 5개 INSERT SQL ============
print("=== 1. 시간 도메인 5개 INSERT SQL ===")
DOMS = [
    ("일자V8",   "일자", "VARCHAR",   8, "YYYYMMDD",          "YYYY-MM-DD", "YYYYMMDD 형식의 8자리 일자 (VARCHAR)"),
    ("일시V14",  "일시", "VARCHAR",  14, "YYYYMMDDHH24MISS",  "YYYY-MM-DD HH:MI:SS", "YYYYMMDDHH24MISS 형식의 14자리 일시 (VARCHAR)"),
    ("일자DT",   "일자", "DATE",      0, "YYYY-MM-DD",        "YYYY-MM-DD", "DATE 타입 일자 (Oracle DATE)"),
    ("일시DT",   "일시", "DATE",      0, "YYYYMMDDHH24MISS",  "YYYY-MM-DD HH:MI:SS", "DATE 타입 일시 (Oracle DATE, 시분초 포함)"),
    ("일시TS",   "일시", "TIMESTAMP", 0, "YYYYMMDDHH24MISS.FF", "YYYY-MM-DD HH:MI:SS.FF", "TIMESTAMP 타입 일시 (Oracle TIMESTAMP)"),
]
sql = ["-- Phase 2-2 시간 도메인 5건 INSERT", "BEGIN;", "",
       "-- 도메인분류 2건 (일자, 일시) 먼저 등록 (FK 충족)"]
for clsf_nm in ("일자", "일시"):
    cid = base64.urlsafe_b64encode(os.urandom(17)).rstrip(b'=').decode()[:22]
    sql.append(
        f"INSERT INTO quality.tb_domain_clsf (domain_clsf_id, domain_clsf_nm, domain_grp_nm, comm_stnd_yn, cret_dt, cret_user_id) "
        f"VALUES ('{cid}','{clsf_nm}','날짜/시간','N',to_char(now(),'YYYYMMDDHH24MISS'),'admin');"
    )
sql.append("")
sql.append("-- 도메인 5건")
for nm, clsf, dt, dl, stor, expr, desc in DOMS:
    did = base64.urlsafe_b64encode(os.urandom(17)).rstrip(b'=').decode()[:22]
    sql.append(
        f"INSERT INTO quality.tb_domain (domain_id, domain_nm, domain_grp_nm, domain_clsf_nm, domain_desc, data_type, data_len, data_decimal_len, stor_fmt, expr_fmt_lst, comm_stnd_yn, aprv_yn, cret_dt, cret_user_id, use_yn) "
        f"VALUES ('{did}','{nm}','날짜/시간','{clsf}','{desc}','{dt}',{dl},0,'{stor}',ARRAY['{expr}'],'N','Y',to_char(now(),'YYYYMMDDHH24MISS'),'admin','Y');"
    )
sql += ["", "SELECT domain_nm, domain_clsf_nm, data_type, data_len, stor_fmt FROM quality.tb_domain WHERE domain_nm IN ('일자V8','일시V14','일자DT','일시DT','일시TS') ORDER BY domain_nm;", "", "COMMIT;"]
APPLY_SQL.write_text("\n".join(sql), encoding="utf-8")
print(f"  → {APPLY_SQL}")

# ============ 2. RAMP 컬럼 정리 BEFORE/AFTER ============
print("\n=== 2. RAMP 컬럼 정리 ===")
wb = load_workbook(RAMP_SCHEMA, read_only=True, data_only=True)
ws = wb["컬럼"]
cols = []
for r in ws.iter_rows(min_row=2, values_only=True):
    if r and r[0] and r[1]:
        kr = str(r[2] or "").strip()
        if kr.endswith(("일자","일시")):
            cols.append({
                "tbl": str(r[0]).strip(),
                "col_en": str(r[1]).strip(),
                "col_kr": kr,
                "dtype": str(r[5] or "").strip(),
                "dlen": str(r[6] or "").strip(),
                "null": str(r[4] or "").strip() if len(r) > 4 else "",
                "pk": str(r[7] or "").strip() if len(r) > 7 else "",
            })
wb.close()
print(f"  일자/일시 컬럼: {len(cols)}")

# 분류
def parse_int(s):
    try: return int(s)
    except: return -1

apply_rows = []  # 적용 (V8 or V14)
suffix_swap = []  # 한글 suffix 변경
pending = []      # 보류

for c in cols:
    L = parse_int(c["dlen"])
    suf = "일자" if c["col_kr"].endswith("일자") else "일시"
    orig_type = f"{c['dtype']}({c['dlen']})"
    note = ""
    if c["dtype"] == "DATETIME":
        # 일시TS
        new_kr = c["col_kr"]
        note = f"기존 {orig_type}"
        if suf == "일자":
            new_kr = new_kr[:-2] + "일시"
            suffix_swap.append({**c, "new_kr": new_kr, "swap": "일자→일시", "domain": "일시TS", "비고": note})
        else:
            apply_rows.append({**c, "new_kr": new_kr, "swap": "", "domain": "일시TS", "비고": note})
    elif c["dtype"] == "DATE":
        # 일자DT
        new_kr = c["col_kr"]
        note = f"기존 {orig_type}" if c["dlen"] not in ("0","") else ""
        if suf == "일시":
            new_kr = new_kr[:-2] + "일자"
            suffix_swap.append({**c, "new_kr": new_kr, "swap": "일시→일자", "domain": "일자DT", "비고": note})
        else:
            apply_rows.append({**c, "new_kr": new_kr, "swap": "", "domain": "일자DT", "비고": note})
    elif suf == "일시":
        # 모든 일시 (길이 무관) → 일시V14
        new_kr = c["col_kr"]
        if L != 14 or c["dtype"] != "VARCHAR":
            note = f"기존 {orig_type}"
        apply_rows.append({**c, "new_kr": new_kr, "swap": "", "domain": "일시V14", "비고": note})
    elif suf == "일자":
        # 일자: L≤8 → 일자V8 / L≥9 → 일시V14 (한글 일시로 변경)
        new_kr = c["col_kr"]
        if L <= 8:
            if L != 8 or c["dtype"] != "VARCHAR":
                note = f"기존 {orig_type}"
            apply_rows.append({**c, "new_kr": new_kr, "swap": "", "domain": "일자V8", "비고": note})
        else:  # L >= 9
            new_kr = new_kr[:-2] + "일시"
            note = f"기존 {orig_type}"
            suffix_swap.append({**c, "new_kr": new_kr, "swap": "일자→일시", "domain": "일시V14", "비고": note})
    else:
        pending.append({**c, "사유": f"분류 불명 ({orig_type})"})

print(f"  적용 (suffix 일치):   {len(apply_rows)}")
print(f"  suffix 변경 (일자↔일시): {len(suffix_swap)}")
print(f"  보류 (V17 등):       {len(pending)}")

# ============ xlsx 생성 ============
THIN = Side(border_style="thin", color="808080")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
HDR_FILL = PatternFill("solid", fgColor="1F4E78")
HDR_FONT = Font(name="맑은 고딕", size=10, color="FFFFFF", bold=True)
CELL_FONT = Font(name="맑은 고딕", size=10)
SECTION_FILL = PatternFill("solid", fgColor="D9E1F2")
SWAP_FILL = PatternFill("solid", fgColor="FFE699")
PEND_FILL = PatternFill("solid", fgColor="F8CBAD")

wb = Workbook()

# 시트1: 표지
ws = wb.active; ws.title = "표지"
ws.column_dimensions["A"].width = 22; ws.column_dimensions["B"].width = 100
ws.merge_cells("A1:B1")
t = ws.cell(row=1, column=1, value="RAMP 시간 컬럼 정리 BEFORE/AFTER (Phase 2-2)")
t.font = Font(name="맑은 고딕", size=18, bold=True, color="1F4E78")
t.alignment = Alignment(horizontal="center", vertical="center")
ws.row_dimensions[1].height = 36

def cr(r, label, value):
    ws.cell(row=r, column=1, value=label).fill = SECTION_FILL
    ws.cell(row=r, column=1).font = Font(name="맑은 고딕", size=10, bold=True)
    ws.cell(row=r, column=1).border = BORDER
    c = ws.cell(row=r, column=2, value=value); c.font = Font(name="맑은 고딕", size=11)
    c.border = BORDER; c.alignment = Alignment(vertical="center", wrap_text=True)

cr(3, "작성일", "2026-05-23")
cr(4, "대상", f"RAMP 시간 suffix 컬럼 {len(cols)}건 (~일자/일시)")
cr(5, "규칙", "길이 ≤8 → 일자V8 / 9~14 → 일시V14 / 그 외 보류")
cr(6, "suffix 변경", "일자↔일시 자동 보정 (예: STRING(14) 시행일자 → 시행일시)")
cr(7, "Cubrid→Oracle", "ALTER 없음. BEFORE/AFTER만 보존")
ws.row_dimensions[8].height = 8
cr(9,  "신규 도메인 (이 작업과 동시 등록)", "일자V8, 일시V14, 일자DT, 일시DT, 일시TS")
cr(10, "적용 (suffix 일치)", f"{len(apply_rows)}건")
cr(11, "suffix 변경 (일자↔일시)", f"{len(suffix_swap)}건")
cr(12, "보류 (V17, V50, DATE, DATETIME 등)", f"{len(pending)}건 — 데이터 확인 후 결정")

# 시트2: 요약
ws2 = wb.create_sheet("요약")
ws2.column_dimensions["A"].width = 26; ws2.column_dimensions["B"].width = 10; ws2.column_dimensions["C"].width = 50
t2 = ws2.cell(row=1, column=1, value="시간 컬럼 정리 통계"); t2.font = Font(name="맑은 고딕", size=14, bold=True, color="1F4E78")
ws2.merge_cells("A1:C1")
r = 3
for cc, h in enumerate(["항목","건수","비고"], 1):
    cell = ws2.cell(row=r, column=cc, value=h); cell.fill = HDR_FILL; cell.font = HDR_FONT
    cell.alignment = Alignment(horizontal="center"); cell.border = BORDER
r += 1
for label, n, note in [
    ("전체 일자/일시 컬럼", len(cols), ""),
    ("일자V8 적용", sum(1 for x in apply_rows if x["domain"]=="일자V8"), ""),
    ("일시V14 적용", sum(1 for x in apply_rows if x["domain"]=="일시V14"), ""),
    ("일자→일시 변경", sum(1 for x in suffix_swap if x["swap"]=="일자→일시"), "길이 9~14인 ~일자"),
    ("일시→일자 변경", sum(1 for x in suffix_swap if x["swap"]=="일시→일자"), "길이 ≤8인 ~일시"),
    ("보류", len(pending), ""),
]:
    ws2.cell(row=r, column=1, value=label).border = BORDER
    ws2.cell(row=r, column=2, value=n).border = BORDER
    ws2.cell(row=r, column=2).alignment = Alignment(horizontal="right")
    ws2.cell(row=r, column=3, value=note).border = BORDER
    r += 1

# 시트3: 적용 전수 — 비고 + 색 구분
ws3 = wb.create_sheet("적용_BEFORE_AFTER")
H = ["No","테이블","컬럼영문","BEFORE 한글","AFTER 한글","BEFORE 타입","BEFORE 길이","AFTER 도메인","변경","비고(원본 보존)"]
W = [5,28,32,22,22,12,8,16,10,28]
for i, w in enumerate(W, 1): ws3.column_dimensions[get_column_letter(i)].width = w
for i, h in enumerate(H, 1):
    c = ws3.cell(row=1, column=i, value=h); c.fill = HDR_FILL; c.font = HDR_FONT
    c.alignment = Alignment(horizontal="center"); c.border = BORDER
ws3.freeze_panes = "C2"

# 색 정의:
#   - 흰색: 표준 그대로 (VARCHAR(8)/(14))
#   - 노란색: suffix 변경 (일자↔일시) — 한글 컬럼명 변경
#   - 주황색: 길이/타입 변경 (이관 시 데이터 트림/변환 가능성)
TYPE_FILL = PatternFill("solid", fgColor="FFC7AA")  # 주황 — 길이/타입 변경

i = 0
for x in apply_rows + suffix_swap:
    i += 1
    row_num = i + 1
    chg = x["swap"] if x["swap"] else ("도메인만" if not x["비고"] else "타입변경")
    vals = [i, x["tbl"], x["col_en"], x["col_kr"], x["new_kr"], x["dtype"], x["dlen"], x["domain"], chg, x["비고"]]
    for j, v in enumerate(vals, 1):
        c = ws3.cell(row=row_num, column=j, value=v); c.font = CELL_FONT; c.border = BORDER
        c.alignment = Alignment(vertical="top", wrap_text=True)
    # 색 우선순위: suffix 변경(노란) > 타입 변경(주황)
    if x["swap"]:
        # 노란 — 한글 컬럼 변경
        ws3.cell(row=row_num, column=4).fill = SWAP_FILL
        ws3.cell(row=row_num, column=5).fill = SWAP_FILL
        ws3.cell(row=row_num, column=9).fill = SWAP_FILL
    if x["비고"]:
        # 주황 — 타입/길이 변경 (이관 주의)
        ws3.cell(row=row_num, column=6).fill = TYPE_FILL
        ws3.cell(row=row_num, column=7).fill = TYPE_FILL
        ws3.cell(row=row_num, column=10).fill = TYPE_FILL

# 시트4: 보류 (데이터 확인 필요)
ws4 = wb.create_sheet("보류_데이터확인필요")
H4 = ["No","테이블","컬럼영문","컬럼한글","타입","길이","Null","PK","사유","결정"]
W4 = [5,28,32,22,12,8,6,5,40,14]
for i, w in enumerate(W4, 1): ws4.column_dimensions[get_column_letter(i)].width = w
for i, h in enumerate(H4, 1):
    c = ws4.cell(row=1, column=i, value=h); c.fill = HDR_FILL; c.font = HDR_FONT
    c.alignment = Alignment(horizontal="center"); c.border = BORDER
ws4.freeze_panes = "A2"
for i, x in enumerate(pending, 1):
    vals = [i, x["tbl"], x["col_en"], x["col_kr"], x["dtype"], x["dlen"], x["null"], x["pk"], x["사유"], ""]
    for j, v in enumerate(vals, 1):
        c = ws4.cell(row=i+1, column=j, value=v); c.font = CELL_FONT; c.border = BORDER
        c.alignment = Alignment(vertical="top", wrap_text=True)
    ws4.cell(row=i+1, column=5).fill = PEND_FILL

wb.save(OUT)
print(f"\n→ {OUT}")
print(f"  시트: 표지 / 요약 / 적용_BEFORE_AFTER ({len(apply_rows)+len(suffix_swap)}) / 보류 ({len(pending)})")
