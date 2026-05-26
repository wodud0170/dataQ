# -*- coding: utf-8 -*-
"""
Phase 1 통계 + 신규 단어 N 적재 후보 추출.
기존 RAMP_BEFORE_AFTER xlsx의 BEFORE_AFTER_전수 시트 + Case2 77 신규약어 + RAMP only 495.
"""
from openpyxl import load_workbook
from pathlib import Path
from collections import Counter, defaultdict
import csv

BASE = Path(r"C:\Users\장재영\Desktop\dataQ\CAMS표준화")
BA = BASE / "04_RAMP분석_2026-05-21" / "RAMP_BEFORE_AFTER_2026-05-21.xlsx"

print("=== Phase 1 BEFORE-AFTER 통계 ===")
wb = load_workbook(BA, read_only=True, data_only=True)
ws = wb["BEFORE_AFTER_전수"]
rows = list(ws.iter_rows(min_row=2, values_only=True))
wb.close()

types = Counter(r[6] for r in rows if r[6])
tables = set(r[0] for r in rows if r[0])
en_change = sum(1 for r in rows if r[2] != r[3])
kr_change = sum(1 for r in rows if r[4] != r[5])
print(f"  변경 컬럼 총수: {len(rows)}")
print(f"  영문 변경: {en_change}  /  한글 변경: {kr_change}")
print(f"  변경 테이블 수: {len(tables)}")
print(f"  변경유형 분포: {dict(types)}")

# Case 1 (영문 변경) 단어별 빈도
print("\n=== Case 1 — 영문 변경 단어 빈도 상위 ===")
en_word_change = Counter()
for r in rows:
    sig = (r[7] or "").strip()  # "SMMR→SMRY(요약)" 식
    if sig: en_word_change[sig] += 1
for sig, n in en_word_change.most_common(15):
    print(f"  {sig:30s} {n}컬럼")

# Case 2 (한글 변경) 단어별 빈도
print("\n=== Case 2 — 한글 변경 단어 빈도 상위 ===")
kr_word_change = Counter()
for r in rows:
    sig = (r[8] or "").strip()
    if sig: kr_word_change[sig] += 1
for sig, n in kr_word_change.most_common(15):
    print(f"  {sig:30s} {n}컬럼")
