# -*- coding: utf-8 -*-
from openpyxl import load_workbook
from pathlib import Path
p = Path(r"C:\Users\장재영\Desktop\dataQ\CAMS표준화\04_RAMP분석_2026-05-21\RAMP_컬럼_BEFORE_AFTER_2026-05-23.xlsx")
wb = load_workbook(p, read_only=True, data_only=True)
print("시트:", wb.sheetnames)
target = None
for sn in wb.sheetnames:
    if "영문" in sn and "매핑" in sn:
        target = sn; break
ws = wb[target]
rows = []
for r in ws.iter_rows(min_row=2, values_only=True):
    if r and r[0] is not None:
        rows.append((r[1], r[2], r[3] or ""))
wb.close()
print(f"\n  순위   변환                          영향  사유")
print("  " + "-"*72)
for i, (sig, n, reason) in enumerate(rows[:30], 1):
    sigfmt = sig[:30] if sig else ""
    print(f"  {i:>3}    {sigfmt:<30}  {n:>4}  {reason}")
print(f"  ... 총 {len(rows)}개 단어 변환")
