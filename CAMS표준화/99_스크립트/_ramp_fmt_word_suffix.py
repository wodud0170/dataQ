# -*- coding: utf-8 -*-
"""
RAMP AFTER 컬럼 패치 + 용어사전 생성.

룰 (행안부 형식단어 545종 전체 기준):
  1. CHAR → VARCHAR (CLAUDE.md 정책)
  2. 한글 끝을 행안부 형식단어 545종 중 '가장 긴 것' 으로 매칭
     - 매칭됨: 영문이 그 형식단어 행안부 약어로 끝나야 함. 안 끝나면 정정 (잘못된 약어는 제거 후 올바른 약어 append)
     - 매칭 안됨 (R8 미종결): D5 6종 (내용/구분/명/값/여부/코드) 중 하나 자동 보충 (한글+영문 둘다)
  3. 영문이 이미 행안부 약어로 끝나면 변경 X
  4. 30자 초과 보고
"""
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from pathlib import Path
from collections import Counter

BASE = Path(r"C:\Users\장재영\Desktop\dataQ\CAMS표준화")
ROOT = BASE / "04_RAMP분석_2026-05-21" / "00_핵심산출"
COL_PATH      = ROOT / "RAMP_업로드_컬럼_2026-05-26.xlsx"         # 원본 (BEFORE)
COL_OUT_PATH  = ROOT / "RAMP_업로드_컬럼_2026-05-26_v2.xlsx"      # AFTER
TERMS_PATH    = ROOT / "RAMP_업로드_용어_2026-05-26.xlsx"
MOIS_SEED_DIR = BASE.parent / "q-center" / "src" / "main" / "resources" / "seed" / "행안부_공통표준"
MOIS_WORDS    = MOIS_SEED_DIR / "행정안전부_공공데이터 공통표준단어.xlsx"
MOIS_TERMS_UPLOAD = [MOIS_SEED_DIR / "용어사전_일괄등록_1.xlsx", MOIS_SEED_DIR / "용어사전_일괄등록_2.xlsx"]

# D5 R8 미종결 보충용 6종 (한글 끝이 행안부 형식단어 545종 어떤 것도 아닐 때)
FALLBACK_FW = [("내용","CN"), ("구분","SE"), ("명","NM"), ("값","VL"), ("여부","YN"), ("코드","CD")]

def s(v): return "" if v is None else str(v).strip()

# ============ 0. 행안부 형식단어 로드 ============
print("=== 0. 행안부 형식단어 545종 로드 ===")
wb = load_workbook(MOIS_WORDS, read_only=True, data_only=True)
ws = wb.active
fw_y = {}   # 한글 → 영문약어 (Y)
all_words = {}  # 한글 → 영문약어 (전체, 형식단어 매칭용)
for r in ws.iter_rows(min_row=2, values_only=True):
    if r and r[0] and r[1]:
        kr, en = s(r[0]), s(r[1])
        all_words[kr] = en
        if r[4] and s(r[4]) == "Y":
            fw_y[kr] = en
wb.close()
# 사용자 정책 D5 추가 형식단어 (행안부 N 이지만 형식단어로 인정)
EXTRA_FW = {"구분":"SE", "값":"VL", "기한":None, "빈도":None, "주기":None, "횟수":"NMTM", "등급":"GRD", "단계":None}
for kr, en in EXTRA_FW.items():
    if kr not in fw_y:
        fw_y[kr] = en if en else all_words.get(kr, "")
fw_sorted = sorted([k for k in fw_y if fw_y[k]], key=len, reverse=True)
print(f"  형식단어 (행안부 545 + D5 추가): {len(fw_sorted)}")

def match_fw(kr):
    """한글 끝과 가장 긴 매칭되는 행안부 형식단어 (한글, 영문)"""
    for f in fw_sorted:
        if kr.endswith(f):
            return (f, fw_y[f])
    return (None, None)

# ============ 1. AFTER 컬럼 로드 (원본) ============
print("\n=== 1. AFTER 컬럼 로드 (원본) ===")
wb = load_workbook(COL_PATH, data_only=True)
ws = wb.active
print(f"  rows: {ws.max_row}")

# ============ 2. CHAR→VARCHAR + 영문 정합 ============
print("\n=== 2. CHAR→VARCHAR + 영문 정합 (행안부 형식단어 기준) ===")
char_to_varchar = 0
patched_kr = 0   # D5 fallback 한글 보충
patched_en = 0   # 영문 약어 보충/정정
already_ok = 0
over30 = []
correct_sample = []
fallback_sample = []

for row in ws.iter_rows(min_row=2):
    if not row[3].value: continue
    # CHAR → VARCHAR
    if row[5].value and str(row[5].value).strip().upper() == "CHAR":
        row[5].value = "VARCHAR"
        char_to_varchar += 1
    en = s(row[3].value)
    kr = s(row[4].value)
    if not en or not kr: continue

    fw_kr, fw_en = match_fw(kr)
    if fw_kr:
        # 한글 끝이 행안부 형식단어 - 영문 약어 정합 확보
        sfx = "_" + fw_en
        if en.endswith(sfx) or en == fw_en:
            already_ok += 1
        else:
            # 영문 끝에 다른 약어가 있으면 떼어내고 올바른 약어로 교체
            # (보수적: 단순히 append 만; 잘못된 약어 제거 룰은 위험)
            # 단, '_NM' 등 다른 형식단어 약어가 끝에 붙어있으면 떼어냄
            other_abrvs = set(fw_y.values())
            parts = en.split("_")
            while len(parts) > 1 and parts[-1] in other_abrvs:
                parts.pop()
            base = "_".join(parts)
            new_en = base + sfx
            if len(new_en) > 30:
                over30.append((s(row[1].value), en, new_en, len(new_en), kr))
            if len(correct_sample) < 10:
                correct_sample.append(f"{en} → {new_en}  ({kr}, fw={fw_kr})")
            row[3].value = new_en
            patched_en += 1
    else:
        # R8 미종결 - D5 6종 fallback 보충 (한글 + 영문)
        # 한글 끝 데이터타입 기반으로 형식단어 선택
        dt = s(row[5].value).upper() if row[5].value else ""
        if dt == "NUMBER":
            fb_kr, fb_en = "값", "VL"
        elif dt in ("DATE","TIMESTAMP"):
            fb_kr, fb_en = "일자" if dt=="DATE" else "일시", "DT" if dt=="DATE" else "TS"
        else:
            fb_kr, fb_en = "내용", "CN"
        new_kr = kr + fb_kr
        sfx = "_" + fb_en
        new_en = en + sfx if not en.endswith(sfx) else en
        if len(new_en) > 30:
            over30.append((s(row[1].value), en, new_en, len(new_en), kr))
        if len(fallback_sample) < 10:
            fallback_sample.append(f"{en}/{kr} → {new_en}/{new_kr}")
        row[3].value = new_en
        row[4].value = new_kr
        patched_kr += 1
        patched_en += 1

print(f"  CHAR → VARCHAR: {char_to_varchar}")
print(f"  영문 약어 정합 (행안부 545종): patched={patched_en}, 이미 OK={already_ok}")
print(f"  D5 fallback 보충 (한글+영문): {patched_kr}")
print(f"  30자 초과: {len(over30)}")
print("\n  정합 샘플:")
for s_ in correct_sample: print(f"    {s_}")
print("\n  fallback 샘플:")
for s_ in fallback_sample: print(f"    {s_}")
if over30:
    print("\n  ⚠️ 30자 초과:")
    for tbl, en, new_en, L, kr in over30[:10]:
        print(f"    {tbl}.{en} → {new_en} ({L}자) [{kr}]")

# ============ 3. AFTER 컬럼 저장 (v2) ============
print(f"\n=== 3. {COL_OUT_PATH.name} 저장 ===")
wb.save(COL_OUT_PATH)
print(f"  → {COL_OUT_PATH}")
wb.close()

# ============ 4. 용어 후보 추출 ============
print("\n=== 4. 용어 후보 추출 ===")
wb = load_workbook(COL_OUT_PATH, read_only=True, data_only=True)
ws = wb.active
terms_cand = {}
for r in ws.iter_rows(min_row=2, values_only=True):
    if not r or not r[3] or not r[4]: continue
    en, kr = s(r[3]), s(r[4])
    dt, dl, ddec = s(r[5]), s(r[6]), s(r[7])
    key = (en, kr)
    if key not in terms_cand:
        terms_cand[key] = (dt, dl, ddec)
wb.close()
print(f"  컬럼 distinct (영문,한글): {len(terms_cand)}")

# ============ 5. 행안부 표준 용어 매칭 ============
print("\n=== 5. 행안부 표준 용어 매칭 ===")
mois = {}
for p in MOIS_TERMS_UPLOAD:
    wb = load_workbook(p, read_only=True, data_only=True)
    ws = wb.active
    for r in ws.iter_rows(min_row=2, values_only=True):
        if r and r[2] and r[4]:
            mois[(s(r[2]), s(r[4]))] = s(r[5])
    wb.close()
print(f"  행안부 용어: {len(mois)}")
mois_hit = sum(1 for en, kr in terms_cand if (kr, en) in mois)
print(f"  행안부 매칭: {mois_hit}")
print(f"  RAMP N: {len(terms_cand) - mois_hit}")

# ============ 6. 도메인 매핑 ============
def map_domain(kr, dt, dl):
    if not dt: return ""
    dt_u = dt.upper()
    if dt_u == "VARCHAR":
        if kr.endswith("여부"): return f"여부V{dl}" if dl else "여부V"
        if kr.endswith("코드"): return f"코드V{dl}" if dl else "코드V"
        if kr.endswith("구분"): return f"구분V{dl}" if dl else "구분V"
        if kr.endswith("명"):   return f"명V{dl}"   if dl else "명V"
        if kr.endswith("값"):   return f"값V{dl}"   if dl else "값V"
        if kr.endswith("내용"): return f"내용V{dl}" if dl else "내용V"
        if kr.endswith("일자") and dl == "8":  return "일자V8"
        if kr.endswith("일시") and dl == "14": return "일시V14"
        if kr.endswith("연도"): return "연도V4"
        return f"내용V{dl}" if dl else "내용V"
    if dt_u == "NUMBER":
        if kr.endswith("값"): return f"값N{dl}"
        if kr.endswith("수"): return f"수N{dl}" if dl else "수N"
        return f"값N{dl}" if dl else "값N"
    if dt_u == "DATE":     return "일자DT"
    if dt_u == "TIMESTAMP": return "일시TS"
    if dt_u == "CLOB":     return "내용L"
    return f"{dt_u}{dl}"

# ============ 7. 용어사전 xlsx 생성 ============
print("\n=== 7. 용어사전 xlsx 생성 ===")
wb = Workbook()
ws = wb.active; ws.title = "Sheet1"
H = ["No","제정차수","용어명","용어설명","용어영문약어명","도메인명","허용값","저장형식","표현형식","코드그룹명","소관기관명","이음동의어목록","요청시스템","표준여부"]
W = [5, 8, 30, 40, 22, 16, 10, 18, 12, 12, 14, 20, 12, 8]
for i, h in enumerate(H, 1):
    c = ws.cell(row=1, column=i, value=h)
    c.font = Font(name="맑은 고딕", size=10, bold=True, color="FFFFFF")
    c.fill = PatternFill("solid", fgColor="1F4E78")
    c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
for i, w in enumerate(W, 1): ws.column_dimensions[get_column_letter(i)].width = w
ws.freeze_panes = "A2"
THIN = Side(border_style="thin", color="808080")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
no = 0
ramp_n_rows = []
for (en, kr), (dt, dl, ddec) in sorted(terms_cand.items(), key=lambda x: (x[0][1], x[0][0])):
    if (kr, en) in mois: continue
    no += 1
    dom = map_domain(kr, dt, dl)
    ramp_n_rows.append([no, "RAMP", kr, "", en, dom, "", "", "", "", "RAMP", "", "RAMP", "N"])
for r_idx, vals in enumerate(ramp_n_rows, 2):
    for c_idx, v in enumerate(vals, 1):
        c = ws.cell(row=r_idx, column=c_idx, value=v)
        c.font = Font(name="맑은 고딕", size=10)
        c.alignment = Alignment(vertical="top", wrap_text=True)
        c.border = BORDER
wb.save(TERMS_PATH)
print(f"  → {TERMS_PATH}  (RAMP N: {len(ramp_n_rows)}건)")
