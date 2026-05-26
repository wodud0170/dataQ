# -*- coding: utf-8 -*-
"""
CAMS 144건 (space@2026-05-19 적재분 중 '질의' 1건 제외) 백업 + DELETE.
이전 skip 79건 RAMP 추가 INSERT.

최종 DB 목표: 행안부 Y 3,277 + RAMP N 562 = 3,839
"""
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from pathlib import Path
import subprocess, csv, io, sys, re, os, base64

BASE = Path(r"C:\Users\장재영\Desktop\dataQ\CAMS표준화")
DEC = BASE / "04_RAMP분석_2026-05-21" / "RAMP_단어결정_2026-05-23.xlsx"
RAMP_DICT = BASE / "01_원본자료" / "RMS4-DE14-04.데이터표준화_단어사전-V2.0.xlsx"
MOIS_DICT = BASE.parent / "q-center" / "src" / "main" / "resources" / "seed" / "행안부_공통표준" / "행정안전부_공공데이터 공통표준단어.xlsx"
BACKUP_XLSX = BASE / "04_RAMP분석_2026-05-21" / "CAMS_단어_백업_2026-05-23.xlsx"
APPLY_SQL = BASE / "04_RAMP분석_2026-05-21" / "Phase1_purge_cams_apply.sql"

R5 = re.compile(r"^[A-Z0-9]+$")

# ============ 1. CAMS 144건 백업 ============
print("=== 1. CAMS 백업 ===")
r = subprocess.run(["docker","exec","-i","dataq-db","psql","-U","admin","-d","postgres",
    "-c", """COPY (SELECT word_id, word_nm, word_eng_abrv_nm, coalesce(word_eng_nm,''),
             coalesce(word_desc,''), coalesce(word_clsf_yn,''), coalesce(domain_clsf_nm,''),
             coalesce(comm_stnd_yn,''), coalesce(aprv_yn,''), cret_dt, cret_user_id
             FROM quality.tb_word WHERE comm_stnd_yn='N' AND cret_user_id='space'
             AND cret_dt LIKE '20260519%' ORDER BY word_nm) TO STDOUT WITH (FORMAT csv, FORCE_QUOTE *)"""],
    capture_output=True, encoding="utf-8")
cams_rows = [row for row in csv.reader(io.StringIO(r.stdout)) if row]
print(f"  CAMS 백업 대상: {len(cams_rows)}건")

wb = Workbook()
ws = wb.active; ws.title = "CAMS_단어_144"
H = ["word_id","word_nm","word_eng_abrv_nm","word_eng_nm","word_desc","word_clsf_yn","domain_clsf_nm","comm_stnd_yn","aprv_yn","cret_dt","cret_user_id"]
THIN = Side(border_style="thin", color="808080")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
for i, h in enumerate(H, 1):
    c = ws.cell(row=1, column=i, value=h)
    c.fill = PatternFill("solid", fgColor="C00000"); c.font = Font(name="맑은 고딕", size=10, color="FFFFFF", bold=True)
    c.alignment = Alignment(horizontal="center"); c.border = BORDER
    ws.column_dimensions[get_column_letter(i)].width = 22 if i in (1,4,5) else 14
ws.freeze_panes = "A2"
for i, row in enumerate(cams_rows, 2):
    for j, v in enumerate(row, 1):
        c = ws.cell(row=i, column=j, value=v); c.font = Font(name="맑은 고딕", size=10); c.border = BORDER
        c.alignment = Alignment(vertical="top", wrap_text=True)
wb.save(BACKUP_XLSX)
print(f"  → {BACKUP_XLSX}")

# ============ 2. 추가 INSERT 79건 추출 ============
print("\n=== 2. 79건 (이전 skip) 재추출 ===")
# A 30건
A_rows = []
wb = load_workbook(DEC, read_only=True, data_only=True)
for r in wb["Case2_신규약어"].iter_rows(min_row=2, values_only=True):
    if not r or not r[0]: continue
    nm = str(r[1] or "").strip()
    eng = str(r[3] or "").strip()
    desc = str(r[4] or "").strip()
    decide = str(r[12] or "").strip()
    if decide and R5.match(decide) and 3 <= len(decide) <= 6:
        A_rows.append({"nm": nm, "abrv": decide, "eng": eng, "desc": desc, "frmt":"", "dmn":""})
wb.close()

# RAMP only 529건 + RAMP 사전 전체 (CLASS·스캔·컷 포함)
mois_abrv=set(); mois_nm=set()
wb = load_workbook(MOIS_DICT, read_only=True, data_only=True)
for r in wb["Sheet"].iter_rows(min_row=2, values_only=True):
    if r and r[0] and r[1]:
        mois_abrv.add(str(r[1]).strip()); mois_nm.add(str(r[0]).strip())
wb.close()

ramp_dict = {}  # nm -> meta
wb = load_workbook(RAMP_DICT, read_only=True, data_only=True)
for r in wb["단어사전"].iter_rows(min_row=2, values_only=True):
    if r and r[2] and r[3]:
        nm = str(r[2]).strip(); abrv = str(r[3]).strip()
        ramp_dict[nm] = {
            "nm": nm, "abrv": abrv,
            "eng": str(r[4] or "").strip().replace("_x000D_",""),
            "desc": str(r[5] or "").strip().replace("_x000D_",""),
            "frmt": str(r[6] or "").strip(),
            "dmn": str(r[7] or "").strip(),
        }
wb.close()

# 79건 = 이전 skip 분 (한글 중복 64 + CLASS·스캔·컷 3 + 동일 12)
# A_rows + RAMP only를 다시 만들고 → 사용자 SKIP_NMS와 ABRV_OVERRIDE 적용
# 이번엔 CAMS가 빠질 예정이므로 충돌 검사는 SKIP한 79건만 추가
# 단순화: 위 _phase1_n_insert.py 와 동일 로직으로 559건 산출 후, 이전 INSERT 된 480건 제외 = 79건
ABRV_OVERRIDE = {"디렉토리": "DIR", "분": "MINUTE"}
SKIP_NMS = set()  # CAMS 삭제할 거니까 skip 없음. CLASS·스캔·컷도 등록.

ramp_only = [v for k,v in ramp_dict.items() if k not in mois_nm and v["abrv"] not in mois_abrv]
all_insert = A_rows + ramp_only
for x in all_insert:
    if x["nm"] in ABRV_OVERRIDE:
        x["abrv"] = ABRV_OVERRIDE[x["nm"]]

# 이미 INSERT 된 단어 한글 셋 확보 (admin 적재)
r = subprocess.run(["docker","exec","-i","dataq-db","psql","-U","admin","-d","postgres",
    "-c", """COPY (SELECT word_nm FROM quality.tb_word WHERE comm_stnd_yn='N' AND cret_user_id='admin')
             TO STDOUT WITH (FORMAT csv, FORCE_QUOTE *)"""],
    capture_output=True, encoding="utf-8")
admin_nms = set(row[0] for row in csv.reader(io.StringIO(r.stdout)) if row)
print(f"  이미 INSERT (admin): {len(admin_nms)}")

# 추가 INSERT 대상 = all_insert 중 admin에 없는 것
to_add = [x for x in all_insert if x["nm"] not in admin_nms]
print(f"  추가 INSERT 대상: {len(to_add)}")

# CLASS·스캔·컷이 RAMP 사전에 있다면 추가
EXTRA = ["CLASS", "스캔", "컷"]
for nm in EXTRA:
    if nm in ramp_dict and ramp_dict[nm]["nm"] not in admin_nms:
        # 이미 to_add 에 있는지 확인
        if not any(x["nm"] == nm for x in to_add):
            to_add.append(ramp_dict[nm])
            print(f"  추가: {nm}")

# ============ 3. SQL 생성 ============
print(f"\n=== 3. SQL 생성 ===")
sql_lines = [
    "-- CAMS 144건 DELETE + RAMP 추가 INSERT",
    f"-- DELETE: {len(cams_rows)}건 / INSERT: {len(to_add)}건",
    "BEGIN;",
    "",
    "-- 1) CAMS 144건 DELETE",
    "DELETE FROM quality.tb_word",
    "WHERE comm_stnd_yn='N' AND cret_user_id='space' AND cret_dt LIKE '20260519%';",
    "",
    "-- 2) 추가 RAMP 단어 INSERT (CAMS 충돌 없으니 모두 가능)",
]
for w in to_add:
    nm = w["nm"].replace("'","''")
    abrv = w["abrv"].replace("'","''")
    eng = w["eng"].replace("'","''")
    desc = (w["desc"] or w["nm"]).replace("'","''")
    frmt = w["frmt"] if w["frmt"] in ("Y","N") else "N"
    dmn = w["dmn"].replace("'","''")
    wid = base64.urlsafe_b64encode(os.urandom(17)).rstrip(b'=').decode()[:22]
    sql_lines.append(f"INSERT INTO quality.tb_word (word_id,word_nm,word_eng_abrv_nm,word_eng_nm,word_desc,word_clsf_yn,domain_clsf_nm,comm_stnd_yn,aprv_yn,cret_dt,cret_user_id,use_yn) VALUES "
                    f"('{wid}','{nm}','{abrv}','{eng}','{desc}','{frmt}','{dmn}','N','Y',to_char(now(),'YYYYMMDDHH24MISS'),'admin','Y');")

sql_lines += [
    "",
    "-- 검증",
    "SELECT 'tb_word Y' as t, count(*) FROM quality.tb_word WHERE comm_stnd_yn='Y';",
    "SELECT 'tb_word N' as t, count(*) FROM quality.tb_word WHERE comm_stnd_yn='N';",
    "SELECT 'CAMS 남은 (space)' as t, count(*) FROM quality.tb_word WHERE cret_user_id='space';",
    "",
    "COMMIT;",
]
APPLY_SQL.write_text("\n".join(sql_lines), encoding="utf-8")
print(f"  → {APPLY_SQL}")
print(f"\n예상 최종: tb_word Y=3,277 / N=483(기존)+{len(to_add)}(추가)={483+len(to_add)}")
