# -*- coding: utf-8 -*-
"""
Phase 2-2 D1 — 다중도메인 한글명 그룹 통일 (최장 우선).
정책:
  - 그룹별 최장 도메인 산출
  - 행안부 매칭 시도 (같은 한글 도메인분류 + 같은 타입 + 같은 길이)
  - 매칭 시 행안부 흡수, 없으면 기관표준 도메인 신규 추가 (D6 정책)
  - 데이터 트림 절대 X
"""
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from pathlib import Path
from collections import defaultdict, Counter
import subprocess, csv, io, sys, os, base64, re

BASE = Path(r"C:\Users\장재영\Desktop\dataQ\CAMS표준화")
RAMP_SCHEMA = BASE / "CAMS_RAMP_통합" / "ramp기관스키마정보.xlsx"
OUT_XLSX = BASE / "04_RAMP분석_2026-05-21" / "00_핵심산출" / "Phase2_D1_도메인통일_2026-05-23.xlsx"
OUT_SQL = BASE / "04_RAMP분석_2026-05-21" / "99_실행SQL_기록" / "Phase2_D1_domain_insert.sql"

# ============ 1. RAMP 컬럼 로드 ============
print("=== 1. RAMP 컬럼 로드 ===")
wb = load_workbook(RAMP_SCHEMA, read_only=True, data_only=True)
ws = wb["컬럼"]
cols = []
for r in ws.iter_rows(min_row=2, values_only=True):
    if r and r[0] and r[1]:
        kr = str(r[2] or "").strip()
        if kr:
            cols.append({
                "tbl": str(r[0]).strip(),
                "col_en": str(r[1]).strip(),
                "col_kr": kr,
                "dtype": str(r[5] or "").strip(),
                "dlen": str(r[6] or "").strip(),
            })
wb.close()
print(f"  RAMP 컬럼: {len(cols)}")

# 한글 그룹화
grp = defaultdict(list)
for c in cols:
    grp[c["col_kr"]].append(c)

# 다중도메인 그룹
def parse_int(s):
    try: return int(s)
    except: return -1

# 시간 컬럼은 D3에서 이미 처리됨 → 제외
TIME_SUFFIX = ("일자","일시","년월","연월","시분초","시분","시각","연도","월일","월")
def is_time(kr):
    return any(kr.endswith(s) for s in TIME_SUFFIX)

multi = []
for kr, lst in grp.items():
    if is_time(kr): continue  # D3 처리 완료
    domains = set((c["dtype"], parse_int(c["dlen"])) for c in lst)
    if len(domains) > 1:
        multi.append({"kr": kr, "cols": lst, "domains": domains})

multi.sort(key=lambda x: -len(x["cols"]))
print(f"  다중도메인 (시간 제외): {len(multi)}종")

# ============ 2. 행안부 도메인 로드 ============
print("\n=== 2. 행안부 도메인 로드 ===")
r = subprocess.run(["docker","exec","-i","dataq-db","psql","-U","admin","-d","postgres",
    "-c", """COPY (SELECT domain_nm, domain_clsf_nm, data_type, coalesce(data_len::text,'0')
             FROM quality.tb_domain WHERE comm_stnd_yn='Y') TO STDOUT WITH (FORMAT csv, FORCE_QUOTE *)"""],
    capture_output=True, encoding="utf-8")
mois_dom = []
mois_by_clsf = defaultdict(list)  # 분류명 → [(nm, type, len)]
for row in csv.reader(io.StringIO(r.stdout)):
    if row:
        mois_dom.append(row)
        mois_by_clsf[row[1]].append((row[0], row[2], parse_int(row[3])))
print(f"  행안부 도메인: {len(mois_dom)}")

# 기관표준 도메인 (N) — Phase 2-2 시간 5건 포함
r = subprocess.run(["docker","exec","-i","dataq-db","psql","-U","admin","-d","postgres",
    "-c", """COPY (SELECT domain_nm, domain_clsf_nm, data_type, coalesce(data_len::text,'0')
             FROM quality.tb_domain WHERE comm_stnd_yn='N') TO STDOUT WITH (FORMAT csv, FORCE_QUOTE *)"""],
    capture_output=True, encoding="utf-8")
n_dom = list(csv.reader(io.StringIO(r.stdout)))
print(f"  기관표준 도메인(N): {len(n_dom)}")
all_existing_nms = set(row[0] for row in mois_dom) | set(row[0] for row in n_dom)

# ============ 3. 처리 ============
print("\n=== 3. 처리 ===")
# 한글명에서 형식단어(분류명) 추출 — 마지막 토큰
def extract_clsf(kr):
    # 길이 짧은 형식단어 우선 (예: "주소", "코드", "번호", "명")
    # 행안부 분류명에 있는 것 중 kr 끝과 일치
    candidates = []
    for c in mois_by_clsf.keys():
        if kr.endswith(c):
            candidates.append(c)
    if candidates:
        return max(candidates, key=len)
    return None

# 타입약자 매핑
TYPE_ABBR = {"VARCHAR":"V", "VARCHAR2":"V", "STRING":"V", "CHAR":"C", "NUMERIC":"N",
             "NUMBER":"N", "INTEGER":"N", "DATE":"D", "DATETIME":"D", "TIMESTAMP":"T", "CLOB":"L"}

def gen_domain_name(clsf, dt, dl):
    abbr = TYPE_ABBR.get(dt.upper(), dt[:1])
    if dl > 0:
        return f"{clsf}{abbr}{dl}"
    return f"{clsf}{abbr}"

absorb = []   # 행안부 흡수 (자동 적용)
new_dom = []  # 신규 기관표준 도메인
skip = []     # 처리 불가 (분류 추출 실패 등)
to_register = {}  # 신규 도메인명 → (clsf, dt, dl)

for x in multi:
    kr = x["kr"]
    # 최장 도메인 (NUMERIC 우선순위 등은 무시하고 길이로만)
    longest = max(x["domains"], key=lambda d: d[1])
    long_dt, long_dl = longest
    clsf = extract_clsf(kr)
    if not clsf:
        skip.append({**x, "사유": "행안부 분류명 매칭 실패"})
        continue
    # 행안부 같은 분류명 + 같은 타입 + 같은 길이 매칭
    matched = None
    for nm, dt2, dl2 in mois_by_clsf[clsf]:
        if dt2.upper() == long_dt.upper() and dl2 == long_dl:
            matched = nm; break
    if matched:
        absorb.append({**x, "longest": longest, "clsf": clsf, "matched_dom": matched})
    else:
        # 기관표준 신규 도메인 필요
        new_nm = gen_domain_name(clsf, long_dt, long_dl)
        # 중복 방지
        suffix_i = 1
        base_nm = new_nm
        while new_nm in all_existing_nms or new_nm in to_register:
            suffix_i += 1
            new_nm = f"{base_nm}_{suffix_i}"
        to_register[new_nm] = (clsf, long_dt, long_dl)
        new_dom.append({**x, "longest": longest, "clsf": clsf, "new_dom": new_nm})

print(f"  자동 흡수 (행안부): {len(absorb)}")
print(f"  신규 기관표준 도메인: {len(new_dom)} (unique: {len(to_register)})")
print(f"  처리 불가: {len(skip)}")

# ============ 4. SQL 생성 ============
print(f"\n=== 4. SQL 생성 ({len(to_register)} INSERT) ===")
sql = ["-- Phase 2-2 D1 — 기관표준 도메인 신규 INSERT (다중도메인 통일용)",
       "BEGIN;", ""]

# 새 분류명 필요 여부 점검
existing_clsf = set(c for c in mois_by_clsf.keys()) | set(row[1] for row in n_dom)
new_clsfs = set(v[0] for v in to_register.values()) - existing_clsf
if new_clsfs:
    sql.append(f"-- 신규 도메인분류 {len(new_clsfs)}건")
    for cn in sorted(new_clsfs):
        cid = base64.urlsafe_b64encode(os.urandom(17)).rstrip(b'=').decode()[:22]
        # 그룹 추정 (분류명에서 — 일단 '기타' 또는 매칭 시도)
        grp_name = "기타"
        sql.append(f"INSERT INTO quality.tb_domain_clsf (domain_clsf_id, domain_clsf_nm, domain_grp_nm, comm_stnd_yn, cret_dt, cret_user_id) "
                   f"VALUES ('{cid}','{cn}','{grp_name}','N',to_char(now(),'YYYYMMDDHH24MISS'),'admin');")
    sql.append("")

sql.append(f"-- 신규 기관표준 도메인 {len(to_register)}건")
for nm, (clsf, dt, dl) in sorted(to_register.items()):
    did = base64.urlsafe_b64encode(os.urandom(17)).rstrip(b'=').decode()[:22]
    desc = f"{clsf} {dt}({dl}) — RAMP 최장 통일용 기관표준"
    sql.append(
        f"INSERT INTO quality.tb_domain (domain_id, domain_nm, domain_grp_nm, domain_clsf_nm, domain_desc, data_type, data_len, data_decimal_len, stor_fmt, expr_fmt_lst, comm_stnd_yn, aprv_yn, cret_dt, cret_user_id, use_yn) "
        f"VALUES ('{did}','{nm}','기타','{clsf}','{desc.replace(chr(39),chr(39)+chr(39))}','{dt}',{dl},0,'',ARRAY[]::text[],'N','Y',to_char(now(),'YYYYMMDDHH24MISS'),'admin','Y');"
    )

sql += ["", f"SELECT count(*) AS new_inserted FROM quality.tb_domain WHERE comm_stnd_yn='N' AND cret_dt LIKE to_char(now(),'YYYYMMDD')||'%';",
        "", "COMMIT;"]
OUT_SQL.write_text("\n".join(sql), encoding="utf-8")
print(f"  → {OUT_SQL}")

# ============ 5. xlsx ============
print(f"\n=== 5. xlsx 생성 ===")
wb = Workbook()
THIN = Side(border_style="thin", color="808080")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
HDR_FILL = PatternFill("solid", fgColor="1F4E78")
HDR_FONT = Font(name="맑은 고딕", size=10, color="FFFFFF", bold=True)
CELL_FONT = Font(name="맑은 고딕", size=10)
SECTION_FILL = PatternFill("solid", fgColor="D9E1F2")
ABSORB_FILL = PatternFill("solid", fgColor="C6E0B4")
NEW_FILL = PatternFill("solid", fgColor="FFE699")

# 표지
ws = wb.active; ws.title = "표지"
ws.column_dimensions["A"].width = 22; ws.column_dimensions["B"].width = 100
ws.merge_cells("A1:B1")
t = ws.cell(row=1, column=1, value="RAMP 도메인 통일 D1 (최장 + 행안부 흡수/기관표준 추가)")
t.font = Font(name="맑은 고딕", size=18, bold=True, color="1F4E78")
t.alignment = Alignment(horizontal="center", vertical="center")
ws.row_dimensions[1].height = 36

def cr(r, label, value):
    ws.cell(row=r, column=1, value=label).fill = SECTION_FILL
    ws.cell(row=r, column=1).font = Font(name="맑은 고딕", size=10, bold=True)
    ws.cell(row=r, column=1).border = BORDER
    c = ws.cell(row=r, column=2, value=value); c.font = Font(name="맑은 고딕", size=11)
    c.border = BORDER; c.alignment = Alignment(vertical="center", wrap_text=True)
cr(3, "작성일", "2026-05-23")
cr(4, "대상", f"RAMP 다중도메인 한글명 그룹 {len(multi)}종 (시간 제외)")
cr(5, "정책", "최장 우선 + 행안부 매칭 시 흡수, 없으면 기관표준 도메인 신규 (D6) — 데이터 트림 X")
cr(6, "행안부 흡수 (자동)", f"{len(absorb)}건")
cr(7, "기관표준 신규 도메인", f"{len(to_register)}건 INSERT")
cr(8, "처리 불가 (분류 추출 실패)", f"{len(skip)}건")

# 시트2: 행안부 흡수
ws2 = wb.create_sheet("행안부흡수")
H = ["No","한글명","컬럼수","최장 도메인","행안부 매칭 도메인","컬럼 분포"]
W = [5,20,8,18,18,55]
for i, w in enumerate(W, 1): ws2.column_dimensions[get_column_letter(i)].width = w
for i, h in enumerate(H, 1):
    c = ws2.cell(row=1, column=i, value=h); c.fill = HDR_FILL; c.font = HDR_FONT
    c.alignment = Alignment(horizontal="center"); c.border = BORDER
ws2.freeze_panes = "A2"
for i, x in enumerate(absorb, 1):
    dist = " | ".join(f"{dt}({dl})" for dt, dl in sorted(x["domains"], key=lambda d: -d[1]))
    long_str = f"{x['longest'][0]}({x['longest'][1]})"
    vals = [i, x["kr"], len(x["cols"]), long_str, x["matched_dom"], dist]
    for j, v in enumerate(vals, 1):
        c = ws2.cell(row=i+1, column=j, value=v); c.font = CELL_FONT; c.border = BORDER
        c.alignment = Alignment(vertical="top", wrap_text=True)
    ws2.cell(row=i+1, column=5).fill = ABSORB_FILL

# 시트3: 기관표준 신규
ws3 = wb.create_sheet("기관표준_신규도메인")
H3 = ["No","한글명","컬럼수","최장 도메인","신규 기관표준 도메인","분류명","컬럼 분포"]
W3 = [5,20,8,18,22,12,55]
for i, w in enumerate(W3, 1): ws3.column_dimensions[get_column_letter(i)].width = w
for i, h in enumerate(H3, 1):
    c = ws3.cell(row=1, column=i, value=h); c.fill = HDR_FILL; c.font = HDR_FONT
    c.alignment = Alignment(horizontal="center"); c.border = BORDER
ws3.freeze_panes = "A2"
for i, x in enumerate(new_dom, 1):
    dist = " | ".join(f"{dt}({dl})" for dt, dl in sorted(x["domains"], key=lambda d: -d[1]))
    long_str = f"{x['longest'][0]}({x['longest'][1]})"
    vals = [i, x["kr"], len(x["cols"]), long_str, x["new_dom"], x["clsf"], dist]
    for j, v in enumerate(vals, 1):
        c = ws3.cell(row=i+1, column=j, value=v); c.font = CELL_FONT; c.border = BORDER
        c.alignment = Alignment(vertical="top", wrap_text=True)
    ws3.cell(row=i+1, column=5).fill = NEW_FILL

# 시트4: 처리 불가
ws4 = wb.create_sheet("처리불가")
H4 = ["No","한글명","컬럼수","도메인 분포","사유"]
W4 = [5,22,8,55,40]
for i, w in enumerate(W4, 1): ws4.column_dimensions[get_column_letter(i)].width = w
for i, h in enumerate(H4, 1):
    c = ws4.cell(row=1, column=i, value=h); c.fill = HDR_FILL; c.font = HDR_FONT
    c.alignment = Alignment(horizontal="center"); c.border = BORDER
ws4.freeze_panes = "A2"
for i, x in enumerate(skip, 1):
    dist = " | ".join(f"{dt}({dl})" for dt, dl in sorted(x["domains"], key=lambda d: -d[1]))
    vals = [i, x["kr"], len(x["cols"]), dist, x["사유"]]
    for j, v in enumerate(vals, 1):
        c = ws4.cell(row=i+1, column=j, value=v); c.font = CELL_FONT; c.border = BORDER
        c.alignment = Alignment(vertical="top", wrap_text=True)

wb.save(OUT_XLSX)
print(f"\n→ {OUT_XLSX}")
print(f"  시트: 표지 / 행안부흡수 ({len(absorb)}) / 기관표준_신규도메인 ({len(new_dom)}) / 처리불가 ({len(skip)})")
