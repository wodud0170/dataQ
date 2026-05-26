# -*- coding: utf-8 -*-
# 컬럼명_통일_결정.xlsx 생성
# 규칙: (영문컬럼명 + 데이터타입 + 데이터길이) 완전일치 = 같은 컬럼.
#       코멘트가 불일치하는 그룹을 결정 대상으로 추출.
import openpyxl, csv
from collections import defaultdict, Counter
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

wb = openpyxl.load_workbook('CAMS_SCHEMA_원본.xlsx', read_only=True, data_only=True)
rows = list(wb['컬럼정의'].iter_rows(values_only=True))[1:]
g = defaultdict(list)   # (영문명,타입,길이) -> [코멘트...]
for r in rows:
    col = (r[2] or '').strip()
    if not col:
        continue
    g[(col, r[4] or '', r[5])].append((r[3] or '').strip())

# 이미 결정된 컬럼명 (컬럼명_표준규칙.tsv)
decided = {}
try:
    with open('컬럼명_표준규칙.tsv', encoding='utf-8') as f:
        for r in csv.DictReader(f, delimiter='\t'):
            decided[r['영문컬럼명'].strip()] = r['표준한글명'].strip()
except FileNotFoundError:
    pass

groups = []
for k, v in g.items():
    if k[0] in decided:                # 컬럼명_표준규칙.tsv 에 결정 완료 → 제외
        continue
    cs = [x for x in v if x]
    uniq = set(cs)
    if len(v) > 1 and len(uniq) > 1:   # 2개+ 컬럼 & 코멘트 불일치
        groups.append((k, v, cs))
groups.sort(key=lambda x: -len(x[1]))

out = Workbook(); ws = out.active; ws.title = '컬럼명통일결정'
hdr = ['No', '영문컬럼명', '데이터타입', '길이', '컬럼수', '코멘트종수',
       '코멘트 변형 (빈도순)', '제안 표준한글명', '결정 표준한글명']
ws.append(hdr)
for i, (k, v, cs) in enumerate(groups, 1):
    col, dt, dl = k
    freq = Counter(cs).most_common()
    variants = ' | '.join('%s(%d)' % (c, n) for c, n in freq)
    empty = len(v) - len(cs)
    if empty:
        variants += ' | [코멘트없음](%d)' % empty
    suggest = freq[0][0] if freq else ''
    ws.append([i, col, dt, dl if dl is not None else '', len(v), len(set(cs)),
               variants, suggest, ''])
fill = PatternFill('solid', fgColor='305496')
for c in ws[1]:
    c.font = Font(bold=True, color='FFFFFF'); c.fill = fill
    c.alignment = Alignment(horizontal='center')
for i, w in enumerate([6, 22, 14, 8, 9, 10, 80, 26, 26], 1):
    ws.column_dimensions[chr(64 + i)].width = w
ws.freeze_panes = 'A2'
out.save('컬럼명_통일_결정.xlsx')
print('컬럼명_통일_결정.xlsx 생성 — 결정대상 %d그룹 (컬럼수 내림차순)' % len(groups))
print('  이미 결정된 컬럼: %s' % ', '.join(decided.keys()))
