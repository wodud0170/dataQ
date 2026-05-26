"""RAMP 단어사전·스키마 구조 파악."""
from openpyxl import load_workbook
from pathlib import Path

BASE = Path(r"C:\Users\장재영\Desktop\dataQ\CAMS표준화")
DICT = BASE / "01_원본자료" / "RMS4-DE14-04.데이터표준화_단어사전-V2.0.xlsx"
SCHEMA = BASE / "CAMS_RAMP_통합" / "ramp기관스키마정보.xlsx"

for label, path in [("DICT", DICT), ("SCHEMA", SCHEMA)]:
    print(f"\n========== {label}: {path.name} ==========")
    wb = load_workbook(path, read_only=True, data_only=True)
    for sn in wb.sheetnames:
        ws = wb[sn]
        print(f"\n  sheet: '{sn}'  (max_row: {ws.max_row}, max_col: {ws.max_column})")
        rows = list(ws.iter_rows(min_row=1, max_row=8, values_only=True))
        for i, row in enumerate(rows, 1):
            cells = [str(c)[:30] if c is not None else "" for c in row[:12]]
            print(f"    r{i}: {cells}")
    wb.close()
