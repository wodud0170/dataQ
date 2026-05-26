# -*- coding: utf-8 -*-
"""
RAMP 컬럼 BEFORE/AFTER 매핑정의서.
  BEFORE = ramp기관스키마정보.xlsx (RAMP 원본 Cubrid)
  AFTER  = RAMP_업로드_컬럼_2026-05-26.xlsx (사용자 확정 업로드본, 절대 건들지 않음 / read-only)
  매칭 키 = (테이블명, 컬럼순서)
"""
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from pathlib import Path
from collections import Counter

BASE = Path(r"C:\Users\장재영\Desktop\dataQ\CAMS표준화")
ROOT = BASE / "04_RAMP분석_2026-05-21" / "00_핵심산출"
BEFORE_PATH = BASE / "CAMS_RAMP_통합" / "ramp기관스키마정보_보정.xlsx"
AFTER_PATH  = ROOT / "RAMP_업로드_컬럼_2026-05-26_v3.xlsx"
OUT         = ROOT / "RAMP_컬럼_매핑정의서_2026-05-26.xlsx"

def s(v): return "" if v is None else str(v).strip()

# ============ BEFORE 로드 ============
print("=== BEFORE (ramp기관스키마정보) ===")
wb = load_workbook(BEFORE_PATH, read_only=True, data_only=True)
before = {}
for r in wb["컬럼"].iter_rows(min_row=2, values_only=True):
    if not r or not r[0] or not r[1]: continue
    tbl = s(r[0])
    ord_n = s(r[10])
    if not ord_n: continue
    key = (tbl, ord_n)
    if key in before: continue   # 중복 첫번째만
    before[key] = {
        "tbl": tbl, "en": s(r[1]), "kr": s(r[2]),
        "null": s(r[4]), "dtype": s(r[5]), "dlen": s(r[6]),
        "pk": s(r[7]), "fk": s(r[8]), "ord": ord_n,
    }
wb.close()
print(f"  BEFORE: {len(before)}")

# ============ AFTER 로드 ============
print("\n=== AFTER (RAMP_업로드_컬럼) ===")
wb = load_workbook(AFTER_PATH, read_only=True, data_only=True)
after = {}
for r in wb["컬럼"].iter_rows(min_row=2, values_only=True):
    if not r or not r[1]: continue
    tbl = s(r[1])
    ord_n = s(r[8])
    if not ord_n: continue
    key = (tbl, ord_n)
    after[key] = {
        "owner": s(r[0]), "tbl": tbl, "tbl_kr": s(r[2]),
        "en": s(r[3]), "kr": s(r[4]),
        "dtype": s(r[5]), "dlen": s(r[6]), "ddec": s(r[7]),
        "ord": ord_n, "null": s(r[9]), "pk": s(r[10]), "fk": s(r[11]),
        "default": s(r[12]), "ref_tbl": s(r[13]), "ref_col": s(r[14]),
    }
wb.close()
print(f"  AFTER: {len(after)}")

# ============ JOIN ============
all_keys = set(before.keys()) | set(after.keys())
only_before = set(before.keys()) - set(after.keys())
only_after = set(after.keys()) - set(before.keys())
print(f"\n  매칭: {len(set(before.keys()) & set(after.keys()))}")
print(f"  BEFORE only: {len(only_before)}")
print(f"  AFTER  only: {len(only_after)}")

rows = []
for key in sorted(all_keys, key=lambda k: (k[0], int(k[1]) if k[1].isdigit() else 0)):
    b = before.get(key, {})
    a = after.get(key, {})
    tbl = b.get("tbl") or a.get("tbl") or key[0]
    bef_en, aft_en   = b.get("en",""),   a.get("en","")
    bef_kr, aft_kr   = b.get("kr",""),   a.get("kr","")
    bef_dt, aft_dt   = b.get("dtype",""),a.get("dtype","")
    bef_dl, aft_dl   = b.get("dlen",""), a.get("dlen","")
    # 변경 유형
    types = []
    if bef_en.upper() != aft_en.upper(): types.append("EN")
    if bef_kr != aft_kr:                 types.append("KR")
    if bef_dt.upper() != aft_dt.upper(): types.append("TYPE")
    if str(bef_dl) != str(aft_dl):       types.append("LEN")
    chg = "+".join(types) if types else "KEEP"
    if not b: chg = "AFTER_ONLY"
    if not a: chg = "BEFORE_ONLY"
    rows.append({
        "tbl": tbl, "ord": key[1],
        "bef_en": bef_en, "aft_en": aft_en,
        "bef_kr": bef_kr, "aft_kr": aft_kr,
        "bef_dt": bef_dt, "bef_dl": bef_dl,
        "aft_dt": aft_dt, "aft_dl": aft_dl, "aft_ddec": a.get("ddec",""),
        "pk": a.get("pk", b.get("pk","")), "fk": a.get("fk",""),
        "null": a.get("null", b.get("null","")),
        "chg": chg,
    })

type_cnt = Counter(r["chg"] for r in rows)
print(f"\n  변경 유형: {dict(type_cnt)}")
upper = sum(1 for r in rows if r["aft_en"] == r["aft_en"].upper())
print(f"  AFTER 영문 대문자: {upper}/{len(rows)}")

# ============ xlsx 생성 ============
print("\n=== xlsx 생성 ===")
wb = Workbook()
THIN = Side(border_style="thin", color="808080")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
HDR_FILL = PatternFill("solid", fgColor="1F4E78")
HDR_FONT = Font(name="맑은 고딕", size=10, color="FFFFFF", bold=True)
CELL_FONT = Font(name="맑은 고딕", size=10)
SECTION_FILL = PatternFill("solid", fgColor="D9E1F2")
EN_FILL = PatternFill("solid", fgColor="FFE699")
KR_FILL = PatternFill("solid", fgColor="C6E0B4")
TYPE_FILL = PatternFill("solid", fgColor="F8CBAD")
KEEP_FILL = PatternFill("solid", fgColor="F2F2F2")
ONLY_FILL = PatternFill("solid", fgColor="FFC7CE")

# 표지
ws = wb.active; ws.title = "표지"
ws.column_dimensions["A"].width = 24; ws.column_dimensions["B"].width = 100
ws.merge_cells("A1:B1")
t = ws.cell(row=1, column=1, value="RAMP 컬럼 BEFORE/AFTER 매핑정의서")
t.font = Font(name="맑은 고딕", size=18, bold=True, color="1F4E78")
t.alignment = Alignment(horizontal="center", vertical="center")
ws.row_dimensions[1].height = 36

def cr(r, label, value):
    ws.cell(row=r, column=1, value=label).fill = SECTION_FILL
    ws.cell(row=r, column=1).font = Font(name="맑은 고딕", size=10, bold=True)
    ws.cell(row=r, column=1).border = BORDER
    c = ws.cell(row=r, column=2, value=value); c.font = Font(name="맑은 고딕", size=11)
    c.border = BORDER; c.alignment = Alignment(vertical="center", wrap_text=True)

cr(3, "작성일", "2026-05-26")
cr(4, "대상", f"RAMP 컬럼 {len(rows)}건")
cr(5, "BEFORE", f"ramp기관스키마정보.xlsx (RAMP 원본 Cubrid) — {len(before)}건")
cr(6, "AFTER",  f"RAMP_업로드_컬럼_2026-05-26.xlsx (확정 업로드본, read-only) — {len(after)}건")
cr(7, "매칭 키", "(테이블명, 컬럼순서)")
cr(8, "매칭 결과", f"매칭 {len(set(before.keys()) & set(after.keys()))} / BEFORE only {len(only_before)} / AFTER only {len(only_after)}")
cr(9, "변경 유형", "EN / KR / TYPE / LEN / KEEP / BEFORE_ONLY / AFTER_ONLY")

# 요약
ws2 = wb.create_sheet("요약")
ws2.column_dimensions["A"].width = 30; ws2.column_dimensions["B"].width = 12; ws2.column_dimensions["C"].width = 60
t2 = ws2.cell(row=1, column=1, value="변경 유형별 통계"); t2.font = Font(name="맑은 고딕", size=14, bold=True, color="1F4E78")
ws2.merge_cells("A1:C1")
rr = 3
for col, h in enumerate(["변경 유형","건수","설명"], 1):
    cell = ws2.cell(row=rr, column=col, value=h); cell.fill = HDR_FILL; cell.font = HDR_FONT
    cell.alignment = Alignment(horizontal="center"); cell.border = BORDER
rr += 1
exp = {"KEEP":"변경 없음","EN":"영문만","KR":"한글만","TYPE":"타입만","LEN":"길이만",
       "EN+KR":"영문+한글","EN+TYPE":"영문+타입","EN+LEN":"영문+길이","KR+TYPE":"한글+타입",
       "TYPE+LEN":"타입+길이","EN+KR+TYPE":"영문+한글+타입","EN+KR+LEN":"영문+한글+길이",
       "EN+TYPE+LEN":"영문+타입+길이","KR+TYPE+LEN":"한글+타입+길이","EN+KR+TYPE+LEN":"모두",
       "BEFORE_ONLY":"BEFORE 만 존재 (AFTER 누락)","AFTER_ONLY":"AFTER 만 존재 (신규)"}
for ct, n in type_cnt.most_common():
    ws2.cell(row=rr, column=1, value=ct).border = BORDER
    ws2.cell(row=rr, column=2, value=n).border = BORDER
    ws2.cell(row=rr, column=2).alignment = Alignment(horizontal="right")
    ws2.cell(row=rr, column=3, value=exp.get(ct,"")).border = BORDER
    rr += 1
rr += 1
ws2.cell(row=rr, column=1, value="AFTER 영문 대문자 일치").border = BORDER
ws2.cell(row=rr, column=2, value=f"{upper}/{len(rows)}").border = BORDER

# 매핑정의서
ws3 = wb.create_sheet("매핑정의서")
H = ["No","테이블","순서","BEFORE 영문","AFTER 영문","BEFORE 한글","AFTER 한글",
     "BEFORE 타입","BEFORE 길이","AFTER 타입","AFTER 길이","AFTER 소수","PK","FK","NULL","변경유형"]
W = [5,26,5,24,24,28,28,12,8,12,8,6,4,4,5,16]
for i, w in enumerate(W, 1): ws3.column_dimensions[get_column_letter(i)].width = w
for i, h in enumerate(H, 1):
    c = ws3.cell(row=1, column=i, value=h); c.fill = HDR_FILL; c.font = HDR_FONT
    c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True); c.border = BORDER
ws3.row_dimensions[1].height = 30
ws3.freeze_panes = "D2"

for i, r in enumerate(rows, 1):
    rn = i + 1
    vals = [i, r["tbl"], r["ord"], r["bef_en"], r["aft_en"], r["bef_kr"], r["aft_kr"],
            r["bef_dt"], r["bef_dl"], r["aft_dt"], r["aft_dl"], r["aft_ddec"],
            r["pk"], r["fk"], r["null"], r["chg"]]
    for j, v in enumerate(vals, 1):
        cell = ws3.cell(row=rn, column=j, value=v); cell.font = CELL_FONT; cell.border = BORDER
        cell.alignment = Alignment(vertical="top", wrap_text=True)
    if r["chg"] in ("BEFORE_ONLY","AFTER_ONLY"):
        for j in range(1, 17): ws3.cell(row=rn, column=j).fill = ONLY_FILL
    else:
        if r["chg"] == "KEEP":
            ws3.cell(row=rn, column=16).fill = KEEP_FILL
        if "EN" in r["chg"]:
            ws3.cell(row=rn, column=4).fill = EN_FILL
            ws3.cell(row=rn, column=5).fill = EN_FILL
        if "KR" in r["chg"]:
            ws3.cell(row=rn, column=6).fill = KR_FILL
            ws3.cell(row=rn, column=7).fill = KR_FILL
        if "TYPE" in r["chg"] or "LEN" in r["chg"]:
            for j in [8,9,10,11]: ws3.cell(row=rn, column=j).fill = TYPE_FILL

wb.save(OUT)
print(f"\n→ {OUT}")
print(f"  시트: 표지 / 요약 / 매핑정의서 ({len(rows)}행)")
