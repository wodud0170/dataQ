# -*- coding: utf-8 -*-
"""
RAMP_업로드_컬럼_v2.xlsx (확정 AFTER) → DB tb_data_model_attr 동기화.
매칭 키: (obj_nm, attr_ord)
UPDATE: attr_nm, attr_nm_kr, data_type, data_len
"""
from openpyxl import load_workbook
from pathlib import Path
import subprocess

BASE = Path(r"C:\Users\장재영\Desktop\dataQ\CAMS표준화")
COL_V2 = BASE / "04_RAMP분석_2026-05-21" / "00_핵심산출" / "RAMP_업로드_컬럼_2026-05-26_v2.xlsx"
SQL_OUT = BASE / "04_RAMP분석_2026-05-21" / "99_실행SQL_기록" / "Phase3_model_attr_update.sql"

def esc(s):
    if s is None: return ""
    return str(s).replace("'", "''")

def s(v): return "" if v is None else str(v).strip()

print("=== 1. v2 컬럼 로드 ===")
wb = load_workbook(COL_V2, read_only=True, data_only=True)
ws = wb.active
cols = []
for r in ws.iter_rows(min_row=2, values_only=True):
    if not r or not r[1] or not r[3]: continue
    cols.append({
        "tbl": s(r[1]), "en": s(r[3]), "kr": s(r[4]),
        "dtype": s(r[5]), "dlen": s(r[6]), "ddec": s(r[7]),
        "ord": s(r[8]),
    })
wb.close()
print(f"  v2 컬럼: {len(cols)}")

print("\n=== 2. UPDATE SQL 생성 ===")
lines = [
    "-- Phase 3: tb_data_model_attr 영문 보충 + CHAR→VARCHAR 동기화",
    "-- 매칭 키: (dm_id, obj_nm, attr_ord)",
    "BEGIN;",
    "",
    "-- 변경 전 상태",
    "SELECT count(*) AS before_total FROM quality.tb_data_model_attr WHERE dm_id=(SELECT dm_id FROM quality.tb_data_model WHERE dm_nm='RAMP' LIMIT 1);",
    "",
]
for c in cols:
    if not c["ord"].isdigit(): continue
    dlen = c["dlen"] if c["dlen"].isdigit() else "0"
    ddec = c["ddec"] if c["ddec"].isdigit() else "0"
    lines.append(
        f"UPDATE quality.tb_data_model_attr SET "
        f"attr_nm='{esc(c['en'])}', attr_nm_kr='{esc(c['kr'])}', "
        f"data_type='{esc(c['dtype'])}', data_len={dlen}, data_decimal_len={ddec} "
        f"WHERE dm_id=(SELECT dm_id FROM quality.tb_data_model WHERE dm_nm='RAMP' LIMIT 1) "
        f"AND obj_nm='{esc(c['tbl'])}' AND attr_ord={c['ord']};"
    )

lines += [
    "",
    "-- 검증",
    "SELECT 'attrs' AS t, count(*) FROM quality.tb_data_model_attr WHERE dm_id=(SELECT dm_id FROM quality.tb_data_model WHERE dm_nm='RAMP' LIMIT 1);",
    "SELECT 'CHAR 잔존' AS t, count(*) FROM quality.tb_data_model_attr WHERE dm_id=(SELECT dm_id FROM quality.tb_data_model WHERE dm_nm='RAMP' LIMIT 1) AND data_type='CHAR';",
    "SELECT '영문 _CN 끝' AS t, count(*) FROM quality.tb_data_model_attr WHERE dm_id=(SELECT dm_id FROM quality.tb_data_model WHERE dm_nm='RAMP' LIMIT 1) AND attr_nm LIKE '%\\_CN' ESCAPE '\\';",
    "SELECT attr_nm, attr_nm_kr, data_type, data_len FROM quality.tb_data_model_attr WHERE attr_nm_kr='비고내용' LIMIT 3;",
    "SELECT attr_nm, attr_nm_kr, data_type FROM quality.tb_data_model_attr WHERE attr_nm_kr='권한그룹설명' LIMIT 1;",
    "",
    "COMMIT;",
]
SQL_OUT.write_text("\n".join(lines), encoding="utf-8")
print(f"  UPDATE 문: {sum(1 for L in lines if L.startswith('UPDATE'))}")
print(f"  → {SQL_OUT}")
