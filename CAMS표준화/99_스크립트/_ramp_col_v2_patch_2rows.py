# -*- coding: utf-8 -*-
"""
AFTER 컬럼 v2 에 누락된 2건 INSERT + 같은 테이블 내 그 이후 ord 1씩 밀기.
입력: RAMP_업로드_컬럼_2026-05-26_v2.xlsx
출력: RAMP_업로드_컬럼_2026-05-26_v3.xlsx
"""
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from pathlib import Path

BASE = Path(r"C:\Users\장재영\Desktop\dataQ\CAMS표준화")
ROOT = BASE / "04_RAMP분석_2026-05-21" / "00_핵심산출"
V2 = ROOT / "RAMP_업로드_컬럼_2026-05-26_v2.xlsx"
V3 = ROOT / "RAMP_업로드_컬럼_2026-05-26_v3.xlsx"

# 추가 2건
INSERTS = [
    {
        "owner":"RAMP", "tbl":"tb_tkfoldermng", "tbl_kr":"",
        "en":"INSP_CN", "kr":"검수내용",
        "dtype":"VARCHAR", "dlen":"500", "ddec":"0",
        "ord":6, "null":"Y", "pk":"N", "fk":"N",
    },
    {
        "owner":"RAMP", "tbl":"tb_tkorgacptnplan", "tbl_kr":"",
        "en":"ACPTN_RQSTR_NM", "kr":"인수요청자명",
        "dtype":"VARCHAR", "dlen":"40", "ddec":"0",
        "ord":12, "null":"Y", "pk":"N", "fk":"N",
    },
]

print("=== v2 로드 ===")
wb = load_workbook(V2, data_only=True)
ws = wb.active
header = [c.value for c in ws[1]]
print(f"  헤더: {header}")
print(f"  rows: {ws.max_row}")

# 모든 행 메모리에
all_rows = []
for r in ws.iter_rows(min_row=2, values_only=True):
    if not r or not r[1]: continue
    all_rows.append(list(r))
wb.close()
print(f"  로드: {len(all_rows)}건")

# 같은 테이블 내 ord 재배치 + 2건 INSERT
print("\n=== 2건 INSERT + 영향 받는 ord 재배치 ===")
# 영향 테이블별로 그룹화
patched_tbls = set()
for ins in INSERTS:
    tbl = ins["tbl"]
    target_ord = ins["ord"]
    patched_tbls.add(tbl)
    # 같은 테이블의 ord >= target_ord 인 행은 +1
    for row in all_rows:
        if row[1] == tbl:
            cur_ord = int(row[8]) if str(row[8]).isdigit() else 0
            if cur_ord >= target_ord:
                row[8] = cur_ord + 1
    # 새 행 INSERT
    new_row = [
        ins["owner"], ins["tbl"], ins["tbl_kr"],
        ins["en"], ins["kr"],
        ins["dtype"], ins["dlen"], ins["ddec"],
        ins["ord"], ins["null"], ins["pk"], ins["fk"],
        None, None, None, None,
    ]
    all_rows.append(new_row)
    print(f"  INSERT: {ins['tbl']} / {ins['en']} / {ins['kr']} (ord={ins['ord']})")

# 정렬: (테이블, ord)
all_rows.sort(key=lambda r: (str(r[1] or ""), int(r[8]) if str(r[8]).isdigit() else 0))

# v3 저장
print(f"\n=== v3 저장 ===")
wb = Workbook()
ws = wb.active; ws.title = "컬럼"
THIN = Side(border_style="thin", color="808080")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
HDR_FILL = PatternFill("solid", fgColor="1F4E78")
HDR_FONT = Font(name="맑은 고딕", size=10, color="FFFFFF", bold=True)
CELL_FONT = Font(name="맑은 고딕", size=10)
W = [12, 28, 30, 24, 28, 12, 8, 8, 8, 8, 6, 6, 12, 24, 22, 10]
for i, w in enumerate(W, 1): ws.column_dimensions[get_column_letter(i)].width = w
for i, h in enumerate(header, 1):
    c = ws.cell(row=1, column=i, value=h)
    c.fill = HDR_FILL; c.font = HDR_FONT
    c.alignment = Alignment(horizontal="center"); c.border = BORDER
ws.freeze_panes = "C2"
for r_idx, row in enumerate(all_rows, 2):
    for c_idx, v in enumerate(row, 1):
        cell = ws.cell(row=r_idx, column=c_idx, value=v)
        cell.font = CELL_FONT; cell.border = BORDER
        cell.alignment = Alignment(vertical="top", wrap_text=True)
wb.save(V3)
print(f"  rows: {len(all_rows)}")
print(f"  → {V3}")
