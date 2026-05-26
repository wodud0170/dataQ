# -*- coding: utf-8 -*-
"""행안부 원본 xlsx + 일괄등록 xlsx 구조·플래그 점검."""
from openpyxl import load_workbook
from pathlib import Path

BASE = Path(r"C:\Users\장재영\Desktop\dataQ\q-center\src\main\resources\seed\행안부_공통표준")
ORIG = BASE / "행정안전부_공공데이터 공통표준단어.xlsx"
UPLD = BASE / "단어사전_일괄등록.xlsx"

for label, path in [("ORIG", ORIG), ("UPLOAD", UPLD)]:
    print(f"\n=== {label}: {path.name} ===")
    wb = load_workbook(path, read_only=True, data_only=True)
    for sn in wb.sheetnames:
        ws = wb[sn]
        print(f"\n  sheet '{sn}'  (max_row: {ws.max_row}, max_col: {ws.max_column})")
        rows = list(ws.iter_rows(min_row=1, max_row=6, values_only=True))
        for i, row in enumerate(rows, 1):
            cells = [str(c)[:35] if c is not None else "" for c in (row or [])[:15]]
            print(f"    r{i}: {cells}")
    wb.close()
