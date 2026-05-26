# -*- coding: utf-8 -*-
"""
Phase 2-2 D7 — 도메인 매칭 실패 285건 잔여 보강:
  1. R8 미종결 잔여 한글 → 형식단어 추가 (D5 보충)
  2. 분류 + 도메인 자동 INSERT
"""
from openpyxl import load_workbook
from pathlib import Path
from collections import defaultdict, Counter
import subprocess, csv, io, os, base64, sys

BASE = Path(r"C:\Users\장재영\Desktop\dataQ\CAMS표준화")
RAMP_SCHEMA = BASE / "CAMS_RAMP_통합" / "ramp기관스키마정보.xlsx"
COL_BA = BASE / "04_RAMP분석_2026-05-21" / "00_핵심산출" / "RAMP_컬럼_BEFORE_AFTER_2026-05-23.xlsx"
TIME_BA = BASE / "04_RAMP분석_2026-05-21" / "00_핵심산출" / "Phase2_시간컬럼_BEFORE_AFTER_2026-05-23.xlsx"
D5_XLSX = BASE / "04_RAMP분석_2026-05-21" / "00_핵심산출" / "Phase2_D5_R8형식단어_2026-05-23.xlsx"
OUT_SQL = BASE / "04_RAMP분석_2026-05-21" / "99_실행SQL_기록" / "Phase2_D7_fillgap.sql"

# ============ D5 결과 로드 (BEFORE 한글 → AFTER 한글 매핑) ============
print("=== D5 결과 로드 ===")
d5_map = {}  # BEFORE → (AFTER, 형식단어, 분류, 타입, 길이)
wb = load_workbook(D5_XLSX, read_only=True, data_only=True)
ws = wb["새용어_BEFORE_AFTER"]
# 헤더: No, BEFORE 한글, 자동 형식단어, AFTER 한글, 최장 타입, 최장 길이, 신규 도메인, ...
for r in ws.iter_rows(min_row=2, values_only=True):
    if r and r[1]:
        bef = str(r[1]); fw = str(r[2] or ""); aft = str(r[3] or ""); dt = str(r[4] or ""); dl = str(r[5] or ""); dom = str(r[6] or "")
        d5_map[bef] = {"aft": aft, "fw": fw, "dt": dt, "dl": dl, "dom": dom}
wb.close()
print(f"  D5 매핑: {len(d5_map)}")

# 형식단어 + EXTRA + D5 신규 분류
r = subprocess.run(["docker","exec","-i","dataq-db","psql","-U","admin","-d","postgres",
    "-c", """COPY (SELECT word_nm FROM quality.tb_word WHERE word_clsf_yn='Y') TO STDOUT WITH (FORMAT csv, FORCE_QUOTE *)"""],
    capture_output=True, encoding="utf-8")
fw_check = set(row[0] for row in csv.reader(io.StringIO(r.stdout)) if row)
EXTRA = {"구분","값","경로","상태","메시지","식별자","파일","순번","번호","자","이름","내역","유형",
         "범위","위치","방법","수단","목록","사항","항목","기간","결과","대상","주체","주소","권한",
         "역할","사유","율","률","비율","기한","빈도","주기","횟수","등급","단계","ID"}
fw_check.update(EXTRA)

# 기존 도메인 + 분류
r = subprocess.run(["docker","exec","-i","dataq-db","psql","-U","admin","-d","postgres",
    "-c", """COPY (SELECT domain_nm, domain_clsf_nm, data_type, coalesce(data_len::text,'0')
             FROM quality.tb_domain) TO STDOUT WITH (FORMAT csv, FORCE_QUOTE *)"""],
    capture_output=True, encoding="utf-8")
exist_by_key = {}; exist_nm = set(); exist_clsf = set()
for row in csv.reader(io.StringIO(r.stdout)):
    if row:
        clsf, dt, dl = row[1], row[2].upper(), row[3]
        L = int(dl) if dl.isdigit() else 0
        exist_by_key[(clsf, dt, L)] = row[0]
        exist_nm.add(row[0]); exist_clsf.add(clsf)
r = subprocess.run(["docker","exec","-i","dataq-db","psql","-U","admin","-d","postgres",
    "-c", "COPY (SELECT domain_clsf_nm FROM quality.tb_domain_clsf) TO STDOUT WITH (FORMAT csv, FORCE_QUOTE *)"],
    capture_output=True, encoding="utf-8")
for row in csv.reader(io.StringIO(r.stdout)):
    if row: exist_clsf.add(row[0])

# ============ AFTER 한글 적용 후 R8 미종결 잔여 검출 ============
TIME_SUFFIX = ("일자","일시","년월","연월","시분초","시분","시각","연도","월일","월")
TIME_DOM = set()
wb = load_workbook(TIME_BA, read_only=True, data_only=True)
for r in wb["적용_BEFORE_AFTER"].iter_rows(min_row=2, values_only=True):
    if r and r[1]: TIME_DOM.add((str(r[1]), str(r[2])))
wb.close()

col_map = {}
wb = load_workbook(COL_BA, read_only=True, data_only=True)
for r in wb["컬럼BEFORE_AFTER"].iter_rows(min_row=2, values_only=True):
    if r and r[1] and r[2]:
        col_map[(str(r[1]), str(r[2]))] = str(r[5] or r[4] or "")
wb.close()

wb = load_workbook(RAMP_SCHEMA, read_only=True, data_only=True)
gap_grp = defaultdict(list)  # AFTER 한글 (D5 적용) → [(dt, dl)]
seen=set()
for r in wb["컬럼"].iter_rows(min_row=2, values_only=True):
    if r and r[0] and r[1]:
        k=(str(r[0]).strip(), str(r[1]).strip())
        if k in seen: continue
        seen.add(k)
        if k in TIME_DOM: continue
        kr = col_map.get(k, str(r[2] or "").strip())
        if not kr: continue
        # D5 AFTER 한글 적용
        if kr in d5_map:
            kr = d5_map[kr]["aft"]
        # R8 검사
        if not any(kr.endswith(f) for f in fw_check):
            dt = str(r[5] or "").upper() if len(r) > 5 else ""
            dl = str(r[6] or "") if len(r) > 6 else "0"
            gap_grp[kr].append((dt, dl))
wb.close()

print(f"\n=== AFTER 한글 잔여 R8 미종결: {len(gap_grp)}종 ===")
for kr, doms in sorted(gap_grp.items(), key=lambda x: -len(x[1]))[:15]:
    dist = Counter(doms).most_common(3)
    print(f"  {kr:25s} {len(doms)}컬럼  {dist}")

# ============ 자동 형식단어 부여 (그룹별 단일, NUMERIC→값, VARCHAR≥1000→내용/명) ============
def auto_fw(dt, L):
    dt = dt.upper()
    if dt in ("VARCHAR","VARCHAR2","STRING"):
        if L > 4000: return "내용"
        if L >= 100: return "내용"
        if L >= 11:  return "명"
        if L <= 3:   return "구분"
        return "명"
    if dt == "CHAR":
        if L == 1: return "여부"
        if L <= 4: return "코드"
        return "구분"
    if dt in ("NUMERIC","NUMBER","INTEGER","DECIMAL"): return "값"
    if dt == "CLOB": return "내용"
    return "값"

def parse_int(s):
    try: return int(s)
    except: return 0

TYPE_ABBR = {"VARCHAR":"V","VARCHAR2":"V","STRING":"V","CHAR":"C","NUMERIC":"N","NUMBER":"N","INTEGER":"N","DECIMAL":"N","CLOB":"L"}
def map_oracle(dt, L):
    if dt in ("VARCHAR","VARCHAR2","STRING") and L > 4000: return "CLOB", 0
    if dt == "STRING": return "VARCHAR", L
    return dt, L

# 잔여 보강 — 그룹별 단일 형식단어 + 분류/도메인
new_clsfs = set()
new_doms = {}
for kr, doms in gap_grp.items():
    longest = max(doms, key=lambda d: parse_int(d[1]))
    dt, dl = longest
    L = parse_int(dl)
    fw = auto_fw(dt, L)
    clsf = fw  # 형식단어가 분류명
    if clsf not in exist_clsf:
        new_clsfs.add(clsf)
    o_dt, o_L = map_oracle(dt, L)
    if (clsf, o_dt.upper(), o_L) not in exist_by_key:
        abbr = TYPE_ABBR.get(o_dt.upper(), "X")
        if o_dt.upper() == "CLOB":
            nm = f"{clsf}L"
        else:
            nm = f"{clsf}{abbr}{o_L}" if o_L > 0 else f"{clsf}{abbr}"
        if nm not in exist_nm and nm not in new_doms:
            new_doms[nm] = (clsf, o_dt, o_L)

print(f"\n=== 추가 분류: {len(new_clsfs)}, 추가 도메인: {len(new_doms)} ===")

# ============ SQL ============
def esc(s): return str(s).replace("'", "''") if s else ""
sql = ["-- Phase 2-2 D7 — 잔여 R8 미종결 분류/도메인 추가", "BEGIN;", ""]
if new_clsfs:
    for cn in sorted(new_clsfs):
        cid = base64.urlsafe_b64encode(os.urandom(17)).rstrip(b'=').decode()[:22]
        sql.append(f"INSERT INTO quality.tb_domain_clsf (domain_clsf_id, domain_clsf_nm, domain_grp_nm, comm_stnd_yn, cret_dt, cret_user_id) "
                   f"VALUES ('{cid}','{esc(cn)}','기타','N',to_char(now(),'YYYYMMDDHH24MISS'),'admin');")
    sql.append("")
for nm, (clsf, dt, L) in sorted(new_doms.items()):
    did = base64.urlsafe_b64encode(os.urandom(17)).rstrip(b'=').decode()[:22]
    desc = f"{clsf} {dt}({L}) — D7 보강"
    sql.append(
        f"INSERT INTO quality.tb_domain (domain_id, domain_nm, domain_grp_nm, domain_clsf_nm, domain_desc, data_type, data_len, data_decimal_len, stor_fmt, expr_fmt_lst, comm_stnd_yn, aprv_yn, cret_dt, cret_user_id, use_yn) "
        f"VALUES ('{did}','{esc(nm)}','기타','{esc(clsf)}','{esc(desc)}','{esc(dt)}',{L},0,'',ARRAY[]::text[],'N','Y',to_char(now(),'YYYYMMDDHH24MISS'),'admin','Y');"
    )

sql += ["", "COMMIT;"]
OUT_SQL.write_text("\n".join(sql), encoding="utf-8")
print(f"\n→ {OUT_SQL}")
