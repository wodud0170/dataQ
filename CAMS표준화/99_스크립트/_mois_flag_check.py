# -*- coding: utf-8 -*-
"""행안부 원본 xlsx 의 개정구분명(폐기/변경) 플래그 점검 + 업로드본·DB 적재본 비교."""
from openpyxl import load_workbook
from pathlib import Path
from collections import Counter, defaultdict

BASE = Path(r"C:\Users\장재영\Desktop\dataQ\q-center\src\main\resources\seed\행안부_공통표준")
ORIG = BASE / "행정안전부_공공데이터 공통표준단어.xlsx"
UPLD = BASE / "단어사전_일괄등록.xlsx"

# 1) 원본 — 개정구분명 분포
wb = load_workbook(ORIG, read_only=True, data_only=True)
ws = wb["Sheet"]
orig = []  # (단어명, 약어, 영문명, 제정차수, 개정구분명, 개정항목, 개정사유)
for r in ws.iter_rows(min_row=2, values_only=True):
    if r and r[0]:
        orig.append({
            "nm": (r[0] or "").strip(),
            "abrv": (r[1] or "").strip(),
            "eng": (r[2] or "").strip(),
            "차수": (r[8] or "").strip(),
            "구분": (r[9] or "").strip(),
            "항목": (r[10] or "").strip(),
            "사유": (r[11] or "").strip(),
        })
wb.close()

print(f"=== 원본 행안부 단어수: {len(orig)} ===\n")

cnt = Counter(w["구분"] for w in orig)
print("[개정구분명 분포]")
for k, v in cnt.most_common():
    print(f"  '{k}': {v}")

ch = [w for w in orig if w["구분"]]
print(f"\n[개정 표기 있는 단어: {len(ch)}건] 상위 15개")
for w in ch[:15]:
    print(f"  {w['nm']:25s} {w['구분']:15s} {w['항목']:20s} {w['사유'][:40]}")

차수 = Counter(w["차수"] for w in orig)
print(f"\n[제정차수 분포]")
for k, v in 차수.most_common():
    print(f"  '{k}': {v}")

# 2) 업로드용 — 단어 셋 추출
wb = load_workbook(UPLD, read_only=True, data_only=True)
ws = wb["Sheet1"]
upld = set()
for r in ws.iter_rows(min_row=2, values_only=True):
    if r and r[2]:
        upld.add((r[2] or "").strip())
wb.close()

orig_nm = set(w["nm"] for w in orig)
폐기_nm = set(w["nm"] for w in orig if "폐기" in w["구분"])
변경_nm = set(w["nm"] for w in orig if "변경" in w["구분"])

print(f"\n=== 비교: 원본 {len(orig_nm)} vs 업로드 {len(upld)} ===")
print(f"  원본only (업로드에 없음): {len(orig_nm - upld)}")
print(f"  업로드only (원본에 없음): {len(upld - orig_nm)}")
print(f"  공통: {len(orig_nm & upld)}")

print(f"\n=== 폐기 단어({len(폐기_nm)}건)가 업로드에 포함됐나? ===")
폐기_in_upld = 폐기_nm & upld
print(f"  업로드에 포함된 폐기 단어: {len(폐기_in_upld)} / {len(폐기_nm)}")
for nm in sorted(폐기_in_upld)[:20]:
    w = next(w for w in orig if w["nm"] == nm)
    print(f"  ✗ {nm:25s} 사유: {w['사유'][:50]}")

print(f"\n=== 변경 단어({len(변경_nm)}건)가 업로드에 포함됐나? ===")
변경_in_upld = 변경_nm & upld
print(f"  업로드에 포함된 변경 단어: {len(변경_in_upld)} / {len(변경_nm)}")
for nm in sorted(변경_in_upld)[:10]:
    w = next(w for w in orig if w["nm"] == nm)
    print(f"  {nm:25s} 항목:{w['항목']:15s} 사유:{w['사유'][:50]}")
