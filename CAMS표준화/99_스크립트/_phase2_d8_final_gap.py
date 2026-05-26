# -*- coding: utf-8 -*-
"""
Phase 2-2 D8 — 용어사전 매칭 실패 134건 잔여 보강.
  - 한글 그룹 → 분류·타입·길이 → 부족한 도메인 자동 INSERT
  - 분류 추출 실패한 한글은 별도 보고
"""
from openpyxl import load_workbook
from pathlib import Path
from collections import defaultdict, Counter
import subprocess, csv, io, os, base64

BASE = Path(r"C:\Users\장재영\Desktop\dataQ\CAMS표준화")
RAMP_SCHEMA = BASE / "CAMS_RAMP_통합" / "ramp기관스키마정보.xlsx"
COL_BA = BASE / "04_RAMP분석_2026-05-21" / "00_핵심산출" / "RAMP_컬럼_BEFORE_AFTER_2026-05-23.xlsx"
TIME_BA = BASE / "04_RAMP분석_2026-05-21" / "00_핵심산출" / "Phase2_시간컬럼_BEFORE_AFTER_2026-05-23.xlsx"
D5_XLSX = BASE / "04_RAMP분석_2026-05-21" / "00_핵심산출" / "Phase2_D5_R8형식단어_2026-05-23.xlsx"
OUT_SQL = BASE / "04_RAMP분석_2026-05-21" / "99_실행SQL_기록" / "Phase2_D8_final_gap.sql"

# 형식단어 + EXTRA
r = subprocess.run(["docker","exec","-i","dataq-db","psql","-U","admin","-d","postgres",
    "-c", """COPY (SELECT word_nm, coalesce(domain_clsf_nm,'') FROM quality.tb_word
             WHERE word_clsf_yn='Y') TO STDOUT WITH (FORMAT csv, FORCE_QUOTE *)"""],
    capture_output=True, encoding="utf-8")
fw_map = {row[0]: row[1] for row in csv.reader(io.StringIO(r.stdout)) if row}
EXTRA = {"구분":"구분","값":"값","경로":"경로","상태":"상태","메시지":"메시지","식별자":"식별자",
         "파일":"파일","순번":"순번","번호":"번호","자":"자","이름":"이름","내역":"내역","유형":"유형",
         "범위":"범위","위치":"위치","방법":"방법","수단":"수단","목록":"목록","사항":"사항","항목":"항목",
         "기간":"기간","결과":"결과","대상":"대상","주체":"주체","주소":"주소","권한":"권한",
         "역할":"역할","사유":"사유","율":"율","률":"률","ID":"ID","기한":"기한","빈도":"빈도",
         "주기":"주기","횟수":"횟수","등급":"등급","단계":"단계"}
for k,v in EXTRA.items(): fw_map.setdefault(k, v)

def extract_clsf(kr):
    cand = [(fw, fw_map[fw]) for fw in fw_map if kr.endswith(fw)]
    valid = [c for c in cand if c[1]]
    return max(valid, key=lambda x: len(x[0]))[1] if valid else None

# 기존 도메인
r = subprocess.run(["docker","exec","-i","dataq-db","psql","-U","admin","-d","postgres",
    "-c", "COPY (SELECT domain_nm, domain_clsf_nm, upper(data_type), coalesce(data_len::text,'0') FROM quality.tb_domain) TO STDOUT WITH (FORMAT csv, FORCE_QUOTE *)"],
    capture_output=True, encoding="utf-8")
exist_by_key = {}; exist_nm = set(); exist_clsf = set()
for row in csv.reader(io.StringIO(r.stdout)):
    if row:
        clsf, dt, dl = row[1], row[2], row[3]
        L = int(dl) if dl.isdigit() else 0
        exist_by_key[(clsf, dt, L)] = row[0]
        exist_nm.add(row[0]); exist_clsf.add(clsf)
r = subprocess.run(["docker","exec","-i","dataq-db","psql","-U","admin","-d","postgres",
    "-c", "COPY (SELECT domain_clsf_nm FROM quality.tb_domain_clsf) TO STDOUT WITH (FORMAT csv, FORCE_QUOTE *)"],
    capture_output=True, encoding="utf-8")
for row in csv.reader(io.StringIO(r.stdout)):
    if row: exist_clsf.add(row[0])

# D5 매핑
d5_after = {}
wb = load_workbook(D5_XLSX, read_only=True, data_only=True)
for r in wb["새용어_BEFORE_AFTER"].iter_rows(min_row=2, values_only=True):
    if r and r[1] and r[3]: d5_after[str(r[1])] = str(r[3])
wb.close()

# Phase 2-1 AFTER + 시간 제외
col_map = {}
wb = load_workbook(COL_BA, read_only=True, data_only=True)
for r in wb["컬럼BEFORE_AFTER"].iter_rows(min_row=2, values_only=True):
    if r and r[1] and r[2]:
        col_map[(str(r[1]),str(r[2]))] = str(r[5] or r[4] or "")
wb.close()
TIME = set()
wb = load_workbook(TIME_BA, read_only=True, data_only=True)
for r in wb["적용_BEFORE_AFTER"].iter_rows(min_row=2, values_only=True):
    if r and r[1]: TIME.add((str(r[1]), str(r[2])))
wb.close()

# RAMP 한글 그룹 (AFTER 적용)
grp = defaultdict(list)
wb = load_workbook(RAMP_SCHEMA, read_only=True, data_only=True)
seen=set()
for r in wb["컬럼"].iter_rows(min_row=2, values_only=True):
    if r and r[0] and r[1]:
        k=(str(r[0]).strip(), str(r[1]).strip())
        if k in seen: continue
        seen.add(k)
        if k in TIME: continue
        kr = col_map.get(k, str(r[2] or "").strip())
        if not kr: continue
        if kr in d5_after: kr = d5_after[kr]
        dt = str(r[5] or "").upper(); dl = str(r[6] or "0")
        grp[kr].append((dt, dl))
wb.close()

def parse_int(s):
    try: return int(s)
    except: return 0
def map_oracle(dt, L):
    if dt in ("VARCHAR","VARCHAR2","STRING") and L>4000: return "CLOB", 0
    if dt == "STRING": return "VARCHAR", L
    return dt, L
TYPE_ABBR = {"VARCHAR":"V","CHAR":"C","NUMERIC":"N","NUMBER":"N","INTEGER":"N","DECIMAL":"N","CLOB":"L"}

# 매칭 실패 찾기 — 분류 추출 실패 시 "값" 형식단어 자동 부여
failed_clsf = []  # 사용자 결정으로 "값" 적용
failed_dom = []   # 도메인 없음
new_doms = {}
new_clsfs = set()
for kr, doms in grp.items():
    longest = max(doms, key=lambda d: parse_int(d[1]))
    dt, dl = longest; L = parse_int(dl)
    o_dt, o_L = map_oracle(dt, L)
    clsf = extract_clsf(kr)
    if not clsf:
        # 분류 추출 실패 → "값" 부여 (사용자 결정 2026-05-23)
        clsf = "값"
        failed_clsf.append({"kr":kr, "cnt":len(doms), "longest":(dt,dl), "applied_fw":"값"})
    if (clsf, o_dt, o_L) in exist_by_key: continue
    abbr = TYPE_ABBR.get(o_dt, "X")
    nm = f"{clsf}L" if o_dt=="CLOB" else (f"{clsf}{abbr}{o_L}" if o_L>0 else f"{clsf}{abbr}")
    if nm in exist_nm or nm in new_doms:
        i=2
        while f"{nm}_{i}" in exist_nm or f"{nm}_{i}" in new_doms: i+=1
        nm = f"{nm}_{i}"
    new_doms[nm] = (clsf, o_dt, o_L)
    if clsf not in exist_clsf: new_clsfs.add(clsf)
    failed_dom.append({"kr":kr, "cnt":len(doms), "clsf":clsf, "dt":o_dt, "L":o_L, "nm":nm})

print(f"분류 추출 실패: {len(failed_clsf)}건")
print(f"도메인 없음 → 신규 추가: {len(failed_dom)}건")
print(f"신규 도메인: {len(new_doms)}")
print(f"신규 분류: {len(new_clsfs)}")

print("\n=== 분류 추출 실패 예시 (상위 15) ===")
for f in sorted(failed_clsf, key=lambda x: -x["cnt"])[:15]:
    print(f"  {f['kr']:25s} {f['cnt']}컬럼 / 최장 {f['longest']}")

print("\n=== 신규 도메인 예시 (상위 15) ===")
for f in sorted(failed_dom, key=lambda x: -x["cnt"])[:15]:
    print(f"  {f['kr']:25s} {f['cnt']}컬럼 → {f['nm']} ({f['clsf']},{f['dt']},{f['L']})")

# SQL
def esc(s): return str(s).replace("'", "''") if s else ""
sql = ["-- Phase 2-2 D8 — 매칭 실패 134건 잔여 보강", "BEGIN;", ""]
if new_clsfs:
    for cn in sorted(new_clsfs):
        cid = base64.urlsafe_b64encode(os.urandom(17)).rstrip(b'=').decode()[:22]
        sql.append(f"INSERT INTO quality.tb_domain_clsf (domain_clsf_id, domain_clsf_nm, domain_grp_nm, comm_stnd_yn, cret_dt, cret_user_id) "
                   f"VALUES ('{cid}','{esc(cn)}','기타','N',to_char(now(),'YYYYMMDDHH24MISS'),'admin');")
    sql.append("")
for nm, (clsf, dt, L) in sorted(new_doms.items()):
    did = base64.urlsafe_b64encode(os.urandom(17)).rstrip(b'=').decode()[:22]
    sql.append(
        f"INSERT INTO quality.tb_domain (domain_id, domain_nm, domain_grp_nm, domain_clsf_nm, domain_desc, data_type, data_len, data_decimal_len, stor_fmt, expr_fmt_lst, comm_stnd_yn, aprv_yn, cret_dt, cret_user_id, use_yn) "
        f"VALUES ('{did}','{esc(nm)}','기타','{esc(clsf)}','{esc(clsf)} {dt}({L}) — D8 보강','{esc(dt)}',{L},0,'',ARRAY[]::text[],'N','Y',to_char(now(),'YYYYMMDDHH24MISS'),'admin','Y');"
    )
sql += ["", "COMMIT;"]
OUT_SQL.write_text("\n".join(sql), encoding="utf-8")
print(f"\n→ {OUT_SQL}")
