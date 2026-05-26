"""
RAMP 단어 1573개를 4분류로 나눠 TB_WORD 양식 xlsx 1개로 생성 (검토용).
출력: CAMS표준화/RAMP_단어_검토_2026-05-21.xlsx
  시트:
    1. 요약
    2. 완전일치_797   (SKIP)
    3. Case1_한글동일_영문약어다름_204  (SKIP, 별도 컬럼명 일괄변환)
    4. Case2_영문약어동일_한글다름_77   (★ 결정필요, MOIS+RAMP 양쪽 + 영향컬럼)
    5. 순수RAMPonly_495                 (★ TB_WORD INSERT 대상)
"""
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from pathlib import Path
from collections import defaultdict, Counter
import re

BASE = Path(r"C:\Users\장재영\Desktop\dataQ\CAMS표준화")
DICT = BASE / "01_원본자료" / "RMS4-DE14-04.데이터표준화_단어사전-V2.0.xlsx"
SCHEMA = BASE / "CAMS_RAMP_통합" / "ramp기관스키마정보.xlsx"
MOIS_WORD = BASE.parent / "q-center" / "src" / "main" / "resources" / "seed" / "행안부_공통표준" / "행정안전부_공공데이터 공통표준단어.xlsx"
OUT = BASE / "04_RAMP분석_2026-05-21" / "RAMP_단어_검토_2026-05-21.xlsx"

# ============ RAMP 사전 ============
wb = load_workbook(DICT, read_only=True, data_only=True)
ws = wb["단어사전"]
ramp = []
for r in ws.iter_rows(min_row=2, values_only=True):
    if not r or not r[2]: continue
    ramp.append({
        "단어명": (r[2] or "").strip(),
        "영문약어": (r[3] or "").strip().upper(),
        "영문명": (r[4] or "").strip(),
        "설명": (r[5] or "").replace("_x000D_", " ").strip(),
        "형식단어": r[6] or "",
        "도메인": r[7] or "",
        "이음동의어": r[8] or "",
        "금칙어": r[9] or "",
    })
wb.close()

# ============ MOIS 사전 ============
wb = load_workbook(MOIS_WORD, read_only=True, data_only=True)
ws = wb[wb.sheetnames[0]]
head = [c for c in next(ws.iter_rows(min_row=1, max_row=1, values_only=True))]
def idx(patterns):
    for i, h in enumerate(head):
        for p in patterns:
            if p in str(h or ""): return i
    return -1
ci = {
    "단어명": idx(["공통표준단어명", "단어명"]),
    "영문약어": idx(["영문약어"]),
    "영문명": idx(["영문명"]),
    "설명": idx(["설명"]),
    "형식단어": idx(["형식단어"]),
    "도메인": idx(["도메인분류"]),
    "이음동의어": idx(["이음동의어"]),
    "금칙어": idx(["금칙어"]),
}
mois = []
for r in ws.iter_rows(min_row=2, values_only=True):
    if not r or not r[ci["단어명"]]: continue
    mois.append({
        "단어명": str(r[ci["단어명"]]).strip(),
        "영문약어": str(r[ci["영문약어"]] or "").strip().upper(),
        "영문명": str(r[ci["영문명"]] or "").strip(),
        "설명": str(r[ci["설명"]] or "").replace("_x000D_", " ").strip(),
        "형식단어": r[ci["형식단어"]] or "",
        "도메인": r[ci["도메인"]] or "",
    })
wb.close()

mois_by_abbr = {m["영문약어"]: m for m in mois if m["영문약어"]}
mois_by_name = {m["단어명"]: m for m in mois}
mois_abbrs = set(mois_by_abbr.keys())
mois_names = set(mois_by_name.keys())

# ============ RAMP 스키마 — Case 2 영향 컬럼 ============
wb = load_workbook(SCHEMA, read_only=True, data_only=True)
ws = wb["컬럼"]
schema_cols = []
for r in ws.iter_rows(min_row=2, values_only=True):
    if not r or not r[1]: continue
    schema_cols.append({
        "tbl": (r[0] or "").strip(),
        "en": (r[1] or "").strip().lower(),
        "kr": (r[2] or "").strip(),
    })
wb.close()

# 토큰 -> 컬럼 매핑
token_cols = defaultdict(list)
for c in schema_cols:
    for tk in re.split(r"_+", c["en"]):
        if tk:
            token_cols[tk.upper()].append(f"{c['tbl']}.{c['en']}({c['kr']})")

# ============ 4 분류 ============
match_full = []     # 완전일치
case1 = []          # 한글동일 영문약어다름
case2 = []          # 영문약어동일 한글다름
pure_only = []      # 순수 RAMP only

for d in ramp:
    a = d["영문약어"]; n = d["단어명"]
    in_mois_abbr = a in mois_abbrs
    in_mois_name = n in mois_names
    if not a: continue
    if in_mois_abbr and in_mois_name:
        mm_by_abbr = mois_by_abbr[a]
        mm_by_name = mois_by_name[n]
        if mm_by_abbr["단어명"] == n:
            match_full.append((d, mm_by_abbr))
        else:
            # 영문약어 같은데 한글 다름 = Case 2
            # 한글이 MOIS에 있긴 함 (다른 영문약어로) = Case 1 도 됨 → 이쪽은 양쪽 동시 충돌
            # 우선순위: Case 2 분류
            case2.append((d, mm_by_abbr, mm_by_name))
    elif in_mois_abbr and not in_mois_name:
        case2.append((d, mois_by_abbr[a], None))
    elif not in_mois_abbr and in_mois_name:
        case1.append((d, mois_by_name[n]))
    else:
        pure_only.append(d)

print(f"완전일치 {len(match_full)}, Case1 {len(case1)}, Case2 {len(case2)}, 순수 RAMP only {len(pure_only)}")
print(f"합계 {len(match_full)+len(case1)+len(case2)+len(pure_only)} / {len(ramp)}")

# ============ xlsx 작성 ============
wb = Workbook()

# 스타일
hdr_font = Font(bold=True, color="FFFFFF")
hdr_fill = PatternFill("solid", fgColor="305496")
hdr_align = Alignment(horizontal="center", vertical="center")
sub_fill = PatternFill("solid", fgColor="D9E1F2")

def style_header(ws, ncol):
    for c in range(1, ncol+1):
        cell = ws.cell(row=1, column=c)
        cell.font = hdr_font; cell.fill = hdr_fill; cell.alignment = hdr_align

def auto_width(ws, widths):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

# 시트 1: 요약
ws = wb.active
ws.title = "요약"
ws.append(["구분", "건수", "처리방향", "비고"])
ws.append(["전체 RAMP 사전", len(ramp), "", "기준"])
ws.append(["① 완전일치 (영문약어+한글)", len(match_full), "SKIP", "MOIS에 이미 있음 — 행안부 우선이라 그대로 사용"])
ws.append(["② Case1: 한글동일/영문약어 다름", len(case1), "SKIP + 컬럼 rename", "MOIS 영문약어로 RAMP 컬럼명 일괄 변환 필요 (별도 작업)"])
ws.append(["③ Case2: 영문약어동일/한글 다름", len(case2), "★ 결정필요", "의미 충돌. 새 영문약어 부여 / 보류 / 통일 중 선택"])
ws.append(["④ 순수 RAMP only", len(pure_only), "★ TB_WORD INSERT", "comm_stnd_yn='N' 으로 등록 대상"])
ws.append([])
ws.append(["MOIS 사전 (참고)", len(mois), "", "행안부 공통표준단어 (comm_stnd_yn='Y')"])
ws.append(["MOIS only (RAMP 미사용)", len(mois_abbrs - {d['영문약어'] for d in ramp}), "", "RAMP 신규 도입 후보"])

style_header(ws, 4)
auto_width(ws, [38, 10, 22, 80])
for row in ws.iter_rows(min_row=2, max_col=4):
    for cell in row:
        cell.alignment = Alignment(vertical="center", wrap_text=True)
ws.row_dimensions[1].height = 22

# 시트 2: 완전일치 (참고용, 데이터만)
ws = wb.create_sheet("완전일치_797_SKIP")
ws.append(["단어명", "영문약어", "영문명", "형식단어", "도메인", "설명(RAMP)"])
for d, m in match_full:
    ws.append([d["단어명"], d["영문약어"], d["영문명"], d["형식단어"], d["도메인"], d["설명"][:200]])
style_header(ws, 6)
auto_width(ws, [20, 18, 35, 12, 20, 60])

# 시트 3: Case1
ws = wb.create_sheet("Case1_한글동일_영문다름_204")
ws.append(["한글 단어명", "RAMP 영문약어", "MOIS 영문약어 (정답)", "변경 액션",
           "RAMP 영문명", "MOIS 영문명", "RAMP 설명"])
case1_sorted = sorted(case1, key=lambda x: x[0]["단어명"])
for d, m in case1_sorted:
    ws.append([
        d["단어명"], d["영문약어"], m["영문약어"],
        f"RAMP {d['영문약어']} → {m['영문약어']} 로 컬럼 영문약어 일괄 변경",
        d["영문명"], m["영문명"], d["설명"][:100]
    ])
style_header(ws, 7)
auto_width(ws, [18, 15, 18, 50, 30, 30, 50])

# 시트 4: Case2 (영문약어 동일, 한글 다름) — 영향 컬럼 포함
ws = wb.create_sheet("Case2_영문동일_한글다름_77")
ws.append(["영문약어", "RAMP 한글", "MOIS 한글 (행안부 표준)",
           "RAMP 영문명", "MOIS 영문명",
           "RAMP 설명", "MOIS 설명",
           "영향 컬럼수", "샘플 컬럼 3개"])
case2_sorted = sorted(case2, key=lambda x: x[0]["영문약어"])
for d, m_abbr, m_name in case2_sorted:
    a = d["영문약어"]
    cols = token_cols.get(a, [])
    sample = "; ".join(cols[:3])
    ws.append([
        a, d["단어명"], m_abbr["단어명"],
        d["영문명"], m_abbr["영문명"],
        d["설명"][:80], m_abbr["설명"][:80],
        len(cols), sample
    ])
style_header(ws, 9)
auto_width(ws, [13, 18, 22, 25, 25, 50, 50, 12, 70])

# 시트 5: 순수 RAMP only (INSERT 대상)
ws = wb.create_sheet("순수RAMPonly_495_INSERT")
ws.append(["단어명", "영문약어", "영문명", "단어설명",
           "형식단어여부", "도메인분류명", "이음동의어목록", "금칙어목록",
           "표준여부(comm_stnd_yn)", "비고"])
pure_sorted = sorted(pure_only, key=lambda d: d["단어명"])
for d in pure_sorted:
    ws.append([
        d["단어명"], d["영문약어"], d["영문명"], d["설명"][:200],
        d["형식단어"] or "N",
        d["도메인"], d["이음동의어"], d["금칙어"],
        "N",
        "RAMP 기관 특화"
    ])
style_header(ws, 10)
auto_width(ws, [20, 15, 35, 60, 12, 18, 20, 15, 20, 18])

wb.save(OUT)
print(f"\nSAVED: {OUT}")
print(f"size: {OUT.stat().st_size:,} bytes")
