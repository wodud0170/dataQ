# -*- coding: utf-8 -*-
"""행안부 용어 원본 + 일괄등록 xlsx + DB tb_terms 구조 점검."""
from openpyxl import load_workbook
from pathlib import Path

BASE = Path(r"C:\Users\장재영\Desktop\dataQ\q-center\src\main\resources\seed\행안부_공통표준")
ORIG = BASE / "행정안전부_공공데이터 공통표준용어.xlsx"
UP1 = BASE / "용어사전_일괄등록_1.xlsx"
UP2 = BASE / "용어사전_일괄등록_2.xlsx"

for label, path in [("ORIG", ORIG), ("UP1", UP1), ("UP2", UP2)]:
    print(f"\n=== {label}: {path.name} ===")
    wb = load_workbook(path, read_only=True, data_only=True)
    for sn in wb.sheetnames:
        ws = wb[sn]
        print(f"\n  sheet '{sn}'  (max_row: {ws.max_row}, max_col: {ws.max_column})")
        rows = list(ws.iter_rows(min_row=1, max_row=4, values_only=True))
        for i, row in enumerate(rows, 1):
            cells = [str(c)[:32] if c is not None else "" for c in (row or [])[:16]]
            print(f"    r{i}: {cells}")
    wb.close()
