"""
가설 검증: RG_DOCUMENT/RG_DETAIL 이 모든 유형의 마스터이고,
유형별 메타 테이블들은 BSID/DSID를 키로 공유하는 부속인가?
"""
from openpyxl import load_workbook
from pathlib import Path
from collections import defaultdict

BASE = Path(r"C:\Users\장재영\Desktop\dataQ\CAMS표준화")
SCHEMA = BASE / "01_원본자료" / "CAMS_SCHEMA_원본.xlsx"

wb = load_workbook(SCHEMA, read_only=True, data_only=True)
cols = []
for r in wb["컬럼정의"].iter_rows(min_row=2, values_only=True):
    if r and r[0]:
        cols.append({
            "tbl": (r[0] or "").strip(),
            "tbl_cmt": (r[1] or "").strip(),
            "col": (r[2] or "").strip().upper(),
            "col_cmt": (r[3] or "").strip(),
            "pk": (r[8] or "").strip(),
            "fk": (r[9] or "").strip(),
        })
wb.close()

# 유형별 메인 후보 (사용자 인지 기반)
candidates = {
    "일반문서(메인)":  ["RG_DOCUMENT", "RG_DETAIL"],
    "정부간행물 메타": ["RG_GOVTINFO", "RG_GOVTINFO_CR", "RG_GOVTMASTER"],
    "총독부 메타":     ["RG_OLD_GOVTINFO", "RG_OLD_GOVTINFO_CR"],
    "행정박물 메타":   ["RG_ADMIN_ADDINFO", "RG_TADMIN_ADDINFO"],
    "시청각 메타":     ["RG_SSENSES", "RG_SSENSESKEEP", "RG_SSENESELIST",
                       "SV_ADDITION_LANGUAGE", "SV_ARCHIVE_ADDITION_ITEM"],
    "해외기록 메타":   ["RG_POSS_MEDIAINFO"],
    "구술 메타":       ["RG_ORAL_ARCHIVE", "RG_ORAL_DOCUMENT",
                       "RG_ORAL_DOCUMENT_ITEM", "RG_ORAL_KEYWORD", "RG_ORAL_CONCERNED"],
    "회의록 메타":     ["RG_SREPORT_INFO"],
}

key_cols = ["BSID", "DSID", "MBSID"]

cols_by_tbl = defaultdict(list)
for c in cols:
    cols_by_tbl[c["tbl"]].append(c)

print(f"\n{'='*100}")
print(f"{'유형':<18}{'테이블':<28}{'PK 컬럼':<30}{'BSID':<6}{'DSID':<6}{'기타키':<20}")
print(f"{'-'*100}")

for grp, tbls in candidates.items():
    for tn in tbls:
        cs = cols_by_tbl.get(tn, [])
        if not cs:
            print(f"{grp:<18}{tn:<28}{'(없음)':<30}")
            continue
        pks = [c["col"] for c in cs if c["pk"] == "Y"]
        has_bsid = "BSID" in [c["col"] for c in cs]
        has_dsid = "DSID" in [c["col"] for c in cs]
        other_keys = [p for p in pks if p not in key_cols]
        print(f"{grp:<18}{tn:<28}{'; '.join(pks):<30}{'Y' if has_bsid else '-':<6}{'Y' if has_dsid else '-':<6}{'; '.join(other_keys[:3])[:20]:<20}")
    print()

print("\n=== 가설 검증 ===")
print("[가설] RG_DOCUMENT/RG_DETAIL = 모든 기록물 마스터")
print("       유형별 메타 = BSID/DSID 공유 부속 (1:1 또는 1:N)")
print()
print("[증거 기준]")
print("  1) 유형 메타 테이블이 BSID 만 보유 → 철 단위 1:1 부속")
print("  2) BSID+DSID 둘 다 → 건 단위 1:1 부속")
print("  3) 자체 PK + BSID/DSID 없음 → 별도 키체계 (독립 마스터)")
