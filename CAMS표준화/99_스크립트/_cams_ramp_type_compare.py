"""
CAMS 유형별 구조 ↔ RAMP tb_rd* 대응 구조 비교.

확인 항목:
  - RAMP에 유형별 부속 테이블이 있는지 (정부간행물·행정박물·시청각·구술 등)
  - RAMP가 1상속 패턴인지, 단일테이블에 type_cd 컬럼만으로 분기하는지
  - 마이그레이션 시 매핑 방향
"""
from openpyxl import load_workbook
from pathlib import Path
from collections import defaultdict

BASE = Path(r"C:\Users\장재영\Desktop\dataQ\CAMS표준화")
SCHEMA = BASE / "CAMS_RAMP_통합" / "ramp기관스키마정보.xlsx"

wb = load_workbook(SCHEMA, read_only=True, data_only=True)
ramp_tables = []
for r in wb["테이블 목록"].iter_rows(min_row=2, values_only=True):
    if r and r[0]:
        ramp_tables.append({"tbl": (r[0] or "").strip(), "cmt": (r[1] or "").strip() if len(r) > 1 else ""})

ramp_cols = []
for r in wb["컬럼"].iter_rows(min_row=2, values_only=True):
    if r and r[1]:
        ramp_cols.append({
            "tbl": (r[0] or "").strip(),
            "en": (r[1] or "").strip().lower(),
            "kr": (r[2] or "").strip(),
            "pk": (r[7] or "").strip(),
        })
wb.close()

cols_by_tbl = defaultdict(list)
for c in ramp_cols:
    cols_by_tbl[c["tbl"]].append(c)

# === 유형별 키워드로 RAMP 테이블 검색 ===
KEYWORDS = {
    "정부간행물": ["govt", "publication", "publ", "gov_pub", "pblc"],
    "총독부": ["oldgov", "japanese", "jeongryung"],
    "행정박물": ["adarchival", "adobj", "admin_obj", "admin"],
    "구술": ["oral", "intervw"],
    "시청각": ["av_", "video", "audio", "phtg", "photo"],
    "해외기록": ["foreign", "overseas", "overs"],
    "회의록": ["meeting", "cnfrn"],
}

print(f"RAMP 테이블 수: {len(ramp_tables)}")
print(f"\n=== 유형별 RAMP 테이블 매칭 ===\n")
for tp, kws in KEYWORDS.items():
    matched = []
    for t in ramp_tables:
        tn = t["tbl"].lower()
        cm = t["cmt"]
        if any(k in tn for k in kws):
            matched.append(t)
            continue
        if any(tp in cm for tp in [tp]):
            matched.append(t)
    print(f"[{tp}] {len(matched)} 후보")
    for t in matched[:8]:
        cs = cols_by_tbl.get(t["tbl"], [])
        pk = [c["en"] for c in cs if c["pk"] == "Y"]
        has_fls = any(c["en"] == "fls_id" for c in cs)
        has_ritm = any(c["en"] == "ritm_id" for c in cs)
        marker = ""
        if has_fls: marker += " [fls_id]"
        if has_ritm: marker += " [ritm_id]"
        print(f"  {t['tbl']:<30} — {t['cmt']:<30}  PK={'; '.join(pk)[:25]}{marker}")
    if len(matched) > 8:
        print(f"  ... 외 {len(matched)-8}")
    print()

# === tb_rdfolder / tb_rdrecord 의 유형 분기 컬럼 ===
print(f"\n=== tb_rdfolder 유형 분기 컬럼 ===")
for c in cols_by_tbl.get("tb_rdfolder", []):
    en = c["en"]; kr = c["kr"]
    if "type" in en or "kind" in en or "유형" in kr or "종류" in kr or "dscd" in en or "구분" in kr:
        print(f"  {en} ({kr})")

print(f"\n=== tb_rdrecord 유형 분기 컬럼 ===")
for c in cols_by_tbl.get("tb_rdrecord", []):
    en = c["en"]; kr = c["kr"]
    if "type" in en or "kind" in en or "유형" in kr or "종류" in kr or "dscd" in en or "구분" in kr:
        print(f"  {en} ({kr})")
