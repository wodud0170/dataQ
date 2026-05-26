"""
CAMS 기록물 유형 분석 — 2단계: 코드 테이블 조회 + 유형별 부속 테이블 패턴 분류.

수행:
  1) 코드 테이블 (CM_CODE / RG001, RG008 등) 찾아서 코드값·코드명 추출
  2) 유형별 후보 테이블의 키 구조 (BSID/DSID/자체키) 분석
  3) 4 패턴 분류
     - A: RG_DETAIL 단일 (코드로만 분기, 동일 메타)
     - B: RG_DETAIL + 부속 1:1 (BSID/DSID 공유)
     - C: 별도 키 (BSID 없이 자체 ID, 예: 구술 GROUP_ID)
     - D: 임시·복본 (작업영역, 메인과 동기화)
  4) 유형별 컬럼 수·메타 차이 매트릭스

출력: 05_CAMS기록물유형_2026-05-21/_02_패턴분류.xlsx
"""
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from pathlib import Path
from collections import defaultdict, Counter
import re

BASE = Path(r"C:\Users\장재영\Desktop\dataQ\CAMS표준화")
SCHEMA = BASE / "01_원본자료" / "CAMS_SCHEMA_원본.xlsx"
OUT = BASE / "05_CAMS기록물유형_2026-05-21" / "_02_패턴분류.xlsx"

# Load
wb = load_workbook(SCHEMA, read_only=True, data_only=True)
tables = []
for r in wb["테이블"].iter_rows(min_row=2, values_only=True):
    if r and r[1]:
        tables.append({"tbl": (r[1] or "").strip(), "cmt": (r[2] or "").strip()})

cols = []
for r in wb["컬럼정의"].iter_rows(min_row=2, values_only=True):
    if r and r[0]:
        cols.append({
            "tbl": (r[0] or "").strip(),
            "tbl_cmt": (r[1] or "").strip(),
            "col": (r[2] or "").strip(),
            "col_cmt": (r[3] or "").strip(),
            "type": (r[4] or "").strip(),
            "len": r[5],
            "notnull": (r[6] or "").strip(),
            "pk": (r[8] or "").strip(),
            "fk": (r[9] or "").strip(),
        })
wb.close()

cols_by_tbl = defaultdict(list)
for c in cols:
    cols_by_tbl[c["tbl"]].append(c)

# === 1) 코드 테이블 찾기 ===
code_tables = [t for t in tables if re.search(r"CODE|CM_COMMONCODE|CMN.*CD", t["tbl"].upper())]
print(f"코드 테이블 후보: {len(code_tables)}")
for t in code_tables[:20]:
    print(f"  {t['tbl']} — {t['cmt']}")

# === 2) 유형별 후보 테이블 정의 ===
# 1단계 결과 + 추가 휴리스틱
type_groups = {
    "01_일반문서": ["RG_DOCUMENT", "RG_DETAIL", "RG_APPENDFILE", "RG_DOCTUK", "RG_DETTUK"],
    "02_시청각": ["RG_SSENSES", "RG_SSENSESKEEP", "RG_SSENESELIST", "RG_REQ_AVP", "RG_REQ_AVP_ERROR",
                "SV_ADDITION_LANGUAGE", "SV_ARCHIVE_ADDITION_ITEM"],
    "04_총독부": ["RG_OLD_GOVTINFO", "RG_OLD_GOVTINFO_CR", "CM_OLDGOV_DIVSYS", "CM_OLDGOV_DIVTABLE",
                "CAMS_TEMP_ORGTYPEJ"],
    "05_정부간행물": ["RG_GOVTINFO", "RG_GOVTINFO_CR", "RG_GOVTMASTER",
                  "RG_GOVTADDFILE_REG_REQ", "RG_GOVTADDFILE_REG_REQ_INFO",
                  "RG_GOVTFILE_REG_REQ", "RG_GOVTFILE_REG_REQ_INFO",
                  "RG_SGOVTPUB"],
    "06_해외기록": ["RG_POSS_MEDIAINFO", "CM_FOREIGN", "CM_OVERS_DIVSYS", "CM_OVERS_DIVTABLE",
                 "SV_FOREIGN_MEDIA_ARRANGE_HIS", "SV_FOREIGN_MEDIA_ARRANGE_LIST",
                 "SV_REQUEST_FOREIGN_MEDIA_LIST"],
    "07_역사기록": [],  # CAMS에서 명확히 찾기 어려움 — 추후 식별
    "08_행정박물": ["RG_ADMIN_ADDINFO", "RG_TADMIN_ADDINFO", "RG_SHANGJONGLIST",
                "RG_NONFIXED", "RG_TAKEOVER_LIST", "RG_TAKEOVER_PAPER",
                "RG_ORGAN_TYPE_STTST", "RG_ORGAN_TYPE_STTST_2",
                "CM_ADMIN_CLASS", "CM_GOVART_DIVSYS"],
    "구술": ["RG_ORAL_ARCHIVE", "RG_ORAL_DOCUMENT", "RG_ORAL_DOCUMENT_ITEM",
            "RG_ORAL_KEYWORD", "RG_ORAL_CONCERNED"],
    "회의록": ["RG_SREPORT_INFO", "TB_CNFRN_REPORT_FORM"],
}

# 각 테이블의 핵심 키 식별
def key_profile(tn):
    cs = cols_by_tbl.get(tn, [])
    keys = [c["col"] for c in cs if c["pk"] == "Y"]
    return keys

# 각 그룹의 키 매트릭스
print("\n=== 유형별 키 구조 ===")
group_rows = []
for grp, tbls in type_groups.items():
    for tn in tbls:
        cs = cols_by_tbl.get(tn, [])
        if not cs:
            group_rows.append({"group": grp, "tbl": tn, "cmt": "(스키마에 없음)", "n_col": 0,
                               "pks": "", "has_bsid": "", "has_dsid": "", "has_mbsid": ""})
            continue
        keys = key_profile(tn)
        has_bsid = any(c["col"].upper() == "BSID" for c in cs)
        has_dsid = any(c["col"].upper() == "DSID" for c in cs)
        has_mbsid = any(c["col"].upper() == "MBSID" for c in cs)
        cmt = next((t["cmt"] for t in tables if t["tbl"] == tn), "")
        group_rows.append({
            "group": grp, "tbl": tn, "cmt": cmt, "n_col": len(cs),
            "pks": "; ".join(keys),
            "has_bsid": "Y" if has_bsid else "",
            "has_dsid": "Y" if has_dsid else "",
            "has_mbsid": "Y" if has_mbsid else "",
        })

# === 3) 패턴 분류 ===
# 패턴 A : BSID + DSID 보유 (메인 1:1 with RG_DETAIL)
# 패턴 B : BSID 만 보유 (RG_DOCUMENT 와 1:1)
# 패턴 C : BSID/DSID 모두 없음 + 자체 PK (별도 키체계)
# 패턴 D : 코드 테이블 / 임시 / 통계 등 — 메타 보조
def classify(row):
    if row["has_bsid"] and row["has_dsid"]:
        return "A_철건1:1"
    if row["has_bsid"] and not row["has_dsid"]:
        return "B_철단위"
    if not row["has_bsid"] and not row["has_dsid"]:
        # 코드/임시/통계는 D
        tn = row["tbl"]
        if tn.startswith("CM_") or "TEMP" in tn or "STTST" in tn or "REQ" in tn:
            return "D_부속메타"
        return "C_별도키"
    return "?"

for r in group_rows:
    r["pattern"] = classify(r)

# === 4) DOCTYPE / ARCAVETYPE / ORGTYPE 코드 테이블 데이터 찾기 ===
# CAMS는 CM_COMMONCODE 류일 가능성 — 컬럼명에 RG001, RG008 등이 데이터 값으로 들어감
code_value_cols = []
for c in cols:
    if c["tbl"].upper().startswith("CM_") and any(k in c["col"].upper() for k in ["CD", "CODE", "GUBUN"]):
        code_value_cols.append(c)

# === xlsx ===
wb = Workbook()
hdr_font = Font(bold=True, color="FFFFFF")
hdr_fill = PatternFill("solid", fgColor="305496")
hdr_align = Alignment(horizontal="center", vertical="center")
fills = {
    "A_철건1:1": PatternFill("solid", fgColor="C6E0B4"),
    "B_철단위":  PatternFill("solid", fgColor="DDEBF7"),
    "C_별도키":  PatternFill("solid", fgColor="FFD966"),
    "D_부속메타": PatternFill("solid", fgColor="F4B084"),
}

def style_header(ws, ncol):
    for c in range(1, ncol+1):
        cell = ws.cell(row=1, column=c)
        cell.font = hdr_font; cell.fill = hdr_fill; cell.alignment = hdr_align

def auto_width(ws, widths):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

# 시트 1: 유형 코드 정의 (DOCTYPE / ARCAVETYPE / ORGTYPE)
ws = wb.active
ws.title = "유형코드_정의"
ws.append(["분기 컬럼", "코드그룹", "코드", "코드명", "비고"])
ws.append(["DOCTYPE",     "RG001", "A", "문서대장",     "기록물 형태"])
ws.append(["DOCTYPE",     "RG001", "B", "도면",         ""])
ws.append(["DOCTYPE",     "RG001", "C", "사진/필름",    ""])
ws.append(["DOCTYPE",     "RG001", "D", "녹음/동영상",  ""])
ws.append(["DOCTYPE",     "RG001", "E", "카드",         ""])
ws.append(["DOCTYPE",     "RG001", "G", "국무회의록",   "F 미사용?"])
ws.append([])
ws.append(["ARCAVETYPE",  "", "01", "일반문서",   "기록물 구분"])
ws.append(["ARCAVETYPE",  "", "02", "시청각",     ""])
ws.append(["ARCAVETYPE",  "", "03", "(미사용)",   "코멘트에 표기 없음"])
ws.append(["ARCAVETYPE",  "", "04", "총독부",     ""])
ws.append(["ARCAVETYPE",  "", "05", "정부간행물", ""])
ws.append(["ARCAVETYPE",  "", "06", "해외기록물", ""])
ws.append(["ARCAVETYPE",  "", "07", "역사기록물", "RG_OLD_GOVTINFO 와 다른 별도 그룹?"])
ws.append(["ARCAVETYPE",  "", "08", "행정박물",   ""])
ws.append([])
ws.append(["ORGTYPE",     "RG008", "Z", "파일(사진/필름류)", "RG_DOCUMENT.ORGTYPE 코멘트에서"])
style_header(ws, 5)
auto_width(ws, [15, 12, 8, 22, 30])

# 시트 2: 유형 × 테이블 + 패턴
ws = wb.create_sheet("유형별_테이블_패턴")
ws.append(["유형(ARCAVETYPE)", "테이블", "코멘트", "컬럼수", "PK들",
           "BSID", "DSID", "MBSID", "패턴분류"])
for r in group_rows:
    ws.append([r["group"], r["tbl"], r["cmt"], r["n_col"], r["pks"],
               r["has_bsid"], r["has_dsid"], r["has_mbsid"], r["pattern"]])
    # 색상
    last = ws.max_row
    fill = fills.get(r["pattern"])
    if fill:
        ws.cell(row=last, column=9).fill = fill
style_header(ws, 9)
auto_width(ws, [15, 28, 40, 8, 22, 7, 7, 8, 14])
ws.freeze_panes = "A2"

# 시트 3: 패턴별 요약
ws = wb.create_sheet("패턴_요약")
ws.append(["패턴", "정의", "의미", "통합 처리방향"])
ws.append(["A_철건1:1", "BSID+DSID 둘 다", "RG_DETAIL 과 동일 그라뉴 — 1:1 부속 메타",
          "통합 DB 에서 부속 컬럼을 RG_DETAIL 에 흡수 또는 별도 1:1 유지"])
ws.append(["B_철단위", "BSID 만 (DSID 없음)", "철 단위 추가 정보 (목차·이력)",
          "RG_DOCUMENT 부속 — 동일하게 처리"])
ws.append(["C_별도키", "BSID/DSID 없음 + 자체 PK", "구술기록 같은 별도 키체계",
          "★ 통합 시 별도 ID 매핑 테이블 또는 BSID 공유로 통합 결정"])
ws.append(["D_부속메타", "코드/임시/통계 테이블", "메타·관리 영역",
          "마이그레이션 시 별도 처리"])
style_header(ws, 4)
auto_width(ws, [14, 28, 35, 50])
for row in ws.iter_rows(min_row=2, max_col=4):
    for cell in row:
        cell.alignment = Alignment(vertical="center", wrap_text=True)
    row[0].fill = fills.get(row[0].value, PatternFill())

# 시트 4: 패턴별 카운트
ws = wb.create_sheet("패턴별_카운트")
ws.append(["유형", "A_철건1:1", "B_철단위", "C_별도키", "D_부속메타", "(없음)", "합계"])
group_pattern_count = defaultdict(lambda: defaultdict(int))
for r in group_rows:
    group_pattern_count[r["group"]][r["pattern"]] += 1
    if r["cmt"] == "(스키마에 없음)":
        group_pattern_count[r["group"]]["없음"] += 1
for grp, pc in group_pattern_count.items():
    ws.append([grp, pc["A_철건1:1"], pc["B_철단위"], pc["C_별도키"], pc["D_부속메타"], pc["없음"],
               sum(pc.values()) - pc.get("없음", 0)])
style_header(ws, 7)
auto_width(ws, [18, 14, 14, 14, 14, 10, 10])

# 시트 5: 코드 테이블 후보
ws = wb.create_sheet("코드테이블_후보")
ws.append(["테이블", "코멘트", "컬럼수"])
for t in code_tables:
    ws.append([t["tbl"], t["cmt"], len(cols_by_tbl.get(t["tbl"], []))])
style_header(ws, 3)
auto_width(ws, [30, 50, 10])

wb.save(OUT)
print(f"\nSAVED: {OUT}")

# 콘솔 요약
print(f"\n=== 패턴별 카운트 ===")
total_pattern = Counter()
for r in group_rows:
    if r["cmt"] != "(스키마에 없음)":
        total_pattern[r["pattern"]] += 1
for p, c in total_pattern.most_common():
    print(f"  {p}: {c}")
