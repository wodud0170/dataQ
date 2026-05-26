# -*- coding: utf-8 -*-
"""
Phase 2-1/2-2 BEFORE-AFTER 정합성 검증 + RAMP 모델 AFTER UPDATE.

검증 항목:
  1. Phase 2-1 컬럼 BEFORE/AFTER 변환 룰
  2. Phase 2-2 D3 시간 컬럼 적용
  3. Phase 2-2 D5 R8 보충
  4. RAMP 모델 (tb_data_model_attr) BEFORE 상태 vs AFTER 적용

작업:
  - 정합성 검증 결과 보고
  - tb_data_model_attr UPDATE (AFTER 영문/한글)
"""
from openpyxl import load_workbook
from pathlib import Path
from collections import defaultdict, Counter
import subprocess, csv, io, os, base64

BASE = Path(r"C:\Users\장재영\Desktop\dataQ\CAMS표준화")
COL_BA = BASE / "04_RAMP분석_2026-05-21" / "00_핵심산출" / "RAMP_컬럼_BEFORE_AFTER_2026-05-23.xlsx"
TIME_BA = BASE / "04_RAMP분석_2026-05-21" / "00_핵심산출" / "Phase2_시간컬럼_BEFORE_AFTER_2026-05-23.xlsx"
D5_XLSX = BASE / "04_RAMP분석_2026-05-21" / "00_핵심산출" / "Phase2_D5_R8형식단어_2026-05-23.xlsx"
RAMP_SCHEMA = BASE / "CAMS_RAMP_통합" / "ramp기관스키마정보.xlsx"
OUT_SQL = BASE / "04_RAMP분석_2026-05-21" / "99_실행SQL_기록" / "Phase2_model_after_update.sql"

# ============ 1. 산출물 로드 ============
print("=== 1. 산출물 로드 ===")

# Phase 2-1
col_after = {}  # (tbl, col_en BEFORE) → (col_en AFTER, col_kr AFTER)
wb = load_workbook(COL_BA, read_only=True, data_only=True)
for r in wb["컬럼BEFORE_AFTER"].iter_rows(min_row=2, values_only=True):
    if r and r[1] and r[2]:
        col_after[(str(r[1]), str(r[2]))] = (str(r[3] or r[2]), str(r[5] or r[4] or ""))
wb.close()
print(f"  Phase 2-1 변환: {len(col_after)}")

# Phase 2-2 D3 시간
time_after = {}  # (tbl, col_en) → (new_en, new_kr, dom)
wb = load_workbook(TIME_BA, read_only=True, data_only=True)
for r in wb["적용_BEFORE_AFTER"].iter_rows(min_row=2, values_only=True):
    if r and r[1] and r[2]:
        time_after[(str(r[1]), str(r[2]))] = (str(r[2]), str(r[4] or r[3] or ""), str(r[7] or ""))
wb.close()
print(f"  Phase 2-2 D3: {len(time_after)}")

# Phase 2-2 D5
d5_kr = {}
wb = load_workbook(D5_XLSX, read_only=True, data_only=True)
for r in wb["새용어_BEFORE_AFTER"].iter_rows(min_row=2, values_only=True):
    if r and r[1] and r[3]:
        d5_kr[str(r[1])] = str(r[3])
wb.close()
print(f"  Phase 2-2 D5: {len(d5_kr)}")

# ============ 2. RAMP 원본 로드 + AFTER 적용 ============
wb = load_workbook(RAMP_SCHEMA, read_only=True, data_only=True)
ws = wb["컬럼"]
updates = []  # (tbl, col_en BEFORE, col_en AFTER, col_kr AFTER)
seen = set()
for r in ws.iter_rows(min_row=2, values_only=True):
    if r and r[0] and r[1]:
        key = (str(r[0]).strip(), str(r[1]).strip())
        if key in seen: continue
        seen.add(key)
        bef_en = key[1]
        bef_kr = str(r[2] or "").strip()
        # Phase 2-1 적용
        aft_en, aft_kr = col_after.get(key, (bef_en, bef_kr))
        if not aft_kr: aft_kr = bef_kr
        # D5 보충 적용 (한글)
        if aft_kr in d5_kr: aft_kr = d5_kr[aft_kr]
        # D3 시간 적용 (한글)
        if key in time_after:
            _, t_kr, _ = time_after[key]
            if t_kr: aft_kr = t_kr
        # 공백 정규화
        aft_kr = aft_kr.replace(" ", "")
        # 영문 대문자
        aft_en = aft_en.upper()
        if aft_en != bef_en or aft_kr != bef_kr:
            updates.append((key[0], bef_en, aft_en, aft_kr))
wb.close()
print(f"\n=== 2. AFTER UPDATE 대상: {len(updates)} ===")

# ============ 3. 정합성 검증 ============
# 변환 유형 카운트
en_change = sum(1 for u in updates if u[1] != u[2])
kr_change = sum(1 for u in updates if u[1].upper() != u[2] or True)  # 이건 모든 row
en_only = sum(1 for u in updates if u[1].upper() == u[2])  # 대문자 변환만
print(f"  영문 변경 (단어 약어 또는 대문자): {en_change}")
print(f"  대문자 변환만: {en_only}")

# ============ 4. SQL 생성 ============
def esc(s): return str(s).replace("'", "''") if s else ""
sql = ["-- Phase 2 RAMP 모델 AFTER UPDATE", "BEGIN;", ""]
for tbl, bef_en, aft_en, aft_kr in updates:
    # tb_data_model_attr 의 (dm_id RAMP, obj_owner=RAMP, obj_nm=tbl, attr_nm=bef_en) → UPDATE attr_nm, attr_nm_kr
    sql.append(
        f"UPDATE quality.tb_data_model_attr SET attr_nm='{esc(aft_en)}', attr_nm_kr='{esc(aft_kr)}' "
        f"WHERE dm_id=(SELECT dm_id FROM quality.tb_data_model WHERE dm_nm='RAMP' LIMIT 1) "
        f"AND obj_nm='{esc(tbl)}' AND attr_nm='{esc(bef_en)}';"
    )
sql += ["",
        "SELECT count(*) as updated FROM quality.tb_data_model_attr WHERE dm_id=(SELECT dm_id FROM quality.tb_data_model WHERE dm_nm='RAMP' LIMIT 1) AND attr_nm = upper(attr_nm);",
        "",
        "COMMIT;"]
OUT_SQL.write_text("\n".join(sql), encoding="utf-8")
print(f"\n→ {OUT_SQL}")
print(f"  UPDATE: {len(updates)}건")
