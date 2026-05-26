# -*- coding: utf-8 -*-
"""
Phase 2-2 도메인 진단 — 회의 결정용.

시트:
  1. 표지
  2. 요약
  3. 한글명_도메인분포   — 같은 한글명에 여러 도메인 (다중매핑)
  4. D3_시간컬럼_전수    — ~일자/일시, STRING으로 저장된 것 (테이블.컬럼)
  5. D4_순번계열_전수    — 순번 컬럼 (테이블.컬럼 + 도메인) 회의 결정용
  6. 최장적용_미리보기   — D1 최장 우선 적용 시 결과
  7. 행안부도메인_매칭참고 — 행안부 표준도메인 126개 매칭 시도
"""
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from pathlib import Path
from collections import defaultdict, Counter
import subprocess, csv, io, sys, re

BASE = Path(r"C:\Users\장재영\Desktop\dataQ\CAMS표준화")
RAMP_SCHEMA = BASE / "CAMS_RAMP_통합" / "ramp기관스키마정보.xlsx"
OUT = BASE / "04_RAMP분석_2026-05-21" / "00_핵심산출" / "Phase2_도메인진단_2026-05-23.xlsx"

# ============ 데이터 로드 ============
print("=== 로드 ===")
wb = load_workbook(RAMP_SCHEMA, read_only=True, data_only=True)
ws = wb["컬럼"]
cols = []
for r in ws.iter_rows(min_row=2, values_only=True):
    if r and r[0] and r[1]:
        cols.append({
            "tbl": str(r[0]).strip(),
            "col_en": str(r[1]).strip(),
            "col_kr": str(r[2] or "").strip(),
            "desc": str(r[3] or "").strip() if len(r) > 3 else "",
            "null": str(r[4] or "").strip() if len(r) > 4 else "",
            "dtype": str(r[5] or "").strip() if len(r) > 5 else "",
            "dlen": str(r[6] or "").strip() if len(r) > 6 else "",
            "pk": str(r[7] or "").strip() if len(r) > 7 else "",
        })
wb.close()
print(f"  RAMP 컬럼: {len(cols)}")

# 행안부 도메인 (DB)
mois_dom = {}  # nm → {grp, clsf, dtype, dlen}
r = subprocess.run(["docker","exec","-i","dataq-db","psql","-U","admin","-d","postgres",
    "-c", """COPY (SELECT domain_nm, coalesce(domain_grp_nm,''), coalesce(domain_clsf_nm,''),
             coalesce(data_type,''), coalesce(data_len::text,'')
             FROM quality.tb_domain WHERE comm_stnd_yn='Y') TO STDOUT WITH (FORMAT csv, FORCE_QUOTE *)"""],
    capture_output=True, encoding="utf-8")
for row in csv.reader(io.StringIO(r.stdout)):
    if row:
        mois_dom[row[0]] = {"grp": row[1], "clsf": row[2], "dtype": row[3], "dlen": row[4]}
print(f"  행안부 도메인: {len(mois_dom)}")

# ============ 분석 ============
print("\n=== 분석 ===")

# 1) 한글명 그룹화 → 도메인 분포
grp_kr = defaultdict(list)  # 한글명 → [(tbl, col_en, dtype, dlen, pk)]
for c in cols:
    if c["col_kr"]:
        grp_kr[c["col_kr"]].append(c)

# 다중 도메인 그룹 (같은 한글, 다른 도메인)
multi_domain = []
for kr, lst in grp_kr.items():
    domains = set((c["dtype"], c["dlen"]) for c in lst)
    if len(domains) > 1:
        # 도메인 분포 카운트
        dom_cnt = Counter((c["dtype"], c["dlen"]) for c in lst)
        multi_domain.append({
            "kr": kr,
            "col_cnt": len(lst),
            "dom_cnt": len(domains),
            "dom_dist": dom_cnt,
            "sample_cols": [(c["tbl"], c["col_en"], c["dtype"], c["dlen"], c["pk"]) for c in lst],
        })
multi_domain.sort(key=lambda x: -x["col_cnt"])
print(f"  같은 한글 다중도메인: {len(multi_domain)} 종")

# 2) 시간 컬럼 (~일자, ~일시, ~년월일 등)
TIME_SUFFIX = ("일자","일시","년월","시각","연월","시분초","월일")
time_cols = []
for c in cols:
    if any(c["col_kr"].endswith(suf) for suf in TIME_SUFFIX):
        time_cols.append(c)
# 타입별 분포
time_dtype_cnt = Counter(c["dtype"] for c in time_cols)
print(f"  시간 컬럼 (suffix 매칭): {len(time_cols)}  / 타입 분포: {dict(time_dtype_cnt)}")

# 3) 순번 계열
seq_cols = [c for c in cols if "순번" in c["col_kr"]]
seq_dom_cnt = Counter((c["dtype"], c["dlen"]) for c in seq_cols)
print(f"  순번 계열 컬럼: {len(seq_cols)}  / 도메인: {len(seq_dom_cnt)}")

# 4) 최장 적용 미리보기
longest = {}  # 한글 → 최장 도메인
def parse_len(s):
    try: return int(s)
    except: return 0
DTYPE_ORDER = ["STRING","VARCHAR","CHAR","NUMERIC","NUMBER","INTEGER","DECIMAL","DATETIME","DATE","TIMESTAMP","CLOB","BLOB"]
def dtype_rank(dt):
    for i, t in enumerate(DTYPE_ORDER):
        if t in dt.upper(): return i
    return 99
for kr, lst in grp_kr.items():
    # 최장 = 길이 큰 것. 단, STRING > NUMERIC 등 타입 우선순위는 보존
    best = max(lst, key=lambda c: (parse_len(c["dlen"]), -dtype_rank(c["dtype"])))
    longest[kr] = (best["dtype"], best["dlen"])

# ============ xlsx 생성 ============
print("\n=== xlsx 생성 ===")
wb = Workbook()
THIN = Side(border_style="thin", color="808080")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
HDR_FILL = PatternFill("solid", fgColor="1F4E78")
HDR_FONT = Font(name="맑은 고딕", size=10, color="FFFFFF", bold=True)
CELL_FONT = Font(name="맑은 고딕", size=10)
SECTION_FILL = PatternFill("solid", fgColor="D9E1F2")
DECISION_FILL = PatternFill("solid", fgColor="FFF2CC")
HIGH_FILL = PatternFill("solid", fgColor="F8CBAD")

# 시트 1: 표지
ws = wb.active; ws.title = "표지"
ws.column_dimensions["A"].width = 22; ws.column_dimensions["B"].width = 100
ws.merge_cells("A1:B1")
t = ws.cell(row=1, column=1, value="RAMP 도메인 진단 — Phase 2-2 회의 결정용")
t.font = Font(name="맑은 고딕", size=18, bold=True, color="1F4E78")
t.alignment = Alignment(horizontal="center", vertical="center")
ws.row_dimensions[1].height = 36

def cr(r, label, value):
    ws.cell(row=r, column=1, value=label).fill = SECTION_FILL
    ws.cell(row=r, column=1).font = Font(name="맑은 고딕", size=10, bold=True)
    ws.cell(row=r, column=1).border = BORDER
    c = ws.cell(row=r, column=2, value=value); c.font = Font(name="맑은 고딕", size=11)
    c.border = BORDER; c.alignment = Alignment(vertical="center", wrap_text=True)

cr(3, "작성일", "2026-05-23")
cr(4, "대상", f"RAMP 컬럼 {len(cols):,}건 도메인 분석")
cr(5, "정책 — D1", "도메인 통일 기준: 최장 우선. 행안부 충돌은 별도 검토")
cr(6, "정책 — D5", "R8 형식단어 미종결 → 형식단어 보충하여 용어 변경")
cr(7, "정책 — D6", "행안부 도메인 강제 흡수 X, 기관표준 도메인(N) 추가 가능")
cr(8, "정책 — Cubrid→Oracle", "ALTER 없음. 신규 생성이라 통일 자유도 큼")
ws.row_dimensions[9].height = 8
cr(10, "다중도메인 한글 그룹", f"{len(multi_domain)}종 — 통일 대상 (시트3)")
cr(11, "시간 컬럼 (~일자/일시)", f"{len(time_cols)}건 / 타입 분포 {dict(time_dtype_cnt)} (시트4)")
cr(12, "순번 계열", f"{len(seq_cols)}건 / 도메인 {len(seq_dom_cnt)}종 (시트5)")
cr(13, "행안부 표준도메인", f"{len(mois_dom)}개 (참고 시트7)")
ws.row_dimensions[14].height = 8
cr(15, "결정 필요 — D3", "시간 컬럼 일괄 DATE/DATETIME 정정 (사용자 의향) — 시트4 검토 후 확정")
cr(16, "결정 필요 — D4", "순번 계열 통일 vs 분리 (사용자: 여러 용어로 쪼갤 가능성) — 시트5 검토 후 회의")

# 시트 2: 요약
ws2 = wb.create_sheet("요약")
ws2.column_dimensions["A"].width = 28; ws2.column_dimensions["B"].width = 12; ws2.column_dimensions["C"].width = 50
t = ws2.cell(row=1, column=1, value="도메인 진단 통계"); t.font = Font(name="맑은 고딕", size=14, bold=True, color="1F4E78")
ws2.merge_cells("A1:C1")
r = 3
for c, h in enumerate(["항목","건수","비고"], 1):
    cell = ws2.cell(row=r, column=c, value=h); cell.fill = HDR_FILL; cell.font = HDR_FONT
    cell.alignment = Alignment(horizontal="center"); cell.border = BORDER
r += 1
stats = [
    ("RAMP 컬럼 전체", len(cols), ""),
    ("한글명 종수", len(grp_kr), ""),
    ("다중도메인 한글 그룹", len(multi_domain), "같은 한글, 다른 도메인 — 통일 대상"),
    ("├ 도메인 2개", sum(1 for x in multi_domain if x["dom_cnt"]==2), ""),
    ("├ 도메인 3개", sum(1 for x in multi_domain if x["dom_cnt"]==3), ""),
    ("├ 도메인 4개 이상", sum(1 for x in multi_domain if x["dom_cnt"]>=4), ""),
    ("시간 suffix 컬럼", len(time_cols), "일자/일시/년월 등"),
    ("├ STRING 저장", time_dtype_cnt.get("STRING",0), "D3 정정 대상"),
    ("├ DATE", time_dtype_cnt.get("DATE",0), ""),
    ("├ DATETIME", time_dtype_cnt.get("DATETIME",0), ""),
    ("├ 기타", sum(v for k,v in time_dtype_cnt.items() if k not in ("STRING","DATE","DATETIME")), ""),
    ("순번 컬럼", len(seq_cols), "D4 결정 대상"),
    ("├ 도메인 종수", len(seq_dom_cnt), ""),
    ("행안부 표준도메인", len(mois_dom), "참고용 (Y)"),
]
for label, n, note in stats:
    ws2.cell(row=r, column=1, value=label).border = BORDER
    ws2.cell(row=r, column=2, value=n).border = BORDER
    ws2.cell(row=r, column=2).alignment = Alignment(horizontal="right")
    ws2.cell(row=r, column=3, value=note).border = BORDER
    r += 1

# 시트 3: 다중도메인 한글 그룹
ws3 = wb.create_sheet("한글명_도메인분포")
H = ["No","한글명","컬럼수","도메인수","도메인 분포","최장 후보","행안부 충돌?","결정 도메인","결정사유"]
W = [5,20,8,8,46,16,12,16,30]
for i, w in enumerate(W, 1): ws3.column_dimensions[get_column_letter(i)].width = w
for i, h in enumerate(H, 1):
    c = ws3.cell(row=1, column=i, value=h); c.fill = HDR_FILL; c.font = HDR_FONT
    c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True); c.border = BORDER
ws3.row_dimensions[1].height = 30
ws3.freeze_panes = "C2"

for i, x in enumerate(multi_domain, 1):
    row = i + 1
    dist = " | ".join(f"{dt}({dl})×{n}" for (dt,dl), n in x["dom_dist"].most_common())
    long_dom = f"{longest[x['kr']][0]}({longest[x['kr']][1]})"
    # 행안부 충돌 — 행안부 도메인 분류명에 한글명이 포함되는지
    conflict = "확인 필요" if x["kr"] in mois_dom else ""
    values = [i, x["kr"], x["col_cnt"], x["dom_cnt"], dist, long_dom, conflict, "", ""]
    for j, v in enumerate(values, 1):
        c = ws3.cell(row=row, column=j, value=v); c.font = CELL_FONT; c.border = BORDER
        c.alignment = Alignment(vertical="top", wrap_text=True)
    if x["dom_cnt"] >= 4: ws3.cell(row=row, column=4).fill = HIGH_FILL
    ws3.cell(row=row, column=8).fill = DECISION_FILL
    ws3.cell(row=row, column=9).fill = DECISION_FILL

# 시트 4: D3 시간 컬럼 전수
ws4 = wb.create_sheet("D3_시간컬럼")
H4 = ["No","테이블","컬럼영문","컬럼한글","타입","길이","Null","PK","제안 도메인","결정"]
W4 = [5,28,32,28,10,8,6,5,18,14]
for i, w in enumerate(W4, 1): ws4.column_dimensions[get_column_letter(i)].width = w
for i, h in enumerate(H4, 1):
    c = ws4.cell(row=1, column=i, value=h); c.fill = HDR_FILL; c.font = HDR_FONT
    c.alignment = Alignment(horizontal="center"); c.border = BORDER
ws4.freeze_panes = "A2"

for i, c in enumerate(time_cols, 1):
    # 제안 도메인
    if c["col_kr"].endswith("일자"): suggest = "DATE"
    elif c["col_kr"].endswith(("일시","연월일시","시분초","시각")): suggest = "TIMESTAMP(0)"
    elif c["col_kr"].endswith(("년월","연월")): suggest = "STRING(6) yyyymm"
    elif c["col_kr"].endswith("월일"): suggest = "STRING(4) mmdd"
    else: suggest = "확인 필요"
    values = [i, c["tbl"], c["col_en"], c["col_kr"], c["dtype"], c["dlen"], c["null"], c["pk"], suggest, ""]
    for j, v in enumerate(values, 1):
        cell = ws4.cell(row=i+1, column=j, value=v); cell.font = CELL_FONT; cell.border = BORDER
        cell.alignment = Alignment(vertical="top", wrap_text=True)
    if c["dtype"] == "STRING":
        ws4.cell(row=i+1, column=5).fill = HIGH_FILL  # STRING 강조
    ws4.cell(row=i+1, column=10).fill = DECISION_FILL

# 시트 5: D4 순번 계열 전수
ws5 = wb.create_sheet("D4_순번계열")
H5 = ["No","테이블","컬럼영문","컬럼한글","타입","길이","Null","PK","용어후보(분리)","결정"]
W5 = [5,28,32,22,10,8,6,5,20,14]
for i, w in enumerate(W5, 1): ws5.column_dimensions[get_column_letter(i)].width = w
for i, h in enumerate(H5, 1):
    c = ws5.cell(row=1, column=i, value=h); c.fill = HDR_FILL; c.font = HDR_FONT
    c.alignment = Alignment(horizontal="center"); c.border = BORDER
ws5.freeze_panes = "A2"

# 타입+길이별 정렬 (회의 결정용)
seq_sorted = sorted(seq_cols, key=lambda c: (c["dtype"], parse_len(c["dlen"]), c["col_kr"]))
for i, c in enumerate(seq_sorted, 1):
    # 용어후보 — 도메인별 그룹 추정
    if c["dtype"] == "STRING":
        candidate = "합성ID순번"
    elif c["dtype"] in ("NUMERIC","NUMBER","INTEGER"):
        L = parse_len(c["dlen"])
        if L <= 3: candidate = "그룹순번(소)"
        elif L <= 10: candidate = "일반순번"
        else: candidate = "대용량순번"
    else: candidate = "?"
    values = [i, c["tbl"], c["col_en"], c["col_kr"], c["dtype"], c["dlen"], c["null"], c["pk"], candidate, ""]
    for j, v in enumerate(values, 1):
        cell = ws5.cell(row=i+1, column=j, value=v); cell.font = CELL_FONT; cell.border = BORDER
        cell.alignment = Alignment(vertical="top", wrap_text=True)
    ws5.cell(row=i+1, column=10).fill = DECISION_FILL

# 시트 6: 최장 적용 미리보기 (다중도메인만)
ws6 = wb.create_sheet("최장적용_미리보기")
H6 = ["No","한글명","컬럼수","현 분포","최장 적용 후","영향"]
W6 = [5,22,8,52,16,30]
for i, w in enumerate(W6, 1): ws6.column_dimensions[get_column_letter(i)].width = w
for i, h in enumerate(H6, 1):
    c = ws6.cell(row=1, column=i, value=h); c.fill = HDR_FILL; c.font = HDR_FONT
    c.alignment = Alignment(horizontal="center"); c.border = BORDER
ws6.freeze_panes = "A2"
for i, x in enumerate(multi_domain, 1):
    dist = " | ".join(f"{dt}({dl})×{n}" for (dt,dl), n in x["dom_dist"].most_common())
    long_dom = f"{longest[x['kr']][0]}({longest[x['kr']][1]})"
    # 변경 영향
    main_dom = x["dom_dist"].most_common(1)[0][0]
    if main_dom == longest[x['kr']]:
        impact = "현 최빈 = 최장 (영향 적음)"
    else:
        impact = f"최빈 {main_dom[0]}({main_dom[1]}) → {long_dom} 변경 영향"
    for j, v in enumerate([i, x["kr"], x["col_cnt"], dist, long_dom, impact], 1):
        cell = ws6.cell(row=i+1, column=j, value=v); cell.font = CELL_FONT; cell.border = BORDER
        cell.alignment = Alignment(vertical="top", wrap_text=True)

# 시트 7: 행안부 도메인 매칭 참고
ws7 = wb.create_sheet("행안부도메인_매칭")
H7 = ["행안부 도메인명","그룹","분류","타입","길이","RAMP 한글명 매칭"]
W7 = [22,12,12,12,8,40]
for i, w in enumerate(W7, 1): ws7.column_dimensions[get_column_letter(i)].width = w
for i, h in enumerate(H7, 1):
    c = ws7.cell(row=1, column=i, value=h); c.fill = HDR_FILL; c.font = HDR_FONT
    c.alignment = Alignment(horizontal="center"); c.border = BORDER
ws7.freeze_panes = "A2"
r = 2
for nm, m in sorted(mois_dom.items()):
    # RAMP 컬럼 한글에 이 도메인명이 포함되는 케이스 (느슨한 매칭)
    matched = [kr for kr in grp_kr if nm in kr][:5]
    matches = " | ".join(matched) + (f" ... 외 다수" if len(matched)==5 else "")
    for j, v in enumerate([nm, m["grp"], m["clsf"], m["dtype"], m["dlen"], matches], 1):
        cell = ws7.cell(row=r, column=j, value=v); cell.font = CELL_FONT; cell.border = BORDER
        cell.alignment = Alignment(vertical="top", wrap_text=True)
    r += 1

wb.save(OUT)
print(f"\n→ {OUT}")
print(f"  시트: 표지/요약/한글명_도메인분포({len(multi_domain)})/D3_시간컬럼({len(time_cols)})/D4_순번계열({len(seq_cols)})/최장적용_미리보기({len(multi_domain)})/행안부도메인_매칭({len(mois_dom)})")
