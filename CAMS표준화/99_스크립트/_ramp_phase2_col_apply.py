# -*- coding: utf-8 -*-
"""
Phase 2-1 — RAMP 컬럼 영문명/한글 BEFORE/AFTER 정리.

매핑 사전 (정합성 기준):
  - 행안부 Y 3,277건 (단어명 → 영문약어)
  - RAMP N 562건 (DB 적재분)

변환 규칙 (각 RAMP 컬럼 토큰별):
  - 토큰의 RAMP 사전 한글이 행안부 단어에 있음 → 행안부 영문약어로 변환
  - RAMP only → RAMP 약어 그대로
  - 사용자 결정 케이스:
    * A 30 (배치율 등): 새 영문약어 적용 (DPRT→DPRTV 등)
    * B-1 4 (DB·FAX·ISBN·ISSN): 영문은 그대로, 한글만 행안부 한글로
    * B-2 4 (감독관 등): 통합 대상 단어로 분해 변경
    * B-3 3 (순번·전자·질의): 영문/한글 모두 그대로 (KEEP)

산출: RAMP_컬럼_BEFORE_AFTER_2026-05-23.xlsx
  - 표지·요약·컬럼별BEFORE_AFTER·변경단어매핑·KEEP
"""
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from pathlib import Path
from collections import defaultdict, Counter
import subprocess, csv, io, sys, re

BASE = Path(r"C:\Users\장재영\Desktop\dataQ\CAMS표준화")
RAMP_SCHEMA = BASE / "CAMS_RAMP_통합" / "ramp기관스키마정보.xlsx"
RAMP_DICT = BASE / "01_원본자료" / "RMS4-DE14-04.데이터표준화_단어사전-V2.0.xlsx"
MOIS_DICT = BASE.parent / "q-center" / "src" / "main" / "resources" / "seed" / "행안부_공통표준" / "행정안전부_공공데이터 공통표준단어.xlsx"
DEC = BASE / "04_RAMP분석_2026-05-21" / "RAMP_단어결정_2026-05-23.xlsx"
OUT = BASE / "04_RAMP분석_2026-05-21" / "RAMP_컬럼_BEFORE_AFTER_2026-05-23.xlsx"

R5 = re.compile(r"^[A-Z0-9]+$")
TOKEN_RE = re.compile(r"[A-Za-z][A-Za-z0-9]*")

# ============ 사전 로드 ============
print("=== 사전 로드 ===")

# 1) 행안부 Y: 단어명 → 영문약어
mois_nm2abrv = {}
mois_abrv2nm = {}
wb = load_workbook(MOIS_DICT, read_only=True, data_only=True)
for r in wb["Sheet"].iter_rows(min_row=2, values_only=True):
    if r and r[0] and r[1]:
        nm = str(r[0]).strip(); abrv = str(r[1]).strip().upper()
        차수 = str(r[8] or "").strip()
        구분 = str(r[9] or "").strip()
        is_dead = ("폐기" in 차수 and "폐기후제정" not in 차수) or 구분 == "폐기"
        if not is_dead:
            mois_nm2abrv[nm] = abrv
            mois_abrv2nm[abrv] = nm
wb.close()
# B-3에서 삭제된 3개 제외
for abrv in ("SEQ","ELCT","SQL"):
    if abrv in mois_abrv2nm:
        nm = mois_abrv2nm.pop(abrv)
        mois_nm2abrv.pop(nm, None)
print(f"  MOIS Y: {len(mois_nm2abrv)}")

# 2) RAMP 사전 1,573: 한글 → 영문약어 (원본)
ramp_nm2abrv = {}
ramp_abrv2nm = {}
wb = load_workbook(RAMP_DICT, read_only=True, data_only=True)
for r in wb["단어사전"].iter_rows(min_row=2, values_only=True):
    if r and r[2] and r[3]:
        nm = str(r[2]).strip(); abrv = str(r[3]).strip().upper()
        ramp_nm2abrv[nm] = abrv
        ramp_abrv2nm[abrv] = nm
wb.close()
print(f"  RAMP 사전: {len(ramp_nm2abrv)}")

# 3) 사용자 결정 (Case2_신규약어 시트)
USER_NEW_ABRV = {"디렉토리": "DIR", "분": "MINUTE"}  # 충돌 보정
wb = load_workbook(DEC, read_only=True, data_only=True)
for r in wb["Case2_신규약어"].iter_rows(min_row=2, values_only=True):
    if not r or not r[0]: continue
    nm = str(r[1] or "").strip()
    decide = str(r[12] or "").strip()
    if decide and R5.match(decide) and 3 <= len(decide) <= 6 and nm not in USER_NEW_ABRV:
        USER_NEW_ABRV[nm] = decide
wb.close()
print(f"  사용자 결정 새 약어: {len(USER_NEW_ABRV)}")

# 4) B-1·B-2·B-3 특수 케이스
B1_HAN = {"DB":"데이터베이스", "FAX":"팩스", "ISBN":"국제표준도서번호", "ISSN":"국제표준연속간행물번호"}
B2_MERGE = {"감독관":"감독", "손망실":"손실", "응모자":"응모", "폐쇄회로감시장치":"CCTV"}
B3_KEEP = {"순번","전자","질의"}  # 영문/한글 그대로

# 5) 최종 변환 사전: RAMP 토큰 → AFTER 토큰
# 우선순위: B-1·B-2·B-3 > 사용자 결정 > 행안부 흡수 > RAMP 그대로
TOKEN_CONVERT = {}  # BEFORE 영문약어 → AFTER 영문약어
for nm, abrv in ramp_nm2abrv.items():
    if nm in B3_KEEP:
        TOKEN_CONVERT[abrv] = abrv  # KEEP
    elif nm in B1_HAN:
        TOKEN_CONVERT[abrv] = abrv  # 영문 그대로
    elif nm in B2_MERGE:
        # 통합 대상 단어의 약어로
        merge_nm = B2_MERGE[nm]
        # 행안부에 통합 대상 있는지
        if merge_nm in mois_nm2abrv:
            TOKEN_CONVERT[abrv] = mois_nm2abrv[merge_nm]
        elif merge_nm in ramp_nm2abrv:
            TOKEN_CONVERT[abrv] = ramp_nm2abrv[merge_nm]
        else:
            TOKEN_CONVERT[abrv] = abrv  # 못 찾으면 그대로
    elif nm in USER_NEW_ABRV:
        TOKEN_CONVERT[abrv] = USER_NEW_ABRV[nm]
    elif nm in mois_nm2abrv:
        # Case 1 (한글 같음) → 행안부 약어
        TOKEN_CONVERT[abrv] = mois_nm2abrv[nm]
    else:
        # RAMP only / Case2 기타 → 그대로
        TOKEN_CONVERT[abrv] = abrv

# 한글 변환 사전: RAMP 토큰 → AFTER 한글
TOKEN_TO_HAN = {}
for nm, abrv in ramp_nm2abrv.items():
    if nm in B1_HAN:
        TOKEN_TO_HAN[abrv] = B1_HAN[nm]
    elif nm in B2_MERGE:
        TOKEN_TO_HAN[abrv] = B2_MERGE[nm]
    elif nm in B3_KEEP:
        TOKEN_TO_HAN[abrv] = nm
    elif nm in mois_nm2abrv:
        # 행안부 흡수: 행안부 한글 그대로 (RAMP·행안부 한글 같음 = 변경 X)
        TOKEN_TO_HAN[abrv] = nm
    else:
        TOKEN_TO_HAN[abrv] = nm

# ============ RAMP 컬럼 로드 + 변환 ============
print("\n=== RAMP 컬럼 변환 ===")
wb = load_workbook(RAMP_SCHEMA, read_only=True, data_only=True)
ws = wb["컬럼"]
cols = []
for r in ws.iter_rows(min_row=2, values_only=True):
    if r and r[0] and r[1]:
        cols.append({
            "tbl": str(r[0]).strip(),
            "col_en": str(r[1]).strip(),
            "col_kr": str(r[2] or "").strip(),
            "desc": str(r[3] or "").strip() if len(r) > 3 else "",
            "null": str(r[4] or "").strip() if len(r) > 4 else "",
            "dtype": str(r[5] or "").strip() if len(r) > 5 else "",
            "dlen": str(r[6] or "").strip() if len(r) > 6 else "",
            "pk": str(r[7] or "").strip() if len(r) > 7 else "",
        })
wb.close()
print(f"  RAMP 컬럼: {len(cols)}")

# 변환 함수
def convert_en(en):
    """영문 컬럼명 토큰 분해 후 변환"""
    out_tokens = []; changes = []
    for m in TOKEN_RE.finditer(en):
        t = m.group()
        tu = t.upper()
        # 원래 케이스 보존 — RAMP 컬럼은 소문자 위주
        is_lower = t.islower()
        if tu in TOKEN_CONVERT:
            new = TOKEN_CONVERT[tu]
            if new != tu:
                changes.append((t, new))
            out_tokens.append((m.start(), m.end(), new.lower() if is_lower else new))
        else:
            out_tokens.append((m.start(), m.end(), t))
    # 토큰 위치별로 다시 합치기
    result = en
    for s, e, new in reversed(out_tokens):
        result = result[:s] + new + result[e:]
    return result, changes

def convert_kr(kr):
    """한글 코멘트 변환 — RAMP 단어가 들어가있으면 매핑된 한글로 치환 (B-1·B-2 우선)"""
    new = kr
    changes = []
    # B-1: 영문 한글로 변경
    for ramp_h, mois_h in B1_HAN.items():
        if ramp_h in new:
            new = new.replace(ramp_h, mois_h)
            changes.append((ramp_h, mois_h))
    # B-2: 통합 대상으로 변경
    for ramp_h, merge_h in B2_MERGE.items():
        if ramp_h in new:
            new = new.replace(ramp_h, merge_h)
            changes.append((ramp_h, merge_h))
    return new, changes

# 변환 적용
changed = []
keep = []
for c in cols:
    new_en, en_chg = convert_en(c["col_en"])
    new_kr, kr_chg = convert_kr(c["col_kr"])
    if new_en != c["col_en"] or new_kr != c["col_kr"]:
        # 변경 유형
        types = []
        if new_en != c["col_en"]: types.append("EN")
        if new_kr != c["col_kr"]: types.append("KR")
        ctype = "+".join(types)
        en_sig = "; ".join(f"{a}→{b}" for a,b in en_chg) if en_chg else ""
        kr_sig = "; ".join(f"{a}→{b}" for a,b in kr_chg) if kr_chg else ""
        changed.append({**c, "new_en": new_en, "new_kr": new_kr, "type": ctype, "en_sig": en_sig, "kr_sig": kr_sig})
    else:
        keep.append(c)

print(f"  변경: {len(changed)}  /  KEEP: {len(keep)}")

# 통계
type_cnt = Counter(c["type"] for c in changed)
table_cnt = Counter(c["tbl"] for c in changed)
en_word_cnt = Counter()
for c in changed:
    if c["en_sig"]:
        for sig in c["en_sig"].split("; "):
            en_word_cnt[sig] += 1
kr_word_cnt = Counter()
for c in changed:
    if c["kr_sig"]:
        for sig in c["kr_sig"].split("; "):
            kr_word_cnt[sig] += 1

print(f"\n  변경 유형: {dict(type_cnt)}")
print(f"  영향 테이블: {len(table_cnt)}")
print(f"  영문 변경 단어 종수: {len(en_word_cnt)}")
print(f"  한글 변경 단어 종수: {len(kr_word_cnt)}")

# ============ xlsx 생성 ============
print("\n=== xlsx 생성 ===")
wb = Workbook()
THIN = Side(border_style="thin", color="808080")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
HDR_FILL = PatternFill("solid", fgColor="1F4E78")
HDR_FONT = Font(name="맑은 고딕", size=10, color="FFFFFF", bold=True)
CELL_FONT = Font(name="맑은 고딕", size=10)
SECTION_FILL = PatternFill("solid", fgColor="D9E1F2")
EN_FILL = PatternFill("solid", fgColor="FFE699")
KR_FILL = PatternFill("solid", fgColor="C6E0B4")
BOTH_FILL = PatternFill("solid", fgColor="F8CBAD")

# 시트 1: 표지
ws = wb.active; ws.title = "표지"
ws.column_dimensions["A"].width = 22; ws.column_dimensions["B"].width = 100
ws.merge_cells("A1:B1")
t = ws.cell(row=1, column=1, value="RAMP 컬럼 영문명/한글 BEFORE/AFTER (Phase 2-1)")
t.font = Font(name="맑은 고딕", size=18, bold=True, color="1F4E78")
t.alignment = Alignment(horizontal="center", vertical="center")
ws.row_dimensions[1].height = 36

def cr(r, label, value):
    ws.cell(row=r, column=1, value=label).fill = SECTION_FILL
    ws.cell(row=r, column=1).font = Font(name="맑은 고딕", size=10, bold=True)
    ws.cell(row=r, column=1).border = BORDER
    c = ws.cell(row=r, column=2, value=value); c.font = Font(name="맑은 고딕", size=11)
    c.border = BORDER; c.alignment = Alignment(vertical="center", wrap_text=True)

cr(3, "작성일", "2026-05-23")
cr(4, "작성자", "장재영")
cr(5, "대상", f"RAMP 스키마 컬럼 {len(cols)}건")
cr(6, "정책", "Phase 1 단어 결정 반영 — 행안부 우선 + RAMP N 562건 흡수")
cr(7, "변환 규칙", "행안부 Y와 한글 일치 → 행안부 약어. RAMP only → 그대로. 사용자 결정 적용 (A·B-1·B-2·B-3·디렉토리·분)")
cr(8, "Cubrid → Oracle 재생성", "ALTER 없음. BEFORE/AFTER 추적만")
ws.row_dimensions[9].height = 8
cr(10, "변경 컬럼", f"{len(changed):,}건 / 전체 {len(cols):,}건")
cr(11, "KEEP", f"{len(keep):,}건 (행안부 흡수·RAMP only·B-3)")
cr(12, "변경 유형 — EN만", f"{type_cnt.get('EN',0):,}건 (영문약어만)")
cr(13, "변경 유형 — KR만", f"{type_cnt.get('KR',0):,}건 (한글만 — B-1·B-2)")
cr(14, "변경 유형 — EN+KR", f"{type_cnt.get('EN+KR',0):,}건 (양쪽)")
cr(15, "영향 테이블", f"{len(table_cnt):,}개")

# 시트 2: 요약
ws2 = wb.create_sheet("요약")
ws2.column_dimensions["A"].width = 26; ws2.column_dimensions["B"].width = 12
ws2.column_dimensions["C"].width = 60
t = ws2.cell(row=1, column=1, value="Phase 2-1 컬럼 변경 통계")
t.font = Font(name="맑은 고딕", size=14, bold=True, color="1F4E78")
ws2.merge_cells("A1:C1")

r = 3
for c, h in enumerate(["항목","건수","비고"], 1):
    cell = ws2.cell(row=r, column=c, value=h); cell.fill = HDR_FILL; cell.font = HDR_FONT
    cell.alignment = Alignment(horizontal="center"); cell.border = BORDER
r += 1
rows = [
    ("전체 컬럼", len(cols), ""),
    ("변경", len(changed), ""),
    ("├ EN만 변경", type_cnt.get("EN",0), "영문약어 변경 (Case1 흡수 + A 새약어)"),
    ("├ KR만 변경", type_cnt.get("KR",0), "B-1 (영문은 그대로, 한글만)"),
    ("└ EN+KR 양쪽", type_cnt.get("EN+KR",0), "B-2 (단어 통합) + EN 영향 시 KR 동반"),
    ("KEEP (변경 없음)", len(keep), "행안부 흡수 + RAMP only 깨끗 + B-3"),
    ("영향 테이블", len(table_cnt), ""),
    ("영문 변경 단어 종수", len(en_word_cnt), ""),
    ("한글 변경 단어 종수", len(kr_word_cnt), ""),
]
for label, n, note in rows:
    ws2.cell(row=r, column=1, value=label).border = BORDER
    ws2.cell(row=r, column=2, value=n).border = BORDER
    ws2.cell(row=r, column=2).alignment = Alignment(horizontal="right")
    ws2.cell(row=r, column=3, value=note).border = BORDER
    r += 1

# 시트 3: 컬럼별 BEFORE/AFTER
ws3 = wb.create_sheet("컬럼BEFORE_AFTER")
H = ["No","테이블","BEFORE 영문","AFTER 영문","BEFORE 한글","AFTER 한글","변경유형","영문 변경 단어","한글 변경 단어","타입","길이","NULL","PK"]
W = [5,24,28,28,28,28,10,28,28,10,8,8,5]
for i, w in enumerate(W, 1):
    ws3.column_dimensions[get_column_letter(i)].width = w
for i, h in enumerate(H, 1):
    c = ws3.cell(row=1, column=i, value=h); c.fill = HDR_FILL; c.font = HDR_FONT
    c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True); c.border = BORDER
ws3.row_dimensions[1].height = 30
ws3.freeze_panes = "C2"

for i, c in enumerate(changed, 1):
    row = i + 1
    fill = {"EN": EN_FILL, "KR": KR_FILL, "EN+KR": BOTH_FILL}.get(c["type"], None)
    values = [i, c["tbl"], c["col_en"], c["new_en"], c["col_kr"], c["new_kr"],
              c["type"], c["en_sig"], c["kr_sig"], c["dtype"], c["dlen"], c["null"], c["pk"]]
    for j, v in enumerate(values, 1):
        cell = ws3.cell(row=row, column=j, value=v); cell.font = CELL_FONT; cell.border = BORDER
        cell.alignment = Alignment(vertical="top", wrap_text=True)
    if fill:
        ws3.cell(row=row, column=7).fill = fill

# 시트 4: 변경 단어 매핑 (영문)
ws4 = wb.create_sheet("변경단어매핑_영문")
H4 = ["No","변환 (BEFORE → AFTER)","영향 컬럼수","사유"]
W4 = [5,32,12,40]
for i, w in enumerate(W4, 1): ws4.column_dimensions[get_column_letter(i)].width = w
for i, h in enumerate(H4, 1):
    c = ws4.cell(row=1, column=i, value=h); c.fill = HDR_FILL; c.font = HDR_FONT
    c.alignment = Alignment(horizontal="center"); c.border = BORDER
ws4.freeze_panes = "A2"
for i, (sig, n) in enumerate(en_word_cnt.most_common(), 1):
    # 사유 추정 (행안부 흡수 / 새약어 / B-2 통합)
    before_after = sig.split("→") if "→" in sig else [sig, ""]
    bef = before_after[0]; aft = before_after[1] if len(before_after)>1 else ""
    bef_nm = ramp_abrv2nm.get(bef.upper(), "?")
    sub = ""
    if bef_nm in B2_MERGE: sub = f"B-2 통합 ({bef_nm}→{B2_MERGE[bef_nm]})"
    elif bef_nm in USER_NEW_ABRV: sub = f"A 새 약어 (사용자 결정)"
    elif bef_nm in mois_nm2abrv: sub = "Case1 — 행안부 약어 흡수"
    else: sub = ""
    for j, v in enumerate([i, sig, n, sub], 1):
        cell = ws4.cell(row=i+1, column=j, value=v); cell.font = CELL_FONT; cell.border = BORDER
        cell.alignment = Alignment(vertical="top", wrap_text=True)

# 시트 5: 변경 단어 매핑 (한글)
ws5 = wb.create_sheet("변경단어매핑_한글")
H5 = ["No","변환 (BEFORE → AFTER)","영향 컬럼수","사유"]
for i, w in enumerate(W4, 1): ws5.column_dimensions[get_column_letter(i)].width = w
for i, h in enumerate(H5, 1):
    c = ws5.cell(row=1, column=i, value=h); c.fill = HDR_FILL; c.font = HDR_FONT
    c.alignment = Alignment(horizontal="center"); c.border = BORDER
ws5.freeze_panes = "A2"
for i, (sig, n) in enumerate(kr_word_cnt.most_common(), 1):
    before_after = sig.split("→") if "→" in sig else [sig, ""]
    bef = before_after[0]
    sub = ""
    if bef in B1_HAN: sub = "B-1 — 행안부 한글 흡수"
    elif bef in B2_MERGE: sub = "B-2 — 다른 단어로 통합"
    for j, v in enumerate([i, sig, n, sub], 1):
        cell = ws5.cell(row=i+1, column=j, value=v); cell.font = CELL_FONT; cell.border = BORDER
        cell.alignment = Alignment(vertical="top", wrap_text=True)

# 시트 6: KEEP 참고 (테이블별 카운트만)
ws6 = wb.create_sheet("KEEP_테이블별")
H6 = ["No","테이블","KEEP 컬럼수"]
for i, w in enumerate([5,32,12], 1): ws6.column_dimensions[get_column_letter(i)].width = w
for i, h in enumerate(H6, 1):
    c = ws6.cell(row=1, column=i, value=h); c.fill = HDR_FILL; c.font = HDR_FONT
    c.alignment = Alignment(horizontal="center"); c.border = BORDER
ws6.freeze_panes = "A2"
keep_tbl = Counter(c["tbl"] for c in keep)
for i, (tbl, n) in enumerate(keep_tbl.most_common(), 1):
    for j, v in enumerate([i, tbl, n], 1):
        cell = ws6.cell(row=i+1, column=j, value=v); cell.font = CELL_FONT; cell.border = BORDER

wb.save(OUT)
print(f"\n→ {OUT}")
print(f"  시트: 표지 / 요약 / 컬럼BEFORE_AFTER ({len(changed)}) / 변경단어매핑_영문 ({len(en_word_cnt)}) / 변경단어매핑_한글 ({len(kr_word_cnt)}) / KEEP_테이블별 ({len(keep_tbl)})")
