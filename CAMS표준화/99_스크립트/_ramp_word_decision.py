# -*- coding: utf-8 -*-
"""
RAMP 단어 정렬 — 신규 약어 부여 작업표 (Phase 1).

정책 (사용자 확정):
  - 의미 분석 X. 영문약어 충돌만 처리.
  - 한글 같음 + 약어 같음 (완전일치): 행안부 그대로 → 추가 등록 0
  - 한글 같음 + 약어 다름 (Case 1): RAMP 컬럼 약어 → 행안부 약어로 변경
  - 한글 다름 + 약어 같음 (Case 2): RAMP 단어를 새 영문약어 부여 후 N 등록
  - 한글 다름 + 약어 다름 (RAMP only): 그대로 N 등록. 약어 충돌 시만 새 약어.

산출: RAMP_단어결정_2026-05-23.xlsx
  시트:
    1. 표지
    2. 요약
    3. Case2_신규약어 (77)   ← 사용자 결정 ①
    4. RAMPonly_등록 (495)    ← 사용자 결정 ② (충돌분만)
    5. Case1_약어변경 (204)   ← 자동 적용, 참고용
    6. 완전일치 (797)         ← 참고용
"""
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from pathlib import Path
from collections import defaultdict, Counter
import re

BASE = Path(r"C:\Users\장재영\Desktop\dataQ\CAMS표준화")
RAMP_DICT = BASE / "01_원본자료" / "RMS4-DE14-04.데이터표준화_단어사전-V2.0.xlsx"
RAMP_SCHEMA = BASE / "CAMS_RAMP_통합" / "ramp기관스키마정보.xlsx"
MOIS_DICT = BASE.parent / "q-center" / "src" / "main" / "resources" / "seed" / "행안부_공통표준" / "행정안전부_공공데이터 공통표준단어.xlsx"
E_TSV = BASE / "04_RAMP분석_2026-05-21" / "_분석산출물_tsv" / "_E_ramp_vs_mois.tsv"
OUT = BASE / "04_RAMP분석_2026-05-21" / "RAMP_단어결정_2026-05-23.xlsx"

# ============ 데이터 수집 ============
print("=== 데이터 수집 ===")

# MOIS 사전 (영문약어 set + 메타)
mois_by_abrv = {}
mois_by_nm = {}
wb = load_workbook(MOIS_DICT, read_only=True, data_only=True)
ws = wb["Sheet"]
for r in ws.iter_rows(min_row=2, values_only=True):
    if r and r[0] and r[1]:
        nm = str(r[0]).strip(); abrv = str(r[1]).strip()
        meta = {
            "nm": nm, "abrv": abrv,
            "eng": str(r[2] or "").strip(),
            "desc": str(r[3] or "").strip(),
            "frmt": str(r[4] or "").strip(),
        }
        mois_by_abrv[abrv] = meta
        mois_by_nm[nm] = meta
wb.close()
mois_abrv_set = set(mois_by_abrv.keys())
mois_nm_set = set(mois_by_nm.keys())
print(f"  MOIS 단어: {len(mois_by_nm)} (약어 unique {len(mois_abrv_set)})")

# RAMP 사전
ramp_words = []  # {nm, abrv, eng, desc, frmt, dmn}
wb = load_workbook(RAMP_DICT, read_only=True, data_only=True)
ws = wb["단어사전"]
for r in ws.iter_rows(min_row=2, values_only=True):
    if r and r[2] and r[3]:
        ramp_words.append({
            "nm": str(r[2]).strip(),
            "abrv": str(r[3]).strip(),
            "eng": str(r[4] or "").strip().replace("_x000D_",""),
            "desc": str(r[5] or "").strip().replace("_x000D_",""),
            "frmt": str(r[6] or "").strip(),
            "dmn": str(r[7] or "").strip(),
        })
wb.close()
print(f"  RAMP 사전: {len(ramp_words)}")

# 4분류
완전일치 = []  # 한글=, 약어=
case1 = []     # 한글=, 약어≠
case2 = []     # 한글≠, 약어=
ramp_only = [] # 한글≠, 약어≠
for w in ramp_words:
    nm, abrv = w["nm"], w["abrv"]
    if nm in mois_by_nm and mois_by_nm[nm]["abrv"] == abrv:
        완전일치.append(w)
    elif nm in mois_by_nm and mois_by_nm[nm]["abrv"] != abrv:
        case1.append(w)
    elif nm not in mois_by_nm and abrv in mois_by_abrv:
        case2.append(w)
    else:
        ramp_only.append(w)
print(f"  완전일치: {len(완전일치)}  Case1: {len(case1)}  Case2: {len(case2)}  RAMPonly: {len(ramp_only)}")

# RAMP 스키마 — 토큰 → 컬럼 사용
TOKEN_RE = re.compile(r"[A-Z][A-Z0-9]*")
token_cols = defaultdict(list)
wb = load_workbook(RAMP_SCHEMA, read_only=True, data_only=True)
ws = wb["컬럼"]
for r in ws.iter_rows(min_row=2, values_only=True):
    if r and r[1]:
        col = str(r[1]).strip()
        for t in TOKEN_RE.findall(col.upper()):
            token_cols[t].append({
                "tbl": str(r[0] or ""), "col": col,
                "kr": str(r[2] or ""),
                "type": str(r[5] or ""), "len": str(r[6] or ""),
            })
wb.close()

# ============ 신규 영문약어 후보 생성 ============
# 점유된 약어 = MOIS + RAMP 기존 약어
RAMP_abrv_set = set(w["abrv"] for w in ramp_words)
TAKEN = mois_abrv_set | RAMP_abrv_set

def suggest_abrv(eng_name, base_abrv):
    """영문명 기반 새 약어 후보 — 자음골격 3~6자, 중복 회피"""
    if not eng_name:
        eng_name = base_abrv
    s = re.sub(r"[^A-Za-z]", "", eng_name).upper()
    if not s: s = base_abrv
    # 후보 패턴들
    cands = []
    # 1) 첫글자 + 자음만 (5자)
    out = [s[0]]
    for c in s[1:]:
        if c not in "AEIOU": out.append(c)
        if len(out) >= 5: break
    cands.append("".join(out))
    # 2) 4자 자음
    out = [s[0]]
    for c in s[1:]:
        if c not in "AEIOU": out.append(c)
        if len(out) >= 4: break
    cands.append("".join(out))
    # 3) 6자 자음
    out = [s[0]]
    for c in s[1:]:
        if c not in "AEIOU": out.append(c)
        if len(out) >= 6: break
    cands.append("".join(out))
    # 4) base + V (variant 표시)
    cands.append(base_abrv + "V")
    # 5) 6자 단순 prefix
    cands.append(s[:6])
    cands.append(s[:5])
    cands.append(s[:4])
    # 중복 없는 첫 후보 반환
    used = set(TAKEN)
    for c in cands:
        if 3 <= len(c) <= 6 and c not in used and c.isalnum():
            return c
    # 끝까지 못 찾으면 base + 숫자
    for i in range(2, 10):
        c = base_abrv + str(i)
        if c not in used and len(c) <= 6:
            return c
    return base_abrv + "X"

# ============ xlsx 생성 ============
print("\n=== xlsx 생성 ===")
wb = Workbook()

THIN = Side(border_style="thin", color="808080")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
HDR_FILL = PatternFill("solid", fgColor="1F4E78")
HDR_FONT = Font(name="맑은 고딕", size=10, color="FFFFFF", bold=True)
CELL_FONT = Font(name="맑은 고딕", size=10)
DECISION_FILL = PatternFill("solid", fgColor="FFF2CC")
SECTION_FILL = PatternFill("solid", fgColor="D9E1F2")
CONFLICT_FILL = PatternFill("solid", fgColor="F8CBAD")

# ===== 시트 1: 표지 =====
ws = wb.active
ws.title = "표지"
ws.column_dimensions["A"].width = 22
ws.column_dimensions["B"].width = 90

ws.merge_cells("A1:B1")
t = ws.cell(row=1, column=1, value="RAMP 단어 정렬 — 신규 영문약어 부여 작업표")
t.font = Font(name="맑은 고딕", size=18, bold=True, color="1F4E78")
t.alignment = Alignment(horizontal="center", vertical="center")
ws.row_dimensions[1].height = 36

def cr(r, label, value):
    ws.cell(row=r, column=1, value=label).fill = SECTION_FILL
    ws.cell(row=r, column=1).font = Font(name="맑은 고딕", size=10, bold=True)
    ws.cell(row=r, column=1).border = BORDER
    c = ws.cell(row=r, column=2, value=value)
    c.font = Font(name="맑은 고딕", size=11)
    c.border = BORDER
    c.alignment = Alignment(vertical="center", wrap_text=True)

cr(3,  "작성일",  "2026-05-23")
cr(4,  "작성자",  "장재영")
cr(5,  "대상",    f"RAMP 단어 {len(ramp_words)}건 (행안부 {len(mois_by_nm)}건과 정렬)")
cr(6,  "정책",    "행안부 우선. 영문약어 충돌 시만 RAMP에 새 약어 부여. 의미 분석 안 함")
cr(7,  "적재 DB", "RAMP 신규 단어 = quality.tb_word, comm_stnd_yn='N' (기관표준)")
cr(8,  "신규 약어 규칙", "R5: 대문자+숫자, 3~6자, MOIS·RAMP 기존 약어와 중복 금지")
ws.row_dimensions[9].height = 8
cr(10, "분류 ①", f"완전일치 {len(완전일치)}건 — 행안부 그대로 사용. 추가 등록 0")
cr(11, "분류 ②", f"Case1 {len(case1)}건 — RAMP 컬럼 약어 → 행안부 약어로 변경 (BEFORE/AFTER). 자동 적용")
cr(12, "분류 ③", f"Case2 {len(case2)}건 — RAMP 단어를 새 영문약어 부여 후 N 등록 ★사용자 결정 ①")
cr(13, "분류 ④", f"RAMPonly {len(ramp_only)}건 — 그대로 N 등록. 약어 충돌 시 새 약어 ★사용자 결정 ②")
ws.row_dimensions[14].height = 8
cr(15, "사용자 결정 ①", f"Case2 시트 — 77건 신규 영문약어 확정 (자동 후보 수용 또는 수정)")
cr(16, "사용자 결정 ②", f"RAMPonly 시트 — 약어 충돌분만 새 영문약어 확정")

# ===== 시트 2: 요약 =====
ws2 = wb.create_sheet("요약")
ws2.column_dimensions["A"].width = 22
ws2.column_dimensions["B"].width = 12
ws2.column_dimensions["C"].width = 18
ws2.column_dimensions["D"].width = 55

t = ws2.cell(row=1, column=1, value="RAMP 단어 4분류 통계")
t.font = Font(name="맑은 고딕", size=14, bold=True, color="1F4E78")
ws2.merge_cells("A1:D1")

r = 3
for col, h in enumerate(["분류","건수","사용자 결정","처리"], 1):
    c = ws2.cell(row=r, column=col, value=h); c.fill = HDR_FILL; c.font = HDR_FONT
    c.alignment = Alignment(horizontal="center"); c.border = BORDER

# RAMPonly 약어 충돌 건수
ramponly_conflict = [w for w in ramp_only if w["abrv"] in mois_abrv_set]
print(f"  RAMPonly 중 약어 충돌: {len(ramponly_conflict)}")

stats = [
    ("완전일치 (한글=·약어=)", len(완전일치), "—", "행안부 그대로. 추가 등록 0"),
    ("Case1 (한글=·약어≠)", len(case1), "—", "RAMP 컬럼 약어 → 행안부 약어 (자동)"),
    ("Case2 (한글≠·약어=)", len(case2), f"{len(case2)}건 모두", "★ 새 영문약어 확정 후 N 등록"),
    ("RAMPonly (한글≠·약어≠)", len(ramp_only), f"{len(ramponly_conflict)}건 (충돌분만)", f"충돌 {len(ramponly_conflict)}건 새 약어 / 나머지 {len(ramp_only)-len(ramponly_conflict)}건 그대로 N 등록"),
    ("─ 합계 ─", len(ramp_words), f"{len(case2)+len(ramponly_conflict)}건", ""),
]
r += 1
for label, n, decide, proc in stats:
    ws2.cell(row=r, column=1, value=label).border = BORDER
    ws2.cell(row=r, column=2, value=n).border = BORDER
    ws2.cell(row=r, column=2).alignment = Alignment(horizontal="right")
    ws2.cell(row=r, column=3, value=decide).border = BORDER
    ws2.cell(row=r, column=4, value=proc).border = BORDER
    r += 1

# ===== 시트 3: Case2 신규약어 (77) =====
ws3 = wb.create_sheet("Case2_신규약어")
HEADERS3 = ["No","RAMP 한글","RAMP 기존약어\n(행안부 충돌)","RAMP 영문명","RAMP 설명",
            "행안부 동일약어\n의 한글","행안부 영문명","행안부 설명",
            "영향 컬럼수","영향 테이블수","사용 컬럼 샘플",
            "신규 약어 후보\n(자동 생성)","확정 영문약어","비고"]
WIDTHS3 = [5,16,14,22,46,16,22,46,10,10,38,14,14,28]
for i, w in enumerate(WIDTHS3, 1):
    ws3.column_dimensions[get_column_letter(i)].width = w
for i, h in enumerate(HEADERS3, 1):
    c = ws3.cell(row=1, column=i, value=h)
    c.fill = HDR_FILL; c.font = HDR_FONT
    c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    c.border = BORDER
ws3.row_dimensions[1].height = 40
ws3.freeze_panes = "C2"

for i, w in enumerate(case2, 1):
    row = i + 1
    mois = mois_by_abrv[w["abrv"]]
    cols = token_cols.get(w["abrv"], [])
    n_col = len(cols)
    n_tbl = len(set(c["tbl"] for c in cols))
    sample = "\n".join(f"- {c['tbl']}.{c['col']} ({c['kr']})" for c in cols[:5])
    if len(cols) > 5: sample += f"\n... 외 {len(cols)-5}개"
    suggest = suggest_abrv(w["eng"], w["abrv"])
    values = [
        i, w["nm"], w["abrv"], w["eng"], w["desc"],
        mois["nm"], mois["eng"], mois["desc"],
        n_col, n_tbl, sample,
        suggest, "", "",
    ]
    for j, v in enumerate(values, 1):
        c = ws3.cell(row=row, column=j, value=v)
        c.font = CELL_FONT; c.border = BORDER
        c.alignment = Alignment(vertical="top", wrap_text=True)
    # RAMP 기존약어 = 충돌 색
    ws3.cell(row=row, column=3).fill = CONFLICT_FILL
    # 결정칸 노란
    ws3.cell(row=row, column=13).fill = DECISION_FILL
    ws3.row_dimensions[row].height = max(48, 14*(1+min(5, n_col)))

# ===== 시트 4: RAMPonly 등록 (495) =====
ws4 = wb.create_sheet("RAMPonly_등록")
HEADERS4 = ["No","RAMP 한글","RAMP 약어","행안부 약어 충돌?",
            "RAMP 영문명","RAMP 설명","RAMP 도메인분류",
            "영향 컬럼수","영향 테이블수","사용 컬럼 샘플",
            "신규 약어 후보\n(충돌시만)","확정 영문약어","비고"]
WIDTHS4 = [5,16,12,12,22,46,12,10,10,38,14,14,28]
for i, w in enumerate(WIDTHS4, 1):
    ws4.column_dimensions[get_column_letter(i)].width = w
for i, h in enumerate(HEADERS4, 1):
    c = ws4.cell(row=1, column=i, value=h)
    c.fill = HDR_FILL; c.font = HDR_FONT
    c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    c.border = BORDER
ws4.row_dimensions[1].height = 40
ws4.freeze_panes = "C2"

# 충돌분 먼저 정렬
ramp_only_sorted = sorted(ramp_only, key=lambda w: (w["abrv"] not in mois_abrv_set, w["nm"]))

for i, w in enumerate(ramp_only_sorted, 1):
    row = i + 1
    conflict = w["abrv"] in mois_abrv_set
    cols = token_cols.get(w["abrv"], [])
    n_col = len(cols)
    n_tbl = len(set(c["tbl"] for c in cols))
    sample = "\n".join(f"- {c['tbl']}.{c['col']} ({c['kr']})" for c in cols[:5])
    if len(cols) > 5: sample += f"\n... 외 {len(cols)-5}개"
    suggest = suggest_abrv(w["eng"], w["abrv"]) if conflict else ""
    default_abrv = "" if conflict else w["abrv"]  # 충돌 없으면 RAMP 약어 그대로 채워둠
    values = [
        i, w["nm"], w["abrv"], "🔴 충돌" if conflict else "✓",
        w["eng"], w["desc"], w["dmn"],
        n_col, n_tbl, sample,
        suggest, default_abrv, "",
    ]
    for j, v in enumerate(values, 1):
        c = ws4.cell(row=row, column=j, value=v)
        c.font = CELL_FONT; c.border = BORDER
        c.alignment = Alignment(vertical="top", wrap_text=True)
    if conflict:
        ws4.cell(row=row, column=4).fill = CONFLICT_FILL
        ws4.cell(row=row, column=12).fill = DECISION_FILL
    ws4.row_dimensions[row].height = max(40, 14*(1+min(5, n_col)))

# ===== 시트 5: Case1 약어변경 (참고) =====
ws5 = wb.create_sheet("Case1_약어변경")
HEADERS5 = ["No","한글","RAMP 약어","행안부 약어","행안부 영문명","행안부 설명","영향 컬럼수"]
WIDTHS5 = [5,18,12,12,22,50,10]
for i, w in enumerate(WIDTHS5, 1):
    ws5.column_dimensions[get_column_letter(i)].width = w
for i, h in enumerate(HEADERS5, 1):
    c = ws5.cell(row=1, column=i, value=h)
    c.fill = HDR_FILL; c.font = HDR_FONT
    c.alignment = Alignment(horizontal="center"); c.border = BORDER
ws5.freeze_panes = "A2"

for i, w in enumerate(case1, 1):
    row = i + 1
    mois = mois_by_nm[w["nm"]]
    cols = token_cols.get(w["abrv"], [])
    values = [i, w["nm"], w["abrv"], mois["abrv"], mois["eng"], mois["desc"], len(cols)]
    for j, v in enumerate(values, 1):
        c = ws5.cell(row=row, column=j, value=v)
        c.font = CELL_FONT; c.border = BORDER
        c.alignment = Alignment(vertical="top", wrap_text=True)

# ===== 시트 6: 완전일치 (참고) =====
ws6 = wb.create_sheet("완전일치")
HEADERS6 = ["No","한글","영문약어","RAMP 영문명","행안부 영문명","행안부 설명"]
WIDTHS6 = [5,18,12,22,22,50]
for i, w in enumerate(WIDTHS6, 1):
    ws6.column_dimensions[get_column_letter(i)].width = w
for i, h in enumerate(HEADERS6, 1):
    c = ws6.cell(row=1, column=i, value=h)
    c.fill = HDR_FILL; c.font = HDR_FONT
    c.alignment = Alignment(horizontal="center"); c.border = BORDER
ws6.freeze_panes = "A2"
for i, w in enumerate(완전일치, 1):
    row = i + 1
    mois = mois_by_nm[w["nm"]]
    values = [i, w["nm"], w["abrv"], w["eng"], mois["eng"], mois["desc"]]
    for j, v in enumerate(values, 1):
        c = ws6.cell(row=row, column=j, value=v)
        c.font = CELL_FONT; c.border = BORDER

wb.save(OUT)
print(f"\n→ {OUT}")
print(f"  시트: 표지 / 요약 / Case2_신규약어({len(case2)}) / RAMPonly_등록({len(ramp_only)}, 충돌 {len(ramponly_conflict)}) / Case1({len(case1)}) / 완전일치({len(완전일치)})")
print(f"\n=== 사용자 결정 필요 ===")
print(f"  ① Case2 시트 — {len(case2)}건 신규 영문약어 확정 (자동 후보 수용/수정)")
print(f"  ② RAMPonly 시트 — 충돌 {len(ramponly_conflict)}건 신규 영문약어 확정")
print(f"  나머지 {len(완전일치)+len(case1)+(len(ramp_only)-len(ramponly_conflict))}건 = 결정 불필요 (자동 처리)")
