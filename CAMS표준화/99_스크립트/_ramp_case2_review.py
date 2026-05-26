# -*- coding: utf-8 -*-
"""
Case 2 (영문약어동일 / 한글다름) 77건 검토 xlsx 생성 — 회의 자료 퀄리티.

시트 구성:
  1. 표지       — 작업 배경·정책·결정 절차
  2. 요약       — 77건 통계 + 위험도 분포
  3. 단어별검토 — 77행 메인 결정 시트 (PK)
  4. 사용컬럼상세 — 영향 컬럼 전수 (참고)
"""
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from pathlib import Path
from collections import defaultdict, Counter
import re

BASE = Path(r"C:\Users\장재영\Desktop\dataQ\CAMS표준화")
RAMP_DICT = BASE / "01_원본자료" / "RMS4-DE14-04.데이터표준화_단어사전-V2.0.xlsx"
RAMP_SCHEMA = BASE / "CAMS_RAMP_통합" / "ramp기관스키마정보.xlsx"
MOIS_DICT = BASE.parent / "q-center" / "src" / "main" / "resources" / "seed" / "행안부_공통표준" / "행정안전부_공공데이터 공통표준단어.xlsx"
E_TSV = BASE / "04_RAMP분석_2026-05-21" / "_분석산출물_tsv" / "_E_ramp_vs_mois.tsv"
OUT = BASE / "04_RAMP분석_2026-05-21" / "RAMP_Case2_단어검토_2026-05-23.xlsx"

# ============ 데이터 수집 ============
print("=== 데이터 수집 ===")

# 1) Case 2 목록 (77건)
case2 = []  # {abrv, ramp_nm, mois_nm}
with open(E_TSV, encoding="utf-8") as f:
    header = f.readline()
    for ln in f:
        p = ln.rstrip("\n").split("\t")
        if p[0] == "영문약어동일_단어명다름":
            case2.append({"abrv": p[1], "ramp_nm": p[2], "mois_nm": p[4]})
print(f"  Case 2: {len(case2)}건")

# 2) RAMP 사전 — 단어 메타 (한글→{eng, desc, frmt, domain})
# 시트 '단어사전': No(0) 제정차수(1) 단어명(2) 단어영문약어명(3) 단어영문명(4) 단어설명(5) 형식단어여부(6) 도메인분류명(7)
ramp_dict = {}
wb = load_workbook(RAMP_DICT, read_only=True, data_only=True)
ws = wb["단어사전"]
for r in ws.iter_rows(min_row=2, values_only=True):
    if r and r[2] and r[3]:
        nm = str(r[2]).strip()
        abrv = str(r[3]).strip()
        ramp_dict[(nm, abrv)] = {
            "eng": str(r[4] or "").strip().replace("_x000D_",""),
            "desc": str(r[5] or "").strip().replace("_x000D_",""),
            "frmt": str(r[6] or "").strip(),
            "dmn": str(r[7] or "").strip(),
        }
wb.close()
print(f"  RAMP 사전 행: {len(ramp_dict)}")

# 3) MOIS 사전 — 영문약어→{nm, eng, desc, frmt, dmn}
mois_dict = {}
wb = load_workbook(MOIS_DICT, read_only=True, data_only=True)
ws = wb["Sheet"]
for r in ws.iter_rows(min_row=2, values_only=True):
    if r and r[1]:
        mois_dict[str(r[1]).strip()] = {
            "nm": str(r[0] or "").strip(),
            "eng": str(r[2] or "").strip(),
            "desc": str(r[3] or "").strip(),
            "frmt": str(r[4] or "").strip(),
            "dmn": str(r[5] or "").strip(),
        }
wb.close()
print(f"  MOIS 사전: {len(mois_dict)}")

# 4) RAMP 스키마 — 약어 토큰 → 사용 컬럼
TOKEN_RE = re.compile(r"[A-Z][A-Z0-9]*")
def tokens(s):
    if not s: return []
    return TOKEN_RE.findall(s.upper())

token_cols = defaultdict(list)  # abrv → [(table, col, kr, type, len)]
wb = load_workbook(RAMP_SCHEMA, read_only=True, data_only=True)
ws = wb["컬럼"] if "컬럼" in wb.sheetnames else wb.worksheets[1]
print(f"  RAMP 스키마 시트: {ws.title}, rows={ws.max_row}")
# 헤더 파악
hd = [str(c.value).strip() if c.value else "" for c in next(ws.iter_rows(min_row=1, max_row=1))[0:20]]
print(f"  스키마 헤더: {hd[:8]}")
# 컬럼명 컬럼 추정
def find_sc(*names):
    for nm in names:
        for i, h in enumerate(hd):
            if nm in h: return i
    return -1
ic_tbl = find_sc("테이블"); ic_col_en = find_sc("컬럼영문","영문컬럼","컬럼명")
ic_col_kr = find_sc("컬럼한글","한글컬럼","컬럼논리","논리명")
ic_type = find_sc("데이터타입","타입"); ic_len = find_sc("길이","크기")
print(f"  스키마 인덱스: tbl={ic_tbl} col_en={ic_col_en} col_kr={ic_col_kr} type={ic_type} len={ic_len}")

for r in ws.iter_rows(min_row=2, values_only=True):
    if r and r[ic_col_en]:
        col = str(r[ic_col_en]).strip()
        for t in tokens(col):
            token_cols[t].append({
                "tbl": str(r[ic_tbl] or ""),
                "col": col,
                "kr": str(r[ic_col_kr] or "") if ic_col_kr>=0 else "",
                "type": str(r[ic_type] or "") if ic_type>=0 else "",
                "len": str(r[ic_len] or "") if ic_len>=0 else "",
            })
wb.close()
print(f"  스키마 토큰 수집: {len(token_cols)}")

# ============ 위험도 산정 ============
def is_record_domain(text):
    """기록물 도메인 핵심 키워드 포함 여부 — 의미 깨짐 위험"""
    kw = ["기록물","문서","보존","이관","인수","서고","매체","NEO","RFID","기록","아카이브"]
    return any(k in text for k in kw)

def risk_level(c2):
    """위험도 산정: 높음(흡수금지) / 중간(검토) / 낮음(흡수가능)"""
    ramp = c2["ramp_nm"]; mois = c2["mois_nm"]
    # 1) 동의어 가능성 (의미 비슷)
    syns = [("순번","순서"),("체계","스키마"),("규칙","규정"),("관계","관련")]
    if (ramp, mois) in syns or (mois, ramp) in syns:
        return ("중간", "RAMP·행안부 의미가 유사 동의어 — 흡수 검토")
    # 2) 의미 완전 다름 + RAMP가 기록물 도메인
    if is_record_domain(ramp):
        return ("높음", f"RAMP '{ramp}' 가 기록물 도메인 핵심어 — 흡수 시 의미 손실, 새 약어 부여 권장")
    # 3) 약어가 일반 영단어 (FAX, DB, SQL 등)
    if c2["abrv"] in ("FAX","DB","SQL","OTP","MMS","FTP","HTTP","ID","API","CSV","XML","JSON"):
        return ("높음", f"일반 영단어 약어 — RAMP·행안부가 다른 의미. 컨텍스트로 분기 필요")
    # 4) 의미 완전 다름
    return ("중간", "의미 다름 — 결정 필요")

def suggest_new_abrv(ramp_nm, ramp_eng):
    """RAMP 의미를 유지하는 새 약어 후보 — 영문명 기반 자음 골격"""
    if not ramp_eng: return ""
    # 영문명에서 모음 제거, 첫글자·자음 위주 4~6자
    s = re.sub(r"[^A-Za-z]", "", ramp_eng).upper()
    if len(s) <= 5: return s
    # 첫글자 + 자음
    out = [s[0]]
    for c in s[1:]:
        if c not in "AEIOU": out.append(c)
        if len(out) >= 5: break
    return "".join(out)

# ============ xlsx 생성 ============
print("\n=== xlsx 생성 ===")
wb = Workbook()

# ----- 스타일 -----
THIN = Side(border_style="thin", color="808080")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
HDR_FILL = PatternFill("solid", fgColor="1F4E78")
HDR_FONT = Font(name="맑은 고딕", size=10, color="FFFFFF", bold=True)
CELL_FONT = Font(name="맑은 고딕", size=10)
DECISION_FILL = PatternFill("solid", fgColor="FFF2CC")
RISK_HIGH = PatternFill("solid", fgColor="F8CBAD")
RISK_MID = PatternFill("solid", fgColor="FFE699")
RISK_LOW = PatternFill("solid", fgColor="C6E0B4")
SECTION_FILL = PatternFill("solid", fgColor="D9E1F2")

# ===== 시트 1: 표지 =====
ws = wb.active
ws.title = "표지"
ws.column_dimensions["A"].width = 22
ws.column_dimensions["B"].width = 90

def cover_row(ws, r, label, value, bold=False):
    ws.cell(row=r, column=1, value=label).fill = SECTION_FILL
    ws.cell(row=r, column=1).font = Font(name="맑은 고딕", size=10, bold=True)
    ws.cell(row=r, column=1).border = BORDER
    ws.cell(row=r, column=1).alignment = Alignment(vertical="center")
    c = ws.cell(row=r, column=2, value=value)
    c.font = Font(name="맑은 고딕", size=11, bold=bold)
    c.border = BORDER
    c.alignment = Alignment(vertical="center", wrap_text=True)

ws.merge_cells("A1:B1")
t = ws.cell(row=1, column=1, value="RAMP 단어 정렬 — Case 2 (영문약어 충돌) 검토표")
t.font = Font(name="맑은 고딕", size=18, bold=True, color="1F4E78")
t.alignment = Alignment(horizontal="center", vertical="center")
ws.row_dimensions[1].height = 36

cover_row(ws, 3,  "작성일",       "2026-05-23")
cover_row(ws, 4,  "작성자",       "장재영")
cover_row(ws, 5,  "대상",          f"RAMP·MOIS Case 2 단어 {len(case2)}건 (영문약어 동일 / 한글 의미 다름)")
cover_row(ws, 6,  "검토 목적",     "RAMP 사전을 행안부 표준으로 흡수할 때 의미가 손실되는 단어를 식별, 흡수/새약어부여/유지 결정")
cover_row(ws, 7,  "검토 절차",     "단어별검토 시트에서 1건씩 [결정] 칸에 (1)흡수 (2)새약어 (3)유지 중 선택 + [결정사유] 기재")

ws.row_dimensions[8].height = 8
cover_row(ws, 9,  "정책 — R1",     "행안부 표준사전 불변 — 행안부 단어·용어·도메인은 읽기 전용 (이음동의어 추가 금지)")
cover_row(ws, 10, "정책 — R2",     "행안부 표준 1순위 / 없을 때만 기관표준(comm_stnd_yn='N') 신규 등록")
cover_row(ws, 11, "정책 — R5",     "약어 = 대문자+숫자만 (R5), 3~6자, 부적절 표현 금지, 행안부 약어와 중복 금지")
cover_row(ws, 12, "정책 — DB",     "RAMP/CAMS 정렬 결과 신규 단어·용어·도메인은 모두 comm_stnd_yn='N'으로 적재")

ws.row_dimensions[13].height = 8
cover_row(ws, 14, "위험도 — 높음", "흡수 시 의미 손실 큼. 새 약어 부여 강력 권장 (예: PRSR 보존→압력, RMT 포맷→송금)")
cover_row(ws, 15, "위험도 — 중간", "동의어 가능성. 컨텍스트 검토 필요")
cover_row(ws, 16, "위험도 — 낮음", "흡수 가능성 높음")

ws.row_dimensions[17].height = 8
cover_row(ws, 18, "후속 단계",    "Phase 2 — 용어 통일/분리 결정 → Phase 3 — 용어사전 재정의 (모두 comm_stnd_yn='N')")
cover_row(ws, 19, "ALTER 적용",  "본 단계에서는 ALTER 불필요. RAMP DB(Cubrid)는 신규 Oracle로 재생성하므로 BEFORE-AFTER 추적만 충실")

# ===== 시트 2: 요약 =====
ws2 = wb.create_sheet("요약")
ws2.column_dimensions["A"].width = 22
ws2.column_dimensions["B"].width = 15
ws2.column_dimensions["C"].width = 60

t = ws2.cell(row=1, column=1, value="Case 2 통계 요약")
t.font = Font(name="맑은 고딕", size=14, bold=True, color="1F4E78")
ws2.merge_cells("A1:C1")

# 위험도 산정
risks = []
for c2 in case2:
    lvl, rsn = risk_level(c2)
    risks.append(lvl)
risk_cnt = Counter(risks)

r = 3
ws2.cell(row=r, column=1, value="항목").fill = HDR_FILL; ws2.cell(row=r,column=1).font = HDR_FONT
ws2.cell(row=r, column=2, value="건수").fill = HDR_FILL; ws2.cell(row=r,column=2).font = HDR_FONT
ws2.cell(row=r, column=3, value="비고").fill = HDR_FILL; ws2.cell(row=r,column=3).font = HDR_FONT
for c in (1,2,3): ws2.cell(row=r, column=c).alignment = Alignment(horizontal="center")
r += 1
rows_data = [
    ("Case 2 총 건수", len(case2), "영문약어 동일 / RAMP 한글 단어명 다름"),
    ("위험도 — 높음", risk_cnt.get("높음",0), "흡수 금지, 새 약어 부여"),
    ("위험도 — 중간", risk_cnt.get("중간",0), "검토 필요"),
    ("위험도 — 낮음", risk_cnt.get("낮음",0), "흡수 가능"),
    ("영향 컬럼 총수 (참고)", sum(len(token_cols.get(c2["abrv"], [])) for c2 in case2), "RAMP 스키마에서 해당 약어 사용 컬럼"),
]
for label, val, note in rows_data:
    ws2.cell(row=r, column=1, value=label).border = BORDER
    ws2.cell(row=r, column=2, value=val).border = BORDER
    ws2.cell(row=r, column=2).alignment = Alignment(horizontal="right")
    ws2.cell(row=r, column=3, value=note).border = BORDER
    r += 1

# ===== 시트 3: 단어별 검토 (메인) =====
ws3 = wb.create_sheet("단어별검토")
HEADERS = [
    "No", "위험도",
    "RAMP 영문약어", "RAMP 한글", "RAMP 영문명", "RAMP 설명",
    "MOIS 영문약어", "MOIS 한글", "MOIS 영문명", "MOIS 설명", "MOIS 도메인분류",
    "영향 컬럼수", "영향 테이블수", "사용 컬럼 샘플",
    "권고", "새약어 후보",
    "결정", "결정 사유",
]
WIDTHS = [5, 10,
          13, 16, 22, 50,
          13, 16, 22, 50, 14,
          10, 10, 36,
          22, 14,
          16, 32]
for i, w in enumerate(WIDTHS, 1):
    ws3.column_dimensions[get_column_letter(i)].width = w

# 헤더
for i, h in enumerate(HEADERS, 1):
    c = ws3.cell(row=1, column=i, value=h)
    c.fill = HDR_FILL; c.font = HDR_FONT
    c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    c.border = BORDER
ws3.row_dimensions[1].height = 34
ws3.freeze_panes = "C2"

# 행 데이터
for i, c2 in enumerate(case2, 1):
    row = i + 1
    abrv = c2["abrv"]
    ramp_meta = ramp_dict.get((c2["ramp_nm"], abrv), {})
    mois_meta = mois_dict.get(abrv, {})
    cols = token_cols.get(abrv, [])
    n_col = len(cols)
    n_tbl = len(set(c["tbl"] for c in cols))
    sample = "\n".join(f"- {c['tbl']}.{c['col']} ({c['kr']})" for c in cols[:5])
    if len(cols) > 5: sample += f"\n... 외 {len(cols)-5}개"
    lvl, rsn = risk_level(c2)
    suggest = suggest_new_abrv(c2["ramp_nm"], ramp_meta.get("eng",""))
    # 권고: 위험도에 따라
    if lvl == "높음":
        recomm = f"새 약어 권장\n({suggest} 등)"
    elif lvl == "중간":
        recomm = "흡수 vs 새약어\n검토 필요"
    else:
        recomm = "흡수 가능"
    values = [
        i, lvl,
        abrv, c2["ramp_nm"], ramp_meta.get("eng",""), ramp_meta.get("desc",""),
        abrv, c2["mois_nm"], mois_meta.get("eng",""), mois_meta.get("desc",""), mois_meta.get("frmt","")+("/"+mois_meta.get("dmn","") if mois_meta.get("dmn") else ""),
        n_col, n_tbl, sample,
        recomm, suggest,
        "", "",
    ]
    for j, v in enumerate(values, 1):
        cell = ws3.cell(row=row, column=j, value=v)
        cell.font = CELL_FONT
        cell.border = BORDER
        cell.alignment = Alignment(vertical="top", wrap_text=True)
    # 위험도 색칠
    risk_fill = {"높음": RISK_HIGH, "중간": RISK_MID, "낮음": RISK_LOW}.get(lvl, None)
    if risk_fill:
        ws3.cell(row=row, column=2).fill = risk_fill
    # 결정 칸 하이라이트 (P=17, Q=18)
    ws3.cell(row=row, column=17).fill = DECISION_FILL
    ws3.cell(row=row, column=18).fill = DECISION_FILL
    # 행 높이
    ws3.row_dimensions[row].height = max(60, 14 * (1 + min(5, n_col)))

# 결정 드롭다운 (Q열 = 17)
dv = DataValidation(type="list", formula1='"흡수,새약어,유지,분리"', allow_blank=True)
dv.add(f"Q2:Q{len(case2)+1}")
ws3.add_data_validation(dv)

# ===== 시트 4: 사용 컬럼 상세 =====
ws4 = wb.create_sheet("사용컬럼상세")
H4 = ["영문약어","RAMP 한글","위험도","테이블","컬럼영문","컬럼한글","타입","길이"]
W4 = [12, 16, 10, 24, 32, 32, 12, 8]
for i, w in enumerate(W4, 1):
    ws4.column_dimensions[get_column_letter(i)].width = w
for i, h in enumerate(H4, 1):
    c = ws4.cell(row=1, column=i, value=h)
    c.fill = HDR_FILL; c.font = HDR_FONT
    c.alignment = Alignment(horizontal="center")
    c.border = BORDER
ws4.freeze_panes = "A2"
r = 2
for c2 in case2:
    abrv = c2["abrv"]
    lvl, _ = risk_level(c2)
    cols = token_cols.get(abrv, [])
    for col in cols:
        vals = [abrv, c2["ramp_nm"], lvl, col["tbl"], col["col"], col["kr"], col["type"], col["len"]]
        for j, v in enumerate(vals, 1):
            cell = ws4.cell(row=r, column=j, value=v)
            cell.font = CELL_FONT
            cell.border = BORDER
        risk_fill = {"높음": RISK_HIGH, "중간": RISK_MID}.get(lvl, None)
        if risk_fill: ws4.cell(row=r, column=3).fill = risk_fill
        r += 1
print(f"  사용컬럼상세 행: {r-2}")

# 저장
wb.save(OUT)
print(f"\n→ {OUT}")
print(f"  시트: 표지 / 요약 / 단어별검토 ({len(case2)}행) / 사용컬럼상세 ({r-2}행)")
print(f"  위험도 분포: 높음 {risk_cnt.get('높음',0)} / 중간 {risk_cnt.get('중간',0)} / 낮음 {risk_cnt.get('낮음',0)}")
