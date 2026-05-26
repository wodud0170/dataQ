# -*- coding: utf-8 -*-
# 단어_일괄등록_CAMS.xlsx 생성 — 시스템 단어 일괄등록 템플릿(word_template.xlsx) 형식
# 13번째 컬럼(예시 코멘트)은 검토용 — 업로드는 12번째까지만 읽으므로 무시됨
# 예시 코멘트는 분해(토큰) 기준 매칭 — substring 오매칭(용기↔사용기간) 방지
import csv, re, openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

# 1) 행안부 표준단어 + 코멘트정정 + B 등록안 → 분해 사전
res = {}
with open('_w.tsv', encoding='utf-8') as f:
    for ln in f:
        p = ln.rstrip('\n').split('\t')
        if len(p) == 2 and p[0]:
            res.setdefault(p[0], p[0])
with open('코멘트정정_규칙.tsv', encoding='utf-8') as f:
    for r in csv.DictReader(f, delimiter='\t'):
        res[r['변형어']] = r['표준어(행안부)']
rows = list(csv.DictReader(open('단어_등록안.tsv', encoding='utf-8'), delimiter='\t'))
for r in rows:
    res.setdefault(r['한글'], r['한글'])
maxlen = max(len(w) for w in res)

def seg(ch):
    i = 0; toks = []
    while i < len(ch):
        hit = None
        for L in range(min(maxlen, len(ch) - i), 0, -1):
            if ch[i:i+L] in res:
                hit = ch[i:i+L]; break
        if hit:
            toks.append(res[hit]); i += len(hit)
        else:
            i += 1
    return toks

# 2) CAMS 코멘트를 분해 → 단어별 등장 코멘트 인덱스
src = openpyxl.load_workbook('CAMS_SCHEMA_원본.xlsx', read_only=True, data_only=True)
idx = {}
for row in list(src['컬럼정의'].iter_rows(values_only=True))[1:]:
    c = (row[3] or '').strip()
    if not c:
        continue
    toks = set()
    for chunk in re.split(r'[\s()\[\]{}/,~\-_·:;.]+', c):
        if chunk:
            toks.update(seg(chunk))
    for t in toks:
        idx.setdefault(t, [])
        if c not in idx[t]:
            idx[t].append(c)

def samples(word, limit=4):
    return ' / '.join(idx.get(word, [])[:limit])

# 3) xlsx 생성
wb = Workbook(); ws = wb.active; ws.title = '단어'
hdr = ['No', '제정차수', '단어명', '단어영문약어명', '단어영문명', '단어설명', '형식단어여부',
       '도메인분류명', '이음동의어목록', '금칙어목록', '요청시스템', '표준여부', '예시 코멘트(검토용)']
ws.append(hdr)
for i, r in enumerate(rows, 1):
    ws.append([i, '', r['한글'], r['영문약어'], r['영문명'], r['비고'], 'N',
               '', '', '', '', 'N', samples(r['한글'])])
fill = PatternFill('solid', fgColor='305496')
for c in ws[1]:
    c.font = Font(bold=True, color='FFFFFF'); c.fill = fill
    c.alignment = Alignment(horizontal='center')
for i, w in enumerate([6, 9, 16, 18, 22, 38, 11, 14, 16, 12, 12, 9, 70], 1):
    ws.column_dimensions[chr(64 + i)].width = w
ws.freeze_panes = 'A2'
wb.save('단어_일괄등록_CAMS.xlsx')
print('단어_일괄등록_CAMS.xlsx 생성 — B %d종 (예시 코멘트 = 분해 토큰 기준)' % len(rows))
