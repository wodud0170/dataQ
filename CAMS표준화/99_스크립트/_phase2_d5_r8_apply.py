# -*- coding: utf-8 -*-
"""
Phase 2-2 D5 — R8 미종결 + D1 처리불가 통합 처리.
정책:
  - 데이터타입 기반 형식단어 자동 부여 (STRING→VARCHAR계열, NUMERIC→NUMBER계열)
  - 새 한글 = 원본 + 형식단어
  - 도메인은 기관표준으로 신규 등록 (D6)
  - 컬럼이 여러 도메인이면 도메인별로 다른 용어 분리
"""
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from pathlib import Path
from collections import defaultdict, Counter
import subprocess, csv, io, sys, os, base64

BASE = Path(r"C:\Users\장재영\Desktop\dataQ\CAMS표준화")
RAMP_SCHEMA = BASE / "CAMS_RAMP_통합" / "ramp기관스키마정보.xlsx"
OUT_XLSX = BASE / "04_RAMP분석_2026-05-21" / "00_핵심산출" / "Phase2_D5_R8형식단어_2026-05-23.xlsx"
OUT_SQL = BASE / "04_RAMP분석_2026-05-21" / "99_실행SQL_기록" / "Phase2_D5_domain_insert.sql"

# ============ 1. 행안부 형식단어 + 도메인분류 ============
print("=== 1. 행안부 형식단어 카탈로그 ===")
r = subprocess.run(["docker","exec","-i","dataq-db","psql","-U","admin","-d","postgres",
    "-c", """COPY (SELECT word_nm, coalesce(domain_clsf_nm,'') FROM quality.tb_word
             WHERE word_clsf_yn='Y') TO STDOUT WITH (FORMAT csv, FORCE_QUOTE *)"""],
    capture_output=True, encoding="utf-8")
format_words = {}  # 형식단어명 → 도메인분류명
for row in csv.reader(io.StringIO(r.stdout)):
    if row and row[0]:
        format_words[row[0]] = row[1]
print(f"  행안부 형식단어: {len(format_words)}")

# ============ 2. 자동 형식단어 부여 룰 ============
def auto_format_word(dtype, dlen):
    """데이터타입+길이 기반 형식단어 추천"""
    dt = dtype.upper()
    L = int(dlen) if dlen and str(dlen).isdigit() else 0
    if dt in ("VARCHAR","VARCHAR2","STRING"):
        if L >= 1000: return "내용"
        if L >= 100:  return "내용"
        if L >= 11:   return "명"
        if L >= 4:    return "명"
        if L <= 3:    return "구분"
        return "명"
    if dt == "CHAR":
        if L == 1:    return "여부"
        if L <= 4:    return "코드"
        return "구분"
    if dt in ("NUMERIC","NUMBER","INTEGER","DECIMAL"):
        return "값"  # NUMERIC 전부 값 (사용자 결정)
    if dt == "DATE":      return "일자"
    if dt == "DATETIME":  return "일시"
    if dt == "TIMESTAMP": return "일시"
    if dt == "CLOB":      return "내용"
    return "값"

def map_to_oracle(dtype, dlen):
    """Oracle 매핑: VARCHAR > 4000 → CLOB"""
    dt = dtype.upper()
    L = int(dlen) if dlen and str(dlen).isdigit() else 0
    if dt in ("VARCHAR","VARCHAR2","STRING") and L > 4000:
        return ("CLOB", 0, f"기존 STRING({L}) — 데이터 확인 후 VARCHAR2(4000) 다운가능")
    return (dt if dt != "STRING" else "VARCHAR", L, "")

# ============ 3. RAMP 컬럼 분석 — R8 미종결 단어 추출 ============
print("\n=== 3. RAMP R8 미종결 추출 ===")
wb = load_workbook(RAMP_SCHEMA, read_only=True, data_only=True)
ws = wb["컬럼"]
cols = []
for r in ws.iter_rows(min_row=2, values_only=True):
    if r and r[0] and r[1]:
        kr = str(r[2] or "").strip()
        if kr:
            cols.append({
                "tbl": str(r[0]).strip(),
                "col_en": str(r[1]).strip(),
                "col_kr": kr,
                "dtype": str(r[5] or "").strip(),
                "dlen": str(r[6] or "").strip(),
            })
wb.close()

# 시간 제외 (D3 처리됨)
TIME_SUFFIX = ("일자","일시","년월","연월","시분초","시분","시각","연도","월일","월")
def is_time(kr): return any(kr.endswith(s) for s in TIME_SUFFIX)

# R8 검사 사전 보강 — 행안부 'N' 이지만 실제 형식단어 역할 하는 단어
EXTRA_FW = {"구분","값","경로","상태","메시지","식별자","파일","순번","번호","자",
            "이름","내역","유형","범위","위치","방법","수단","목록","사항","항목",
            "기간","결과","대상","리스트","주체","주소","권한","역할","사유",
            "율","률","비율",
            # 사용자 결정 (2026-05-23): 측정 단위 명확한 명사 형식단어 인정
            "기한","빈도","주기","횟수","등급","단계"}
# 위 시간 suffix는 D3 처리
fw_check = set(format_words.keys()) | EXTRA_FW

# R8 미종결 = 한글명 끝이 형식단어가 아닌 경우
r8_cols = []
for c in cols:
    if is_time(c["col_kr"]): continue
    ends_with_fw = any(c["col_kr"].endswith(fw) for fw in fw_check)
    if not ends_with_fw:
        r8_cols.append(c)
print(f"  R8 미종결 컬럼: {len(r8_cols)}  (검사 사전 {len(fw_check)}단어)")

# 한글명 그룹화
r8_grp = defaultdict(list)
for c in r8_cols:
    r8_grp[c["col_kr"]].append(c)
print(f"  R8 미종결 단어: {len(r8_grp)}")

# ============ 4. 그룹별 1 형식단어 + 최장 도메인 (D1 정책) ============
print("\n=== 4. 형식단어 자동 부여 (그룹별 단일) ===")

# 의미 패턴 가산 (한글명 끝 단어 → 권장 형식단어)
MEANING_HINT = [
    (("비고","설명","사유","내역","상세설명","요약"), "내용"),
    (("이름","명칭",), "명"),
    (("개수","건수","행수","번수","수량","총수","량"), "수"),
    (("율","률","비율"), "율"),
    (("값",), "값"),
]
def meaning_hint(kr):
    for keys, fw in MEANING_HINT:
        if any(kr.endswith(k) for k in keys):
            return fw
    return None

def parse_int(s):
    try: return int(s)
    except: return 0

new_terms = []
for kr, lst in r8_grp.items():
    longest = max(((c["dtype"], c["dlen"]) for c in lst), key=lambda d: parse_int(d[1]))
    long_dt, long_dl = longest
    fw = meaning_hint(kr) or auto_format_word(long_dt, long_dl)
    new_kr = kr + fw
    clsf = format_words.get(fw, fw)
    # Oracle 매핑 (>4000 → CLOB)
    oracle_dt, oracle_dl, oracle_note = map_to_oracle(long_dt, long_dl)
    new_terms.append({
        "old_kr": kr,
        "new_kr": new_kr,
        "format_word": fw,
        "dtype": oracle_dt,
        "dlen": oracle_dl,
        "orig_type": long_dt,
        "orig_len": long_dl,
        "비고": oracle_note,
        "clsf": clsf,
        "cols": lst,
        "dom_dist": Counter((c["dtype"], c["dlen"]) for c in lst),
    })

print(f"  새 용어 (그룹별 단일): {len(new_terms)}")

# ============ 5. 신규 도메인 산출 ============
print("\n=== 5. 신규 도메인 ===")
# 기존 도메인 nm/clsf 셋
r = subprocess.run(["docker","exec","-i","dataq-db","psql","-U","admin","-d","postgres",
    "-c", "COPY (SELECT domain_nm, domain_clsf_nm FROM quality.tb_domain) TO STDOUT WITH (FORMAT csv, FORCE_QUOTE *)"],
    capture_output=True, encoding="utf-8")
existing_dom = set()
existing_clsf = set()
for row in csv.reader(io.StringIO(r.stdout)):
    if row: existing_dom.add(row[0]); existing_clsf.add(row[1])

# 기존 분류
r = subprocess.run(["docker","exec","-i","dataq-db","psql","-U","admin","-d","postgres",
    "-c", "COPY (SELECT domain_clsf_nm FROM quality.tb_domain_clsf) TO STDOUT WITH (FORMAT csv, FORCE_QUOTE *)"],
    capture_output=True, encoding="utf-8")
for row in csv.reader(io.StringIO(r.stdout)):
    if row: existing_clsf.add(row[0])

TYPE_ABBR = {"VARCHAR":"V","VARCHAR2":"V","STRING":"V","CHAR":"C","NUMERIC":"N","NUMBER":"N","INTEGER":"N","DECIMAL":"N","DATE":"D","DATETIME":"D","TIMESTAMP":"T","CLOB":"L","BLOB":"B"}
def gen_dom_nm(clsf, dt, dl):
    abbr = TYPE_ABBR.get(dt.upper(), "X")
    # CLOB는 길이 없음
    if dt.upper() == "CLOB":
        return f"{clsf}L"
    if dl and str(dl).isdigit() and int(dl) > 0:
        return f"{clsf}{abbr}{dl}"
    return f"{clsf}{abbr}"

# 신규 도메인 정리
to_reg_dom = {}  # 도메인명 → (분류, 타입, 길이)
to_reg_clsf = set()  # 새 분류명

for t in new_terms:
    nm = gen_dom_nm(t["clsf"], t["dtype"], t["dlen"])
    if nm in existing_dom: continue
    if nm in to_reg_dom: continue
    to_reg_dom[nm] = (t["clsf"], t["dtype"], t["dlen"])
    if t["clsf"] not in existing_clsf:
        to_reg_clsf.add(t["clsf"])

print(f"  신규 도메인: {len(to_reg_dom)}")
print(f"  신규 분류: {len(to_reg_clsf)}")

# ============ 6. SQL 생성 ============
print(f"\n=== 6. SQL 생성 ===")
sql = ["-- Phase 2-2 D5 — R8 미종결 형식단어 보충 + 기관표준 도메인 추가", "BEGIN;", ""]

if to_reg_clsf:
    sql.append(f"-- 신규 도메인분류 {len(to_reg_clsf)}건")
    for cn in sorted(to_reg_clsf):
        cid = base64.urlsafe_b64encode(os.urandom(17)).rstrip(b'=').decode()[:22]
        sql.append(f"INSERT INTO quality.tb_domain_clsf (domain_clsf_id, domain_clsf_nm, domain_grp_nm, comm_stnd_yn, cret_dt, cret_user_id) "
                   f"VALUES ('{cid}','{cn}','기타','N',to_char(now(),'YYYYMMDDHH24MISS'),'admin');")
    sql.append("")

sql.append(f"-- 신규 기관표준 도메인 {len(to_reg_dom)}건")
for nm, (clsf, dt, dl) in sorted(to_reg_dom.items()):
    did = base64.urlsafe_b64encode(os.urandom(17)).rstrip(b'=').decode()[:22]
    dl_int = int(dl) if dl and str(dl).isdigit() else 0
    desc = f"{clsf} {dt}({dl_int}) — R8 미종결 보충용"
    sql.append(
        f"INSERT INTO quality.tb_domain (domain_id, domain_nm, domain_grp_nm, domain_clsf_nm, domain_desc, data_type, data_len, data_decimal_len, stor_fmt, expr_fmt_lst, comm_stnd_yn, aprv_yn, cret_dt, cret_user_id, use_yn) "
        f"VALUES ('{did}','{nm}','기타','{clsf}','{desc.replace(chr(39),chr(39)+chr(39))}','{dt}',{dl_int},0,'',ARRAY[]::text[],'N','Y',to_char(now(),'YYYYMMDDHH24MISS'),'admin','Y');"
    )

sql += ["", "COMMIT;"]
OUT_SQL.write_text("\n".join(sql), encoding="utf-8")
print(f"  → {OUT_SQL}")

# ============ 7. xlsx ============
print(f"\n=== 7. xlsx 생성 ===")
wb = Workbook()
THIN = Side(border_style="thin", color="808080")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
HDR_FILL = PatternFill("solid", fgColor="1F4E78")
HDR_FONT = Font(name="맑은 고딕", size=10, color="FFFFFF", bold=True)
CELL_FONT = Font(name="맑은 고딕", size=10)
SECTION_FILL = PatternFill("solid", fgColor="D9E1F2")
NEW_FILL = PatternFill("solid", fgColor="FFE699")

# 표지
ws = wb.active; ws.title = "표지"
ws.column_dimensions["A"].width = 22; ws.column_dimensions["B"].width = 100
ws.merge_cells("A1:B1")
t = ws.cell(row=1, column=1, value="RAMP R8 미종결 형식단어 보충 (Phase 2-2 D5)")
t.font = Font(name="맑은 고딕", size=18, bold=True, color="1F4E78")
t.alignment = Alignment(horizontal="center", vertical="center")
ws.row_dimensions[1].height = 36

def cr(r, label, value):
    ws.cell(row=r, column=1, value=label).fill = SECTION_FILL
    ws.cell(row=r, column=1).font = Font(name="맑은 고딕", size=10, bold=True)
    ws.cell(row=r, column=1).border = BORDER
    c = ws.cell(row=r, column=2, value=value); c.font = Font(name="맑은 고딕", size=11)
    c.border = BORDER; c.alignment = Alignment(vertical="center", wrap_text=True)
cr(3, "작성일", "2026-05-23")
cr(4, "대상", f"R8 미종결 단어 {len(r8_grp)}종 (시간 제외)")
cr(5, "정책", "데이터타입 기반 형식단어 자동 부여 + 기관표준 도메인 신규 (D6)")
cr(6, "도메인별 분리", "같은 한글 단어라도 도메인 다르면 별도 용어로 분리")
cr(7, "신규 용어 수", f"{len(new_terms)}건 (도메인별 분리)")
cr(8, "신규 도메인", f"{len(to_reg_dom)}건 INSERT")
cr(9, "신규 분류", f"{len(to_reg_clsf)}건")

# 시트2: 형식단어 매핑 (새 용어) — 결정칸 추가
ws2 = wb.create_sheet("새용어_BEFORE_AFTER")
H = ["No","BEFORE 한글","자동 형식단어","AFTER 한글","최장 타입","최장 길이","신규 도메인","컬럼수","도메인 분포","사용 컬럼 샘플","결정 형식단어","결정 사유"]
W = [5,20,10,26,10,8,18,8,40,40,12,28]
for i, w in enumerate(W, 1): ws2.column_dimensions[get_column_letter(i)].width = w
for i, h in enumerate(H, 1):
    c = ws2.cell(row=1, column=i, value=h); c.fill = HDR_FILL; c.font = HDR_FONT
    c.alignment = Alignment(horizontal="center"); c.border = BORDER
ws2.freeze_panes = "C2"

DECISION_FILL = PatternFill("solid", fgColor="FFF2CC")
new_terms.sort(key=lambda x: (-len(x["cols"]), x["old_kr"]))
for i, t in enumerate(new_terms, 1):
    dom_nm = gen_dom_nm(t["clsf"], t["dtype"], t["dlen"])
    sample = "\n".join(f"- {c['tbl']}.{c['col_en']}" for c in t["cols"][:3])
    if len(t["cols"]) > 3: sample += f"\n... 외 {len(t['cols'])-3}개"
    dist = " | ".join(f"{dt}({dl})×{n}" for (dt,dl), n in t["dom_dist"].most_common())
    vals = [i, t["old_kr"], t["format_word"], t["new_kr"], t["dtype"], t["dlen"], dom_nm, len(t["cols"]), dist, sample, "", ""]
    for j, v in enumerate(vals, 1):
        c = ws2.cell(row=i+1, column=j, value=v); c.font = CELL_FONT; c.border = BORDER
        c.alignment = Alignment(vertical="top", wrap_text=True)
    ws2.cell(row=i+1, column=3).fill = NEW_FILL
    ws2.cell(row=i+1, column=4).fill = NEW_FILL
    ws2.cell(row=i+1, column=11).fill = DECISION_FILL
    ws2.cell(row=i+1, column=12).fill = DECISION_FILL

# 시트3: 신규 도메인 리스트
ws3 = wb.create_sheet("신규도메인")
H3 = ["No","도메인명","분류","타입","길이","사용 단어수"]
W3 = [5,22,12,10,8,10]
for i, w in enumerate(W3, 1): ws3.column_dimensions[get_column_letter(i)].width = w
for i, h in enumerate(H3, 1):
    c = ws3.cell(row=1, column=i, value=h); c.fill = HDR_FILL; c.font = HDR_FONT
    c.alignment = Alignment(horizontal="center"); c.border = BORDER
ws3.freeze_panes = "A2"
dom_usage = Counter()
for t in new_terms:
    nm = gen_dom_nm(t["clsf"], t["dtype"], t["dlen"])
    dom_usage[nm] += len(t["cols"])
for i, (nm, (clsf, dt, dl)) in enumerate(sorted(to_reg_dom.items()), 1):
    vals = [i, nm, clsf, dt, dl, dom_usage.get(nm, 0)]
    for j, v in enumerate(vals, 1):
        c = ws3.cell(row=i+1, column=j, value=v); c.font = CELL_FONT; c.border = BORDER

# 시트4: 형식단어별 분포
ws4 = wb.create_sheet("형식단어별_분포")
H4 = ["형식단어","적용 용어수","적용 컬럼수"]
for i, w in enumerate([12,12,12], 1): ws4.column_dimensions[get_column_letter(i)].width = w
for i, h in enumerate(H4, 1):
    c = ws4.cell(row=1, column=i, value=h); c.fill = HDR_FILL; c.font = HDR_FONT
    c.alignment = Alignment(horizontal="center"); c.border = BORDER
fw_cnt = Counter()
fw_col_cnt = Counter()
for t in new_terms:
    fw_cnt[t["format_word"]] += 1
    fw_col_cnt[t["format_word"]] += len(t["cols"])
for i, (fw, n) in enumerate(fw_cnt.most_common(), 1):
    for j, v in enumerate([fw, n, fw_col_cnt[fw]], 1):
        c = ws4.cell(row=i+1, column=j, value=v); c.font = CELL_FONT; c.border = BORDER

wb.save(OUT_XLSX)
print(f"\n→ {OUT_XLSX}")
print(f"  시트: 표지 / 새용어_BEFORE_AFTER ({len(new_terms)}) / 신규도메인 ({len(to_reg_dom)}) / 형식단어별_분포 ({len(fw_cnt)})")
print(f"\n=== 형식단어별 분포 ===")
for fw, n in fw_cnt.most_common():
    print(f"  {fw:6s}: 용어 {n:>4} / 컬럼 {fw_col_cnt[fw]:>5}")
