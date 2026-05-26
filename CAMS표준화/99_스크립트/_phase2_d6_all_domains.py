# -*- coding: utf-8 -*-
"""
Phase 2-2 D6 — RAMP 전체 컬럼의 모든 (분류, 타입, 길이) 조합 도메인 등록.
정책: 행안부 + 기존 기관표준 매칭 안 되는 조합은 무조건 기관표준(N) 신규 추가.
     > 4000 VARCHAR → CLOB (L)
"""
from openpyxl import load_workbook
from pathlib import Path
from collections import defaultdict, Counter
import subprocess, csv, io, os, base64, sys

BASE = Path(r"C:\Users\장재영\Desktop\dataQ\CAMS표준화")
RAMP_SCHEMA = BASE / "CAMS_RAMP_통합" / "ramp기관스키마정보.xlsx"
COL_BA = BASE / "04_RAMP분석_2026-05-21" / "00_핵심산출" / "RAMP_컬럼_BEFORE_AFTER_2026-05-23.xlsx"
TIME_BA = BASE / "04_RAMP분석_2026-05-21" / "00_핵심산출" / "Phase2_시간컬럼_BEFORE_AFTER_2026-05-23.xlsx"
OUT_SQL = BASE / "04_RAMP분석_2026-05-21" / "99_실행SQL_기록" / "Phase2_D6_all_domains.sql"

# ============ 1. 형식단어 + 기존 도메인 ============
print("=== 1. 형식단어 + 기존 도메인 로드 ===")
r = subprocess.run(["docker","exec","-i","dataq-db","psql","-U","admin","-d","postgres",
    "-c", """COPY (SELECT word_nm, coalesce(domain_clsf_nm,'') FROM quality.tb_word
             WHERE word_clsf_yn='Y') TO STDOUT WITH (FORMAT csv, FORCE_QUOTE *)"""],
    capture_output=True, encoding="utf-8")
fw_map = {row[0]: row[1] for row in csv.reader(io.StringIO(r.stdout)) if row}
EXTRA = {"구분":"구분","값":"값","경로":"경로","상태":"상태","메시지":"메시지","식별자":"식별자",
         "파일":"파일","순번":"순번","번호":"번호","자":"자","이름":"이름","내역":"내역","유형":"유형",
         "범위":"범위","위치":"위치","방법":"방법","수단":"수단","목록":"목록","사항":"사항","항목":"항목",
         "기간":"기간","결과":"결과","대상":"대상","주체":"주체","주소":"주소","권한":"권한",
         "역할":"역할","사유":"사유","율":"율","률":"률","ID":"ID"}
fw_map.update(EXTRA)
print(f"  형식단어 카탈로그: {len(fw_map)}")

def extract_clsf(kr):
    cand = [(fw, fw_map[fw]) for fw in fw_map if kr.endswith(fw)]
    valid = [c for c in cand if c[1]]
    if valid: return max(valid, key=lambda x: len(x[0]))[1]
    return None

# 기존 도메인
r = subprocess.run(["docker","exec","-i","dataq-db","psql","-U","admin","-d","postgres",
    "-c", """COPY (SELECT domain_nm, domain_clsf_nm, data_type, coalesce(data_len::text,'0'), comm_stnd_yn
             FROM quality.tb_domain) TO STDOUT WITH (FORMAT csv, FORCE_QUOTE *)"""],
    capture_output=True, encoding="utf-8")
exist_doms = set()
exist_by_key = {}  # (clsf, type, len) → nm
exist_nm = set()
for row in csv.reader(io.StringIO(r.stdout)):
    if row:
        clsf, dt, dl = row[1], row[2].upper(), row[3]
        L = int(dl) if dl.isdigit() else 0
        exist_by_key[(clsf, dt, L)] = row[0]
        exist_nm.add(row[0])
print(f"  기존 도메인: {len(exist_nm)}")

# 기존 분류 + 그룹
r = subprocess.run(["docker","exec","-i","dataq-db","psql","-U","admin","-d","postgres",
    "-c", "COPY (SELECT domain_clsf_nm FROM quality.tb_domain_clsf) TO STDOUT WITH (FORMAT csv, FORCE_QUOTE *)"],
    capture_output=True, encoding="utf-8")
exist_clsf = set(row[0] for row in csv.reader(io.StringIO(r.stdout)) if row)

# ============ 2. RAMP 컬럼 + AFTER 적용 ============
print("\n=== 2. RAMP 컬럼 + AFTER 한글 ===")
col_map = {}
wb = load_workbook(COL_BA, read_only=True, data_only=True)
for r in wb["컬럼BEFORE_AFTER"].iter_rows(min_row=2, values_only=True):
    if r and r[1] and r[2]:
        col_map[(str(r[1]), str(r[2]))] = str(r[5] or r[4] or "")
wb.close()
# 시간 컬럼은 D3에서 이미 도메인 매칭 됨
TIME_DOM = set()
wb = load_workbook(TIME_BA, read_only=True, data_only=True)
for r in wb["적용_BEFORE_AFTER"].iter_rows(min_row=2, values_only=True):
    if r and r[1] and r[2]:
        TIME_DOM.add((str(r[1]), str(r[2])))
wb.close()
print(f"  시간 컬럼 (D3 처리): {len(TIME_DOM)}")

# RAMP 컬럼 로드
wb = load_workbook(RAMP_SCHEMA, read_only=True, data_only=True)
all_cols = []
seen = set()
for r in wb["컬럼"].iter_rows(min_row=2, values_only=True):
    if r and r[0] and r[1]:
        key = (str(r[0]).strip(), str(r[1]).strip())
        if key in seen: continue
        seen.add(key)
        kr = col_map.get(key, str(r[2] or "").strip())  # AFTER 한글
        all_cols.append({
            "tbl": key[0], "col_en": key[1], "col_kr": kr,
            "dtype": str(r[5] or "").strip().upper() if len(r) > 5 else "",
            "dlen": str(r[6] or "").strip() if len(r) > 6 else "",
        })
wb.close()
print(f"  RAMP 컬럼 (dedupe): {len(all_cols)}")

# ============ 3. (분류, 타입, 길이) unique 조합 추출 ============
print("\n=== 3. (분류, 타입, 길이) 조합 추출 ===")
def parse_int(s):
    try: return int(s)
    except: return 0

TYPE_ABBR = {"VARCHAR":"V","VARCHAR2":"V","STRING":"V","CHAR":"C","NUMERIC":"N","NUMBER":"N",
             "INTEGER":"N","DECIMAL":"N","DATE":"D","DATETIME":"D","TIMESTAMP":"T","CLOB":"L","BLOB":"B"}

def map_oracle(dt, L):
    """Oracle 매핑: STRING > 4000 → CLOB"""
    if dt in ("VARCHAR","VARCHAR2","STRING") and L > 4000:
        return "CLOB", 0
    if dt == "STRING":
        return "VARCHAR", L
    return dt, L

combos = defaultdict(int)  # (clsf, dt, L) → count
no_clsf = 0
for c in all_cols:
    if (c["tbl"], c["col_en"]) in TIME_DOM: continue  # 시간 제외
    if not c["col_kr"]: continue
    clsf = extract_clsf(c["col_kr"])
    if not clsf:
        no_clsf += 1
        continue
    L = parse_int(c["dlen"])
    dt, L = map_oracle(c["dtype"], L)
    combos[(clsf, dt, L)] += 1
print(f"  unique 조합: {len(combos)}")
print(f"  분류 추출 실패: {no_clsf}건")

# ============ 4. 신규 도메인 ============
print("\n=== 4. 신규 도메인 ===")
new_doms = []
new_clsfs = set()
def gen_nm(clsf, dt, L):
    abbr = TYPE_ABBR.get(dt.upper(), "X")
    if dt.upper() == "CLOB":
        return f"{clsf}L"
    if L > 0:
        return f"{clsf}{abbr}{L}"
    return f"{clsf}{abbr}"

to_register = []
for (clsf, dt, L), cnt in combos.items():
    if (clsf, dt, L) in exist_by_key: continue
    nm = gen_nm(clsf, dt, L)
    # 이름 충돌 방지
    base = nm; i = 1
    while nm in exist_nm or nm in (x["nm"] for x in to_register):
        i += 1
        nm = f"{base}_{i}"
    to_register.append({"nm": nm, "clsf": clsf, "dt": dt, "len": L, "cnt": cnt})
    if clsf not in exist_clsf: new_clsfs.add(clsf)

to_register.sort(key=lambda x: -x["cnt"])
print(f"  신규 도메인: {len(to_register)}")
print(f"  신규 분류: {len(new_clsfs)}")

# ============ 5. SQL 생성 ============
def esc(s): return str(s).replace("'", "''") if s else ""
sql = ["-- Phase 2-2 D6 — RAMP 전체 (분류,타입,길이) 도메인 등록", "BEGIN;", ""]

if new_clsfs:
    sql.append(f"-- 신규 분류 {len(new_clsfs)}건")
    for cn in sorted(new_clsfs):
        cid = base64.urlsafe_b64encode(os.urandom(17)).rstrip(b'=').decode()[:22]
        sql.append(f"INSERT INTO quality.tb_domain_clsf (domain_clsf_id, domain_clsf_nm, domain_grp_nm, comm_stnd_yn, cret_dt, cret_user_id) "
                   f"VALUES ('{cid}','{esc(cn)}','기타','N',to_char(now(),'YYYYMMDDHH24MISS'),'admin');")
    sql.append("")

sql.append(f"-- 신규 도메인 {len(to_register)}건")
for d in to_register:
    did = base64.urlsafe_b64encode(os.urandom(17)).rstrip(b'=').decode()[:22]
    desc = f"{d['clsf']} {d['dt']}({d['len']}) — 기관표준"
    sql.append(
        f"INSERT INTO quality.tb_domain (domain_id, domain_nm, domain_grp_nm, domain_clsf_nm, domain_desc, data_type, data_len, data_decimal_len, stor_fmt, expr_fmt_lst, comm_stnd_yn, aprv_yn, cret_dt, cret_user_id, use_yn) "
        f"VALUES ('{did}','{esc(d['nm'])}','기타','{esc(d['clsf'])}','{esc(desc)}','{esc(d['dt'])}',{d['len']},0,'',ARRAY[]::text[],'N','Y',to_char(now(),'YYYYMMDDHH24MISS'),'admin','Y');"
    )

sql += ["", "SELECT 'tb_domain N' as t, count(*) FROM quality.tb_domain WHERE comm_stnd_yn='N';", "", "COMMIT;"]
OUT_SQL.write_text("\n".join(sql), encoding="utf-8")
print(f"\n  → {OUT_SQL}")
print(f"  분류 추출 실패: {no_clsf}건 (도메인 등록 안 됨, 별도 처리)")
