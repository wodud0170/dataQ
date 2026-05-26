"""
RAMP 단어사전 ↔ RAMP 스키마 ↔ 행안부 단어 종합 분석.

산출:
  _A_ramp_dict_normalized.tsv       — RAMP 원본 사전 정규화 (1573)
  _B_ramp_schema_tokens.tsv         — RAMP 스키마 컬럼 토큰 분해
  _C_dict_vs_schema_match.tsv       — 사전↔스키마 매칭률
  _D_reverse_dict_full.tsv          — 스키마에서 역산한 토큰 전수 (빈도+사전매칭)
  _E_ramp_vs_mois.tsv               — RAMP vs 행안부 중복/충돌
  _F_summary.txt                    — 요약 통계
"""
from openpyxl import load_workbook
from pathlib import Path
from collections import Counter, defaultdict
import re

BASE = Path(r"C:\Users\장재영\Desktop\dataQ\CAMS표준화")
DICT = BASE / "01_원본자료" / "RMS4-DE14-04.데이터표준화_단어사전-V2.0.xlsx"
SCHEMA = BASE / "CAMS_RAMP_통합" / "ramp기관스키마정보.xlsx"
MOIS_WORD = BASE.parent / "q-center" / "src" / "main" / "resources" / "seed" / "행안부_공통표준" / "행정안전부_공공데이터 공통표준단어.xlsx"

OUT = BASE / "04_RAMP분석_2026-05-21" / "_분석산출물_tsv"
OUT.mkdir(exist_ok=True)

# ============ 1) RAMP 사전 로드 ============
print("=== 1) RAMP 사전 로드 ===")
wb = load_workbook(DICT, read_only=True, data_only=True)
ws = wb["단어사전"]
rows = list(ws.iter_rows(min_row=2, values_only=True))
wb.close()

ramp_dict = []  # list of dicts
for r in rows:
    if not r or not r[2]:  # 단어명 비어있으면 skip
        continue
    ramp_dict.append({
        "no": r[0],
        "차수": r[1] or "",
        "단어명": (r[2] or "").strip(),
        "영문약어": (r[3] or "").strip().upper(),
        "영문명": (r[4] or "").strip(),
        "설명": (r[5] or "").replace("_x000D_", " ").strip(),
        "형식단어": r[6] or "",
        "도메인": r[7] or "",
        "이음동의어": r[8] or "",
        "금칙어": r[9] or "",
        "표준여부": r[11] or "",
    })

with open(OUT/"_A_ramp_dict_normalized.tsv", "w", encoding="utf-8") as f:
    f.write("단어명\t영문약어\t영문명\t형식단어\t도메인\t표준여부\t설명\n")
    for d in ramp_dict:
        f.write(f"{d['단어명']}\t{d['영문약어']}\t{d['영문명']}\t{d['형식단어']}\t{d['도메인']}\t{d['표준여부']}\t{d['설명'][:50]}\n")

print(f"  RAMP 사전 행: {len(ramp_dict)}")
std_y = sum(1 for d in ramp_dict if d["표준여부"] == "Y")
print(f"  표준여부 Y: {std_y}, N/blank: {len(ramp_dict) - std_y}")

# 사전을 영문약어 기준 dict로 — 중복 키도 확인
dict_by_abbr = defaultdict(list)
for d in ramp_dict:
    if d["영문약어"]:
        dict_by_abbr[d["영문약어"]].append(d)
abbr_dup = {k: v for k, v in dict_by_abbr.items() if len(v) > 1}
print(f"  영문약어 중복 키: {len(abbr_dup)}")
if abbr_dup:
    for k, lst in list(abbr_dup.items())[:5]:
        print(f"    {k}: {[d['단어명'] for d in lst]}")

# ============ 2) RAMP 스키마 컬럼 로드 + 토큰 분해 ============
print("\n=== 2) RAMP 스키마 로드 ===")
wb = load_workbook(SCHEMA, read_only=True, data_only=True)
ws = wb["컬럼"]
schema_rows = list(ws.iter_rows(min_row=2, values_only=True))
wb.close()

# 컬럼 — (테이블, 영문, 한글)
schema_cols = []
for r in schema_rows:
    if not r or not r[1]:
        continue
    schema_cols.append({
        "tbl": (r[0] or "").strip(),
        "en": (r[1] or "").strip().lower(),
        "kr": (r[2] or "").strip(),
        "type": r[5] or "",
        "len": r[6] or "",
    })
print(f"  스키마 컬럼: {len(schema_cols)}")

# 영문 토큰 분해 (snake_case)
def tokenize_en(s):
    s = s.strip().lower()
    if not s: return []
    return [t for t in re.split(r"[_]+", s) if t]

# 영문 토큰 빈도 + 토큰→소속 컬럼 매핑
token_freq = Counter()
token_cols = defaultdict(list)  # token -> [(tbl, en_col)]
for c in schema_cols:
    for tk in tokenize_en(c["en"]):
        token_freq[tk] += 1
        token_cols[tk].append((c["tbl"], c["en"]))
print(f"  unique 영문 토큰: {len(token_freq)}")

with open(OUT/"_B_ramp_schema_tokens.tsv", "w", encoding="utf-8") as f:
    f.write("token\tfreq\tsample_cols\n")
    for tk, cnt in token_freq.most_common():
        sample = "; ".join(f"{t}.{c}" for t, c in token_cols[tk][:3])
        f.write(f"{tk.upper()}\t{cnt}\t{sample}\n")

# ============ 3) 사전 ↔ 스키마 매칭률 ============
print("\n=== 3) 사전 ↔ 스키마 매칭률 ===")
dict_abbrs = set(d["영문약어"] for d in ramp_dict if d["영문약어"])
schema_tokens = set(tk.upper() for tk in token_freq.keys())

both = dict_abbrs & schema_tokens
dict_only = dict_abbrs - schema_tokens   # 사전엔 있는데 스키마 미사용
schema_only = schema_tokens - dict_abbrs # 스키마엔 있는데 사전 미등록

print(f"  사전 영문약어: {len(dict_abbrs)}")
print(f"  스키마 토큰: {len(schema_tokens)}")
print(f"  교집합: {len(both)}")
print(f"  사전 only (스키마 미사용): {len(dict_only)}")
print(f"  스키마 only (사전 미등록): {len(schema_only)}")
print(f"  사전 활용률: {len(both)/len(dict_abbrs)*100:.1f}%")
print(f"  스키마 표준 적합률: {len(both)/len(schema_tokens)*100:.1f}%")

with open(OUT/"_C_dict_vs_schema_match.tsv", "w", encoding="utf-8") as f:
    f.write("구분\t영문약어\t단어명\t스키마등장수\t비고\n")
    for tk in sorted(both):
        # 사전 정보
        ds = dict_by_abbr.get(tk, [])
        d0 = ds[0] if ds else {}
        cnt = token_freq[tk.lower()]
        f.write(f"매칭\t{tk}\t{d0.get('단어명','')}\t{cnt}\t\n")
    for tk in sorted(dict_only):
        ds = dict_by_abbr.get(tk, [])
        d0 = ds[0] if ds else {}
        f.write(f"사전only\t{tk}\t{d0.get('단어명','')}\t0\t스키마 미사용\n")
    for tk in sorted(schema_only):
        cnt = token_freq[tk.lower()]
        sample = token_cols[tk.lower()][:2]
        sample_s = "; ".join(f"{t}.{c}" for t, c in sample)
        f.write(f"스키마only\t{tk}\t\t{cnt}\t{sample_s}\n")

# ============ 4) 스키마 역산 — 스키마 토큰 전수 + 한글 후보 추정 ============
print("\n=== 4) 스키마 역산 단어사전 ===")
# 각 토큰의 한글 후보: 동일 컬럼이 등장한 컬럼들의 한글명에서 토큰 위치 기반 매핑
# 단순 접근: 사전 매칭되는 토큰은 사전 한글명 그대로, 미매칭은 컬럼 한글명 사용

# 컬럼별 영문토큰 / 한글토큰 정렬 추정 (한글은 조사가 없어 토큰 분리가 어려움 — naive)
# token -> 한글 명 후보(빈도)
token_kr_cand = defaultdict(Counter)
for c in schema_cols:
    en_tokens = tokenize_en(c["en"])
    kr = c["kr"]
    if not kr: continue
    # 단일 토큰 컬럼은 한글 = 해당 토큰의 후보
    if len(en_tokens) == 1:
        token_kr_cand[en_tokens[0]][kr] += 1

with open(OUT/"_D_reverse_dict_full.tsv", "w", encoding="utf-8") as f:
    f.write("영문약어\t빈도\t사전등록\t사전단어명\t한글후보_단일컬럼\t샘플컬럼\n")
    for tk, cnt in token_freq.most_common():
        UP = tk.upper()
        in_dict = "Y" if UP in dict_abbrs else "N"
        ds = dict_by_abbr.get(UP, [])
        dict_kr = ds[0]["단어명"] if ds else ""
        kr_cands = token_kr_cand.get(tk, Counter())
        kr_top = "; ".join(f"{k}({v})" for k, v in kr_cands.most_common(3))
        sample = "; ".join(f"{t}.{c}" for t, c in token_cols[tk][:2])
        f.write(f"{UP}\t{cnt}\t{in_dict}\t{dict_kr}\t{kr_top}\t{sample}\n")
print(f"  -> {OUT}/_D_reverse_dict_full.tsv")

# ============ 5) 행안부 사전 vs RAMP 비교 ============
print("\n=== 5) 행안부 사전 vs RAMP 사전 ===")
if not MOIS_WORD.exists():
    print(f"  MOIS 사전 없음: {MOIS_WORD}")
else:
    wb = load_workbook(MOIS_WORD, read_only=True, data_only=True)
    print(f"  MOIS 시트: {wb.sheetnames}")
    # 첫 시트 사용
    ws = wb[wb.sheetnames[0]]
    head = [c for c in next(ws.iter_rows(min_row=1, max_row=1, values_only=True))]
    print(f"  MOIS 헤더: {head}")
    # 헤더에서 단어명/영문약어 컬럼 인덱스 찾기
    def find_col(headers, patterns):
        for i, h in enumerate(headers):
            hs = str(h or "")
            for p in patterns:
                if p in hs:
                    return i
        return -1
    name_col = find_col(head, ["단어명", "표준단어명", "한글명"])
    abbr_col = find_col(head, ["영문약어", "약어명"])
    full_col = find_col(head, ["영문", "영문명"])
    if name_col < 0 or abbr_col < 0:
        print(f"  ! 헤더 매칭 실패: name={name_col}, abbr={abbr_col}")
    mois_rows = list(ws.iter_rows(min_row=2, values_only=True))
    wb.close()
    mois = []
    for r in mois_rows:
        if not r or not r[name_col]: continue
        mois.append({
            "단어명": str(r[name_col]).strip(),
            "영문약어": str(r[abbr_col] or "").strip().upper(),
            "영문명": str(r[full_col] or "").strip() if full_col >= 0 else "",
        })
    print(f"  MOIS 단어: {len(mois)}")

    mois_abbrs = set(m["영문약어"] for m in mois if m["영문약어"])
    mois_names = set(m["단어명"] for m in mois)
    print(f"  MOIS 영문약어 unique: {len(mois_abbrs)}")
    print(f"  MOIS 단어명 unique: {len(mois_names)}")

    # 영문약어 기준 RAMP↔MOIS 매칭
    ramp_abbrs_y = set(d["영문약어"] for d in ramp_dict if d["영문약어"] and d["표준여부"]=="Y")
    ramp_abbrs_all = set(d["영문약어"] for d in ramp_dict if d["영문약어"])
    ramp_names = set(d["단어명"] for d in ramp_dict)

    abbr_overlap = ramp_abbrs_all & mois_abbrs
    name_overlap = ramp_names & mois_names

    # 더 정밀: 같은 영문약어인데 단어명이 다른 → 충돌
    mois_by_abbr = {m["영문약어"]: m for m in mois if m["영문약어"]}
    ramp_by_abbr = {d["영문약어"]: d for d in ramp_dict if d["영문약어"]}

    abbr_match_name_match = 0   # 영문약어/단어명 일치 — 동일
    abbr_match_name_diff = 0    # 영문약어 같은데 단어명 다름 — 충돌
    conflicts = []
    for a in abbr_overlap:
        mr = ramp_by_abbr[a]
        mm = mois_by_abbr[a]
        if mr["단어명"] == mm["단어명"]:
            abbr_match_name_match += 1
        else:
            abbr_match_name_diff += 1
            conflicts.append((a, mr["단어명"], mm["단어명"]))

    # 단어명 같은데 영문약어 다른 케이스 — 잠재적 영문약어 변경 대상
    mois_by_name = {m["단어명"]: m for m in mois}
    ramp_by_name = {d["단어명"]: d for d in ramp_dict}
    name_overlap_abbr_diff = []
    for n in name_overlap:
        mr = ramp_by_name[n]
        mm = mois_by_name[n]
        if mr["영문약어"] and mm["영문약어"] and mr["영문약어"] != mm["영문약어"]:
            name_overlap_abbr_diff.append((n, mr["영문약어"], mm["영문약어"]))

    print(f"\n  [영문약어 일치] RAMP∩MOIS: {len(abbr_overlap)}")
    print(f"    그중 단어명도 같음: {abbr_match_name_match}")
    print(f"    영문약어는 같지만 단어명 다름 (충돌): {abbr_match_name_diff}")
    print(f"  [단어명 일치] RAMP∩MOIS: {len(name_overlap)}")
    print(f"    그중 영문약어 다름 (영문 변경 대상): {len(name_overlap_abbr_diff)}")

    with open(OUT/"_E_ramp_vs_mois.tsv", "w", encoding="utf-8") as f:
        f.write("구분\tRAMP영문약어\tRAMP단어명\tMOIS영문약어\tMOIS단어명\t조치\n")
        for a, rn, mn in conflicts:
            f.write(f"영문약어동일_단어명다름\t{a}\t{rn}\t{a}\t{mn}\tRAMP 단어명→MOIS로 변경 또는 영문약어 변경\n")
        for n, ra, ma in name_overlap_abbr_diff:
            f.write(f"단어명동일_영문약어다름\t{ra}\t{n}\t{ma}\t{n}\tRAMP 영문약어 {ra}→{ma} 로 변경\n")
        # 영문약어/단어명 모두 일치 — 그대로 사용 가능 (대량이라 샘플만)
        for a in sorted(abbr_overlap)[:200]:
            mr = ramp_by_abbr[a]
            mm = mois_by_abbr[a]
            if mr["단어명"] == mm["단어명"]:
                f.write(f"완전일치\t{a}\t{mr['단어명']}\t{a}\t{mm['단어명']}\t유지\n")

    # 행안부에만 있는 단어 (RAMP가 추가 도입할 단어)
    mois_only_abbr = mois_abbrs - ramp_abbrs_all
    ramp_only_abbr = ramp_abbrs_all - mois_abbrs

    # 요약
    with open(OUT/"_F_summary.txt", "w", encoding="utf-8") as f:
        f.write("=== RAMP 단어사전 ↔ RAMP 스키마 ↔ 행안부 사전 분석 요약 ===\n\n")
        f.write(f"## A. 원본 RAMP 사전\n")
        f.write(f"  - 사전 등재 단어수: {len(ramp_dict)}\n")
        f.write(f"  - 표준여부=Y: {std_y}\n")
        f.write(f"  - 표준여부=N/blank: {len(ramp_dict)-std_y}\n")
        f.write(f"  - 영문약어 중복키: {len(abbr_dup)}\n\n")
        f.write(f"## B. RAMP 스키마\n")
        f.write(f"  - 테이블: 365 (별도 시트)\n")
        f.write(f"  - 컬럼: {len(schema_cols)}\n")
        f.write(f"  - 영문 토큰 unique: {len(token_freq)}\n\n")
        f.write(f"## C. 사전 ↔ 스키마 매칭률\n")
        f.write(f"  - 사전 영문약어수: {len(dict_abbrs)}\n")
        f.write(f"  - 스키마 토큰수: {len(schema_tokens)}\n")
        f.write(f"  - 양쪽 모두 등장 (정상): {len(both)}\n")
        f.write(f"  - 사전 only (사전 등재 but 스키마 미사용): {len(dict_only)}\n")
        f.write(f"  - 스키마 only (스키마 사용 but 사전 미등록): {len(schema_only)}\n")
        f.write(f"  - 사전 활용률 = 교집합/사전: {len(both)/len(dict_abbrs)*100:.1f}%\n")
        f.write(f"  - 스키마 표준 적합률 = 교집합/스키마: {len(both)/len(schema_tokens)*100:.1f}%\n\n")
        if MOIS_WORD.exists():
            f.write(f"## D. 행안부(MOIS) 사전 비교\n")
            f.write(f"  - MOIS 단어수: {len(mois)}\n")
            f.write(f"  - MOIS 영문약어 unique: {len(mois_abbrs)}\n\n")
            f.write(f"### D-1. 영문약어 기준 교차\n")
            f.write(f"  - RAMP∩MOIS 영문약어: {len(abbr_overlap)}\n")
            f.write(f"    └ 단어명까지 동일 (완전일치): {abbr_match_name_match}\n")
            f.write(f"    └ 영문약어 같지만 단어명 다름 (★ 충돌): {abbr_match_name_diff}\n")
            f.write(f"  - RAMP only 영문약어: {len(ramp_only_abbr)}\n")
            f.write(f"  - MOIS only 영문약어: {len(mois_only_abbr)}\n\n")
            f.write(f"### D-2. 단어명(한글) 기준 교차\n")
            f.write(f"  - RAMP∩MOIS 단어명: {len(name_overlap)}\n")
            f.write(f"    └ 한글 같지만 영문약어 다름 (★ 영문약어 변경 대상): {len(name_overlap_abbr_diff)}\n\n")
            f.write(f"### D-3. RAMP 측 변경 필요 (행안부 우선 원칙)\n")
            f.write(f"  - 충돌(영문약어 같은데 단어명 다름) {abbr_match_name_diff}건 → RAMP 단어명 또는 영문약어 변경\n")
            f.write(f"  - 영문약어 변경 {len(name_overlap_abbr_diff)}건 → 같은 한글 단어인데 RAMP가 다른 영문 약어 사용 중\n")
            f.write(f"  - 영향 컬럼 추정:\n")
            affected_cols = 0
            for a, _, _ in conflicts:
                affected_cols += token_freq.get(a.lower(), 0)
            for n, ra, _ in name_overlap_abbr_diff:
                affected_cols += token_freq.get(ra.lower(), 0)
            f.write(f"     충돌 영문약어가 등장하는 RAMP 컬럼 추정: {affected_cols} 건\n")

    print(f"  RAMP only: {len(ramp_only_abbr)}, MOIS only: {len(mois_only_abbr)}")
    print(f"\n  -> {OUT}/_E_ramp_vs_mois.tsv, _F_summary.txt")

print("\n=== DONE ===")
