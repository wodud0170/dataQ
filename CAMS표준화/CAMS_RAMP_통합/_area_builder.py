# -*- coding: utf-8 -*-
"""
영역별 매핑 HTML 생성기. _make_html.py 의 W7 전용 로직을 일반화.

사용:
  python _area_builder.py <영역코드> [<영역코드> ...]
  python _area_builder.py W25 W2
"""
import openpyxl, json, re, collections, html, sys, os

ROOT='..'
CAMS_XLSX=os.path.join(ROOT,'CAMS_SCHEMA_원본.xlsx')
RAMP_XLSX='ramp기관스키마정보.xlsx'

# ===== 영역 정의 =====
AREAS={
# W25 (사용자·조직·시스템)는 통합 대상에서 제외 — 곁가지 (메뉴·사용자는 통합 후 신규 구성)
# 단, 조직·부서·기관(처리과·기록관 마스터)은 비즈니스 FK라 키 매핑(K5/K6)으로 별도 처리
'W2':{
 'title':'분류·기능분류',
 'subtitle':'W2 · 단위업무·기능분류·분류체계',
 'sections':[
   {'name':'단위업무 마스터',
    'cams':['CM_CLASS_BASTABLE','CM_UNIT_TASK','CM_UNITBS_SKILLDIV'],
    'ramp':['tb_zzunit','tb_zzorgunit','tb_zzsendunit']},
   {'name':'단위업무 신청·이력',
    'cams':[],  # CAMS 측 단위업무 신청 마스터 별도
    'ramp':['tb_zzunitnewreq','tb_zzunitchgreq','tb_zzunitreq','tb_zzunitmovereq','tb_zzunitmovreq','tb_zzunitsetreq','tb_zzunitdelreq','tb_zzunitchghist','tb_zzorgunitworkhist','tb_zzunitrslt','tb_zzunitprsrtermhist','tb_zzprsrtermhist','tb_zzprsrtermrule','tb_zzprsrrcptrslt']},
   {'name':'기능분류',
    'cams':['CM_CLASS','CM_CLASS_DIVSYS'],
    'ramp':['tb_zzfnctclsf','tb_zzfnctclsfhist','tb_zzfnctclsfchg','tb_zzfnctorg','tb_zzfnctclsfrelinfosys','tb_zzfnctclsfrellaw','tb_zzfnctclsfrelbiz']},
   {'name':'분류체계 (역사·해외·정부간행물·박물)',
    'cams':['CM_OFFDOC_TERMTAB','CM_OFFDOC_TERMEX','CM_64OFFDOC_DIVSYS','CM_HISTDOC_DIVSYS','CM_OLDGOV_DIVSYS','CM_OLDGOV_ORGAN','CM_OLDGOV_DIVTABLE','CM_OVERS_DIVSYS','CM_OVERS_DIVTABLE','CM_GOVART_DIVSYS','CM_SKLDIV_XML_SKLDIV','CM_SKLDIV_XML_ORGAN','CM_SKLDIV_XML_INFOSYS','CM_SKLDIV_XML_LAW','CM_SKLDIV_REFINFOSYS','CM_SKLDIV_REFLAW','CM_SKL_PROC_LIST','CM_SKL_DIV_COD_TRANS_INFO','CM_SKL_DIV_COD_TRANS_ORG','CM_SKL_OFLN_RECP_XML'],
    'ramp':['tb_zzpjtclsf','tb_zzpjtclsfhist','tb_zzpjttypemappng','tb_zzpjtmappng','tb_zzclsf','tb_zzcomcd','tb_zzcomtypecd','tb_zzorgcomstnd','tb_zzstndmng','tb_zzstndmngtype','tb_zzrecordclsfancmnt','tb_zzbrmupldhist','tb_zzbrmsendfilehist','tb_zzsendunitorg']},
 ]},
}

# ===== 스키마 로드 =====
def load_cams():
    wb=openpyxl.load_workbook(CAMS_XLSX,read_only=True,data_only=True)
    rows=list(wb['컬럼정의'].iter_rows(values_only=True))[1:]
    out=collections.defaultdict(list)
    for r in rows:
        if r[0] and r[2]:
            out[r[0]].append({'en':r[2],'cmt':r[3] or '','dt':r[4] or '','dl':r[5] or '','pk':str(r[8] or '').strip().upper()=='Y','tbl':r[0]})
    return dict(out)

def load_ramp():
    wb=openpyxl.load_workbook(RAMP_XLSX,read_only=True,data_only=True)
    rows=[r for r in list(wb['컬럼'].iter_rows(values_only=True))[1:] if r[2]]
    out=collections.defaultdict(list)
    for r in rows:
        out[r[0]].append({'en':r[1],'kr':r[2],'desc':r[3] or '','dt':r[5] or '','dl':r[6] or '','pk':str(r[7] or '').strip().upper()=='Y','tbl':r[0]})
    return dict(out)

# ===== 매칭 알고리즘 =====
def norm(s):
    if not s: return ''
    s=re.sub(r'\([^)]*\)|\[[^\]]*\]|:.*$','',str(s))
    s=re.sub(r'[\s·/_]','',s)
    return s.strip()

SYN=[(r'아이디','ID'),(r'순번','일련번호'),(r'년도','연도'),(r'이름','명'),(r'세부','상세'),(r'사이즈','크기'),(r'메세지','메시지'),(r'타입','유형')]
def canon(s):
    s=norm(s)
    for pat,rep in SYN: s=re.sub(pat,rep,s)
    return s

USER_CONFIRMED={
 ('RG_DOCUMENT','BSID'):('tb_rdfolder','fls_id'),
 ('RG_DETAIL','BSID'):('tb_rdrecord','fls_id'),
 ('RG_DETAIL','DSID'):('tb_rdrecord','ritm_id'),
}

def match_columns(cams_cols, ramp_cols):
    out=[]
    used=set()
    for c in cams_cols:
        key=(c['tbl'],c['en'])
        if key in USER_CONFIRMED:
            target=USER_CONFIRMED[key]
            for r in ramp_cols:
                if r['tbl']==target[0] and r['en']==target[1]:
                    out.append({'cams':c,'ramp':r,'signal':'S6','conf':'A','note':'사용자 확정'})
                    used.add((r['tbl'],r['en'])); break
            continue
        cn=canon(c['cmt']); best=None; best_score=0; best_signal=''
        for r in ramp_cols:
            if (r['tbl'],r['en']) in used: continue
            rn=canon(r['kr'])
            if not cn or not rn: continue
            score=0; sig=[]
            if cn==rn: score+=10; sig.append('S1a')
            elif cn in rn or rn in cn: score+=5; sig.append('S1b')
            elif len(cn)>=2 and len(rn)>=2 and cn[-2:]==rn[-2:]: score+=2; sig.append('S1b')
            if c['dt']=='VARCHAR2' and r['dt']=='STRING' and str(c['dl'])==str(r['dl']): score+=3; sig.append('S3')
            elif c['dt']=='NUMBER' and r['dt'] in ('NUMERIC','INTEGER') and str(c['dl'])==str(r['dl']): score+=3; sig.append('S3')
            elif c['dt']=='DATE' and r['dt'] in ('DATETIME','DATE'): score+=2; sig.append('S3')
            elif c['dt']=='CHAR' and r['dt']=='CHAR': score+=2; sig.append('S3')
            if score>best_score: best_score=score; best=r; best_signal='+'.join(sig)
        if best and best_score>=5:
            if 'S1a' in best_signal and 'S3' in best_signal: conf='A'
            elif 'S1a' in best_signal: conf='B'
            elif 'S1b' in best_signal and 'S3' in best_signal: conf='B'
            else: conf='C'
            out.append({'cams':c,'ramp':best,'signal':best_signal,'conf':conf,'note':''})
            used.add((best['tbl'],best['en']))
        else:
            out.append({'cams':c,'ramp':None,'signal':'','conf':'D','note':'CAMS 단방향'})
    for r in ramp_cols:
        if (r['tbl'],r['en']) not in used:
            out.append({'cams':None,'ramp':r,'signal':'','conf':'D','note':'RAMP 단방향'})
    return out

# ===== HTML 생성 =====
CONF_COLOR={'A':'#16a34a','B':'#ca8a04','C':'#ea580c','D':'#94a3b8'}
CONF_BG   ={'A':'#dcfce7','B':'#fef9c3','C':'#ffedd5','D':'#f1f5f9'}

def esc(s): return html.escape(str(s)) if s else ''

def row_html(r):
    c=r.get('cams'); rmp=r.get('ramp'); conf=r['conf']
    color=CONF_COLOR[conf]; bg=CONF_BG[conf]
    if c:
        pk='🔑 ' if c.get('pk') else ''
        cams=f"{pk}<div class='col-tbl'>{esc(c['tbl'])}</div><div class='col-en'>{esc(c['en'])}</div><div class='col-cmt'>{esc(c['cmt'])}</div><div class='col-dt'>{esc(c['dt'])}({esc(c['dl'])})</div>"
    else: cams="<div class='empty'>—</div>"
    if rmp:
        pk='🔑 ' if rmp.get('pk') else ''
        ramp=f"{pk}<div class='col-tbl'>{esc(rmp['tbl'])}</div><div class='col-en'>{esc(rmp['en'])}</div><div class='col-cmt'>{esc(rmp['kr'])}</div><div class='col-dt'>{esc(rmp['dt'])}({esc(rmp['dl'])})</div>"
    else: ramp="<div class='empty'>—</div>"
    arrow='↔' if c and rmp else ('→' if c else '←')
    return f"<tr style='background:{bg}'><td class='cams-col'>{cams}</td><td class='arrow-col' style='color:{color}'><div class='arrow'>{arrow}</div><div class='conf-badge' style='background:{color};color:white'>{conf}</div><div class='signal'>{esc(r['signal']) or '-'}</div></td><td class='ramp-col'>{ramp}</td><td class='note-col'>{esc(r['note'])}</td></tr>"

def stat_card(rows,label):
    s=collections.Counter(x['conf'] for x in rows)
    return f"<div class='stat-card'><div class='stat-num conf-A'>{s.get('A',0)}</div><div class='stat-num conf-B'>{s.get('B',0)}</div><div class='stat-num conf-C'>{s.get('C',0)}</div><div class='stat-num conf-D'>{s.get('D',0)}</div><div class='stat-label'>{label} · {len(rows)}행</div></div>"

CSS="""
@import url('https://fonts.googleapis.com/css2?family=Noto+Sans+KR:wght@400;500;700&display=swap');
*{box-sizing:border-box}
body{font-family:'Noto Sans KR','Malgun Gothic',sans-serif;color:#1e293b;margin:0;background:#f8fafc;line-height:1.55}
.container{max-width:1280px;margin:0 auto;padding:32px}
h1{font-size:32px;font-weight:700;margin:0 0 8px 0;color:#0f172a}
h2{font-size:24px;font-weight:700;margin:48px 0 16px 0;color:#0f172a;border-left:6px solid #2563eb;padding-left:12px}
h3{font-size:18px;font-weight:700;margin:24px 0 12px 0;color:#1e293b}
.cover{background:linear-gradient(135deg,#1e3a8a 0%,#1e40af 100%);color:white;padding:48px 32px;border-radius:12px;margin-bottom:32px}
.cover h1{color:white;font-size:36px}
.cover .subtitle{color:#bfdbfe;font-size:16px}
.cover-meta{margin-top:24px;font-size:14px;color:#bfdbfe}
.info{background:#eff6ff;border-left:4px solid #2563eb;padding:16px 20px;margin:16px 0;border-radius:6px}
.alert{background:#fef2f2;border-left:4px solid #dc2626;padding:16px 20px;margin:16px 0;border-radius:6px}
.alert-title{font-weight:700;color:#991b1b;margin-bottom:6px}
.stat-bar{display:flex;gap:12px;margin:14px 0;flex-wrap:wrap}
.stat-card{background:white;border-radius:8px;padding:12px 16px;box-shadow:0 1px 3px rgba(0,0,0,0.08);display:flex;gap:8px;align-items:center}
.stat-num{font-size:20px;font-weight:700;padding:3px 8px;border-radius:4px;color:white;min-width:32px;text-align:center}
.stat-num.conf-A{background:#16a34a}.stat-num.conf-B{background:#ca8a04}.stat-num.conf-C{background:#ea580c}.stat-num.conf-D{background:#94a3b8}
.stat-label{font-weight:600;color:#0f172a;margin-left:6px;font-size:13px}
.mapping-table{width:100%;border-collapse:collapse;margin:12px 0 28px 0;background:white;box-shadow:0 1px 3px rgba(0,0,0,0.08);border-radius:8px;overflow:hidden}
.mapping-table thead th{background:#0f172a;color:white;padding:9px 12px;text-align:left;font-size:12.5px;font-weight:600}
.mapping-table tbody td{padding:9px 12px;font-size:12.5px;border-top:1px solid #e2e8f0;vertical-align:top}
.cams-col,.ramp-col{width:36%}.arrow-col{width:12%;text-align:center;font-weight:700}.note-col{width:16%;font-size:11.5px;color:#64748b}
.col-tbl{font-size:10.5px;color:#94a3b8;font-family:'Consolas',monospace;margin-bottom:2px}
.col-en{font-family:'Consolas',monospace;font-weight:700;font-size:12.5px;color:#1e3a8a}
.col-cmt{font-size:12px;margin-top:2px;color:#334155}
.col-dt{font-size:10.5px;color:#64748b;margin-top:2px;font-family:'Consolas',monospace}
.arrow{font-size:18px;margin-bottom:3px}
.conf-badge{display:inline-block;padding:2px 7px;border-radius:9px;font-size:10.5px;font-weight:700;letter-spacing:0.4px}
.signal{font-size:9.5px;color:#64748b;margin-top:2px;font-family:'Consolas',monospace}
.empty{color:#cbd5e1;font-style:italic;text-align:center}
.legend{display:flex;gap:16px;margin:14px 0;padding:12px 16px;background:white;border-radius:8px;font-size:12.5px;flex-wrap:wrap}
.legend-item{display:flex;align-items:center;gap:6px}
.legend-color{width:16px;height:16px;border-radius:3px}
.decision-list{background:white;border-radius:8px;padding:14px 22px;margin:12px 0;box-shadow:0 1px 3px rgba(0,0,0,0.08)}
.section-meta{font-size:12.5px;color:#64748b;margin:4px 0 8px 0}
@media print{body{background:white}.container{max-width:none;padding:12px}.mapping-table{box-shadow:none}}
"""

def build_area_html(code, area, cams_schema, ramp_schema):
    sections_html=''
    grand_stat=collections.Counter()
    for sec in area['sections']:
        cams_cols=[]
        for t in sec['cams']:
            for c in cams_schema.get(t,[]):
                cams_cols.append(c)
        ramp_cols=[]
        for t in sec['ramp']:
            for c in ramp_schema.get(t,[]):
                ramp_cols.append(c)
        rows=match_columns(cams_cols, ramp_cols)
        for r in rows: grand_stat[r['conf']]+=1
        rows_html=''.join(row_html(r) for r in rows)
        meta=f"CAMS {len(cams_cols)}컬럼 ({len(sec['cams'])}테이블) · RAMP {len(ramp_cols)}컬럼 ({len(sec['ramp'])}테이블)"
        sections_html += f"""<h3>{esc(sec['name'])}</h3>
<div class='section-meta'>{meta}</div>
<div class='stat-bar'>{stat_card(rows,sec['name'])}</div>
<table class='mapping-table'><thead><tr>
<th>CAMS (테이블 · 영문명 · 코멘트 · 도메인)</th><th>매핑</th><th>RAMP (테이블 · 영문명 · 한글명 · 도메인)</th><th>비고</th>
</tr></thead><tbody>{rows_html}</tbody></table>
"""
    grand=f"<div class='stat-bar' style='margin:20px 0'><div class='stat-card' style='background:#fef9c3;border:2px solid #ca8a04'><div class='stat-num conf-A'>{grand_stat['A']}</div><div class='stat-num conf-B'>{grand_stat['B']}</div><div class='stat-num conf-C'>{grand_stat['C']}</div><div class='stat-num conf-D'>{grand_stat['D']}</div><div class='stat-label'>{esc(area['title'])} 전체</div></div></div>"
    doc=f"""<!DOCTYPE html><html lang='ko'><head><meta charset='utf-8'><title>{esc(code)} {esc(area['title'])} 매핑 보고서</title><style>{CSS}</style></head><body><div class='container'>
<div class='cover'><h1>{esc(code)} · {esc(area['title'])} 매핑 보고서</h1><div class='subtitle'>{esc(area['subtitle'])}</div><div class='cover-meta'>작성 2026-05-19 · 입력 CAMS_SCHEMA_원본.xlsx · ramp기관스키마정보.xlsx<br>방법론: <code>10_메인테이블_매핑_방법론.md</code> · 업무 매트릭스: <code>11_업무별_테이블그룹_매핑.md</code></div></div>

<div class='alert'><div class='alert-title'>⚠️ 본 보고서는 1차 자동 매핑</div>
W7 메인테이블과 동일 방법론·신뢰도 등급. CAMS 데이터 접근 불가, 메타데이터 기반. 통합팀·업무 이해관계자 검토 필수.
임시(_TEMP) 테이블은 자동 제외.</div>

<div class='info'><strong>전체 요약</strong>{grand}</div>

<div class='legend'>
<div class='legend-item'><span class='legend-color' style='background:#dcfce7'></span><strong style='color:#16a34a'>A 확정</strong> — 사용자 확정 또는 S1a+S3</div>
<div class='legend-item'><span class='legend-color' style='background:#fef9c3'></span><strong style='color:#ca8a04'>B 후보</strong> — S1a 또는 S1b+S3</div>
<div class='legend-item'><span class='legend-color' style='background:#ffedd5'></span><strong style='color:#ea580c'>C 추정</strong> — 약신호만</div>
<div class='legend-item'><span class='legend-color' style='background:#f1f5f9'></span><strong style='color:#94a3b8'>D 미매핑</strong> — 단방향</div>
</div>

<h2>{esc(area['title'])} 세부 매핑</h2>
{sections_html}

<h2>의사결정 포인트</h2>
<div class='decision-list'><ol>
<li><strong>D 등급 미매핑 컬럼</strong>: 통합 후 RAMP에 신규 컬럼으로 추가할지 / 폐기할지 / legacy 매핑테이블로만 보존할지.</li>
<li><strong>C 등급 추정 매핑</strong>: 약신호만 — 업무 이해관계자 전수 검토 필요.</li>
<li><strong>RAMP-only 영역</strong>: CAMS 측 동등 영역이 비어있는 sub-section은 CAMS 측에 매핑할 마스터가 진짜 없는지 재확인 (사전 필터링되었을 수 있음).</li>
<li><strong>흡수 후 컬럼명</strong>: 매핑된 컬럼은 RAMP 명명 채택 (통합 정책).</li>
</ol></div>

<div style='margin-top:48px;padding:16px;border-top:1px solid #e2e8f0;color:#94a3b8;font-size:12px;text-align:center'>CAMS_RAMP_통합/{esc(code)}_{esc(area['title'].replace('·','_'))}.html · 자립형 HTML · 인쇄: Ctrl+P</div>
</div></body></html>"""
    fname=f"{code}_{area['title'].replace('·','_').replace(' ','_')}.html"
    with open(fname,'w',encoding='utf-8') as f: f.write(doc)
    print(f"  → {fname} ({len(doc)//1024} KB, A:{grand_stat['A']} B:{grand_stat['B']} C:{grand_stat['C']} D:{grand_stat['D']})")

# ===== main =====
if __name__=='__main__':
    cams=load_cams(); ramp=load_ramp()
    targets=sys.argv[1:] or list(AREAS.keys())
    for code in targets:
        if code not in AREAS: print(f"[SKIP] {code} 정의 없음"); continue
        print(f"=== {code} {AREAS[code]['title']} ===")
        build_area_html(code, AREAS[code], cams, ramp)
