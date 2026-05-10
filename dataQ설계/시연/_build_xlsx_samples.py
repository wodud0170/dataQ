"""
종합 시연용 엑셀 샘플 파일 생성기.

생성 위치: dataQ설계/시연/data/
파일:
  시연_단어_샘플.xlsx          — 단어 일괄 등록 (5건)
  시연_용어_샘플.xlsx          — 용어 일괄 등록 (5건)
  시연_도메인_샘플.xlsx        — 도메인 일괄 등록 (3건)
  시연_컬럼_샘플.xlsx          — 컬럼 일괄 등록 / 컬럼 멀티 paste 용 (10건)
  시연_컬럼한글명_paste.txt    — 그리드 multi-paste 시연용 (10줄, 한글명만)

각 양식의 컬럼명·순서는 DataQ 일괄 등록 정적 양식과 동일.
"""
import os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

OUT_DIR = os.path.join(os.path.dirname(os.path.abspath(__file__)), "data")
os.makedirs(OUT_DIR, exist_ok=True)

HDR_FILL = PatternFill("solid", fgColor="3949AB")
HDR_FONT = Font(name="맑은 고딕", color="FFFFFF", bold=True, size=11)
BODY_FONT = Font(name="맑은 고딕", size=10)
THIN_BORDER = Border(
    left=Side(style="thin", color="CCCCCC"),
    right=Side(style="thin", color="CCCCCC"),
    top=Side(style="thin", color="CCCCCC"),
    bottom=Side(style="thin", color="CCCCCC"),
)


def make_sheet(wb, title, headers, rows, widths=None):
    ws = wb.active if len(wb.sheetnames) == 1 and wb["Sheet"].max_row == 1 and wb["Sheet"].max_column == 1 else wb.create_sheet()
    ws.title = title
    for col_idx, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=col_idx, value=h)
        c.fill = HDR_FILL
        c.font = HDR_FONT
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = THIN_BORDER
        if widths and col_idx <= len(widths):
            ws.column_dimensions[c.column_letter].width = widths[col_idx - 1]
    for row_idx, row in enumerate(rows, 2):
        for col_idx, val in enumerate(row, 1):
            c = ws.cell(row=row_idx, column=col_idx, value=val)
            c.font = BODY_FONT
            c.alignment = Alignment(vertical="center")
            c.border = THIN_BORDER
    ws.row_dimensions[1].height = 24


# ========== 1. 단어 샘플 ==========
def build_word():
    wb = Workbook()
    headers = ["단어명(한글)", "영문약어명", "영문명", "단어설명", "도메인분류명", "형식단어여부"]
    rows = [
        ["시연단어1",  "DEMO1",   "Demo One",      "시연용 단어 1",     "명",     "N"],
        ["시연단어2",  "DEMO2",   "Demo Two",      "시연용 단어 2",     "명",     "N"],
        ["시연식별자", "DEMOID",  "Demo Identifier", "시연용 식별자",   "식별자", "N"],
        ["시연일자",   "DEMODT",  "Demo Date",     "시연용 일자",       "일자",   "Y"],
        ["시연여부",   "DEMOYN",  "Demo Flag",     "시연용 여부 플래그", "여부",   "Y"],
    ]
    make_sheet(wb, "단어", headers, rows, widths=[14, 14, 22, 28, 14, 12])
    out = os.path.join(OUT_DIR, "시연_단어_샘플.xlsx")
    wb.save(out)
    print(f"saved: {out}")


# ========== 2. 용어 샘플 ==========
def build_terms():
    wb = Workbook()
    headers = ["용어명(한글)", "영문약어명", "용어설명", "도메인명"]
    rows = [
        ["시연회원ID",      "DEMO_MBR_ID",    "시연 회원 식별자",     "회원ID"],
        ["시연회원명",      "DEMO_MBR_NM",    "시연 회원 이름",       "명"],
        ["시연회원전화번호", "DEMO_MBR_TEL_NO", "시연 회원 전화번호",  "전화번호"],
        ["시연가입일자",    "DEMO_JOIN_DT",   "시연 가입일자",        "일자"],
        ["시연회원여부",    "DEMO_MBR_YN",    "시연 회원 여부",       "여부"],
    ]
    make_sheet(wb, "용어", headers, rows, widths=[18, 22, 28, 14])
    out = os.path.join(OUT_DIR, "시연_용어_샘플.xlsx")
    wb.save(out)
    print(f"saved: {out}")


# ========== 3. 도메인 샘플 ==========
def build_domain():
    wb = Workbook()
    headers = ["도메인명", "도메인분류명", "데이터타입", "길이", "소수점", "도메인설명"]
    rows = [
        ["시연식별자20", "식별자", "VARCHAR",   20, 0, "시연용 식별자 20자"],
        ["시연금액",     "금액",   "NUMERIC",   15, 2, "시연용 금액 (15,2)"],
        ["시연플래그",   "여부",   "CHAR",       1, 0, "시연용 Y/N 플래그"],
    ]
    make_sheet(wb, "도메인", headers, rows, widths=[16, 14, 14, 8, 8, 28])
    out = os.path.join(OUT_DIR, "시연_도메인_샘플.xlsx")
    wb.save(out)
    print(f"saved: {out}")


# ========== 4. 컬럼 샘플 (모델 ① 논리 모델용) ==========
def build_columns():
    wb = Workbook()
    headers = [
        "소유자", "테이블영문명", "테이블한글명",
        "컬럼한글명", "컬럼영문명", "데이터타입", "길이", "소수점",
        "NULL허용", "PK", "FK", "디폴트", "순서"
    ]
    rows = [
        ["USER1", "TB_DEMO_MEMBER", "시연회원", "회원ID",        "MBR_ID",       "VARCHAR", 20,  0, "N", "Y", "N", "",   1],
        ["USER1", "TB_DEMO_MEMBER", "시연회원", "회원명",        "MBR_NM",       "VARCHAR", 100, 0, "N", "N", "N", "",   2],
        ["USER1", "TB_DEMO_MEMBER", "시연회원", "회원전화번호",  "MBR_TEL_NO",   "VARCHAR", 13,  0, "Y", "N", "N", "",   3],
        ["USER1", "TB_DEMO_MEMBER", "시연회원", "가입일자",      "JOIN_DT",      "DATE",    0,   0, "Y", "N", "N", "",   4],
        ["USER1", "TB_DEMO_MEMBER", "시연회원", "회원여부",      "MBR_YN",       "CHAR",    1,   0, "N", "N", "N", "Y",  5],
        ["USER1", "TB_DEMO_ORDER",  "시연주문", "주문ID",        "ORD_ID",       "VARCHAR", 20,  0, "N", "Y", "N", "",   1],
        ["USER1", "TB_DEMO_ORDER",  "시연주문", "회원ID",        "MBR_ID",       "VARCHAR", 20,  0, "N", "N", "Y", "",   2],
        ["USER1", "TB_DEMO_ORDER",  "시연주문", "주문일자",      "ORD_DT",       "DATE",    0,   0, "N", "N", "N", "",   3],
        ["USER1", "TB_DEMO_ORDER",  "시연주문", "주문금액",      "ORD_AMT",      "NUMERIC", 15,  2, "N", "N", "N", "0",  4],
        ["USER1", "TB_DEMO_ORDER",  "시연주문", "주문상태코드",  "ORD_STAT_CD",  "VARCHAR", 4,   0, "N", "N", "N", "",   5],
    ]
    widths = [9, 18, 14, 16, 16, 12, 7, 7, 9, 5, 5, 9, 6]
    make_sheet(wb, "컬럼", headers, rows, widths=widths)
    out = os.path.join(OUT_DIR, "시연_컬럼_샘플.xlsx")
    wb.save(out)
    print(f"saved: {out}")


# ========== 5. 멀티 paste 시연용 텍스트 (한글명만) ==========
def build_paste_txt():
    lines = [
        "회원ID",
        "회원명",
        "회원전화번호",
        "회원이메일",
        "가입일자",
        "회원여부",
        "회원등급코드",
        "최종접속일자",
        "비밀번호",
        "회원생년월일",
    ]
    out = os.path.join(OUT_DIR, "시연_컬럼한글명_paste.txt")
    with open(out, "w", encoding="utf-8") as f:
        f.write("\n".join(lines) + "\n")
    print(f"saved: {out}")


if __name__ == "__main__":
    build_word()
    build_terms()
    build_domain()
    build_columns()
    build_paste_txt()
    print(f"\nall samples → {OUT_DIR}")
