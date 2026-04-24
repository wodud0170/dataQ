"""
53번 설계 Phase 5/6 — 컬럼 엑셀 업로드 E2E 테스트

흐름:
  1. 로그인
  2. API로 모델 + 테이블 1건 선행 생성 (업로드는 이미 존재하는 테이블로 매칭)
  3. 컬럼 화면 이동 → 모델 선택
  4. 테스트용 xlsx 생성 (openpyxl, 3행 — 그중 1행 PK, 1행 FK)
  5. [엑셀 업로드] → preview → [등록 실행]
  6. API 검증 — 3개 컬럼 등록, attrNm=TMP_COL_N, PK/FK 반영

주의:
  - FK 참조 테이블/컬럼은 같은 파일 내의 컬럼(PK)을 참조해도 되지만,
    현재 parseAttrWorkbook 은 참조 테이블 존재 여부만 확인.
    — 이 테스트는 같은 테이블의 PK 컬럼을 refObjNmKr 로 지정.
"""
import os
import sys
import time
import traceback
from datetime import datetime

import requests
from openpyxl import Workbook
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import TimeoutException

BASE_URL = "http://localhost:28091"
SCREENSHOT_DIR = os.path.join(os.path.dirname(__file__), "screenshots")
TMP_DIR = os.path.join(os.path.dirname(__file__), "tmp_xlsx")
TS = datetime.now().strftime("%m%d%H%M%S")
MODEL_NAME = "E2E_53_UP_COL_" + TS
TABLE_KR = "주문상세_" + TS
OWNER = "QA"

# 업로드할 컬럼 3건: (한글명, PK, FK, refObjNmKr, refAttrNmKr, 삭제규칙)
COLS = [
    {"kr": "주문번호",   "pk": "Y", "fk": "N", "refTbl": "", "refCol": "", "del": ""},
    {"kr": "상품명",     "pk": "N", "fk": "N", "refTbl": "", "refCol": "", "del": ""},
    {"kr": "수량",       "pk": "N", "fk": "N", "refTbl": "", "refCol": "", "del": ""},
]

os.makedirs(SCREENSHOT_DIR, exist_ok=True)
os.makedirs(TMP_DIR, exist_ok=True)
results = []


def make_driver():
    opts = webdriver.EdgeOptions()
    opts.add_argument("--log-level=3")
    opts.add_experimental_option("excludeSwitches", ["enable-logging"])
    prefs = {"download.default_directory": TMP_DIR}
    opts.add_experimental_option("prefs", prefs)
    d = webdriver.Edge(options=opts)
    d.set_window_size(1600, 1000)
    return d


def shot(d, name):
    p = os.path.join(SCREENSHOT_DIR, "e2e53upc_" + name + ".png")
    d.save_screenshot(p); print(f"  [SHOT] {name}")


def wait_visible(d, by, sel, t=10):
    return WebDriverWait(d, t).until(EC.visibility_of_element_located((by, sel)))


def wait_clickable(d, by, sel, t=10):
    return WebDriverWait(d, t).until(EC.element_to_be_clickable((by, sel)))


def dismiss_swal(d):
    for _ in range(5):
        try:
            b = d.find_element(By.CSS_SELECTOR, ".swal2-confirm")
            if b.is_displayed():
                b.click(); time.sleep(0.5); continue
        except Exception: pass
        break


def login(d, user="space", pw="123"):
    d.get(BASE_URL + "/signin")
    wait_visible(d, By.CSS_SELECTOR, "input[type='text']", 15); time.sleep(1)
    d.find_element(By.CSS_SELECTOR, "input[type='text']").send_keys(user)
    pw_in = d.find_element(By.CSS_SELECTOR, "input[type='password']")
    pw_in.send_keys(pw); pw_in.send_keys(Keys.ENTER)
    WebDriverWait(d, 15).until(lambda drv: "/main" in drv.current_url)
    time.sleep(2)


def _click_el(d, el):
    try: el.click()
    except Exception:
        d.execute_script("arguments[0].scrollIntoView({block:'center'});", el); time.sleep(0.2)
        try: el.click()
        except Exception: d.execute_script("arguments[0].click();", el)


def nav(d, group_id, menu_id):
    dismiss_swal(d)
    items = d.find_elements(By.ID, menu_id)
    need = not items or not items[0].is_displayed()
    if need:
        g = wait_clickable(d, By.ID, group_id, 10)
        _click_el(d, g)
        try: wait_visible(d, By.ID, menu_id, 5)
        except TimeoutException:
            _click_el(d, g); wait_visible(d, By.ID, menu_id, 5)
    m = wait_visible(d, By.ID, menu_id, 10)
    _click_el(d, m); time.sleep(2)


def select_autocomplete(d, scope_el, value):
    ac = scope_el.find_element(By.CSS_SELECTOR, ".v-autocomplete input[type='text']")
    _click_el(d, ac); time.sleep(0.3)
    ac.send_keys(Keys.CONTROL, "a"); ac.send_keys(Keys.DELETE)
    ac.send_keys(value); time.sleep(1.2)
    items = d.find_elements(By.CSS_SELECTOR, ".menuable__content__active .v-list-item") or \
            d.find_elements(By.CSS_SELECTOR, "[role='option']")
    for it in items:
        if value in (it.text or ""):
            _click_el(d, it); time.sleep(0.5); return
    raise RuntimeError(f"'{value}' 매칭 옵션 없음")


def step(name, fn):
    print(f"\n{'=' * 60}\nSTEP: {name}\n{'=' * 60}")
    try:
        fn(); results.append((name, "PASS", None)); print("  >> PASS"); return True
    except Exception as e:
        tb = traceback.format_exc()
        results.append((name, "FAIL", tb)); print(f"  >> FAIL: {e}"); print(tb); return False


# ---------- steps ----------
state = {}


def step1_login(d):
    login(d); shot(d, "01_login")


def step2_prep(d):
    """모델 + 테이블(OWNER, TABLE_KR) 1건 생성 — attrs 업로드는 테이블 매칭 필요"""
    cookies = {c["name"]: c["value"] for c in d.get_cookies()}
    r = requests.post(BASE_URL + "/api/dm/createDataModel", cookies=cookies,
                      json={"dataModelNm": MODEL_NAME, "modelType": "LOGICAL", "ver": "1.0"}, timeout=10)
    r.raise_for_status()
    r2 = requests.post(BASE_URL + "/api/dm/getDataModelStatsList", cookies=cookies, json={}, timeout=10)
    target = next((m for m in r2.json() if m.get("dataModelNm") == MODEL_NAME), None)
    if not target: raise RuntimeError("모델 조회 실패")
    dm_id = target["dataModelId"]
    state["dm"] = dm_id

    # addObj — objOwner 도 같이 지정하려면 updateObj 로 후처리
    r3 = requests.post(BASE_URL + "/api/dm/addObj", cookies=cookies,
                       json={"dataModelId": dm_id, "objNmKr": TABLE_KR, "objOwner": OWNER}, timeout=10)
    r3.raise_for_status()
    # 테이블 조회 — owner 업데이트 필요하면 updateObj
    r4 = requests.get(BASE_URL + "/api/dm/getDataModelObjListByClctId",
                      params={"clctId": dm_id}, cookies=cookies, timeout=10)
    objs = [o for o in r4.json() if o.get("objNmKr") == TABLE_KR]
    if not objs: raise RuntimeError("테이블 생성 실패")
    obj_nm = objs[0]["objNm"]
    if not objs[0].get("objOwner"):
        # updateObj 로 소유자 채워넣기
        requests.post(BASE_URL + "/api/dm/updateObj", cookies=cookies,
                      json={"dataModelId": dm_id, "objNm": obj_nm, "objNmKr": TABLE_KR, "objOwner": OWNER}, timeout=10)
    state["obj"] = obj_nm
    print(f"  [prep] dm={dm_id}, objNm={obj_nm}, OWNER={OWNER}, TABLE_KR={TABLE_KR}")


def step3_build_xlsx(d):
    wb = Workbook()
    sh = wb.active; sh.title = "Sheet1"
    headers = ["소유자", "테이블명(한글)", "컬럼명(한글)", "컬럼 순서",
               "PK여부", "FK여부", "참조 테이블(한글)", "참조 컬럼(한글)", "삭제 규칙"]
    for i, h in enumerate(headers):
        sh.cell(row=1, column=i + 1, value=h)
    for r, c in enumerate(COLS, start=2):
        sh.cell(row=r, column=1, value=OWNER)
        sh.cell(row=r, column=2, value=TABLE_KR)
        sh.cell(row=r, column=3, value=c["kr"])
        sh.cell(row=r, column=4, value=r - 1)
        sh.cell(row=r, column=5, value=c["pk"])
        sh.cell(row=r, column=6, value=c["fk"])
        sh.cell(row=r, column=7, value=c["refTbl"])
        sh.cell(row=r, column=8, value=c["refCol"])
        sh.cell(row=r, column=9, value=c["del"])
    path = os.path.join(TMP_DIR, "upload_attrs_" + TS + ".xlsx")
    wb.save(path)
    state["xlsx"] = path
    print(f"  [xlsx] created {path}")


def step4_upload_and_commit(d):
    nav(d, "dmGroup", "nav_datamodelStatusColumn")
    filter_bar = d.find_element(By.CSS_SELECTOR, ".filterWrapper")
    select_autocomplete(d, filter_bar, MODEL_NAME)
    time.sleep(1)
    shot(d, "04a_model_selected")

    # keep-alive 로 Table/Column 두 컴포넌트 file input 이 공존할 수 있음.
    # 컬럼 업로드 버튼(id=btn-upload-attrs) 이후의 첫 file input 으로 고정.
    file_input = d.find_element(
        By.XPATH,
        "//*[@id='btn-upload-attrs']/following::input[@type='file'][1]"
    )
    d.execute_script("arguments[0].style.display='block';", file_input)
    file_input.send_keys(state["xlsx"])
    # preview 다이얼로그 오픈 대기 (서버 응답 기다림 포함)
    try:
        WebDriverWait(d, 15).until(
            EC.visibility_of_element_located((By.CSS_SELECTOR, ".v-dialog--active"))
        )
    except TimeoutException:
        shot(d, "04b_preview_dialog")
        raise RuntimeError("preview 다이얼로그 오픈 실패")
    shot(d, "04b_preview_dialog")

    commit = WebDriverWait(d, 8).until(
        EC.presence_of_element_located((By.ID, "btn-upload-attrs-commit"))
    )
    d.execute_script("arguments[0].scrollIntoView({block:'center'});", commit)
    time.sleep(0.5)
    d.execute_script("arguments[0].click();", commit)
    time.sleep(3)
    dismiss_swal(d)
    shot(d, "04c_committed")


def step5_verify_api(d):
    cookies = {c["name"]: c["value"] for c in d.get_cookies()}
    dm = state["dm"]
    r = requests.get(BASE_URL + "/api/dm/getDataModelAttrListByClctId",
                     params={"clctId": dm}, cookies=cookies, timeout=10)
    r.raise_for_status()
    attrs = r.json() or []
    mine = [a for a in attrs if a.get("objNm") == state["obj"]]
    print(f"  [verify] 등록된 컬럼 {len(mine)}건: "
          f"{[(a.get('attrNmKr'), a.get('attrNm'), a.get('pkYn'), a.get('fkYn')) for a in mine]}")
    if len(mine) != len(COLS):
        raise RuntimeError(f"컬럼 개수 불일치: 기대 {len(COLS)}, 실제 {len(mine)}")
    for c in COLS:
        m = [a for a in mine if a.get("attrNmKr") == c["kr"]]
        if not m: raise RuntimeError(f"'{c['kr']}' 컬럼 없음")
        a = m[0]
        if not (a.get("attrNm") or "").startswith("TMP_COL_"):
            raise RuntimeError(f"attrNm 형식 오류: {a.get('attrNm')}")
        if c["pk"] == "Y" and a.get("pkYn") != "Y":
            raise RuntimeError(f"'{c['kr']}' pkYn 불일치")
        if c["pk"] == "Y" and a.get("nullableYn") == "Y":
            raise RuntimeError(f"'{c['kr']}' PK 이지만 nullable=Y")
    print("  [verify] OK — 컬럼 3건 등록 + PK/nullable 일치")


def main():
    d = make_driver()
    try:
        if not step("1. 로그인", lambda: step1_login(d)): return
        if not step("2. 모델 + 테이블 선행 생성", lambda: step2_prep(d)): return
        if not step("3. 테스트 xlsx 생성", lambda: step3_build_xlsx(d)): return
        if not step("4. UI 업로드 → preview → commit", lambda: step4_upload_and_commit(d)): return
        step("5. API 로 등록 결과 검증", lambda: step5_verify_api(d))
    finally:
        try: shot(d, "99_final")
        except Exception: pass
        d.quit()

    print(f"\n{'=' * 60}\n결과\n{'=' * 60}")
    pas = sum(1 for _, s, _ in results if s == "PASS")
    fal = len(results) - pas
    for name, status, _ in results:
        print(f"  {'[PASS]' if status == 'PASS' else '[FAIL]'} {name}")
    print(f"\n  총 {len(results)}: PASS {pas}, FAIL {fal}")
    sys.exit(0 if fal == 0 else 1)


if __name__ == "__main__":
    main()
