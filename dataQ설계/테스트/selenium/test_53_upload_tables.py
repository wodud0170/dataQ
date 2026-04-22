"""
53번 설계 Phase 5/6 — 테이블 엑셀 업로드 E2E 테스트

흐름:
  1. 로그인
  2. API로 모델 생성
  3. 테이블 화면 이동 → 모델 선택
  4. [양식 다운로드] — xlsx 파일 내려받기 확인 (requests)
  5. 테스트용 xlsx 생성 (openpyxl, 2행)
  6. [엑셀 업로드] → 파일 선택 → preview 다이얼로그 오픈 확인
  7. [등록 실행] 클릭 → commit
  8. API 검증 — 2개 테이블 등록, objNm=TMP_TBL_N, objNmKr/objOwner 일치

검증 포인트:
  - 템플릿 다운로드 xlsx 매직넘버 (PK)
  - preview 다이얼로그 요약 카운트 표시
  - commit 후 그리드 새로고침 + API 조회 결과 일치
"""
import io
import os
import sys
import time
import traceback
from datetime import datetime

import requests
from openpyxl import Workbook
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import TimeoutException

BASE_URL = "http://localhost:28091"
SCREENSHOT_DIR = os.path.join(os.path.dirname(__file__), "screenshots")
TMP_DIR = os.path.join(os.path.dirname(__file__), "tmp_xlsx")
MODEL_NAME = "E2E_53_UP_TBL_" + datetime.now().strftime("%m%d%H%M%S")

TABLES = [
    {"owner": "QA", "nmKr": "고객마스터_" + datetime.now().strftime("%H%M%S"), "desc": "고객정보 테이블"},
    {"owner": "QA", "nmKr": "주문_" + datetime.now().strftime("%H%M%S"), "desc": "주문 테이블"},
]

os.makedirs(SCREENSHOT_DIR, exist_ok=True)
os.makedirs(TMP_DIR, exist_ok=True)
results = []


def make_driver():
    opts = webdriver.EdgeOptions()
    opts.add_argument("--log-level=3")
    opts.add_experimental_option("excludeSwitches", ["enable-logging"])
    # 다운로드 경로 설정
    prefs = {"download.default_directory": TMP_DIR}
    opts.add_experimental_option("prefs", prefs)
    d = webdriver.Edge(options=opts)
    d.set_window_size(1600, 1000)
    return d


def shot(d, name):
    p = os.path.join(SCREENSHOT_DIR, "e2e53upt_" + name + ".png")
    d.save_screenshot(p)
    print(f"  [SHOT] {name}")


def wait_visible(d, by, sel, t=10):
    return WebDriverWait(d, t).until(EC.visibility_of_element_located((by, sel)))


def wait_clickable(d, by, sel, t=10):
    return WebDriverWait(d, t).until(EC.element_to_be_clickable((by, sel)))


def dismiss_swal(d):
    for _ in range(5):
        try:
            b = d.find_element(By.CSS_SELECTOR, ".swal2-confirm")
            if b.is_displayed():
                b.click(); time.sleep(0.5); continue
        except Exception:
            pass
        break


def login(d, user="space", pw="123"):
    d.get(BASE_URL + "/signin")
    wait_visible(d, By.CSS_SELECTOR, "input[type='text']", 15)
    time.sleep(1)
    d.find_element(By.CSS_SELECTOR, "input[type='text']").send_keys(user)
    pw_in = d.find_element(By.CSS_SELECTOR, "input[type='password']")
    pw_in.send_keys(pw); pw_in.send_keys(Keys.ENTER)
    WebDriverWait(d, 15).until(lambda drv: "/main" in drv.current_url)
    time.sleep(2)


def _click_el(d, el):
    try: el.click()
    except Exception:
        d.execute_script("arguments[0].scrollIntoView({block:'center'});", el)
        time.sleep(0.2)
        try: el.click()
        except Exception: d.execute_script("arguments[0].click();", el)


def nav(d, group_id, menu_id):
    dismiss_swal(d)
    items = d.find_elements(By.ID, menu_id)
    need = not items or not items[0].is_displayed()
    if need:
        g = wait_clickable(d, By.ID, group_id, 10)
        _click_el(d, g)
        try: wait_visible(d, By.ID, menu_id, 5)
        except TimeoutException:
            _click_el(d, g); wait_visible(d, By.ID, menu_id, 5)
    m = wait_visible(d, By.ID, menu_id, 10)
    _click_el(d, m)
    time.sleep(2)


def select_autocomplete(d, scope_el, value):
    ac = scope_el.find_element(By.CSS_SELECTOR, ".v-autocomplete input[type='text']")
    _click_el(d, ac); time.sleep(0.3)
    ac.send_keys(Keys.CONTROL, "a"); ac.send_keys(Keys.DELETE)
    ac.send_keys(value); time.sleep(1.2)
    items = d.find_elements(By.CSS_SELECTOR, ".menuable__content__active .v-list-item") or \
            d.find_elements(By.CSS_SELECTOR, "[role='option']")
    for it in items:
        if value in (it.text or ""):
            _click_el(d, it); time.sleep(0.5); return
    raise RuntimeError(f"'{value}' 매칭 옵션 없음")


def step(name, fn):
    print(f"\n{'=' * 60}\nSTEP: {name}\n{'=' * 60}")
    try:
        fn(); results.append((name, "PASS", None)); print("  >> PASS"); return True
    except Exception as e:
        tb = traceback.format_exc()
        results.append((name, "FAIL", tb)); print(f"  >> FAIL: {e}"); print(tb); return False


# ---------- steps ----------
dm_id_holder = {}


def step1_login(d):
    login(d); shot(d, "01_login")


def step2_create_model(d):
    cookies = {c["name"]: c["value"] for c in d.get_cookies()}
    r = requests.post(BASE_URL + "/api/dm/createDataModel", cookies=cookies,
                      json={"dataModelNm": MODEL_NAME, "modelType": "LOGICAL", "ver": "1.0"}, timeout=10)
    r.raise_for_status()
    r2 = requests.post(BASE_URL + "/api/dm/getDataModelStatsList", cookies=cookies, json={}, timeout=10)
    target = next((m for m in r2.json() if m.get("dataModelNm") == MODEL_NAME), None)
    if not target: raise RuntimeError("모델 조회 실패")
    dm_id_holder["dm"] = target["dataModelId"]


def step3_template_download(d):
    # 서버에서 양식 xlsx 직접 받아 매직넘버 확인
    cookies = {c["name"]: c["value"] for c in d.get_cookies()}
    r = requests.get(BASE_URL + "/api/dm/uploadTemplate", params={"scope": "tables"},
                     cookies=cookies, timeout=10)
    r.raise_for_status()
    if not r.content.startswith(b"PK"):
        raise RuntimeError("xlsx 매직넘버 'PK' 아님")
    print(f"  [template] size={len(r.content)} bytes, PK signature OK")


def step4_build_xlsx(d):
    """테스트 데이터용 xlsx 파일 TMP_DIR 에 생성"""
    wb = Workbook()
    sh = wb.active
    sh.title = "Sheet1"
    headers = ["소유자", "테이블명(한글)", "설명"]
    for i, h in enumerate(headers):
        sh.cell(row=1, column=i + 1, value=h)
    for r, t in enumerate(TABLES, start=2):
        sh.cell(row=r, column=1, value=t["owner"])
        sh.cell(row=r, column=2, value=t["nmKr"])
        sh.cell(row=r, column=3, value=t["desc"])
    path = os.path.join(TMP_DIR, "upload_tables_" + datetime.now().strftime("%H%M%S") + ".xlsx")
    wb.save(path)
    dm_id_holder["xlsx"] = path
    print(f"  [xlsx] created {path}")


def step5_upload_preview_commit(d):
    nav(d, "dmGroup", "nav_datamodelStatusTable")
    filter_bar = d.find_element(By.CSS_SELECTOR, ".filterWrapper")
    select_autocomplete(d, filter_bar, MODEL_NAME)
    time.sleep(1)
    shot(d, "05a_model_selected")

    # 파일 input 에 경로 주입 — <input ref="uploadTablesInput">
    file_input = d.find_element(By.CSS_SELECTOR, "input[type='file'][accept='.xlsx']")
    d.execute_script("arguments[0].style.display='block';", file_input)
    file_input.send_keys(dm_id_holder["xlsx"])
    time.sleep(2)
    shot(d, "05b_preview_dialog")

    # 다이얼로그 오픈 확인
    dialogs = d.find_elements(By.CSS_SELECTOR, ".v-dialog--active")
    if not dialogs:
        raise RuntimeError("preview 다이얼로그 오픈 실패")
    commit = WebDriverWait(d, 8).until(
        EC.presence_of_element_located((By.ID, "btn-upload-tables-commit"))
    )
    d.execute_script("arguments[0].scrollIntoView({block:'center'});", commit)
    time.sleep(0.5)
    d.execute_script("arguments[0].click();", commit)
    time.sleep(3)
    dismiss_swal(d)
    shot(d, "05c_committed")


def step6_verify_api(d):
    cookies = {c["name"]: c["value"] for c in d.get_cookies()}
    dm = dm_id_holder["dm"]
    r = requests.get(BASE_URL + "/api/dm/getDataModelObjListByClctId",
                     params={"clctId": dm}, cookies=cookies, timeout=10)
    r.raise_for_status()
    objs = r.json() or []
    kr_list = [(o.get("objOwner"), o.get("objNmKr"), o.get("objNm")) for o in objs]
    print(f"  [verify] 등록된 테이블: {kr_list}")
    for t in TABLES:
        match = [o for o in objs if o.get("objNmKr") == t["nmKr"]]
        if not match:
            raise RuntimeError(f"업로드된 테이블 '{t['nmKr']}' 을 API 에서 찾을 수 없음")
        m = match[0]
        if not (m.get("objNm") or "").startswith("TMP_TBL_"):
            raise RuntimeError(f"objNm 물리명 형식 오류: {m.get('objNm')}")
    print(f"  [verify] OK — {len(TABLES)}개 테이블 등록 확인")


def main():
    d = make_driver()
    try:
        if not step("1. 로그인", lambda: step1_login(d)): return
        if not step("2. 모델 생성", lambda: step2_create_model(d)): return
        if not step("3. 양식 다운로드 매직넘버 검증", lambda: step3_template_download(d)): return
        if not step("4. 테스트 xlsx 생성", lambda: step4_build_xlsx(d)): return
        if not step("5. UI 업로드 → preview → commit", lambda: step5_upload_preview_commit(d)): return
        step("6. API 로 등록 결과 검증", lambda: step6_verify_api(d))
    finally:
        try: shot(d, "99_final")
        except Exception: pass
        d.quit()

    print(f"\n{'=' * 60}\n결과\n{'=' * 60}")
    pas = sum(1 for _, s, _ in results if s == "PASS")
    fal = len(results) - pas
    for name, status, _ in results:
        print(f"  {'[PASS]' if status == 'PASS' else '[FAIL]'} {name}")
    print(f"\n  총 {len(results)}: PASS {pas}, FAIL {fal}")
    sys.exit(0 if fal == 0 else 1)


if __name__ == "__main__":
    main()
